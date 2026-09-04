import os
import re
import time
import asyncio
import math
import csv
import io
import json
import base64
import secrets
import hashlib
import hmac
import random
import string
import urllib.parse
from pywebpush import webpush, WebPushException
from pathlib import Path
from datetime import datetime, timedelta, date, time as dtime, timezone, tzinfo
from zoneinfo import ZoneInfo
from typing import Optional, List, Dict, Any, Tuple, Literal

from fastapi import FastAPI, Depends, HTTPException, status, Query, Request, Response, Cookie
from fastapi.responses import FileResponse, JSONResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.staticfiles import StaticFiles
from starlette.middleware.trustedhost import TrustedHostMiddleware
from pydantic import BaseModel, EmailStr, ConfigDict, Field, field_validator
from sqlalchemy import (
    create_engine, Column, Integer, String, Float, ForeignKey, Text, Boolean,
    DateTime as SQLDateTime, and_, or_, func, UniqueConstraint, inspect,
    cast, literal, text as sql_text, event
)
from sqlalchemy.orm import declarative_base, sessionmaker, relationship, Session
from sqlalchemy.exc import IntegrityError
from passlib.context import CryptContext
import jwt
from jwt.exceptions import InvalidTokenError as JWTError
from openpyxl import Workbook
from supabase_client import get_supabase_client, SupabaseConfigurationError
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    from google import genai
except Exception:
    genai = None

# -----------------------------------------------------------------------------
# APP / CONFIG
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
# Cauldra deployment structure: the frontend is a separate top-level folder,
# a sibling of backend/ (Railway container: /app/backend/main.py + /app/frontend/).
# index.html, sw.js, assets/, css/ and js/ live under frontend/, never backend/.
PROJECT_DIR = BASE_DIR.parent
FRONTEND_DIR = PROJECT_DIR / "frontend"
INDEX_PATH = FRONTEND_DIR / "index.html"
ASSETS_DIR = FRONTEND_DIR / "assets"
CSS_DIR = FRONTEND_DIR / "css"
JS_DIR = FRONTEND_DIR / "js"
ENVIRONMENT = os.getenv("SUPPLY_AI_ENV", "development").strip().lower()
IS_PRODUCTION = ENVIRONMENT == "production"

def resolve_app_path(value: str, default: Path) -> Path:
    """Resolve configurable data paths without allowing an implicit web-root path."""
    candidate = Path(value).expanduser() if value else default
    return candidate.resolve() if candidate.is_absolute() else (BASE_DIR / candidate).resolve()

# -----------------------------------------------------------------------------
# DATABASE_URL — mandatory, PostgreSQL-only, source of truth for both the app
# and Alembic (see alembic/env.py, which reads this exact same variable).
# Cauldra used to fall back to a local SQLite file when this was unset; that
# fallback is gone entirely — a database engine chosen silently by "was an
# env var present or not" is exactly the kind of environment drift this
# removes. Every environment (development, staging, production, tests) must
# supply a real PostgreSQL DATABASE_URL of its own; see DATABASE_MIGRATION.md
# and .env.example for the recommended per-environment database names
# (cauldra_dev / cauldra_test / cauldra_prod).
# -----------------------------------------------------------------------------
def _normalize_database_url(raw: str) -> str:
    """Normalizes provider-supplied URL forms into what this app's pinned
    driver expects, without rejecting any form the caller deliberately chose:
    - `postgres://` (Heroku-style, and some managed-Postgres dashboards) is
      not a scheme SQLAlchemy recognizes on its own — rewritten to
      `postgresql://` first.
    - A bare `postgresql://` (no `+driver`) leaves the actual DBAPI driver to
      SQLAlchemy's own default resolution, which silently depends on
      whichever of psycopg2/psycopg happens to be installed — rewritten to
      the explicit `postgresql+psycopg://` this app is built and tested
      against (see requirements.txt's pinned `psycopg[binary]`), so the
      driver in use is never ambiguous. A URL that already names a driver
      (`+psycopg`, `+psycopg2`, ...) is left exactly as given."""
    if raw.startswith("postgres://"):
        raw = "postgresql://" + raw[len("postgres://"):]
    if raw.startswith("postgresql://"):
        raw = "postgresql+psycopg://" + raw[len("postgresql://"):]
    return raw

_RAW_DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if not _RAW_DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL is required. Cauldra uses PostgreSQL for all environments. "
        "Set DATABASE_URL to a PostgreSQL connection string, e.g. "
        "postgresql+psycopg://USER:PASSWORD@HOST:5432/cauldra_dev"
    )
DATABASE_URL = _normalize_database_url(_RAW_DATABASE_URL)
if not DATABASE_URL.startswith(("postgresql://", "postgresql+psycopg://", "postgresql+psycopg2://")):
    raise RuntimeError(
        "DATABASE_URL must be a PostgreSQL connection string (postgresql://... or "
        "postgresql+psycopg://...). Cauldra does not support SQLite or any other database engine."
    )
# Set only by a migration/import-only context that manages its own database
# connection lifecycle (see alembic/env.py) — never by the running
# application — so importing main.py's Base metadata for a migration can
# never also trigger main.py's own startup connectivity check as a side
# effect (that check belongs to the app process actually serving traffic).
SKIP_DB_STARTUP_CHECK = os.getenv("SUPPLY_AI_SKIP_DB_STARTUP_CHECK", "false").strip().lower() == "true"
UPLOAD_STORAGE_DIR = resolve_app_path(os.getenv("SUPPLY_AI_UPLOAD_DIR", "uploads"), BASE_DIR / "uploads")
MAX_UPLOAD_BYTES = int(os.getenv("SUPPLY_AI_MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))

SECRET_KEY = os.getenv("SUPPLY_AI_SECRET_KEY", "").strip()
if not SECRET_KEY or len(SECRET_KEY) < 32:
    raise RuntimeError(
        "SUPPLY_AI_SECRET_KEY is required and must be at least 32 characters long. "
        "Generate one with: python -c \"import secrets; print(secrets.token_hex(32))\""
    )

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("SUPPLY_AI_ACCESS_TOKEN_MINUTES", "15"))
REFRESH_TOKEN_EXPIRE_DAYS = int(os.getenv("SUPPLY_AI_REFRESH_TOKEN_DAYS", "30"))
# How long a just-rotated refresh token is still forgiven for, so a second
# browser tab/window presenting the token it already had (rather than the
# brand-new one another tab just rotated it into) recovers the current
# session instead of being logged out. See auth_refresh()'s grace-recovery
# branch. Kept short and narrow on purpose: it only ever forgives the exact
# immediate predecessor of the current session, never anything older.
REFRESH_ROTATION_GRACE_SECONDS = int(os.getenv("SUPPLY_AI_REFRESH_ROTATION_GRACE_SECONDS", "10"))
REFRESH_COOKIE_NAME = os.getenv("SUPPLY_AI_REFRESH_COOKIE_NAME", "cauldra_refresh")
REFRESH_COOKIE_SECURE = os.getenv("SUPPLY_AI_REFRESH_COOKIE_SECURE", "true" if IS_PRODUCTION else "false").strip().lower() == "true"
REFRESH_COOKIE_SAMESITE = os.getenv("SUPPLY_AI_REFRESH_COOKIE_SAMESITE", "lax").strip().lower()
TRUSTED_HOSTS = [x.strip() for x in os.getenv("SUPPLY_AI_TRUSTED_HOSTS", "").split(",") if x.strip()]
RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("SUPPLY_AI_RATE_LIMIT_WINDOW_SECONDS", "300"))
RATE_LIMIT_MAX_FAILURES = int(os.getenv("SUPPLY_AI_RATE_LIMIT_MAX_FAILURES", "8"))
FORGOT_CODE_TTL_SECONDS = int(os.getenv("SUPPLY_AI_FORGOT_CODE_TTL", "600"))
FORGOT_RESEND_SECONDS = int(os.getenv("SUPPLY_AI_FORGOT_RESEND", "60"))
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.5-flash")
TERMII_BASE_URL = os.getenv("TERMII_BASE_URL", "https://api.ng.termii.com").rstrip("/")
RESEND_FROM = os.getenv("RESEND_FROM", "onboarding@resend.dev")
PAYSTACK_SECRET_KEY = os.getenv("PAYSTACK_SECRET_KEY", "").strip()
PAYSTACK_PUBLIC_KEY = os.getenv("PAYSTACK_PUBLIC_KEY", "").strip()
PAYSTACK_CALLBACK_URL = os.getenv("PAYSTACK_CALLBACK_URL", "").strip()
SUPPLY_AI_FRONTEND_URL = os.getenv("SUPPLY_AI_FRONTEND_URL", "").strip().rstrip("/")
# Web Push (VAPID). VAPID_PUBLIC_KEY is safe to hand to the frontend (served
# via GET /push/vapid-public-key) — it's what the browser's
# pushManager.subscribe({applicationServerKey}) call needs. VAPID_PRIVATE_KEY
# must stay server-only; deliver_push_notification() is the only place it is
# ever read. Both are the raw base64url-encoded EC key pair py_vapid
# generates (see scratch key-generation notes in the notification-system
# implementation) — not PEM files, so there's no extra file to manage or
# accidentally commit. Push delivery is a no-op (in-app notifications still
# work normally) whenever either is unset — see deliver_push_notification().
VAPID_PUBLIC_KEY = os.getenv("VAPID_PUBLIC_KEY", "").strip()
VAPID_PRIVATE_KEY = os.getenv("VAPID_PRIVATE_KEY", "").strip()
VAPID_CONTACT_EMAIL = os.getenv("VAPID_CONTACT_EMAIL", "").strip() or "support@cauldra.app"

def _is_placeholder_url(value: str) -> bool:
    value = str(value or "").strip().lower()
    return not value or "your-domain.example" in value or "your-domain" in value

def paystack_callback_url(request: Request) -> str:
    """Return the real Cauldra frontend URL for Paystack's browser return."""
    if SUPPLY_AI_FRONTEND_URL:
        return SUPPLY_AI_FRONTEND_URL + "/"

    origin = str(request.headers.get("origin") or "").strip().rstrip("/")
    configured_origins = {str(x).strip().rstrip("/") for x in ALLOWED_ORIGINS}
    if origin and origin in configured_origins:
        return origin + "/"

    if PAYSTACK_CALLBACK_URL and not _is_placeholder_url(PAYSTACK_CALLBACK_URL):
        return PAYSTACK_CALLBACK_URL

    return str(request.base_url)

# Paystack does not support a guaranteed true zero-amount charge on card
# transactions across all banks/brands. Their own current guidance is a small
# tokenization charge (minimum recommended ~NGN 50) to authenticate the card
# via 2FA, which is then refunded. We follow that officially-supported pattern
# rather than inventing a fake "free" verification. See /subscription/trial/init.
PAYSTACK_TRIAL_VERIFICATION_AMOUNT_KOBO = int(os.getenv("PAYSTACK_TRIAL_VERIFICATION_AMOUNT_KOBO", "5000"))
PAYSTACK_GRACE_PERIOD_DAYS = int(os.getenv("PAYSTACK_GRACE_PERIOD_DAYS", "3"))
APP_NAME = "Cauldra"

allowed_origins_raw = os.getenv("SUPPLY_AI_CORS_ORIGINS", "")
ALLOWED_ORIGINS = [x.strip() for x in allowed_origins_raw.split(",") if x.strip()]

if REFRESH_COOKIE_SAMESITE not in {"lax", "strict", "none"}:
    raise RuntimeError("SUPPLY_AI_REFRESH_COOKIE_SAMESITE must be lax, strict, or none.")
if REFRESH_COOKIE_SAMESITE == "none" and not REFRESH_COOKIE_SECURE:
    raise RuntimeError("SameSite=None refresh cookies require SUPPLY_AI_REFRESH_COOKIE_SECURE=true.")
if IS_PRODUCTION:
    if not REFRESH_COOKIE_SECURE:
        raise RuntimeError("Production requires SUPPLY_AI_REFRESH_COOKIE_SECURE=true.")
    if not TRUSTED_HOSTS or "*" in TRUSTED_HOSTS:
        raise RuntimeError("Production requires explicit SUPPLY_AI_TRUSTED_HOSTS; do not use *.")
    if "*" in ALLOWED_ORIGINS:
        raise RuntimeError("SUPPLY_AI_CORS_ORIGINS must list exact origins in production.")
elif not TRUSTED_HOSTS:
    TRUSTED_HOSTS = ["127.0.0.1", "localhost", "testserver"]

app = FastAPI(
    title="Cauldra Backend",
    version="4.1",
    docs_url=None if IS_PRODUCTION else "/docs",
    redoc_url=None if IS_PRODUCTION else "/redoc",
    openapi_url=None if IS_PRODUCTION else "/openapi.json",
)
app.add_middleware(TrustedHostMiddleware, allowed_hosts=TRUSTED_HOSTS)
if ALLOWED_ORIGINS:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=ALLOWED_ORIGINS,
        allow_credentials=True,
        allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
        allow_headers=["Authorization", "Content-Type", "Accept"],
    )

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=()")
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; base-uri 'self'; frame-ancestors 'none'; form-action 'self'; "
            "img-src 'self' data: blob: https://flagcdn.com; media-src 'self' blob:; worker-src 'self' blob:; connect-src 'self'; "
            "style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; font-src 'self' data:"
        )
        if IS_PRODUCTION:
            response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        return response

# app.add_middleware(SecurityHeadersMiddleware)

# --- Development-only request timing (performance refactor, section 3) -----
# Logs `[perf] METHOD PATH duration_ms status_code` for every request, plus a
# `[slow-api]` warning above SLOW_REQUEST_THRESHOLD_MS — used to find real
# bottlenecks before optimizing rather than guessing. Never logs headers,
# query params, or body (no tokens/secrets), and is skipped entirely in
# production so it can never add overhead or noise there.
SLOW_REQUEST_THRESHOLD_MS = 500

class PerfLoggingMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        start = time.perf_counter()
        response = await call_next(request)
        duration_ms = (time.perf_counter() - start) * 1000
        print(f"[perf] {request.method} {request.url.path} {duration_ms:.0f}ms {response.status_code}")
        if duration_ms > SLOW_REQUEST_THRESHOLD_MS:
            print(f"[slow-api] {request.method} {request.url.path} took {duration_ms:.0f}ms")
        return response

if not IS_PRODUCTION:
    app.add_middleware(PerfLoggingMiddleware)
if ASSETS_DIR.is_dir():
    app.mount("/assets", StaticFiles(directory=str(ASSETS_DIR)), name="assets")
if CSS_DIR.is_dir():
    app.mount("/css", StaticFiles(directory=str(CSS_DIR)), name="css")
if JS_DIR.is_dir():
    app.mount("/js", StaticFiles(directory=str(JS_DIR)), name="js")
if FRONTEND_DIR.is_dir():
    app.mount("/frontend", StaticFiles(directory=str(FRONTEND_DIR)), name="frontend")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# -----------------------------------------------------------------------------
# SQLAlchemy engine — PostgreSQL only. Pool sizing is environment-configurable
# with conservative defaults (see DEPLOYMENT.md): remember total connections
# against the database are approximately workers x (pool_size + max_overflow),
# so raising these without checking the provider's real connection limit can
# exhaust a managed Postgres instance shared across every worker/replica.
# -----------------------------------------------------------------------------
def _positive_int_env(name: str, default: int) -> int:
    """Parses a positive-integer env var with a safe fallback — a missing,
    empty, non-numeric, or non-positive value never crashes startup or
    silently produces a nonsensical pool setting; it just uses `default`."""
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        value = int(raw)
    except ValueError:
        return default
    return value if value > 0 else default

DATABASE_POOL_SIZE = _positive_int_env("DATABASE_POOL_SIZE", 5)
DATABASE_MAX_OVERFLOW = _positive_int_env("DATABASE_MAX_OVERFLOW", 10)
DATABASE_POOL_TIMEOUT = _positive_int_env("DATABASE_POOL_TIMEOUT", 30)
DATABASE_POOL_RECYCLE = _positive_int_env("DATABASE_POOL_RECYCLE", 1800)
DATABASE_CONNECT_TIMEOUT = _positive_int_env("DATABASE_CONNECT_TIMEOUT", 10)

# libpq connection-level timeout (seconds) — bounds how long a single
# connection ATTEMPT can hang (unreachable host, firewall black-hole, etc.)
# before psycopg gives up, independent of pool_timeout (which bounds how
# long a request waits for a pool slot once connections are established).
_engine_connect_args: dict = {"connect_timeout": DATABASE_CONNECT_TIMEOUT}
# DATABASE_SSL_MODE is only ever applied when DATABASE_URL doesn't already
# pin an sslmode itself (Supabase and most managed-Postgres connection
# strings already carry "?sslmode=require") — this must never override a
# provider's own explicit URL setting, only fill in a gap when there isn't one.
_database_ssl_mode = os.getenv("DATABASE_SSL_MODE", "").strip()
if _database_ssl_mode and "sslmode=" not in DATABASE_URL:
    _engine_connect_args["sslmode"] = _database_ssl_mode

engine = create_engine(
    DATABASE_URL,
    connect_args=_engine_connect_args,
    pool_pre_ping=True,       # discards a stale/closed pooled connection instead of handing it to a request
    pool_size=DATABASE_POOL_SIZE,
    max_overflow=DATABASE_MAX_OVERFLOW,
    pool_timeout=DATABASE_POOL_TIMEOUT,
    pool_recycle=DATABASE_POOL_RECYCLE,
)

# Test-only schema pinning. Managed transaction poolers may ignore PGOPTIONS
# and connection-string `options`, so the isolated PostgreSQL harness asks the
# application to SET search_path on every pool checkout and then verifies it.
# Production can never enable this escape hatch, and only generated remediation
# schema names are accepted.
_test_db_search_path = os.getenv("SUPPLY_AI_DB_SEARCH_PATH", "").strip()
if _test_db_search_path:
    if IS_PRODUCTION or not re.fullmatch(r"cauldra_[a-z0-9_]+_[a-f0-9]{12}", _test_db_search_path):
        raise RuntimeError("SUPPLY_AI_DB_SEARCH_PATH is restricted to generated non-production test schemas.")

    @event.listens_for(engine, "checkout")
    def _pin_test_schema(dbapi_connection, _connection_record, _connection_proxy):
        cursor = dbapi_connection.cursor()
        try:
            cursor.execute(f'SET search_path TO "{_test_db_search_path}"')
        finally:
            cursor.close()

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

def _classify_database_connection_error(exc: BaseException) -> str:
    """Maps a raw psycopg/SQLAlchemy connection failure to one short,
    secret-free explanation a human can act on immediately, instead of a raw
    driver traceback. Classification reads only the driver's own error text
    (which never includes the password — confirmed against real Postgres
    auth-failure/timeout/refused errors) — DATABASE_URL itself is never
    inspected or echoed here."""
    description = str(exc).lower()
    if "password authentication failed" in description or "authentication failed" in description:
        return "PostgreSQL authentication failed. Check the username/password in DATABASE_URL."
    if any(marker in description for marker in (
        "could not translate host name", "name or service not known",
        "nodename nor servname", "getaddrinfo failed", "temporary failure in name resolution",
    )):
        return "PostgreSQL host could not be resolved. Check the hostname in DATABASE_URL."
    if "connection refused" in description:
        return "PostgreSQL refused the connection. Verify host, port, database service, and provider connection settings."
    if "timeout" in description or "timed out" in description:
        return "PostgreSQL connection timed out. Check host, network, provider status, or firewall settings."
    if "ssl" in description:
        return "PostgreSQL SSL/connection configuration problem. Check DATABASE_SSL_MODE and provider SSL requirements."
    return "PostgreSQL connection failed. Check DATABASE_URL, network access, and provider status."

def _ping_database() -> None:
    """SELECT 1 against a short-lived connection, released back to the pool
    immediately (the `with` block closes it whether the query succeeds or
    raises) — no transaction is opened beyond what that single implicit
    SELECT needs, and no schema/data is ever touched. Raises on failure;
    every caller below decides for itself what "failure" means to it
    (startup: fatal; GET /health: a 503 response)."""
    with engine.connect() as conn:
        conn.execute(sql_text("SELECT 1"))

def verify_database_connectivity() -> None:
    """One-time startup check — fails fast with a short, secret-free message
    instead of letting the first real request surface a raw driver
    traceback. Never mutates schema; Alembic (`alembic upgrade head`) is the
    only thing that does that — this only proves the configured DATABASE_URL
    is actually reachable before the app starts serving traffic. The
    original exception is preserved via `raise ... from exc` so the full
    driver-level detail is still available in server logs/tracebacks for
    debugging, without it being the primary message a human has to parse."""
    try:
        _ping_database()
    except Exception as exc:
        raise RuntimeError(_classify_database_connection_error(exc)) from exc

def _log_database_backend_identity() -> None:
    """One safe, unambiguous startup log line so a Railway deployment (or
    anyone reading its logs) can immediately confirm which database is
    actually in use, without ever printing a credential. Only host/port/
    database name (from SQLAlchemy's already-parsed engine.url — never the
    raw DATABASE_URL string, which is the only part that could carry a
    password) plus a live current_schema() read. Best-effort: a failure here
    is never fatal on its own — verify_database_connectivity() right above
    this call is what actually gates startup."""
    url = engine.url
    schema = "unknown"
    try:
        with engine.connect() as conn:
            schema = conn.execute(sql_text("SELECT current_schema()")).scalar() or "unknown"
    except Exception:
        pass
    print(f"[startup] Database backend: PostgreSQL")
    print(f"[startup] Database host: {url.host}:{url.port or 5432}")
    print(f"[startup] Database name: {url.database}")
    print(f"[startup] Database schema: {schema}")

if not SKIP_DB_STARTUP_CHECK:
    verify_database_connectivity()
    _log_database_backend_identity()

# -----------------------------------------------------------------------------
# MODELS
# -----------------------------------------------------------------------------
class BusinessProfile(Base):
    __tablename__ = "business_profile"
    id = Column(Integer, primary_key=True, index=True)
    business_code = Column(String, unique=True, index=True, nullable=False)
    company_name = Column(String, default="Cauldra")
    email = Column(String, nullable=True)
    phone = Column(String, nullable=True)
    address = Column(String, nullable=True)
    currency = Column(String, default="USD ($)")
    tax_id = Column(String, nullable=True)
    country = Column(String, nullable=True)
    country_code = Column(String, nullable=True)
    locale = Column(String, default="en-US")
    language = Column(String, default="en")
    timezone = Column(String, default="UTC")
    phone_country_code = Column(String, nullable=True)
    subscription_plan = Column(String, nullable=False, default="starter")
    billing_interval = Column(String, nullable=False, default="monthly")
    subscription_started_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    trial_started_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    brain_relationships_checked_at = Column(SQLDateTime, nullable=True)
    # Business Brain invalidation (performance): GET /business-brain used to
    # run refresh_business_brain()'s full per-product recompute on EVERY
    # read. Now it only recomputes when dirty, set true here and cleared
    # once a refresh completes — see mark_business_brain_dirty() and
    # business_brain() below. Defaults true so any business that predates
    # this column gets one correct refresh on its first read afterward.
    business_brain_dirty = Column(Boolean, default=True, nullable=False)
    business_brain_refreshed_at = Column(SQLDateTime, nullable=True)

    users = relationship("User", back_populates="business_rel", cascade="all, delete-orphan")

class Warehouse(Base):
    __tablename__ = "warehouses"
    __table_args__ = (UniqueConstraint("business_id", "name", name="uq_warehouse_business_name"),)
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    name = Column(String, nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(SQLDateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, index=True, nullable=False)
    password = Column(String, nullable=False)
    previous_password_hash = Column(String, nullable=True)
    role = Column(String, nullable=False, default="staff")
    firstname = Column(String, nullable=True)
    lastname = Column(String, nullable=True)
    position = Column(String, nullable=True)
    email = Column(String, nullable=False)
    phone = Column(String, nullable=False)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    must_change_password = Column(Boolean, default=False, nullable=False)
    disabled = Column(Boolean, default=False, nullable=False)
    auth_version = Column(Integer, default=1, nullable=False)

    business_rel = relationship("BusinessProfile", back_populates="users")
    # No ownership cascade: business data survives user deletion.

class Product(Base):
    __tablename__ = "products"
    id = Column(Integer, primary_key=True, index=True)
    sku = Column(String, index=True, nullable=False)
    barcode = Column(String, index=True, nullable=True)
    name = Column(String, nullable=False)
    category = Column(String, nullable=False)
    size = Column(String, nullable=True)
    quantity = Column(Integer, nullable=False, default=0)
    min_stock_level = Column(Integer, nullable=False, default=5)
    cost_price = Column(Float, nullable=False, default=0.0)
    wholesale_price = Column(Float, default=0.0)
    retail_price = Column(Float, nullable=False, default=0.0)
    warehouse = Column(String, default="Main Central Warehouse")
    initial_stock = Column(Integer, default=0)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    expiry_date = Column(SQLDateTime, nullable=True)
    seasonal_checked_at = Column(SQLDateTime, nullable=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    owner_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    # client_ref / synced_at: offline-sync support (see Expense for the original
    # version of this pattern). client_ref is a client-generated id an offline
    # client attaches before a server id exists; retrying the same create with
    # the same client_ref returns the original product instead of duplicating it.
    client_ref = Column(String, nullable=True, index=True)
    synced_at = Column(SQLDateTime, nullable=True)

class Supplier(Base):
    __tablename__ = "suppliers"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    contact_email = Column(String, nullable=True)
    phone = Column(String, nullable=False)
    lead_time_days = Column(Integer, default=3)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    owner_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)

class PurchaseOrder(Base):
    __tablename__ = "purchase_orders"
    id = Column(Integer, primary_key=True, index=True)
    supplier_id = Column(Integer, ForeignKey("suppliers.id", ondelete="SET NULL"), nullable=True)
    status = Column(String, default="DRAFT")
    total_estimated_cost = Column(Float, default=0.0)
    email_draft = Column(Text)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    owner_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class Expense(Base):
    """A business expense transaction. Deliberately mirrors PurchaseOrder's
    shape (business_id + owner_id + created_at) since it's the same kind of
    business-owned financial-adjacent record.

    Fields present now but not yet used:
    - payment_source: nullable free-text placeholder. Cauldra has no real
      account/payment-source model to reference (only AccountActionRequest,
      which is an unrelated staff-management approval workflow) — adding a
      foreign key to a model that doesn't exist isn't possible, so this stays
      a plain string until that system is built.
    - client_ref / synced_at: present so a future offline-sync feature (also
      not yet built anywhere in this app) can be added without a schema
      migration — client_ref is what a future offline client would generate
      locally before a server ID exists, and is unique per business so a
      retried sync of the same locally-created expense can't double-insert.
      Both are unused by any endpoint today.
    """
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    category = Column(String, nullable=False)
    amount = Column(Float, nullable=False)
    payment_source = Column(String, nullable=True)
    note = Column(String, nullable=True)
    owner_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False, index=True)
    client_ref = Column(String, nullable=True, index=True)
    synced_at = Column(SQLDateTime, nullable=True)
    # Direct Business Day relationship (nullable: rows recorded before this
    # column existed, or recorded with no open day in a rare edge case, have
    # no value here — reporting falls back to the existing timestamp-window
    # matching for those, exactly as it already did for every row before).
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="SET NULL"), nullable=True, index=True)

class WarehouseStock(Base):
    __tablename__ = "warehouse_stocks"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=False, index=True)
    warehouse = Column(String, nullable=False)
    quantity = Column(Integer, nullable=False, default=0)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(SQLDateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class ProductDeletionRequest(Base):
    __tablename__ = "product_deletion_requests"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="SET NULL"), nullable=True)
    product_name = Column(String, nullable=False)
    requested_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    requested_by_name = Column(String, nullable=False)
    status = Column(String, nullable=False, default="PENDING", index=True)
    reason = Column(Text, nullable=True)
    resolved_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    resolved_by_name = Column(String, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    resolved_at = Column(SQLDateTime, nullable=True)

class SaleModel(Base):
    __tablename__ = "sales"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="SET NULL"), nullable=True)
    quantity = Column(Integer, default=0)
    total_price = Column(Float, default=0.0)
    # --- Sale-time snapshots -------------------------------------------------
    # Together these make a sale line independently reconstructable: every
    # financial figure this row contributes to any report is derivable from
    # the row itself, with NO dependency on the live Product row's current
    # price, cost, or name (and none on the Product row still existing).
    #
    # unit_cost_at_sale: the product's cost_price AT THE MOMENT this sale was
    # recorded — set once, at checkout, never touched again. Historical COGS
    # must always read this snapshot, never the product's CURRENT cost_price,
    # so a later cost change can never rewrite old profit figures.
    # unit_price: the actual transaction selling price per unit (which may be
    # a cashier-negotiated price, not the catalog retail/wholesale price).
    # total_price / quantity would approximate it, but storing it explicitly
    # avoids a division and keeps the sale's own stated price authoritative.
    # product_name_snapshot: what the product was called when it was sold —
    # so history stays readable after a rename or a deletion (product_id is
    # ON DELETE SET NULL). Display only; never used in any arithmetic.
    #
    # All three are nullable ONLY for rows recorded before each column
    # existed. Reads fall back conservatively for those legacy rows (see
    # compute_financial_summary / _serialize_sale_transactions) — a read-time
    # fallback for data that predates the snapshot, never a write that would
    # rewrite history with a guessed value.
    unit_cost_at_sale = Column(Float, nullable=True)
    unit_price = Column(Float, nullable=True)
    # Which catalog tier (or negotiation) this line's unit_price actually
    # came from — "retail" | "wholesale" | "negotiated" (see
    # SalesCheckoutItem.price_mode / sales_checkout()). Purely descriptive:
    # unit_price is already the complete, authoritative charged amount on
    # its own, so nothing here is ever re-derived FROM pricing_type — it
    # only answers "which tier" for display/audit, never "how much". NULL
    # for sales recorded before this column existed.
    pricing_type = Column(String, nullable=True)
    product_name_snapshot = Column(String, nullable=True)
    timestamp = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    # client_ref: THE CHECKOUT/TRANSACTION GROUPING KEY. Every SaleModel row
    # created by one /sales/checkout submission shares this exact value, so
    # "how many transactions?" is COUNT(DISTINCT client_ref) while "how many
    # units?" is SUM(quantity) — a 3-product checkout is ONE transaction, not
    # three (see checkout_key_expr / compute_financial_summary). It doubles
    # as the idempotency key: a retried submission carrying a client_ref
    # already recorded returns the original result instead of re-applying
    # inventory and re-creating rows. Generated server-side when the client
    # sends none. synced_at is set only when the request genuinely came from
    # the offline outbox, so it keeps its original narrower meaning.
    client_ref = Column(String, nullable=True, index=True)
    synced_at = Column(SQLDateTime, nullable=True)
    # Direct Business Day relationship (see Expense.business_day_id for the
    # same nullable-for-pre-migration-rows rationale).
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="SET NULL"), nullable=True, index=True)

class SaleTransaction(Base):
    """One CHECKOUT — the header for the one-or-more SaleModel line rows it
    created. Deliberately minimal: it exists to make "one checkout" a real,
    database-enforced thing rather than something inferred, and it carries
    only what identifies the checkout as a whole.

    Its reason for existing is the UNIQUE(business_id, client_ref)
    constraint. Checkout's idempotency guard used to be a plain
    "SELECT ... if found, return early" check, which is a time-of-check /
    time-of-use race: two genuinely concurrent submissions of the same cart
    (a fast double-click dispatching both requests before either commits)
    BOTH passed that check and BOTH recorded a sale — one unit sold,
    reported as two, with the stock decrement of one request silently lost
    to the other's write. A read cannot fix that race; only the database
    can. Inserting this header FIRST means the second writer's INSERT is
    rejected by the unique index no matter how the two requests interleave,
    and that loser then returns the winner's result instead of writing
    anything at all (see sales_checkout).

    Reporting deliberately does NOT depend on this table: transaction counts
    still go through checkout_key_expr() over SaleModel.client_ref, which
    also covers pre-existing rows that have no header. This table is the
    write-side integrity anchor, not a second source of truth."""
    __tablename__ = "sale_transactions"
    __table_args__ = (UniqueConstraint("business_id", "client_ref", name="uq_sale_transaction_business_client_ref"),)
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="SET NULL"), nullable=True, index=True)
    client_ref = Column(String, nullable=False, index=True)
    total_price = Column(Float, nullable=False, default=0.0)
    created_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_by_name = Column(String, nullable=True)
    created_by_role = Column(String, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False, index=True)

class MutationIdempotency(Base):
    """Atomic replay claim for offline-capable non-checkout mutations."""
    __tablename__ = "mutation_idempotency"
    __table_args__ = (
        UniqueConstraint(
            "business_id", "operation", "client_ref",
            name="uq_mutation_idempotency_business_operation_ref",
        ),
    )
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    operation = Column(String, nullable=False)
    client_ref = Column(String, nullable=False)
    request_hash = Column(String, nullable=False)
    status = Column(String, nullable=False, default="processing", index=True)
    response_json = Column(Text, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    completed_at = Column(SQLDateTime, nullable=True)

class BusinessDay(Base):
    """One row per Business Day SESSION — NOT one row per business per
    calendar date. A business may open, close, and open again multiple
    independent sessions on the same business-local date (e.g. #41 opened
    8:00 AM/closed 1:30 PM, then #42 opened 2:15 PM the same day) — each is
    its own row with its own id, never merged or reused. `date` records
    which business-local calendar date a session was opened on for display/
    filtering purposes only; it carries no uniqueness constraint. The one
    real constraint is that at most one session per business can be ACTIVE
    (is_open=True) at any moment — enforced by a partial unique index on
    (business_id) WHERE is_open (see startup migrations) and by
    get_active_business_day()/start_business_day() below, never by date.

    A row's identity never changes across its whole lifecycle (open ->
    closed -> reopened -> closed again) — the intermediate transitions are
    NOT stored here as the source of truth, they live as immutable AuditLog
    rows (business_day_id FK) so the full sequence can never be collapsed or
    lost. This row is a convenience "current state" summary, always
    reconstructable from that log.

    opened_at/opened_by_id: the ORIGINAL open — never overwritten by a reopen.
    closed_at/closed_by_id: the MOST RECENT close (matches the existing
    sales_history/sales_analytics query semantics, which use closed_at as the
    day's upper time boundary — after a reopen+reclose that boundary must
    reflect the latest closure, not the first one). The first closure's exact
    timestamp/actor/snapshot is preserved forever in its own AuditLog row.
    is_open: kept exactly as before (existing queries filter on it) —
    True for OPEN/REOPENED, False for CLOSED. status is the richer lifecycle
    field new code should prefer.
    """
    __tablename__ = "business_days"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    date = Column(String, nullable=False, index=True)
    opened_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    closed_at = Column(SQLDateTime, nullable=True)
    is_open = Column(Boolean, default=True, nullable=False)
    status = Column(String, default="OPEN", nullable=False)  # OPEN | CLOSED | REOPENED
    opened_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    opened_by_name = Column(String, nullable=True)
    opened_by_role = Column(String, nullable=True)
    closed_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    closed_by_name = Column(String, nullable=True)
    closed_by_role = Column(String, nullable=True)
    reopen_count = Column(Integer, default=0, nullable=False)

class BusinessDayReopenRequest(Base):
    """Mirrors AccountActionRequest's shape deliberately — same pending ->
    approved/rejected workflow this codebase already uses, applied to
    Business Day reopening instead of staff account actions. The row itself
    is mutable workflow bookkeeping (status changes PENDING -> resolved);
    the actual historical record of what happened is the AuditLog trail
    (REOPEN_REQUESTED/REOPEN_APPROVED/REOPEN_REJECTED), not this row."""
    __tablename__ = "business_day_reopen_requests"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="CASCADE"), nullable=False, index=True)
    reason = Column(Text, nullable=False)
    requested_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    requested_by_name = Column(String, nullable=True)
    requested_by_role = Column(String, nullable=True)
    status = Column(String, default="PENDING", nullable=False, index=True)  # PENDING | APPROVED | REJECTED
    resolved_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    resolved_by_name = Column(String, nullable=True)
    resolution_note = Column(Text, nullable=True)
    resolved_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class SaleAdjustment(Base):
    """A documented correction to a Sale. The original SaleModel row is never
    modified — this references it and records the delta. Reporting computes
    final totals as original + sum(adjustments), never by mutating history."""
    __tablename__ = "sale_adjustments"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    sale_id = Column(Integer, ForeignKey("sales.id", ondelete="CASCADE"), nullable=False, index=True)
    quantity_delta = Column(Integer, default=0, nullable=False)
    amount_delta = Column(Float, default=0.0, nullable=False)
    reason = Column(Text, nullable=False)
    created_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_by_name = Column(String, nullable=True)
    created_by_role = Column(String, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class ExpenseAdjustment(Base):
    """A documented correction to an Expense — same principle as SaleAdjustment."""
    __tablename__ = "expense_adjustments"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    expense_id = Column(Integer, ForeignKey("expenses.id", ondelete="CASCADE"), nullable=False, index=True)
    amount_delta = Column(Float, default=0.0, nullable=False)
    reason = Column(Text, nullable=False)
    created_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_by_name = Column(String, nullable=True)
    created_by_role = Column(String, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class RefundTransaction(Base):
    """One refund event, possibly spanning several original sale line items
    (e.g. a customer returning 1 Coca Cola and 2 Waters from the same
    receipt in one go). The original SaleModel rows are NEVER modified or
    deleted — a refund is always a separate, additive financial event that
    references them (see RefundLine.original_sale_id). Mirrors
    BusinessDayReopenRequest/SaleAdjustment's own established shape
    (business_id + actor identity + created_at) rather than inventing a new
    convention.

    business_day_id is the day the REFUND itself was performed on — which
    may be a different day than the original sale's own business_day_id
    (see BusinessDay model docstring's session model). A refund's financial
    effect belongs to whatever day it actually happened, exactly like a
    real cash drawer adjustment would.

    original_client_ref: the checkout transaction being refunded, when the
    original sale(s) have one (see SalesCheckoutRequest.client_ref, now
    generated for every checkout, not just offline ones — see
    sales_checkout()). Purely for traceability/display; not a join key any
    query depends on (RefundLine.original_sale_id is the real reference).

    client_ref: an idempotency key for THIS refund submission itself (same
    pattern as SaleModel.client_ref for checkout) — a duplicate submission
    with the same value returns the original result instead of refunding
    twice. Enforced by a partial unique index (business_id, client_ref)
    WHERE client_ref IS NOT NULL, see startup migrations."""
    __tablename__ = "refund_transactions"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="SET NULL"), nullable=True, index=True)
    original_client_ref = Column(String, nullable=True, index=True)
    reason = Column(String, nullable=True)
    note = Column(Text, nullable=True)
    refund_total = Column(Float, nullable=False, default=0.0)
    # NULL means the original legacy sale never recorded cost-at-sale. It is
    # intentionally distinct from a real zero cost and must remain unknown.
    refund_cost_total = Column(Float, nullable=True)
    created_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_by_name = Column(String, nullable=True)
    created_by_role = Column(String, nullable=True)
    client_ref = Column(String, nullable=True, index=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False, index=True)

class RefundLine(Base):
    """One refunded product line within a RefundTransaction. quantity is
    ALWAYS validated against (original sale quantity - already refunded
    quantity for that same sale) before this row is ever created — see
    create_refund() — so SUM(RefundLine.quantity) for a given
    original_sale_id can never exceed that sale's own quantity.

    unit_price/unit_cost are snapshots copied from the ORIGINAL sale at
    refund time (SaleModel.total_price/quantity and
    SaleModel.unit_cost_at_sale) — never the product's CURRENT price/cost —
    so a refund's financial value can never drift from what was actually
    sold, even if the product's pricing changes later (same principle as
    SaleModel.unit_cost_at_sale itself).

    business_day_id is denormalized from the parent RefundTransaction
    (same value) purely so day-level reporting (sales_history,
    compute_financial_summary) can aggregate refunds with a single indexed
    GROUP BY, exactly like Sale/Expense.business_day_id already do —
    never treated as a second source of truth.

    restocked records whether THIS line's quantity was actually returned to
    inventory — independent per line, since a multi-item refund may return
    some goods to stock (customer return) and not others (damaged/consumed)
    in the same submission (see section 8 of the refund spec)."""
    __tablename__ = "refund_lines"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="SET NULL"), nullable=True, index=True)
    refund_transaction_id = Column(Integer, ForeignKey("refund_transactions.id", ondelete="CASCADE"), nullable=False, index=True)
    original_sale_id = Column(Integer, ForeignKey("sales.id", ondelete="SET NULL"), nullable=True, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="SET NULL"), nullable=True)
    product_name_snapshot = Column(String, nullable=False)
    quantity = Column(Integer, nullable=False)
    unit_price = Column(Float, nullable=False)
    unit_cost = Column(Float, nullable=True)
    refund_amount = Column(Float, nullable=False)
    refund_cost = Column(Float, nullable=True)
    restocked = Column(Boolean, nullable=False, default=False)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False, index=True)

class Alert(Base):
    __tablename__ = "alerts"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    type = Column(String, nullable=False)
    title = Column(String, nullable=False)
    message = Column(Text, nullable=False)
    severity = Column(String, default="info")
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    expires_at = Column(SQLDateTime, nullable=True)

class AlertRead(Base):
    __tablename__ = "alert_reads"
    id = Column(Integer, primary_key=True, index=True)
    alert_id = Column(Integer, ForeignKey("alerts.id", ondelete="CASCADE"), nullable=False)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    read_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class AuditLog(Base):
    __tablename__ = "audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    actor_user_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    actor_username = Column(String, nullable=True)
    actor_role = Column(String, nullable=True)
    action = Column(String, nullable=False)
    target_user_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    target_username = Column(String, nullable=True)
    description = Column(Text, nullable=False)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    # Additive, both nullable — every existing row and every existing
    # add_audit() call site is unaffected. business_day_id ties an event to a
    # specific Business Day (the Business Day activity timeline is just a
    # filtered read of this same table, not a separate log). metadata_json
    # carries structured details a plain description can't (closing
    # snapshots, correction old/new values, reopen reason) as a JSON string —
    # this table remains an append-only log; nothing here is ever updated
    # after creation.
    business_day_id = Column(Integer, ForeignKey("business_days.id", ondelete="SET NULL"), nullable=True, index=True)
    metadata_json = Column(Text, nullable=True)

class AccountActionRequest(Base):
    __tablename__ = "account_action_requests"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    target_user_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    target_username = Column(String, nullable=False)
    # Denormalized (captured at request time) rather than looked up live, so it
    # still displays correctly after an approved "delete" removes the target row.
    target_position = Column(String, nullable=True)
    action = Column(String, nullable=False)  # delete / disable
    requested_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    requested_by_name = Column(String, nullable=True)
    requested_by_position = Column(String, nullable=True)
    status = Column(String, default="PENDING", nullable=False, index=True)
    resolved_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    resolved_by_name = Column(String, nullable=True)
    resolved_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class SessionRevocation(Base):
    __tablename__ = "session_revocations"
    id = Column(Integer, primary_key=True, index=True)
    jti = Column(String, unique=True, index=True, nullable=False)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=True)
    expires_at = Column(SQLDateTime, nullable=False)
    revoked_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class RefreshSession(Base):
    __tablename__ = "refresh_sessions"
    id = Column(Integer, primary_key=True, index=True)
    token_hash = Column(String, unique=True, index=True, nullable=False)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    expires_at = Column(SQLDateTime, nullable=False)
    revoked_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    replaced_by_hash = Column(String, nullable=True)

class AuthFailure(Base):
    __tablename__ = "auth_failures"
    id = Column(Integer, primary_key=True, index=True)
    scope = Column(String, index=True, nullable=False)
    key_hash = Column(String, index=True, nullable=False)
    failures = Column(Integer, default=0, nullable=False)
    window_started_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    locked_until = Column(SQLDateTime, nullable=True)

class PasswordRecovery(Base):
    __tablename__ = "password_recoveries"
    id = Column(Integer, primary_key=True, index=True)
    recovery_id = Column(String, unique=True, index=True, nullable=False)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    channel = Column(String, nullable=False)
    code_hash = Column(String, nullable=False)
    expires_at = Column(SQLDateTime, nullable=False)
    resend_after = Column(SQLDateTime, nullable=False)
    attempts = Column(Integer, default=0, nullable=False)
    used = Column(Boolean, default=False, nullable=False)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class PresenceSession(Base):
    __tablename__ = "presence_sessions"
    id = Column(Integer, primary_key=True, index=True)
    session_id = Column(String, unique=True, index=True, nullable=False)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    signed_in_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    last_seen_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    signed_out_at = Column(SQLDateTime, nullable=True)

class PriceMonitorSource(Base):
    __tablename__ = "price_monitor_sources"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False)
    supplier_id = Column(Integer, ForeignKey("suppliers.id", ondelete="SET NULL"), nullable=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="SET NULL"), nullable=True)
    source_type = Column(String, nullable=False)
    source_url = Column(String, nullable=True)
    last_price = Column(Float, nullable=True)
    last_checked_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class PriceHistory(Base):
    __tablename__ = "price_history"
    id = Column(Integer, primary_key=True, index=True)
    source_id = Column(Integer, ForeignKey("price_monitor_sources.id", ondelete="CASCADE"), nullable=False)
    price = Column(Float, nullable=False)
    recorded_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

# The Business Brain is deliberately a small, additive intelligence ledger.
# It stores only evidence calculated from the owning business's operational
# data, which lets Cauldra measure a forecast after its horizon passes instead
# of presenting a transient, untestable "AI insight".
class BusinessBrainPrediction(Base):
    __tablename__ = "business_brain_predictions"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=False, index=True)
    kind = Column(String, nullable=False, default="velocity")  # "velocity" (7-day) or "seasonal" (recurring-week)
    forecast_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False, index=True)
    target_at = Column(SQLDateTime, nullable=False, index=True)
    horizon_days = Column(Integer, nullable=False, default=7)
    predicted_units = Column(Float, nullable=False)
    confidence = Column(Float, nullable=False)
    evidence_json = Column(Text, nullable=False, default="{}")
    model_version = Column(String, nullable=False, default="brain-v1")
    actual_units = Column(Float, nullable=True)
    accuracy_score = Column(Float, nullable=True)
    evaluated_at = Column(SQLDateTime, nullable=True)

class BusinessBrainRecommendation(Base):
    __tablename__ = "business_brain_recommendations"
    __table_args__ = (UniqueConstraint("business_id", "fingerprint", name="uq_brain_recommendation_fingerprint"),)
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=True, index=True)
    fingerprint = Column(String, nullable=False)
    kind = Column(String, nullable=False, index=True)
    priority = Column(String, nullable=False, default="important")
    title = Column(String, nullable=False)
    summary = Column(Text, nullable=False)
    evidence_json = Column(Text, nullable=False, default="{}")
    status = Column(String, nullable=False, default="new", index=True)
    opened_at = Column(SQLDateTime, nullable=True)
    acted_at = Column(SQLDateTime, nullable=True)
    dismissed_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(SQLDateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class BusinessBrainMemory(Base):
    __tablename__ = "business_brain_memories"
    __table_args__ = (UniqueConstraint("business_id", "fingerprint", name="uq_brain_memory_fingerprint"),)
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=True, index=True)
    fingerprint = Column(String, nullable=False)
    statement = Column(Text, nullable=False)
    evidence_json = Column(Text, nullable=False, default="{}")
    confidence = Column(Float, nullable=False)
    last_observed_at = Column(SQLDateTime, nullable=False)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(SQLDateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

# Recurring-week detection, additive to the Business Brain above. One row per
# (business, product) holding only the single best-evidenced week-of-year
# bucket for that product — never a generic "category X = season Y" guess.
# A row is only written once it clears a confidence floor, and is withdrawn
# (deleted) the moment the evidence behind it no longer holds.
class BusinessBrainSeasonalPattern(Base):
    __tablename__ = "business_brain_seasonal_patterns"
    __table_args__ = (UniqueConstraint("business_id", "product_id", name="uq_brain_seasonal_product"),)
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=False, index=True)
    week_of_year = Column(Integer, nullable=False)
    cycles_observed = Column(Integer, nullable=False)
    avg_units_in_week = Column(Float, nullable=False)
    baseline_avg_units = Column(Float, nullable=False)
    lift_ratio = Column(Float, nullable=False)
    consistency = Column(Float, nullable=False)
    confidence = Column(Float, nullable=False)
    last_computed_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

# Cross-product relationships (section 12): two products whose weekly sales
# have moved together (or opposite) with enough shared history to be
# meaningful. Never speculative — a row only exists while the correlation
# behind it clears the evidence floor; product_a_id is always the smaller id
# so a pair is stored once regardless of which product is queried first.
class BusinessBrainRelationship(Base):
    __tablename__ = "business_brain_relationships"
    __table_args__ = (UniqueConstraint("business_id", "product_a_id", "product_b_id", name="uq_brain_relationship_pair"),)
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    product_a_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=False, index=True)
    product_b_id = Column(Integer, ForeignKey("products.id", ondelete="CASCADE"), nullable=False, index=True)
    overlapping_weeks = Column(Integer, nullable=False)
    correlation = Column(Float, nullable=False)
    confidence = Column(Float, nullable=False)
    last_computed_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class GeneralCatalog(Base):
    """Shared PRODUCT IDENTITY only — never inventory. A barcode's entry here
    means "this barcode identifies a product named X, made by brand Y", full
    stop. It intentionally carries none of: business_id/business name/owner,
    cost/wholesale/retail price, quantity/stock/min-stock, warehouse,
    supplier, internal SKU, sales/purchase history, or business notes — a
    row here must never let one business learn anything about another
    business's inventory, pricing, or operations. See /catalog/barcode-lookup
    and upcitemdb_provider.py for how this is populated and read.

    `category` is DEPRECATED: earlier code populated and returned it, but
    category is a business's own inventory choice, not shared product
    identity (a "phone charger" might be "Electronics" to one business and
    "Accessories" to another) — no code should set or rely on it going
    forward. The column stays physically present (nullable-safe default)
    rather than being dropped, so existing PostgreSQL rows and any pending
    Alembic history are never disturbed by a purely behavioral change.
    """
    __tablename__ = "general_catalog"
    id = Column(Integer, primary_key=True, index=True)
    barcode = Column(String, unique=True, nullable=True, index=True)
    catalog_key = Column(String, unique=True, index=True, nullable=False, default="legacy")
    product_name = Column(String, nullable=False)
    category = Column(String, nullable=False, default="General")  # deprecated — see class docstring
    brand = Column(String, nullable=True)
    size = Column(String, nullable=True)
    # "business_submission" (from auto_upsert_general_catalog, any business
    # adding/editing a product) or "upcitemdb" (cached from a live provider
    # lookup) — see upsert_general_catalog_identity(). Never any other value.
    source = Column(String, nullable=False, default="business_submission")
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(SQLDateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class StoredUpload(Base):
    """Metadata for a private, backend-retained document.

    The file bytes are intentionally stored outside the web root; access is
    always mediated by an authenticated, business-scoped endpoint.
    """
    __tablename__ = "stored_uploads"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    uploaded_by_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True, index=True)
    kind = Column(String, nullable=False, index=True)
    original_name = Column(String, nullable=False)
    storage_key = Column(String, unique=True, nullable=False, index=True)
    content_type = Column(String, nullable=False)
    size_bytes = Column(Integer, nullable=False)
    content_hash = Column(String, nullable=False, index=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class AIUsageLedger(Base):
    """Immutable, provider-independent record of billable AI activity."""
    __tablename__ = "ai_usage_ledger"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True, index=True)
    operation_type = Column(String, nullable=False, index=True)
    credits_consumed = Column(Integer, nullable=False, default=0)
    billing_period = Column(String, nullable=False, index=True)
    provider = Column(String, nullable=True)
    model = Column(String, nullable=True)
    success = Column(Boolean, nullable=False, default=False)
    input_tokens = Column(Integer, nullable=True)
    output_tokens = Column(Integer, nullable=True)
    estimated_provider_cost = Column(Float, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class BusinessSubscription(Base):
    __tablename__ = "business_subscriptions"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, unique=True, index=True)
    plan = Column(String, nullable=False)
    billing_interval = Column(String, nullable=False)
    status = Column(String, nullable=False, default="trialing", index=True)
    trial_start_at = Column(SQLDateTime, nullable=True)
    trial_end_at = Column(SQLDateTime, nullable=True)
    paid_at = Column(SQLDateTime, nullable=True)
    current_period_start = Column(SQLDateTime, nullable=True)
    current_period_end = Column(SQLDateTime, nullable=True)
    next_billing_at = Column(SQLDateTime, nullable=True)
    paystack_customer_code = Column(String, nullable=True)
    paystack_subscription_code = Column(String, nullable=True)
    paystack_plan_code = Column(String, nullable=True)
    latest_transaction_reference = Column(String, nullable=True)
    payment_status = Column(String, nullable=True)
    cancelled_at = Column(SQLDateTime, nullable=True)
    card_verified = Column(Boolean, nullable=False, default=False)
    paystack_authorization_code = Column(String, nullable=True)
    card_last4 = Column(String, nullable=True)
    card_type = Column(String, nullable=True)
    card_exp_month = Column(String, nullable=True)
    card_exp_year = Column(String, nullable=True)
    trial_consent_at = Column(SQLDateTime, nullable=True)
    grace_period_ends_at = Column(SQLDateTime, nullable=True)
    cancel_at_period_end = Column(Boolean, nullable=False, default=False)
    # Scheduled downgrade (see schedule_downgrade / cancel_pending_downgrade /
    # the downgrade-aware branch of the recurring-charge webhook below). A
    # downgrade never touches `plan`/`billing_interval` directly — it only
    # ever takes effect once the current, already-paid billing period ends.
    pending_downgrade_plan = Column(String, nullable=True)
    pending_downgrade_billing_interval = Column(String, nullable=True)
    pending_downgrade_effective_at = Column(SQLDateTime, nullable=True)
    pending_downgrade_requested_at = Column(SQLDateTime, nullable=True)
    pending_downgrade_requested_by_user_id = Column(Integer, nullable=True)
    pending_downgrade_paystack_subscription_code = Column(String, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(SQLDateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class PaymentRecord(Base):
    __tablename__ = "payment_records"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    subscription_id = Column(Integer, ForeignKey("business_subscriptions.id", ondelete="SET NULL"), nullable=True, index=True)
    plan = Column(String, nullable=False)
    billing_interval = Column(String, nullable=False)
    amount_kobo = Column(Integer, nullable=False)
    currency = Column(String, nullable=False, default="NGN")
    paystack_reference = Column(String, nullable=False, unique=True, index=True)
    status = Column(String, nullable=False, default="initialized", index=True)
    transaction_metadata = Column(Text, nullable=True)
    purpose = Column(String, nullable=False, default="subscription")
    # A verification-charge refund is asynchronous at Paystack. Local state is
    # deliberately coarser than the provider's (`processing` and
    # `needs-attention` both remain locally pending) so refunded_at can only
    # mean the provider explicitly reported `processed`.
    refund_status = Column(String, nullable=False, default="not_requested", index=True)
    refund_provider_status = Column(String, nullable=True)
    refund_provider_id = Column(String, nullable=True, index=True)
    refund_attempt_count = Column(Integer, nullable=False, default=0)
    refund_requested_at = Column(SQLDateTime, nullable=True)
    refund_updated_at = Column(SQLDateTime, nullable=True)
    refund_last_error = Column(Text, nullable=True)
    paystack_transaction_id = Column(String, nullable=True)
    refunded_at = Column(SQLDateTime, nullable=True)
    paid_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class PaystackWebhookEvent(Base):
    __tablename__ = "paystack_webhook_events"
    id = Column(Integer, primary_key=True, index=True)
    event_key = Column(String, nullable=False, unique=True, index=True)
    event_type = Column(String, nullable=False)
    received_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

class OnboardingAuthorization(Base):
    """A verified, reusable Paystack card authorization captured BEFORE a
    business exists, for the new-business flow: plan -> card verification ->
    registration -> trial. BusinessSubscription/PaymentRecord both require an
    existing business_id, so this is the one new piece of state needed to
    support verifying payment ahead of registration; everything else (Paystack
    helpers, PLAN_CONFIG, trial math, subscription scheduling) reuses the
    existing architecture unchanged. Never stores card numbers/CVV — only the
    safe, reusable authorization info Paystack returns."""
    __tablename__ = "onboarding_authorizations"
    id = Column(Integer, primary_key=True, index=True)
    paystack_reference = Column(String, unique=True, index=True, nullable=False)
    email = Column(String, nullable=False)
    plan = Column(String, nullable=False)
    billing_interval = Column(String, nullable=False)
    amount_kobo = Column(Integer, nullable=False)
    status = Column(String, nullable=False, default="initialized", index=True)  # initialized -> verified -> consumed | failed
    paystack_customer_code = Column(String, nullable=True)
    paystack_authorization_code = Column(String, nullable=True)
    card_last4 = Column(String, nullable=True)
    card_type = Column(String, nullable=True)
    card_exp_month = Column(String, nullable=True)
    card_exp_year = Column(String, nullable=True)
    verified_at = Column(SQLDateTime, nullable=True)
    refund_status = Column(String, nullable=False, default="not_requested", index=True)
    refund_provider_status = Column(String, nullable=True)
    refund_provider_id = Column(String, nullable=True, index=True)
    refund_attempt_count = Column(Integer, nullable=False, default=0)
    refund_requested_at = Column(SQLDateTime, nullable=True)
    refund_updated_at = Column(SQLDateTime, nullable=True)
    refund_last_error = Column(Text, nullable=True)
    paystack_transaction_id = Column(String, nullable=True)
    refunded_at = Column(SQLDateTime, nullable=True)
    consumed_at = Column(SQLDateTime, nullable=True)
    expires_at = Column(SQLDateTime, nullable=False)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

# How long an anonymous guest has to finish the Paystack checkout popup after
# /onboarding/payment/init before the reference is considered abandoned.
ONBOARDING_PAYMENT_SESSION_MINUTES = 60
# How long a *verified* authorization stays usable to complete registration —
# generous enough that a user filling out the registration form isn't rushed,
# bounded so a verified-but-unused authorization can't be replayed indefinitely.
ONBOARDING_AUTHORIZATION_CONSUME_HOURS = 24

class SubscriptionUpgradeQuote(Base):
    """A short-lived, server-authoritative quote for an upgrade's prorated
    price, computed from the business's actual current subscription state.
    The frontend never computes or supplies this amount — it only displays
    what this record says, and payment can only ever be initialized against
    the amount stored here. current_period_end_snapshot lets the checkout
    step detect if the underlying subscription changed after the quote was
    issued (e.g. a race with cancellation or another upgrade), in which case
    the quote is invalidated rather than trusted."""
    __tablename__ = "subscription_upgrade_quotes"
    id = Column(Integer, primary_key=True, index=True)
    quote_reference = Column(String, unique=True, index=True, nullable=False)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    from_plan = Column(String, nullable=False)
    from_interval = Column(String, nullable=False)
    to_plan = Column(String, nullable=False)
    to_interval = Column(String, nullable=False)
    current_price_kobo = Column(Integer, nullable=False)
    new_price_kobo = Column(Integer, nullable=False)
    unused_credit_kobo = Column(Integer, nullable=False)
    amount_due_kobo = Column(Integer, nullable=False)
    current_period_end_snapshot = Column(SQLDateTime, nullable=False)
    status = Column(String, nullable=False, default="issued", index=True)  # issued -> paid | expired | invalidated
    paystack_reference = Column(String, nullable=True, index=True)
    expires_at = Column(SQLDateTime, nullable=False)
    consumed_at = Column(SQLDateTime, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)

# -----------------------------------------------------------------------------
# NOTIFICATIONS (in-app notification center + external push delivery)
# -----------------------------------------------------------------------------
class Notification(Base):
    """The single persisted record behind Cauldra's notification bell/center
    AND external push delivery — see create_notification() below, the one
    function that is ever allowed to insert a row here. A notification may
    be delivered through one or two channels (in_app / push); a channel is a
    delivery detail on THIS row (the in_app/push booleans plus
    push_sent_at), never a reason to create a second row for the same event.

    Recipient-scoped: one row per (event, recipient). A business-wide event
    that reaches three eligible users creates three rows, not one shared
    row, so each user's read state is independent.

    dedup_key (+ stage) is what stops the same condition from re-notifying
    on every request/poll/background sweep — see NOTIFICATION_MANDATORY_
    CATEGORIES and create_notification()'s docstring for the exact rule.
    resolved_at marks an ongoing staged condition (an active stockout, a
    payment failure) as no longer active without deleting the historical
    row, and frees its dedup_key to fire again on a future recurrence."""
    __tablename__ = "notifications"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    recipient_user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    type = Column(String, nullable=False, index=True)
    category = Column(String, nullable=False, index=True)
    severity = Column(String, nullable=False, default="info")  # info | important | critical
    title = Column(String, nullable=False)
    message = Column(Text, nullable=False)
    related_entity_type = Column(String, nullable=True)
    related_entity_id = Column(Integer, nullable=True)
    deep_link = Column(String, nullable=True)
    in_app = Column(Boolean, nullable=False, default=True)
    push = Column(Boolean, nullable=False, default=False)
    is_read = Column(Boolean, nullable=False, default=False, index=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False, index=True)
    read_at = Column(SQLDateTime, nullable=True)
    push_sent_at = Column(SQLDateTime, nullable=True)
    dedup_key = Column(String, nullable=True, index=True)
    stage = Column(String, nullable=True)
    resolved_at = Column(SQLDateTime, nullable=True)

class PushSubscription(Base):
    """One row per browser/device Web Push registration. A user with
    multiple devices/browsers has multiple active rows — every one of them
    receives a push. disabled_at marks a subscription the push service has
    reported as gone (HTTP 404/410 from a delivery attempt) so it is never
    retried again, without deleting the record of it having existed."""
    __tablename__ = "push_subscriptions"
    id = Column(Integer, primary_key=True, index=True)
    business_id = Column(Integer, ForeignKey("business_profile.id", ondelete="CASCADE"), nullable=False, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    endpoint = Column(Text, nullable=False, unique=True)
    p256dh = Column(String, nullable=False)
    auth = Column(String, nullable=False)
    user_agent = Column(String, nullable=True)
    created_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    last_seen_at = Column(SQLDateTime, default=datetime.utcnow, nullable=False)
    disabled_at = Column(SQLDateTime, nullable=True)

class NotificationPreference(Base):
    """Per-user opt-out for OPTIONAL notification categories only (inventory,
    purchase_order, supplier, ai_usage, business_insight, team). Security and
    subscription/billing-critical categories are never stored here — they
    are mandatory and enforced in code (see NOTIFICATION_MANDATORY_CATEGORIES),
    so this table can never be used to silence them, and a row here can never
    grant visibility into a category a user's role isn't already allowed to
    see (see notification_recipients()) — this only ever narrows, never
    widens, who receives what."""
    __tablename__ = "notification_preferences"
    __table_args__ = (UniqueConstraint("user_id", "category", name="ux_notification_preferences_user_category"),)
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    category = Column(String, nullable=False)
    enabled = Column(Boolean, nullable=False, default=True)

UPLOAD_STORAGE_DIR.mkdir(parents=True, exist_ok=True)

# Schema creation, ad-hoc ALTER/CREATE INDEX statements, and one-time legacy
# data backfills used to live here, gated behind AUTO_CREATE_SCHEMA. All of
# it is now Alembic-managed: every column/table/index that block used to
# create is covered by migrations 0001-0011 (verified empirically — running
# the full chain against an empty database produces a schema that matches
# every current SQLAlchemy model column-for-column and index-for-index), and
# every one-time data backfill (Business Day status/business_day_id,
# card_verified grandfathering, subscription-anchor backfill, legacy role
# rename, warehouse registry seed) is now migration 0012_legacy_data_backfill.
# Ordinary application startup performs schema verification (see
# verify_database_connectivity() above) and nothing else — run
# `alembic upgrade head` to apply schema changes.

# Central subscription and AI-cost policy.  None represents an intentionally
# unlimited Enterprise people/location resource, not a large hidden cap.
#
# "enterprise" is displayed to customers as "Premium" (see `label` below) — the
# internal id is kept as "enterprise" for backward compatibility with existing
# database records, Paystack subscriptions, and env-configured plan codes.
# Core is the entry-level, non-AI plan: included_ai_credits=0 means the plan
# has no AI entitlement at all (see require_ai_access below), not just a small
# credit allowance.
PLAN_CONFIG = {
    "core": {"label": "Core", "monthly_price": 5000, "annual_price": 50000, "trial_days": 14,
        "admin": 1, "manager": 1, "staff": 2, "branch": 1, "city": 1, "country": 1, "product": 250, "supplier": 10, "warehouse": 1,
        "purchase_order": 15, "price_monitor": 0, "storage_gb": 2, "included_ai_credits": 0, "ai_overage_unit": 250, "ai_overage_price": 5000},
    "starter": {"label": "Starter", "monthly_price": 20000, "annual_price": 200000, "trial_days": 14,
        "admin": 1, "manager": 1, "staff": 3, "branch": 1, "city": 1, "country": 1, "product": 500, "supplier": 20, "warehouse": 2,
        "purchase_order": 30, "price_monitor": 5, "storage_gb": 5, "included_ai_credits": 500, "ai_overage_unit": 250, "ai_overage_price": 5000},
    "business": {"label": "Business", "monthly_price": 50000, "annual_price": 500000, "trial_days": 14,
        "admin": 2, "manager": 5, "staff": 15, "branch": 5, "city": 5, "country": 2, "product": 2500, "supplier": 100, "warehouse": 10,
        "purchase_order": 200, "price_monitor": 30, "storage_gb": 25, "included_ai_credits": 2500, "ai_overage_unit": 500, "ai_overage_price": 7500},
    "enterprise": {"label": "Premium", "monthly_price": 200000, "annual_price": 2100000, "trial_days": 14,
        "admin": None, "manager": None, "staff": None, "branch": None, "city": None, "country": None, "product": 50000, "supplier": 2000, "warehouse": 100,
        "purchase_order": 5000, "price_monitor": 500, "storage_gb": 250, "included_ai_credits": 15000, "ai_overage_unit": 1000, "ai_overage_price": 10000},
}

PLAN_LIMIT_FIELDS = ["admin", "manager", "staff", "branch", "city", "country", "product", "supplier", "warehouse",
                      "purchase_order", "price_monitor", "storage_gb", "included_ai_credits"]

def plan_public_view(pid: str, cfg: dict) -> dict:
    """Shared shape for both the public /plans catalog and the authenticated
    usage summary's embedded plan catalog, so pricing/limits/savings are computed
    in exactly one place and never duplicated or hardcoded on the frontend."""
    annual_saving = max(0, cfg["monthly_price"] * 12 - cfg["annual_price"])
    return {
        "id": pid, "label": cfg["label"], "monthly_price": cfg["monthly_price"],
        "annual_price": cfg["annual_price"], "annual_saving": annual_saving,
        "trial_days": cfg["trial_days"], "currency": "NGN",
        "ai_included": bool(cfg.get("included_ai_credits")),
        "limits": {f: cfg.get(f) for f in PLAN_LIMIT_FIELDS},
    }

# Strict plan ordering used ONLY to decide whether a requested plan change is
# a genuine upgrade (see /subscription/upgrade-quote). A billing-interval
# change on the SAME plan (e.g. Business Monthly -> Business Annual) is
# intentionally NOT an "upgrade" under this ranking, even though the annual
# price is numerically larger — that stays on the existing /subscription/
# change-plan lateral-move path, unchanged by this feature.
PLAN_RANK = {"core": 0, "starter": 1, "business": 2, "enterprise": 3}
UPGRADE_QUOTE_VALIDITY_MINUTES = 15
UPGRADE_MINIMUM_CHARGE_KOBO = 100  # Paystack requires a positive amount; this is a floor, never the norm

@app.get("/plans")
def list_public_plans():
    """Public, unauthenticated plan catalog for the pre-registration onboarding
    UI (choose-a-plan screen) and the in-app plan picker. Server-authoritative —
    the frontend must never hardcode prices or limits; it renders whatever this
    endpoint returns. Resource limits are plan marketing information, not
    sensitive data — Paystack plan codes are still never exposed here."""
    return {pid: plan_public_view(pid, cfg) for pid, cfg in PLAN_CONFIG.items()}
for _plan_id, _plan in PLAN_CONFIG.items():
    _plan["paystack_monthly_plan_code"] = os.getenv(f"PAYSTACK_{_plan_id.upper()}_MONTHLY_PLAN_CODE", "").strip()
    _plan["paystack_annual_plan_code"] = os.getenv(f"PAYSTACK_{_plan_id.upper()}_ANNUAL_PLAN_CODE", "").strip()
AI_CREDIT_WEIGHTS = {"margin_advisor": 2, "chat": 2, "inventory_insight": 5, "predictive_analysis": 8, "invoice_ocr": 10, "complex_analysis": 10}
# Maps a plan id to the next plan id up — used only to name the next upgrade
# target in messages. enterprise has no plan above it, so it maps to itself.
UPGRADE_PATH = {"core": "starter", "starter": "business", "business": "enterprise", "enterprise": "enterprise"}

EXPENSE_CATEGORIES = [
    "Rent", "Electricity", "Water", "Internet", "Staff Salaries", "Staff Wages",
    "Transportation", "Fuel", "Vehicle Maintenance", "Repairs", "Equipment",
    "Office Supplies", "Packaging", "Cleaning", "Advertising", "Marketing",
    "Software Subscriptions", "Bank Charges", "Taxes", "Government Fees",
    "Insurance", "Security", "Professional Services", "Legal Fees", "Accounting",
    "Inventory Transportation", "Warehouse Expenses", "Store Maintenance",
    "Phone/Airtime", "Delivery Expenses", "Meals/Refreshments", "Travel",
    "Accommodation", "Loan Repayment", "Interest", "Refunds", "Miscellaneous",
]

# Inventory Financial Intelligence — reusable analysis-window configuration.
# Kept in one place rather than scattered as magic numbers throughout the
# calculation engine (see financial_intelligence() below).
FIN_INTEL_MIN_HISTORY_DAYS = 6           # same threshold /products/predictive-forecast already uses for report_ready
FIN_INTEL_STOCKOUT_RISK_DAYS = 7         # "next 7 days" stockout sales-risk window
FIN_INTEL_SLOW_MOVING_LOOKBACK_DAYS = 30  # no sale within this window (or ever) => slow-moving candidate
FIN_INTEL_MARGIN_WINDOW_DAYS = 30        # "monthlyized" margin-pressure impact window
FIN_INTEL_SEVERELY_STALE_DAYS = FIN_INTEL_SLOW_MOVING_LOOKBACK_DAYS * 2

def get_or_create_subscription(db: Session, business: BusinessProfile, commit: bool = True) -> BusinessSubscription:
    """BusinessSubscription is the single source of truth for plan/trial/billing state.
    This backfills a row for any business that predates the subscription table (or
    the rare race where a request lands before registration's own insert commits),
    seeded from the legacy BusinessProfile columns so no existing business is locked out."""
    sub = db.query(BusinessSubscription).filter(BusinessSubscription.business_id == business.id).first()
    if sub:
        return sub
    now = datetime.utcnow()
    plan = (business.subscription_plan or "starter").strip().lower()
    if plan not in PLAN_CONFIG:
        plan = "starter"
    interval = (business.billing_interval or "monthly").strip().lower()
    if interval not in ("monthly", "annual"):
        interval = "monthly"
    trial_start = business.trial_started_at or business.subscription_started_at or now
    trial_end = trial_start + timedelta(days=PLAN_CONFIG[plan]["trial_days"])
    sub = BusinessSubscription(
        business_id=business.id, plan=plan, billing_interval=interval,
        status="trialing" if now < trial_end else "expired",
        trial_start_at=trial_start, trial_end_at=trial_end,
        current_period_start=trial_start, current_period_end=trial_end,
    )
    db.add(sub)
    if commit:
        db.commit()
        db.refresh(sub)
    else:
        db.flush()
    return sub

def subscription_for(db: Session, business: BusinessProfile) -> dict:
    sub = get_or_create_subscription(db, business)
    return PLAN_CONFIG.get((sub.plan or "starter").strip().lower(), PLAN_CONFIG["starter"])

def add_billing_interval(start: datetime, interval: str) -> datetime:
    months = 12 if interval == "annual" else 1
    month = start.month + months
    return datetime(start.year + (month - 1) // 12, (month - 1) % 12 + 1, 1)

def get_subscription(db: Session, business_id: int) -> Optional[BusinessSubscription]:
    return db.query(BusinessSubscription).filter(BusinessSubscription.business_id == business_id).first()

def clear_pending_downgrade(sub: BusinessSubscription):
    """Called at every point a subscription's plan legitimately changes outside
    of the downgrade-apply path itself (a lateral change, a fresh checkout, or a
    genuine upgrade) so a stale scheduled downgrade can never re-apply later and
    silently undo a newer legitimate change (see the webhook's downgrade-aware
    renewal branch, which is the only place a pending downgrade is applied)."""
    sub.pending_downgrade_plan = None
    sub.pending_downgrade_billing_interval = None
    sub.pending_downgrade_effective_at = None
    sub.pending_downgrade_requested_at = None
    sub.pending_downgrade_requested_by_user_id = None
    sub.pending_downgrade_paystack_subscription_code = None

def refresh_subscription_status(db: Session, subscription: Optional[BusinessSubscription]) -> Optional[BusinessSubscription]:
    """Server-authoritative status transitions. Never driven by frontend or client
    clock — only by the server's own now() compared against server-recorded dates."""
    if not subscription:
        return None
    now = datetime.utcnow()
    if subscription.status == "trialing" and subscription.trial_end_at and now >= subscription.trial_end_at:
        # Trial elapsed without a completed paid conversion (either Paystack's
        # scheduled first debit hasn't landed/succeeded yet, or the business never
        # completed checkout). Do not silently keep granting paid access.
        subscription.status = "expired"; subscription.payment_status = "unpaid"
        # This branch only ever executes once per transition (the condition
        # stops matching the instant status changes) — a naturally one-shot
        # hook, the same pattern every other state-transition notification
        # in this file uses.
        create_notification(
            db, business_id=subscription.business_id, category="subscription", severity="critical", type="SUBSCRIPTION_EXPIRED",
            title="Subscription expired", message="Your Cauldra trial has ended. Subscribe to a plan to continue using Cauldra.",
            deep_link="subscription", dedup_key=f"sub_expired:{subscription.business_id}",
        )
        db.commit()
    elif subscription.status == "active" and subscription.current_period_end and now >= subscription.current_period_end:
        # A renewal charge should have arrived via webhook before this point. If it
        # hasn't, treat it as a failed/missed payment and start the grace period
        # rather than either granting free access or cutting them off instantly.
        subscription.status = "past_due"
        subscription.payment_status = "failed"
        subscription.grace_period_ends_at = subscription.grace_period_ends_at or (now + timedelta(days=PAYSTACK_GRACE_PERIOD_DAYS))
        # Same dedup_key the explicit Paystack invoice.payment_failed webhook
        # branch uses — whichever of the two detects the failure first wins;
        # the other is automatically a no-op via create_notification's own
        # dedup check, so this can never double-push for one failure episode.
        create_notification(
            db, business_id=subscription.business_id, category="subscription", severity="critical", type="SUBSCRIPTION_PAYMENT_FAILED",
            title="Subscription payment failed",
            message="We couldn't process your Cauldra subscription payment. Update your billing information to prevent service interruption.",
            deep_link="subscription", dedup_key=f"payment_failed:{subscription.business_id}",
        )
        db.commit()
    elif subscription.status == "past_due" and subscription.grace_period_ends_at and now >= subscription.grace_period_ends_at:
        subscription.status = "expired"
        create_notification(
            db, business_id=subscription.business_id, category="subscription", severity="critical", type="SUBSCRIPTION_EXPIRED",
            title="Subscription expired", message="Your Cauldra subscription has expired. Subscribe to a plan to continue using Cauldra.",
            deep_link="subscription", dedup_key=f"sub_expired:{subscription.business_id}",
        )
        db.commit()
    return subscription

def require_subscription_access(db: Session, user: User):
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    subscription = refresh_subscription_status(db, get_or_create_subscription(db, business))
    if subscription.status == "past_due":
        # Grace period: access continues while we wait for the retried/resolved charge.
        return subscription
    if subscription.status not in {"trialing", "active"}:
        messages = {
            "pending_payment_method": "Add a payment method to start your 14-day free trial or subscribe to a plan.",
            "expired": "Your 14-day trial has ended. Subscribe to continue using Cauldra.",
            "cancelled": "Your subscription has been cancelled. Subscribe to continue using Cauldra.",
        }
        raise HTTPException(status_code=402, detail=messages.get(subscription.status, "Your subscription is not currently active. Choose a plan or subscribe to continue."))
    return subscription

def billing_period_for(db: Session, business: BusinessProfile, now: Optional[datetime] = None) -> tuple[datetime, datetime, str]:
    sub = get_or_create_subscription(db, business)
    now = now or datetime.utcnow()
    start = sub.current_period_start or sub.trial_start_at or now
    annual = (sub.billing_interval or "monthly").lower() == "annual"
    months = 12 if annual else 1
    elapsed_months = max(0, (now.year - start.year) * 12 + now.month - start.month)
    period_index = elapsed_months // months
    period_start_month = start.month + period_index * months
    period_start = datetime(start.year + (period_start_month - 1) // 12, (period_start_month - 1) % 12 + 1, 1)
    next_month = period_start.month + months
    period_end = datetime(period_start.year + (next_month - 1) // 12, (next_month - 1) % 12 + 1, 1)
    return period_start, period_end, period_start.strftime("%Y-%m-%d")

def raise_limit_reached(db: Session, business: BusinessProfile, resource: str, current: int, maximum: int):
    plan = subscription_for(db, business)
    current_plan_id = (get_or_create_subscription(db, business).plan or "starter").lower()
    upgrade_id = UPGRADE_PATH.get(current_plan_id, "enterprise")
    upgrade_label = PLAN_CONFIG.get(upgrade_id, PLAN_CONFIG["enterprise"])["label"]
    raise HTTPException(status_code=409, detail=f"You have reached the {plan['label']} {resource} limit of {maximum}. Current usage: {current}/{maximum}. Upgrade to {upgrade_label} for a higher allowance.")

def check_plan_limit(db: Session, business: BusinessProfile, resource: str, current: int, increment: int = 1):
    maximum = subscription_for(db, business).get(resource)
    if maximum is not None and current + increment > maximum:
        raise_limit_reached(db, business, resource.replace("_", " "), current, maximum)

def check_storage_limit(db: Session, business: BusinessProfile, byte_count: int):
    maximum = subscription_for(db, business)["storage_gb"] * 1024 * 1024 * 1024
    current = db.query(func.coalesce(func.sum(StoredUpload.size_bytes), 0)).filter(StoredUpload.business_id == business.id).scalar() or 0
    if current + byte_count > maximum:
        raise_limit_reached(db, business, "storage", round(current / (1024**3), 2), subscription_for(db, business)["storage_gb"])

def record_ai_usage(db: Session, user: User, operation_type: str, success: bool, provider: str, model: str, **metadata):
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    _, _, period = billing_period_for(db, business)
    credits = AI_CREDIT_WEIGHTS[operation_type] if success else 0
    db.add(AIUsageLedger(business_id=user.business_id, user_id=user.id, operation_type=operation_type, credits_consumed=credits,
                         billing_period=period, provider=provider, model=model, success=success,
                         input_tokens=metadata.get("input_tokens"), output_tokens=metadata.get("output_tokens"), estimated_provider_cost=metadata.get("estimated_provider_cost")))
    return credits

def usage_summary(db: Session, business: BusinessProfile) -> dict:
    sub = get_or_create_subscription(db, business)
    plan = subscription_for(db, business); period_start, period_end, period = billing_period_for(db, business)
    used = db.query(func.coalesce(func.sum(AIUsageLedger.credits_consumed), 0)).filter(AIUsageLedger.business_id == business.id, AIUsageLedger.billing_period == period, AIUsageLedger.success == True).scalar() or 0
    included = plan["included_ai_credits"]; overage = max(0, int(used) - included)
    overage_charge = ((overage + plan["ai_overage_unit"] - 1) // plan["ai_overage_unit"]) * plan["ai_overage_price"] if overage else 0
    trial_end = sub.trial_end_at or (datetime.utcnow() + timedelta(days=plan["trial_days"]))
    current_price = plan["annual_price" if (sub.billing_interval or "monthly") == "annual" else "monthly_price"]
    return {"plan": (sub.plan or "starter").lower(), "plan_label": plan["label"], "billing_interval": sub.billing_interval or "monthly",
            "status": sub.status, "billing_period_start": to_utc_iso(period_start), "billing_period_end": to_utc_iso(period_end),
            "trial_active": sub.status == "trialing", "trial_ends_at": to_utc_iso(trial_end),
            "trial_start_at": to_utc_iso(sub.trial_start_at),
            "trial_days_remaining": max(0, (trial_end - datetime.utcnow()).days) if sub.status == "trialing" else 0,
            "included_ai_credits": included, "used_ai_credits": int(used), "remaining_ai_credits": max(0, included-int(used)),
            "overage_credits": overage, "estimated_overage_charge": overage_charge, "ai_warning": "overage" if overage else ("warning" if used >= included * .8 else None),
            # Server-authoritative billing/subscription-management fields — the
            # frontend only ever displays these, it never decides them.
            "current_price_naira": current_price, "currency": "NGN",
            "card_verified": bool(sub.card_verified), "card_last4": sub.card_last4, "card_type": sub.card_type,
            "card_exp_month": sub.card_exp_month, "card_exp_year": sub.card_exp_year,
            "next_billing_at": to_utc_iso(sub.next_billing_at),
            "current_period_end": to_utc_iso(sub.current_period_end),
            "payment_status": sub.payment_status, "cancel_at_period_end": bool(sub.cancel_at_period_end),
            "cancelled_at": to_utc_iso(sub.cancelled_at),
            "grace_period_ends_at": to_utc_iso(sub.grace_period_ends_at),
            # A scheduled downgrade never changes `plan`/`billing_interval` above —
            # those still reflect the currently-active, already-paid plan. This is
            # purely informational for the UI; the server alone decides if/when it
            # actually applies (see the webhook's downgrade-aware renewal branch).
            "pending_downgrade": ({
                "plan": sub.pending_downgrade_plan,
                "plan_label": PLAN_CONFIG.get(sub.pending_downgrade_plan, {}).get("label", sub.pending_downgrade_plan),
                "billing_interval": sub.pending_downgrade_billing_interval,
                "effective_at": to_utc_iso(sub.pending_downgrade_effective_at),
                "requested_at": to_utc_iso(sub.pending_downgrade_requested_at),
            } if sub.pending_downgrade_plan else None),
            "can_start_trial": (not sub.card_verified) and sub.trial_start_at is None and sub.status == "pending_payment_method",
            "plans": {pid: plan_public_view(pid, cfg) for pid, cfg in PLAN_CONFIG.items()}}

# -----------------------------------------------------------------------------
# CANONICAL COUNTRY CONTEXT
# -----------------------------------------------------------------------------
# Registration and Business Profile country settings are authoritative for the
# entire business. Currency, phone code, and country identity are normalized
# server-side so the client cannot accidentally save a mismatched combination.
COUNTRY_CONTEXTS = {
    'AF': {"name": 'Afghanistan', "code": '+93', "currency": 'AFN (؋)', "timezone": 'Asia/Kabul'},
    'AL': {"name": 'Albania', "code": '+355', "currency": 'ALL (ALL)', "timezone": 'Europe/Tirane'},
    'DZ': {"name": 'Algeria', "code": '+213', "currency": 'DZD (DZD)', "timezone": 'Africa/Algiers'},
    'AS': {"name": 'American Samoa', "code": '+1684', "currency": 'USD ($)', "timezone": 'Pacific/Pago_Pago'},
    'AD': {"name": 'Andorra', "code": '+376', "currency": 'EUR (€)', "timezone": 'Europe/Andorra'},
    'AO': {"name": 'Angola', "code": '+244', "currency": 'AOA (Kz)', "timezone": 'Africa/Luanda'},
    'AI': {"name": 'Anguilla', "code": '+1264', "currency": 'XCD ($)', "timezone": 'America/Anguilla'},
    'AG': {"name": 'Antigua and Barbuda', "code": '+1268', "currency": 'XCD ($)', "timezone": 'America/Antigua'},
    'AR': {"name": 'Argentina', "code": '+54', "currency": 'ARS ($)', "timezone": 'America/Argentina/Buenos_Aires'},
    'AM': {"name": 'Armenia', "code": '+374', "currency": 'AMD (֏)', "timezone": 'Asia/Yerevan'},
    'AW': {"name": 'Aruba', "code": '+297', "currency": 'AWG (AWG)', "timezone": 'America/Aruba'},
    'AU': {"name": 'Australia', "code": '+61', "currency": 'AUD ($)', "timezone": 'Australia/Sydney'},
    'AT': {"name": 'Austria', "code": '+43', "currency": 'EUR (€)', "timezone": 'Europe/Vienna'},
    'AZ': {"name": 'Azerbaijan', "code": '+994', "currency": 'AZN (₼)', "timezone": 'Asia/Baku'},
    'BH': {"name": 'Bahrain', "code": '+973', "currency": 'BHD (BHD)', "timezone": 'Asia/Bahrain'},
    'BD': {"name": 'Bangladesh', "code": '+880', "currency": 'BDT (৳)', "timezone": 'Asia/Dhaka'},
    'BB': {"name": 'Barbados', "code": '+1246', "currency": 'BBD ($)', "timezone": 'America/Barbados'},
    'BY': {"name": 'Belarus', "code": '+375', "currency": 'BYN (BYN)', "timezone": 'Europe/Minsk'},
    'BE': {"name": 'Belgium', "code": '+32', "currency": 'EUR (€)', "timezone": 'Europe/Brussels'},
    'BZ': {"name": 'Belize', "code": '+501', "currency": 'BZD ($)', "timezone": 'America/Belize'},
    'BJ': {"name": 'Benin', "code": '+229', "currency": 'XOF (F CFA)', "timezone": 'Africa/Porto-Novo'},
    'BM': {"name": 'Bermuda', "code": '+1441', "currency": 'BMD ($)', "timezone": 'Atlantic/Bermuda'},
    'BT': {"name": 'Bhutan', "code": '+975', "currency": 'INR (₹)', "timezone": 'Asia/Thimphu'},
    'BO': {"name": 'Bolivia', "code": '+591', "currency": 'BOB (Bs)', "timezone": 'America/La_Paz'},
    'BA': {"name": 'Bosnia and Herzegovina', "code": '+387', "currency": 'BAM (KM)', "timezone": 'Europe/Sarajevo'},
    'BW': {"name": 'Botswana', "code": '+267', "currency": 'BWP (P)', "timezone": 'Africa/Gaborone'},
    'BR': {"name": 'Brazil', "code": '+55', "currency": 'BRL (R$)', "timezone": 'America/Sao_Paulo'},
    'IO': {"name": 'British Indian Ocean Territory', "code": '+246', "currency": 'USD ($)', "timezone": 'Indian/Chagos'},
    'BN': {"name": 'Brunei', "code": '+673', "currency": 'BND ($)', "timezone": 'Asia/Brunei'},
    'BG': {"name": 'Bulgaria', "code": '+359', "currency": 'BGN (BGN)', "timezone": 'Europe/Sofia'},
    'BF': {"name": 'Burkina Faso', "code": '+226', "currency": 'XOF (F CFA)', "timezone": 'Africa/Ouagadougou'},
    'BI': {"name": 'Burundi', "code": '+257', "currency": 'BIF (BIF)', "timezone": 'Africa/Bujumbura'},
    'KH': {"name": 'Cambodia', "code": '+855', "currency": 'KHR (៛)', "timezone": 'Asia/Phnom_Penh'},
    'CM': {"name": 'Cameroon', "code": '+237', "currency": 'XAF (FCFA)', "timezone": 'Africa/Douala'},
    'CA': {"name": 'Canada', "code": '+1', "currency": 'CAD ($)', "timezone": 'America/Toronto'},
    'CV': {"name": 'Cape Verde', "code": '+238', "currency": 'CVE (CVE)', "timezone": 'Atlantic/Cape_Verde'},
    'KY': {"name": 'Cayman Islands', "code": '+1345', "currency": 'KYD ($)', "timezone": 'America/Cayman'},
    'CF': {"name": 'Central African Republic', "code": '+236', "currency": 'XAF (FCFA)', "timezone": 'Africa/Bangui'},
    'TD': {"name": 'Chad', "code": '+235', "currency": 'XAF (FCFA)', "timezone": 'Africa/Ndjamena'},
    'CL': {"name": 'Chile', "code": '+56', "currency": 'CLP ($)', "timezone": 'America/Santiago'},
    'CN': {"name": 'China', "code": '+86', "currency": 'CNY (¥)', "timezone": 'Asia/Shanghai'},
    'CX': {"name": 'Christmas Island', "code": '+61', "currency": 'AUD ($)', "timezone": 'Indian/Christmas'},
    'CC': {"name": 'Cocos (Keeling) Islands', "code": '+61', "currency": 'AUD ($)', "timezone": 'Indian/Cocos'},
    'CO': {"name": 'Colombia', "code": '+57', "currency": 'COP ($)', "timezone": 'America/Bogota'},
    'KM': {"name": 'Comoros', "code": '+269', "currency": 'KMF (CF)', "timezone": 'Indian/Comoro'},
    'CK': {"name": 'Cook Islands', "code": '+682', "currency": 'NZD ($)', "timezone": 'Pacific/Rarotonga'},
    'CR': {"name": 'Costa Rica', "code": '+506', "currency": 'CRC (₡)', "timezone": 'America/Costa_Rica'},
    'HR': {"name": 'Croatia', "code": '+385', "currency": 'EUR (€)', "timezone": 'Europe/Zagreb'},
    'CU': {"name": 'Cuba', "code": '+53', "currency": 'CUP ($)', "timezone": 'America/Havana'},
    'CY': {"name": 'Cyprus', "code": '+357', "currency": 'EUR (€)', "timezone": 'Asia/Nicosia'},
    'CZ': {"name": 'Czechia', "code": '+420', "currency": 'CZK (Kč)', "timezone": 'Europe/Prague'},
    'CD': {"name": 'Democratic Republic of the Congo', "code": '+243', "currency": 'CDF (CDF)', "timezone": 'Africa/Kinshasa'},
    'DK': {"name": 'Denmark', "code": '+45', "currency": 'DKK (kr)', "timezone": 'Europe/Copenhagen'},
    'DJ': {"name": 'Djibouti', "code": '+253', "currency": 'DJF (DJF)', "timezone": 'Africa/Djibouti'},
    'DM': {"name": 'Dominica', "code": '+1767', "currency": 'XCD ($)', "timezone": 'America/Dominica'},
    'DO': {"name": 'Dominican Republic', "code": '+1809', "currency": 'DOP ($)', "timezone": 'America/Santo_Domingo'},
    'TL': {"name": 'Timor-Leste', "code": '+670', "currency": 'USD ($)', "timezone": 'Asia/Dili'},
    'EC': {"name": 'Ecuador', "code": '+593', "currency": 'USD ($)', "timezone": 'America/Guayaquil'},
    'EG': {"name": 'Egypt', "code": '+20', "currency": 'EGP (E£)', "timezone": 'Africa/Cairo'},
    'SV': {"name": 'El Salvador', "code": '+503', "currency": 'USD ($)', "timezone": 'America/El_Salvador'},
    'GQ': {"name": 'Equatorial Guinea', "code": '+240', "currency": 'XAF (FCFA)', "timezone": 'Africa/Malabo'},
    'ER': {"name": 'Eritrea', "code": '+291', "currency": 'ERN (ERN)', "timezone": 'Africa/Asmara'},
    'EE': {"name": 'Estonia', "code": '+372', "currency": 'EUR (€)', "timezone": 'Europe/Tallinn'},
    'ET': {"name": 'Ethiopia', "code": '+251', "currency": 'ETB (ETB)', "timezone": 'Africa/Addis_Ababa'},
    'FK': {"name": 'Falkland Islands', "code": '+500', "currency": 'FKP (£)', "timezone": 'Atlantic/Stanley'},
    'FO': {"name": 'Faroe Islands', "code": '+298', "currency": 'DKK (kr)', "timezone": 'Atlantic/Faroe'},
    'FM': {"name": 'Federated States of Micronesia', "code": '+691', "currency": 'USD ($)', "timezone": 'Pacific/Pohnpei'},
    'FJ': {"name": 'Fiji', "code": '+679', "currency": 'FJD ($)', "timezone": 'Pacific/Fiji'},
    'FI': {"name": 'Finland', "code": '+358', "currency": 'EUR (€)', "timezone": 'Europe/Helsinki'},
    'FR': {"name": 'France', "code": '+33', "currency": 'EUR (€)', "timezone": 'Europe/Paris'},
    'GF': {"name": 'French Guiana', "code": '+594', "currency": 'EUR (€)', "timezone": 'America/Cayenne'},
    'PF': {"name": 'French Polynesia', "code": '+689', "currency": 'XPF (CFPF)', "timezone": 'Pacific/Tahiti'},
    'GA': {"name": 'Gabon', "code": '+241', "currency": 'XAF (FCFA)', "timezone": 'Africa/Libreville'},
    'GE': {"name": 'Georgia', "code": '+995', "currency": 'GEL (₾)', "timezone": 'Asia/Tbilisi'},
    'DE': {"name": 'Germany', "code": '+49', "currency": 'EUR (€)', "timezone": 'Europe/Berlin'},
    'GH': {"name": 'Ghana', "code": '+233', "currency": 'GHS (GH₵)', "timezone": 'Africa/Accra'},
    'GI': {"name": 'Gibraltar', "code": '+350', "currency": 'GIP (£)', "timezone": 'Europe/Gibraltar'},
    'GR': {"name": 'Greece', "code": '+30', "currency": 'EUR (€)', "timezone": 'Europe/Athens'},
    'GL': {"name": 'Greenland', "code": '+299', "currency": 'DKK (kr)', "timezone": 'America/Nuuk'},
    'GD': {"name": 'Grenada', "code": '+1473', "currency": 'XCD ($)', "timezone": 'America/Grenada'},
    'GP': {"name": 'Guadeloupe', "code": '+590', "currency": 'EUR (€)', "timezone": 'America/Guadeloupe'},
    'GU': {"name": 'Guam', "code": '+1671', "currency": 'USD ($)', "timezone": 'Pacific/Guam'},
    'GT': {"name": 'Guatemala', "code": '+502', "currency": 'GTQ (Q)', "timezone": 'America/Guatemala'},
    'GG': {"name": 'Guernsey', "code": '+44', "currency": 'GBP (£)', "timezone": 'Europe/Guernsey'},
    'GN': {"name": 'Guinea', "code": '+224', "currency": 'GNF (FG)', "timezone": 'Africa/Conakry'},
    'GW': {"name": 'Guinea-Bissau', "code": '+245', "currency": 'XOF (F CFA)', "timezone": 'Africa/Bissau'},
    'GY': {"name": 'Guyana', "code": '+592', "currency": 'GYD ($)', "timezone": 'America/Guyana'},
    'HT': {"name": 'Haiti', "code": '+509', "currency": 'HTG (HTG)', "timezone": 'America/Port-au-Prince'},
    'HN': {"name": 'Honduras', "code": '+504', "currency": 'HNL (L)', "timezone": 'America/Tegucigalpa'},
    'HK': {"name": 'Hong Kong', "code": '+852', "currency": 'HKD ($)', "timezone": 'Asia/Hong_Kong'},
    'HU': {"name": 'Hungary', "code": '+36', "currency": 'HUF (Ft)', "timezone": 'Europe/Budapest'},
    'IS': {"name": 'Iceland', "code": '+354', "currency": 'ISK (kr)', "timezone": 'Atlantic/Reykjavik'},
    'IN': {"name": 'India', "code": '+91', "currency": 'INR (₹)', "timezone": 'Asia/Kolkata'},
    'ID': {"name": 'Indonesia', "code": '+62', "currency": 'IDR (Rp)', "timezone": 'Asia/Jakarta'},
    'IR': {"name": 'Iran', "code": '+98', "currency": 'IRR (IRR)', "timezone": 'Asia/Tehran'},
    'IQ': {"name": 'Iraq', "code": '+964', "currency": 'IQD (IQD)', "timezone": 'Asia/Baghdad'},
    'IE': {"name": 'Ireland', "code": '+353', "currency": 'EUR (€)', "timezone": 'Europe/Dublin'},
    'IM': {"name": 'Isle of Man', "code": '+44', "currency": 'GBP (£)', "timezone": 'Europe/Isle_of_Man'},
    'IL': {"name": 'Israel', "code": '+972', "currency": 'ILS (₪)', "timezone": 'Asia/Jerusalem'},
    'IT': {"name": 'Italy', "code": '+39', "currency": 'EUR (€)', "timezone": 'Europe/Rome'},
    'CI': {"name": 'Côte d’Ivoire', "code": '+225', "currency": 'XOF (F CFA)', "timezone": 'Africa/Abidjan'},
    'JM': {"name": 'Jamaica', "code": '+1876', "currency": 'JMD ($)', "timezone": 'America/Jamaica'},
    'JP': {"name": 'Japan', "code": '+81', "currency": 'JPY (¥)', "timezone": 'Asia/Tokyo'},
    'JE': {"name": 'Jersey', "code": '+44', "currency": 'GBP (£)', "timezone": 'Europe/Jersey'},
    'JO': {"name": 'Jordan', "code": '+962', "currency": 'JOD (JOD)', "timezone": 'Asia/Amman'},
    'KZ': {"name": 'Kazakhstan', "code": '+76', "currency": 'KZT (₸)', "timezone": 'Asia/Almaty'},
    'KE': {"name": 'Kenya', "code": '+254', "currency": 'KES (KES)', "timezone": 'Africa/Nairobi'},
    'KI': {"name": 'Kiribati', "code": '+686', "currency": 'AUD ($)', "timezone": 'Pacific/Tarawa'},
    'XK': {"name": 'Kosovo', "code": '+383', "currency": 'EUR (€)', "timezone": 'Europe/Pristina'},
    'KW': {"name": 'Kuwait', "code": '+965', "currency": 'KWD (KWD)', "timezone": 'Asia/Kuwait'},
    'KG': {"name": 'Kyrgyzstan', "code": '+996', "currency": 'KGS (⃀)', "timezone": 'Asia/Bishkek'},
    'LA': {"name": 'Laos', "code": '+856', "currency": 'LAK (₭)', "timezone": 'Asia/Vientiane'},
    'LV': {"name": 'Latvia', "code": '+371', "currency": 'EUR (€)', "timezone": 'Europe/Riga'},
    'LB': {"name": 'Lebanon', "code": '+961', "currency": 'LBP (L£)', "timezone": 'Asia/Beirut'},
    'LS': {"name": 'Lesotho', "code": '+266', "currency": 'ZAR (R)', "timezone": 'Africa/Maseru'},
    'LR': {"name": 'Liberia', "code": '+231', "currency": 'LRD ($)', "timezone": 'Africa/Monrovia'},
    'LY': {"name": 'Libya', "code": '+218', "currency": 'LYD (LYD)', "timezone": 'Africa/Tripoli'},
    'LI': {"name": 'Liechtenstein', "code": '+423', "currency": 'CHF (CHF)', "timezone": 'Europe/Vaduz'},
    'LT': {"name": 'Lithuania', "code": '+370', "currency": 'EUR (€)', "timezone": 'Europe/Vilnius'},
    'LU': {"name": 'Luxembourg', "code": '+352', "currency": 'EUR (€)', "timezone": 'Europe/Luxembourg'},
    'MO': {"name": 'Macau', "code": '+853', "currency": 'MOP (MOP)', "timezone": 'Asia/Macau'},
    'MG': {"name": 'Madagascar', "code": '+261', "currency": 'MGA (Ar)', "timezone": 'Indian/Antananarivo'},
    'MW': {"name": 'Malawi', "code": '+265', "currency": 'MWK (MWK)', "timezone": 'Africa/Blantyre'},
    'MY': {"name": 'Malaysia', "code": '+60', "currency": 'MYR (RM)', "timezone": 'Asia/Kuala_Lumpur'},
    'MV': {"name": 'Maldives', "code": '+960', "currency": 'MVR (MVR)', "timezone": 'Indian/Maldives'},
    'ML': {"name": 'Mali', "code": '+223', "currency": 'XOF (F CFA)', "timezone": 'Africa/Bamako'},
    'MT': {"name": 'Malta', "code": '+356', "currency": 'EUR (€)', "timezone": 'Europe/Malta'},
    'MH': {"name": 'Marshall Islands', "code": '+692', "currency": 'USD ($)', "timezone": 'Pacific/Majuro'},
    'MQ': {"name": 'Martinique', "code": '+596', "currency": 'EUR (€)', "timezone": 'America/Martinique'},
    'MR': {"name": 'Mauritania', "code": '+222', "currency": 'MRU (MRU)', "timezone": 'Africa/Nouakchott'},
    'MU': {"name": 'Mauritius', "code": '+230', "currency": 'MUR (Rs)', "timezone": 'Indian/Mauritius'},
    'YT': {"name": 'Mayotte', "code": '+262', "currency": 'EUR (€)', "timezone": 'Indian/Mayotte'},
    'MX': {"name": 'Mexico', "code": '+52', "currency": 'MXN ($)', "timezone": 'America/Mexico_City'},
    'MD': {"name": 'Moldova', "code": '+373', "currency": 'MDL (MDL)', "timezone": 'Europe/Chisinau'},
    'MC': {"name": 'Monaco', "code": '+377', "currency": 'EUR (€)', "timezone": 'Europe/Monaco'},
    'MN': {"name": 'Mongolia', "code": '+976', "currency": 'MNT (₮)', "timezone": 'Asia/Ulaanbaatar'},
    'ME': {"name": 'Montenegro', "code": '+382', "currency": 'EUR (€)', "timezone": 'Europe/Podgorica'},
    'MS': {"name": 'Montserrat', "code": '+1664', "currency": 'XCD ($)', "timezone": 'America/Montserrat'},
    'MA': {"name": 'Morocco', "code": '+212', "currency": 'MAD (MAD)', "timezone": 'Africa/Casablanca'},
    'MZ': {"name": 'Mozambique', "code": '+258', "currency": 'MZN (MZN)', "timezone": 'Africa/Maputo'},
    'MM': {"name": 'Myanmar', "code": '+95', "currency": 'MMK (K)', "timezone": 'Asia/Yangon'},
    'NA': {"name": 'Namibia', "code": '+264', "currency": 'ZAR (R)', "timezone": 'Africa/Windhoek'},
    'NR': {"name": 'Nauru', "code": '+674', "currency": 'AUD ($)', "timezone": 'Pacific/Nauru'},
    'NP': {"name": 'Nepal', "code": '+977', "currency": 'NPR (Rs)', "timezone": 'Asia/Kathmandu'},
    'NL': {"name": 'Netherlands', "code": '+31', "currency": 'EUR (€)', "timezone": 'Europe/Amsterdam'},
    'NC': {"name": 'New Caledonia', "code": '+687', "currency": 'XPF (CFPF)', "timezone": 'Pacific/Noumea'},
    'NZ': {"name": 'New Zealand', "code": '+64', "currency": 'NZD ($)', "timezone": 'Pacific/Auckland'},
    'NI': {"name": 'Nicaragua', "code": '+505', "currency": 'NIO (C$)', "timezone": 'America/Managua'},
    'NE': {"name": 'Niger', "code": '+227', "currency": 'XOF (F CFA)', "timezone": 'Africa/Niamey'},
    'NG': {"name": 'Nigeria', "code": '+234', "currency": 'NGN (₦)', "timezone": 'Africa/Lagos'},
    'NU': {"name": 'Niue', "code": '+683', "currency": 'NZD ($)', "timezone": 'Pacific/Niue'},
    'NF': {"name": 'Norfolk Island', "code": '+672', "currency": 'AUD ($)', "timezone": 'Pacific/Norfolk'},
    'KP': {"name": 'North Korea', "code": '+850', "currency": 'KPW (₩)', "timezone": 'Asia/Pyongyang'},
    'MP': {"name": 'Northern Mariana Islands', "code": '+1670', "currency": 'USD ($)', "timezone": 'Pacific/Saipan'},
    'NO': {"name": 'Norway', "code": '+47', "currency": 'NOK (kr)', "timezone": 'Europe/Oslo'},
    'OM': {"name": 'Oman', "code": '+968', "currency": 'OMR (OMR)', "timezone": 'Asia/Muscat'},
    'PK': {"name": 'Pakistan', "code": '+92', "currency": 'PKR (Rs)', "timezone": 'Asia/Karachi'},
    'PW': {"name": 'Palau', "code": '+680', "currency": 'USD ($)', "timezone": 'Pacific/Palau'},
    'PS': {"name": 'Palestine', "code": '+970', "currency": 'ILS (₪)', "timezone": 'Asia/Gaza'},
    'PA': {"name": 'Panama', "code": '+507', "currency": 'PAB (PAB)', "timezone": 'America/Panama'},
    'PG': {"name": 'Papua New Guinea', "code": '+675', "currency": 'PGK (PGK)', "timezone": 'Pacific/Port_Moresby'},
    'PY': {"name": 'Paraguay', "code": '+595', "currency": 'PYG (₲)', "timezone": 'America/Asuncion'},
    'PE': {"name": 'Peru', "code": '+51', "currency": 'PEN (PEN)', "timezone": 'America/Lima'},
    'PH': {"name": 'Philippines', "code": '+63', "currency": 'PHP (₱)', "timezone": 'Asia/Manila'},
    'PN': {"name": 'Pitcairn Islands', "code": '+64', "currency": 'NZD ($)', "timezone": 'Pacific/Pitcairn'},
    'PL': {"name": 'Poland', "code": '+48', "currency": 'PLN (zł)', "timezone": 'Europe/Warsaw'},
    'PT': {"name": 'Portugal', "code": '+351', "currency": 'EUR (€)', "timezone": 'Europe/Lisbon'},
    'PR': {"name": 'Puerto Rico', "code": '+1787', "currency": 'USD ($)', "timezone": 'America/Puerto_Rico'},
    'QA': {"name": 'Qatar', "code": '+974', "currency": 'QAR (QAR)', "timezone": 'Asia/Qatar'},
    'MK': {"name": 'North Macedonia', "code": '+389', "currency": 'MKD (MKD)', "timezone": 'Europe/Skopje'},
    'CG': {"name": 'Republic of the Congo', "code": '+242', "currency": 'XAF (FCFA)', "timezone": 'Africa/Brazzaville'},
    'RO': {"name": 'Romania', "code": '+40', "currency": 'RON (lei)', "timezone": 'Europe/Bucharest'},
    'RU': {"name": 'Russia', "code": '+7', "currency": 'RUB (₽)', "timezone": 'Europe/Moscow'},
    'RW': {"name": 'Rwanda', "code": '+250', "currency": 'RWF (RF)', "timezone": 'Africa/Kigali'},
    'RE': {"name": 'Réunion', "code": '+262', "currency": 'EUR (€)', "timezone": 'Indian/Reunion'},
    'SH': {"name": 'Saint Helena', "code": '+290', "currency": 'SHP (£)', "timezone": 'Atlantic/St_Helena'},
    'KN': {"name": 'Saint Kitts and Nevis', "code": '+1869', "currency": 'XCD ($)', "timezone": 'America/St_Kitts'},
    'LC': {"name": 'Saint Lucia', "code": '+1758', "currency": 'XCD ($)', "timezone": 'America/St_Lucia'},
    'PM': {"name": 'Saint Pierre and Miquelon', "code": '+508', "currency": 'EUR (€)', "timezone": 'America/Miquelon'},
    'VC': {"name": 'Saint Vincent and the Grenadines', "code": '+1784', "currency": 'XCD ($)', "timezone": 'America/St_Vincent'},
    'WS': {"name": 'Samoa', "code": '+685', "currency": 'WST (WST)', "timezone": 'Pacific/Apia'},
    'SM': {"name": 'San Marino', "code": '+378', "currency": 'EUR (€)', "timezone": 'Europe/San_Marino'},
    'SA': {"name": 'Saudi Arabia', "code": '+966', "currency": 'SAR (SAR)', "timezone": 'Asia/Riyadh'},
    'SN': {"name": 'Senegal', "code": '+221', "currency": 'XOF (F CFA)', "timezone": 'Africa/Dakar'},
    'RS': {"name": 'Serbia', "code": '+381', "currency": 'RSD (RSD)', "timezone": 'Europe/Belgrade'},
    'SC': {"name": 'Seychelles', "code": '+248', "currency": 'SCR (SCR)', "timezone": 'Indian/Mahe'},
    'SL': {"name": 'Sierra Leone', "code": '+232', "currency": 'SLE (SLE)', "timezone": 'Africa/Freetown'},
    'SG': {"name": 'Singapore', "code": '+65', "currency": 'SGD ($)', "timezone": 'Asia/Singapore'},
    'SK': {"name": 'Slovakia', "code": '+421', "currency": 'EUR (€)', "timezone": 'Europe/Bratislava'},
    'SI': {"name": 'Slovenia', "code": '+386', "currency": 'EUR (€)', "timezone": 'Europe/Ljubljana'},
    'SB': {"name": 'Solomon Islands', "code": '+677', "currency": 'SBD ($)', "timezone": 'Pacific/Guadalcanal'},
    'SO': {"name": 'Somalia', "code": '+252', "currency": 'SOS (SOS)', "timezone": 'Africa/Mogadishu'},
    'ZA': {"name": 'South Africa', "code": '+27', "currency": 'ZAR (R)', "timezone": 'Africa/Johannesburg'},
    'GS': {"name": 'South Georgia', "code": '+500', "currency": 'GBP (£)', "timezone": 'Atlantic/South_Georgia'},
    'KR': {"name": 'South Korea', "code": '+82', "currency": 'KRW (₩)', "timezone": 'Asia/Seoul'},
    'SS': {"name": 'South Sudan', "code": '+211', "currency": 'SSP (£)', "timezone": 'Africa/Juba'},
    'ES': {"name": 'Spain', "code": '+34', "currency": 'EUR (€)', "timezone": 'Europe/Madrid'},
    'LK': {"name": 'Sri Lanka', "code": '+94', "currency": 'LKR (Rs)', "timezone": 'Asia/Colombo'},
    'SD': {"name": 'Sudan', "code": '+249', "currency": 'SDG (SDG)', "timezone": 'Africa/Khartoum'},
    'SR': {"name": 'Suriname', "code": '+597', "currency": 'SRD ($)', "timezone": 'America/Paramaribo'},
    'SJ': {"name": 'Svalbard and Jan Mayen', "code": '+4779', "currency": 'NOK (kr)', "timezone": 'Arctic/Longyearbyen'},
    'SZ': {"name": 'Eswatini', "code": '+268', "currency": 'SZL (SZL)', "timezone": 'Africa/Mbabane'},
    'SE': {"name": 'Sweden', "code": '+46', "currency": 'SEK (kr)', "timezone": 'Europe/Stockholm'},
    'CH': {"name": 'Switzerland', "code": '+41', "currency": 'CHF (CHF)', "timezone": 'Europe/Zurich'},
    'SY': {"name": 'Syria', "code": '+963', "currency": 'SYP (£)', "timezone": 'Asia/Damascus'},
    'ST': {"name": 'São Tomé and Príncipe', "code": '+239', "currency": 'STN (Db)', "timezone": 'Africa/Sao_Tome'},
    'TW': {"name": 'Taiwan', "code": '+886', "currency": 'TWD ($)', "timezone": 'Asia/Taipei'},
    'TJ': {"name": 'Tajikistan', "code": '+992', "currency": 'TJS (TJS)', "timezone": 'Asia/Dushanbe'},
    'TZ': {"name": 'Tanzania', "code": '+255', "currency": 'TZS (TZS)', "timezone": 'Africa/Dar_es_Salaam'},
    'TH': {"name": 'Thailand', "code": '+66', "currency": 'THB (฿)', "timezone": 'Asia/Bangkok'},
    'BS': {"name": 'Bahamas', "code": '+1242', "currency": 'BSD ($)', "timezone": 'America/Nassau'},
    'GM': {"name": 'Gambia', "code": '+220', "currency": 'GMD (GMD)', "timezone": 'Africa/Banjul'},
    'TG': {"name": 'Togo', "code": '+228', "currency": 'XOF (F CFA)', "timezone": 'Africa/Lome'},
    'TK': {"name": 'Tokelau', "code": '+690', "currency": 'NZD ($)', "timezone": 'Pacific/Fakaofo'},
    'TO': {"name": 'Tonga', "code": '+676', "currency": 'TOP (T$)', "timezone": 'Pacific/Tongatapu'},
    'TT': {"name": 'Trinidad and Tobago', "code": '+1868', "currency": 'TTD ($)', "timezone": 'America/Port_of_Spain'},
    'TN': {"name": 'Tunisia', "code": '+216', "currency": 'TND (TND)', "timezone": 'Africa/Tunis'},
    'TR': {"name": 'Türkiye', "code": '+90', "currency": 'TRY (₺)', "timezone": 'Europe/Istanbul'},
    'TM': {"name": 'Turkmenistan', "code": '+993', "currency": 'TMT (TMT)', "timezone": 'Asia/Ashgabat'},
    'TV': {"name": 'Tuvalu', "code": '+688', "currency": 'AUD ($)', "timezone": 'Pacific/Funafuti'},
    'UG': {"name": 'Uganda', "code": '+256', "currency": 'UGX (UGX)', "timezone": 'Africa/Kampala'},
    'UA': {"name": 'Ukraine', "code": '+380', "currency": 'UAH (₴)', "timezone": 'Europe/Kyiv'},
    'AE': {"name": 'United Arab Emirates', "code": '+971', "currency": 'AED (AED)', "timezone": 'Asia/Dubai'},
    'GB': {"name": 'United Kingdom', "code": '+44', "currency": 'GBP (£)', "timezone": 'Europe/London'},
    'US': {"name": 'United States', "code": '+1', "currency": 'USD ($)', "timezone": 'America/New_York'},
    'UY': {"name": 'Uruguay', "code": '+598', "currency": 'UYU ($)', "timezone": 'America/Montevideo'},
    'UZ': {"name": 'Uzbekistan', "code": '+998', "currency": 'UZS (UZS)', "timezone": 'Asia/Tashkent'},
    'VU': {"name": 'Vanuatu', "code": '+678', "currency": 'VUV (VUV)', "timezone": 'Pacific/Efate'},
    'VA': {"name": 'Vatican City', "code": '+39', "currency": 'EUR (€)', "timezone": 'Europe/Vatican'},
    'VE': {"name": 'Venezuela', "code": '+58', "currency": 'VES (VES)', "timezone": 'America/Caracas'},
    'VN': {"name": 'Vietnam', "code": '+84', "currency": 'VND (₫)', "timezone": 'Asia/Ho_Chi_Minh'},
    'WF': {"name": 'Wallis and Futuna', "code": '+681', "currency": 'XPF (CFPF)', "timezone": 'Pacific/Wallis'},
    'EH': {"name": 'Western Sahara', "code": '+212', "currency": 'MAD (MAD)', "timezone": 'Africa/El_Aaiun'},
    'YE': {"name": 'Yemen', "code": '+967', "currency": 'YER (YER)', "timezone": 'Asia/Aden'},
    'ZM': {"name": 'Zambia', "code": '+260', "currency": 'ZMW (ZK)', "timezone": 'Africa/Lusaka'},
    'ZW': {"name": 'Zimbabwe', "code": '+263', "currency": 'USD ($)', "timezone": 'Africa/Harare'},
}

def canonical_country_context(country: Optional[str], country_code: Optional[str]) -> dict:
    code = str(country_code or "").strip().upper()
    if code and code in COUNTRY_CONTEXTS:
        return COUNTRY_CONTEXTS[code]
    name = str(country or "").strip().casefold()
    for ctx in COUNTRY_CONTEXTS.values():
        if ctx["name"].casefold() == name:
            return ctx
    raise HTTPException(status_code=400, detail="Please select a valid country or region.")

def canonical_phone_for_country(raw: Optional[str], phone_country_code: str) -> str:
    digits = re.sub(r"\D", "", str(raw or ""))
    prefix_digits = re.sub(r"\D", "", phone_country_code or "")
    if not digits or not prefix_digits:
        return str(raw or "").strip()
    if digits.startswith(prefix_digits):
        local = digits[len(prefix_digits):]
    else:
        local = digits.lstrip("0")
    return f"{phone_country_code}{local}" if local else phone_country_code

def canonical_business_country_values(country: Optional[str], country_code: Optional[str], language: Optional[str] = "en") -> dict:
    ctx = canonical_country_context(country, country_code)
    lang = str(language or "en").strip().lower() or "en"
    iso2 = country_code.strip().upper() if country_code and country_code.strip().upper() in COUNTRY_CONTEXTS else next(k for k,v in COUNTRY_CONTEXTS.items() if v["name"] == ctx["name"])
    # Keep language independent. The country supplies the region component.
    return {
        "country": ctx["name"],
        "country_code": iso2,
        "currency": ctx["currency"],
        "phone_country_code": ctx["code"],
        "timezone": ctx.get("timezone") or "UTC",
        "locale": f"{lang}-{iso2}",
    }

# -----------------------------------------------------------------------------
# PROVIDERS
# -----------------------------------------------------------------------------
openai_client = None
if OpenAI and os.getenv("OPENAI_API_KEY"):
    openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

gemini_client = None
if genai and os.getenv("GEMINI_API_KEY"):
    gemini_client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

# -----------------------------------------------------------------------------
# DEPENDENCIES / HELPERS
# -----------------------------------------------------------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def normalize_username(value: str) -> str:
    return " ".join((value or "").strip().split()).casefold()

def normalize_phone(value: str) -> str:
    return re.sub(r"\D", "", value or "")

def normalize_barcode(value: Optional[str]) -> Optional[str]:
    """Canonical form for a UPC/EAN/GTIN product barcode.

    Deliberately its OWN function rather than a call to normalize_phone
    above: phone numbers and product barcodes are unrelated identifiers with
    unrelated validity rules, and reusing phone normalization here would tie
    barcode behavior to whatever phone-specific changes happen later. Used
    consistently everywhere a barcode is stored or matched — Product.barcode,
    GeneralCatalog.barcode, the duplicate-in-inventory check, the General
    Catalog lookup, the UPCitemdb request, and POS scan matching — so the
    same physical barcode always normalizes to the same string.

    Returns None (never "") for anything that isn't a plausible barcode, so
    every caller can use a plain `if not barcode:` check. Plausible = 6-14
    digits after stripping non-digit characters (scanners/typed input can
    include stray dashes/spaces) — covering UPC-E(6)/UPC-A(12)/EAN-8(8)/
    EAN-13(13)/GTIN-14(14) without accepting arbitrary short numeric text as
    if it were a real barcode.
    """
    digits = re.sub(r"\D", "", value or "")
    if len(digits) < 6 or len(digits) > 14:
        return None
    return digits

def phones_match(submitted: str, stored: str) -> bool:
    """Digits-only comparison used only to VERIFY an already-authenticated
    account's identity for SMS recovery — the submitted value is never used as
    the send destination. Country-code prefixes can be entered inconsistently
    (with/without '+', with/without the leading 0), so we accept an exact
    digit match or a suffix match of reasonable length, rather than requiring
    byte-for-byte formatting to line up."""
    a = normalize_phone(submitted)
    b = normalize_phone(stored)
    if not a or not b:
        return False
    if a == b:
        return True
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    return len(shorter) >= 7 and longer.endswith(shorter)

def hash_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()

def mutation_payload_hash(payload: dict) -> str:
    return hash_text(json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str))

def claim_idempotent_mutation(
    db: Session, business_id: int, operation: str, client_ref: Optional[str], payload: dict,
) -> tuple[Optional[MutationIdempotency], Optional[dict]]:
    """Atomically claim a mutation or return its previously committed result."""
    if not client_ref:
        return None, None
    request_hash = mutation_payload_hash(payload)

    def resolve_existing() -> tuple[MutationIdempotency, Optional[dict]]:
        existing = db.query(MutationIdempotency).filter(
            MutationIdempotency.business_id == business_id,
            MutationIdempotency.operation == operation,
            MutationIdempotency.client_ref == client_ref,
        ).first()
        if not existing:
            raise HTTPException(status_code=409, detail="This change is already being processed.")
        if not hmac.compare_digest(existing.request_hash, request_hash):
            raise HTTPException(status_code=409, detail="This request key was already used for different data.")
        if existing.status != "completed" or not existing.response_json:
            raise HTTPException(status_code=409, detail="This change is already being processed.")
        try:
            response = json.loads(existing.response_json)
        except Exception:
            raise HTTPException(status_code=409, detail="The stored result for this change is unavailable.")
        response["duplicate"] = True
        return existing, response

    existing = db.query(MutationIdempotency).filter(
        MutationIdempotency.business_id == business_id,
        MutationIdempotency.operation == operation,
        MutationIdempotency.client_ref == client_ref,
    ).first()
    if existing:
        return resolve_existing()

    claim = MutationIdempotency(
        business_id=business_id, operation=operation, client_ref=client_ref,
        request_hash=request_hash, status="processing",
    )
    db.add(claim)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        return resolve_existing()
    return claim, None

def complete_idempotent_mutation(claim: Optional[MutationIdempotency], response: dict):
    if not claim:
        return
    claim.status = "completed"
    claim.response_json = json.dumps(response, sort_keys=True, separators=(",", ":"), default=str)
    claim.completed_at = datetime.utcnow()

def decode_base64_upload(file_data: str, allowed_content_types: set[str]) -> tuple[bytes, str]:
    """Decode a browser data URL without trusting its name or MIME declaration."""
    if not isinstance(file_data, str) or not file_data.startswith("data:") or "," not in file_data:
        raise HTTPException(status_code=422, detail="The uploaded file could not be read.")
    header, encoded = file_data.split(",", 1)
    match = re.fullmatch(r"data:([A-Za-z0-9.+/-]+);base64", header.strip(), flags=re.IGNORECASE)
    if not match:
        raise HTTPException(status_code=422, detail="The uploaded file format is not supported.")
    content_type = match.group(1).lower()
    if content_type not in allowed_content_types:
        raise HTTPException(status_code=422, detail="This file type is not allowed.")
    try:
        raw = base64.b64decode(encoded, validate=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="The uploaded file could not be decoded.") from exc
    if not raw or len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=422, detail=f"Upload must be between 1 byte and {MAX_UPLOAD_BYTES} bytes.")
    return raw, content_type

def safe_upload_name(value: str, fallback: str) -> str:
    name = Path(str(value or "")).name.strip()
    name = re.sub(r"[^A-Za-z0-9._ -]+", "_", name).strip(" .")
    return (name or fallback)[:180]

def persist_upload(db: Session, user: User, kind: str, original_name: str, content_type: str, raw: bytes) -> StoredUpload:
    """Save bytes privately and create an auditable, business-scoped record."""
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    check_storage_limit(db, business, len(raw))
    suffix = Path(original_name).suffix.lower()
    row = StoredUpload(
        business_id=user.business_id,
        uploaded_by_id=user.id,
        kind=kind,
        original_name=safe_upload_name(original_name, f"{kind}-upload"),
        storage_key="pending",
        content_type=content_type,
        size_bytes=len(raw),
        content_hash=hashlib.sha256(raw).hexdigest(),
    )
    db.add(row)
    db.flush()
    key = f"{user.business_id}/{row.id}-{secrets.token_urlsafe(18)}{suffix}"
    target = (UPLOAD_STORAGE_DIR / key).resolve()
    if UPLOAD_STORAGE_DIR not in target.parents:
        raise HTTPException(status_code=500, detail="The upload storage location is invalid.")
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(target.suffix + ".tmp")
    try:
        temporary.write_bytes(raw)
        os.replace(temporary, target)
    except OSError as exc:
        temporary.unlink(missing_ok=True)
        raise HTTPException(status_code=503, detail="The server could not securely store the uploaded file.") from exc
    row.storage_key = key
    return row

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(password: str, hashed: str) -> bool:
    # --- TEMPORARY DEV DIAGNOSTIC (auth investigation) -----------------------
    # Distinguishes "pwd_context.verify() returned False" from "pwd_context.verify()
    # raised an exception that got silently swallowed" — the two look identical
    # to the caller today. Logs no password, no hash, no token. Safe to remove
    # once the real-browser login failure is diagnosed.
    try:
        return pwd_context.verify(password, hashed)
    except Exception as exc:
        print(
            f"[auth-diag] AUTH PASSWORD VERIFICATION EXCEPTION "
            f"exception_type={type(exc).__name__} exception_message={exc!r} "
            f"hash_present={bool(hashed)} hash_prefix={(hashed or '')[:4]!r} "
            f"pid={os.getpid()}"
        )
        return False
    # --- END TEMPORARY DEV DIAGNOSTIC ----------------------------------------

def validate_password_strength(password: str):
    if len(password) > 72:
        raise HTTPException(status_code=400, detail="Password must be 72 characters or fewer.")
    if len(password) < 8 or not re.search(r"[A-Z]", password) or not re.search(r"[a-z]", password) or not re.search(r"\d", password):
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters and include uppercase, lowercase, and a number.")

def generate_dynamic_business_code(db: Session, business_name: str) -> str:
    words = business_name.strip().split()
    initials = ((words[0][0] + words[1][0]) if len(words) >= 2 else (words[0][:2] if words else "BX")).upper()
    for _ in range(50):
        code = f"{initials}-{random.randint(1000,9999)}-{random.randint(10,99)}"
        if not db.query(BusinessProfile).filter(BusinessProfile.business_code == code).first():
            return code
    raise HTTPException(status_code=500, detail="We could not create a unique Business ID. Please try again.")

# -----------------------------------------------------------------------------
# CENTRAL FINANCIAL CALCULATION LAYER
# -----------------------------------------------------------------------------
# Single source of truth for revenue / COGS / gross profit / expenses / net
# profit for any date range. Sales History period totals, Expense History
# period totals, the Dashboard, the Profit page, and reports must all call
# financial_summary_for_period() (or resolve_financial_period() +
# compute_financial_summary() directly) rather than independently
# recalculating anything — that is what guarantees they always agree.
FINANCIAL_PERIODS = {"today", "yesterday", "week", "previous_week", "month", "previous_month", "year", "previous_year", "custom", "all"}

def business_local_zoneinfo(business: "BusinessProfile") -> tzinfo:
    # ZoneInfo(...) looks up the IANA tz database on disk (system package, or
    # the "tzdata" pip package as a fallback) — it is NOT self-contained. A
    # slim/minimal deployment image without that database raises
    # ZoneInfoNotFoundError for ANY key, including "UTC" (it is not
    # hardcoded — Python looks it up the same way as every other zone). If
    # that fallback timezone.utc call also throws, it would 500 every
    # request that reaches this function. datetime.timezone.utc is a stdlib
    # constant that can never fail to resolve, so it's used here as the
    # last-resort fallback to guarantee the app degrades to UTC display
    # instead of crashing. If businesses are seeing UTC instead of their
    # real configured timezone, install the "tzdata" package (pip install
    # tzdata) in the deployment environment — that's the actual fix for
    # correct (not just non-crashing) business-local times.
    try:
        return ZoneInfo(business.timezone or "UTC") if business else timezone.utc
    except Exception:
        return timezone.utc

def to_utc_iso(dt: Optional[datetime]) -> Optional[str]:
    """Serialize a real point-in-time datetime for the frontend.

    Every timestamp in this app is stored as a NAIVE datetime that is
    already UTC (datetime.utcnow(), no tzinfo). Calling bare .isoformat()
    on that produces a string with no timezone marker at all (e.g.
    "2026-08-22T07:40:12"). Per the JS Date Time String Format spec, a
    browser's `new Date(...)` treats a marker-less string as LOCAL time —
    so a device already set to the business's own timezone (the common
    case) ends up reinterpreting an already-UTC value as if it were
    already local time, silently dropping the UTC offset entirely (this
    is exactly the "Opened 07:40 instead of 08:40" bug). Appending "Z"
    makes the instant unambiguous so any timezone-aware formatter
    (formatBusinessTime/DateTime on the frontend) can convert it to the
    business's configured timezone correctly. Use this for any real
    point-in-time DateTime field being sent to a client — NOT for plain
    calendar dates (e.g. product expiry_date, BusinessDay.date), which
    have no time-of-day component and must not be reinterpreted as a
    UTC instant."""
    if dt is None:
        return None
    return dt.isoformat() + "Z"

# -----------------------------------------------------------------------------
# CSV EXPORT — one shared serializer for every backend export endpoint.
# -----------------------------------------------------------------------------
def _csv_sanitize_cell(value) -> str:
    """Neutralizes spreadsheet formula-injection triggers so opening an
    exported CSV in Excel/Sheets/Numbers can never execute a formula from a
    business's own data (a product name, note, or category that happens to
    start with '=', '+', '-', or '@'). Mirrors the frontend's
    sanitizeCsvCell() exactly, so client-side and server-side exports apply
    identical protection. An ordinary negative/signed NUMBER (e.g. -42.5, a
    real amount_delta) is deliberately left untouched — only a leading
    +/-/=/@ on a value that is NOT actually numeric is treated as risky."""
    if value is None:
        text = ""
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    else:
        text = str(value)
    trimmed = text.strip()
    if trimmed and trimmed[0] in ("=", "+", "-", "@", "\t", "\r"):
        try:
            float(trimmed)
            is_plain_number = True
        except ValueError:
            is_plain_number = False
        if not is_plain_number:
            text = "'" + text
    return text

def safe_csv_filename(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", name or "").strip("_") or "cauldra_export.csv"

def build_csv_response(filename: str, header: List[str], rows: List[list]) -> Response:
    """The one place every export endpoint turns (header, rows) into an
    HTTP response — escaping (via Python's csv module), formula-injection
    sanitization, the UTF-8 BOM Excel needs to render non-ASCII text
    correctly, and the download headers are always handled identically
    here, so no individual endpoint re-implements CSV serialization."""
    buffer = io.StringIO()
    buffer.write("﻿")  # UTF-8 BOM
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow([_csv_sanitize_cell(h) for h in header])
    for row in rows:
        writer.writerow([_csv_sanitize_cell(c) for c in row])
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_csv_filename(filename)}"'},
    )

# -----------------------------------------------------------------------------
# EXCEL (.xlsx) EXPORT — one shared workbook builder for every export.
#
# CSV stays the plain, universal, data-only export (see above). This builds
# the "polished business report" counterpart: real typed cells, a styled
# header row, sensible column widths, wrapped long text, AutoFilter, a
# frozen header row, and — where a feature naturally has more than one
# related table (e.g. Profit's summary + expense breakdown + top products)
# — multiple worksheets, all from the exact same underlying data every CSV
# export already uses (see each /export endpoint and the frontend's
# buildXCsvSpec()-style functions: one spec, two renderers).
# -----------------------------------------------------------------------------
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSX_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Cauldra dark-slate
XLSX_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
XLSX_TITLE_FONT = Font(bold=True, size=14, color="1E293B")
XLSX_META_FONT = Font(italic=True, size=9, color="64748B")
XLSX_NUMBER_FORMATS = {
    "number": "#,##0",
    "decimal": "0.00",
    "currency": "#,##0.00",
    "percent": "0.00%",
    "date": "yyyy-mm-dd",
    "datetime": "yyyy-mm-dd hh:mm",
    "time": "hh:mm:ss",
}
XLSX_MAX_COL_WIDTH = 60
XLSX_MIN_COL_WIDTH = 8
XLSX_WRAP_THRESHOLD = 40  # text cells longer than this get wrap_text + a wider column
# Abuse/DoS guards for the generic, client-payload-driven endpoint — a
# reasonable ceiling given every real export in this app tops out at a few
# thousand rows (see each CSV export's own row caps).
XLSX_MAX_SHEETS = 10
XLSX_MAX_ROWS_PER_SHEET = 5000
XLSX_MAX_COLUMNS = 40
XLSX_MAX_CELL_TEXT = 4000

def _xlsx_sheet_name(name: str, used: set) -> str:
    """Excel sheet names: <=31 chars, no [ ] : * ? / \\ , and unique per workbook."""
    cleaned = re.sub(r'[\[\]:*?/\\]', '', (name or "Sheet")).strip()[:31] or "Sheet"
    candidate = cleaned
    n = 2
    while candidate.casefold() in used:
        suffix = f" ({n})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.casefold())
    return candidate

def _xlsx_cell_value(raw, col_type: str):
    """Converts one already-resolved value into what actually gets written to
    the cell — a real number/bool/datetime for typed columns (so Excel can
    sort/filter/calculate on it), or a formula-injection-safe string for
    everything else. ISO date/datetime strings (what every backend endpoint
    already sends the frontend) are parsed into real datetime objects here so
    the XLSX cell carries an actual date, not text that merely looks like one."""
    if raw is None or raw == "":
        return None
    if col_type in ("number", "decimal", "currency", "percent"):
        try:
            return float(raw)
        except (TypeError, ValueError):
            return _csv_sanitize_cell(raw)
    if col_type == "bool":
        return bool(raw)
    if col_type in ("date", "datetime"):
        if isinstance(raw, (datetime, date)):
            return raw
        try:
            text = str(raw).rstrip("Z")
            return datetime.fromisoformat(text)
        except ValueError:
            return _csv_sanitize_cell(raw)
    if col_type == "time":
        if isinstance(raw, dtime):
            return raw
        try:
            return datetime.strptime(str(raw), "%H:%M:%S").time()
        except ValueError:
            return _csv_sanitize_cell(raw)
    # "text" (default) — the same formula-injection guard CSV uses, since a
    # plain string cell assigned a leading "=" is exactly as dangerous in a
    # real .xlsx as in a .csv (openpyxl/Excel both treat a leading "=" on a
    # string value as a formula to evaluate).
    return _csv_sanitize_cell(raw)

def build_xlsx_workbook(report_title: Optional[str], sheets: List[dict]) -> bytes:
    """sheets: [{"name": str, "columns": [{"key","label","type","width"}],
    "rows": [[...], ...], "title": optional per-sheet title,
    "metadata": optional [[label, value], ...] context lines}]"""
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for sheet in sheets:
        ws = wb.create_sheet(title=_xlsx_sheet_name(sheet.get("name"), used_names))
        columns = sheet["columns"]
        rows = sheet["rows"]
        title = sheet.get("title") or report_title
        metadata = sheet.get("metadata") or []

        current_row = 1
        if title:
            ws.cell(row=current_row, column=1, value=str(title)[:XLSX_MAX_CELL_TEXT]).font = XLSX_TITLE_FONT
            current_row += 1
        for label, value in metadata:
            ws.cell(row=current_row, column=1, value=f"{label}: {value}"[:XLSX_MAX_CELL_TEXT]).font = XLSX_META_FONT
            current_row += 1
        if title or metadata:
            current_row += 1  # blank spacer before the table — never inside the header row itself

        header_row = current_row
        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=c_idx, value=col["label"])
            cell.font = XLSX_HEADER_FONT
            cell.fill = XLSX_HEADER_FILL
            cell.alignment = Alignment(vertical="center")

        col_widths = [len(str(c["label"])) for c in columns]
        for r_offset, row in enumerate(rows):
            r_idx = header_row + 1 + r_offset
            for c_idx, col in enumerate(columns, start=1):
                raw = row[c_idx - 1] if c_idx - 1 < len(row) else None
                value = _xlsx_cell_value(raw, col.get("type", "text"))
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                fmt = XLSX_NUMBER_FORMATS.get(col.get("type", "text"))
                if fmt and isinstance(value, (int, float, datetime, date, dtime)) and not isinstance(value, bool):
                    cell.number_format = fmt
                text_len = len(str(raw)) if raw is not None else 0
                if col.get("type", "text") == "text" and text_len > XLSX_WRAP_THRESHOLD:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                col_widths[c_idx - 1] = max(col_widths[c_idx - 1], min(text_len, XLSX_MAX_COL_WIDTH))

        for c_idx, col in enumerate(columns, start=1):
            letter = get_column_letter(c_idx)
            width = col.get("width") or max(XLSX_MIN_COL_WIDTH, min(XLSX_MAX_COL_WIDTH, col_widths[c_idx - 1] + 2))
            ws.column_dimensions[letter].width = width

        last_row = header_row + len(rows)
        last_col = len(columns)
        if rows and last_col:
            table_ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
            try:
                table_name = re.sub(r"[^A-Za-z0-9_]", "_", f"tbl_{ws.title}")[:30].lstrip("0123456789_") or "tbl_data"
                table = Table(displayName=table_name, ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                ws.add_table(table)
            except Exception:
                ws.auto_filter.ref = table_ref  # still get filtering even if the named Table object fails

        # Freeze the header row (and everything above it) so it stays visible
        # while scrolling a long export — never freezes any data column.
        ws.freeze_panes = f"A{header_row + 1}"
        ws.print_title_rows = f"{header_row}:{header_row}"
        ws.page_setup.orientation = "landscape" if last_col > 6 else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_xlsx_response(filename: str, report_title: Optional[str], sheets: List[dict]) -> Response:
    content = build_xlsx_workbook(report_title, sheets)
    safe_name = safe_csv_filename(filename)
    if not safe_name.lower().endswith(".xlsx"):
        safe_name = re.sub(r"\.csv$", "", safe_name, flags=re.IGNORECASE) + ".xlsx"
    return Response(
        content=content,
        media_type=XLSX_MIME_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )

def resolve_financial_period(business: "BusinessProfile", period: str, custom_start: Optional[str] = None, custom_end: Optional[str] = None) -> Tuple[Optional[datetime], Optional[datetime]]:
    """Returns (start_utc, end_utc) as naive UTC datetimes — matching how
    every timestamp in this app is already stored (datetime.utcnow(), no
    tzinfo) — bounding the requested period in the BUSINESS'S OWN local
    timezone. 'Today' means the business's calendar day, not whatever day it
    happens to be in UTC, per the business's stored BusinessProfile.timezone.
    period="all" returns (None, None), meaning "no date filter at all"."""
    if period not in FINANCIAL_PERIODS:
        raise HTTPException(status_code=400, detail=f"Unknown period '{period}'. Use one of: {', '.join(sorted(FINANCIAL_PERIODS))}.")
    if period == "all":
        return None, None
    tz = business_local_zoneinfo(business)
    now_local = datetime.utcnow().replace(tzinfo=timezone.utc).astimezone(tz)
    today_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)

    if period == "today":
        start_local, end_local = today_local, today_local + timedelta(days=1)
    elif period == "yesterday":
        start_local, end_local = today_local - timedelta(days=1), today_local
    elif period == "week":
        start_local = today_local - timedelta(days=today_local.weekday())
        end_local = start_local + timedelta(days=7)
    elif period == "previous_week":
        this_week_start = today_local - timedelta(days=today_local.weekday())
        start_local, end_local = this_week_start - timedelta(days=7), this_week_start
    elif period == "month":
        start_local = today_local.replace(day=1)
        end_local = (start_local + timedelta(days=32)).replace(day=1)
    elif period == "previous_month":
        this_month_start = today_local.replace(day=1)
        end_local = this_month_start
        start_local = (this_month_start - timedelta(days=1)).replace(day=1)
    elif period == "year":
        start_local = today_local.replace(month=1, day=1)
        end_local = start_local.replace(year=start_local.year + 1)
    elif period == "previous_year":
        this_year_start = today_local.replace(month=1, day=1)
        start_local = this_year_start.replace(year=this_year_start.year - 1)
        end_local = this_year_start
    else:  # custom
        # The backend is the final authority on this rule, independent of
        # whatever the frontend's Apply-button flow already enforced: a
        # custom range is never guessed from a single boundary. Exact
        # messages match what the frontend shows for the same conditions
        # (see the Custom Date Range spec) so a direct API caller gets the
        # same clear guidance a UI user would.
        if not custom_start and not custom_end:
            raise HTTPException(status_code=400, detail="Select a start date and end date.")
        if not custom_start:
            raise HTTPException(status_code=400, detail="Select a start date.")
        if not custom_end:
            raise HTTPException(status_code=400, detail="Select an end date.")
        try:
            start_local = datetime.strptime(custom_start, "%Y-%m-%d").replace(tzinfo=tz)
            end_local = datetime.strptime(custom_end, "%Y-%m-%d").replace(tzinfo=tz) + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Custom dates must be in YYYY-MM-DD format.")
        if end_local <= start_local:
            raise HTTPException(status_code=400, detail="End date cannot be earlier than start date.")

    return start_local.astimezone(timezone.utc).replace(tzinfo=None), end_local.astimezone(timezone.utc).replace(tzinfo=None)

def checkout_key_expr():
    """The SQL expression identifying which CHECKOUT a sale line belongs to —
    the single definition every "how many transactions?" count in this app
    uses, so they can never disagree with each other.

    SaleModel is a LINE ITEM (one row per product), not a whole checkout, so
    COUNT(SaleModel.id) counts lines and reports a 3-product checkout as
    three transactions. The correct count is
    COUNT(DISTINCT checkout_key_expr()).

    client_ref is that key: every line created by one /sales/checkout call
    shares it (see SaleModel.client_ref). Rows recorded before checkout
    started stamping one have client_ref NULL — each of those is genuinely
    its own single-line transaction, so it falls back to a per-row unique
    value. The 'S' prefix keeps that synthetic key from ever colliding with
    a real client_ref that happened to be all digits.

    cast(...) + literal string concat compiles to `'S' || CAST(id AS VARCHAR)`
    on both SQLite and PostgreSQL — verified against both dialects; no
    dialect-specific branching needed."""
    return func.coalesce(SaleModel.client_ref, literal("S") + cast(SaleModel.id, String))

def compute_financial_summary(db: Session, business_id: int, start_utc: Optional[datetime], end_utc: Optional[datetime]) -> Dict[str, Any]:
    """Revenue, COGS, gross/net profit and margin for [start_utc, end_utc) —
    or across all time if both are None — computed with DB-side aggregation
    (not a Python loop over every row) so this stays fast regardless of how
    many years of history a business has accumulated.

    This is the ONE authoritative profit calculation in the app — Dashboard,
    the Profit page, and every export/report call this (via
    financial_summary_for_period) rather than computing their own figures,
    so they can never disagree.

        Gross Profit = Sales Revenue - Cost of Goods Sold
        Net Profit   = Gross Profit - Business Expenses

    COGS uses only each sale's OWN unit_cost_at_sale snapshot. A NULL legacy
    snapshot is unknown — never replaced by today's Product.cost_price and
    never silently treated as zero. Revenue/units and the known-cost subtotal
    remain exact, while total COGS/profit are returned as NULL whenever the
    selected period contains unknown cost evidence.

    Audited corrections (SaleAdjustment/ExpenseAdjustment) are included by
    the ORIGINAL record's period, not the correction's own creation date —
    a correction recorded today against last month's sale still affects
    last month's reported profit, exactly like the "original vs corrected"
    model everywhere else in the app (it never moves history into whatever
    period the fix happened to be filed in).

    Refunds (RefundLine) are the one deliberate exception to that rule: a
    refund is attributed to the period IT was performed in, not the
    original sale's period — refunding a sale from last month today reduces
    THIS month's net figures, exactly like a real cash-drawer adjustment
    would (see RefundTransaction model docstring / section 5 of the refund
    spec). refund_amount/refund_cost are themselves snapshots of the
    original sale's own price/cost (see RefundLine model docstring), so
    reversing them here never reintroduces a current-price/current-cost
    dependency."""
    sales_filters = [SaleModel.business_id == business_id]
    expense_filters = [Expense.business_id == business_id]
    sale_adj_filters = [SaleAdjustment.business_id == business_id]
    expense_adj_filters = [ExpenseAdjustment.business_id == business_id]
    if start_utc is not None and end_utc is not None:
        sales_filters += [SaleModel.timestamp >= start_utc, SaleModel.timestamp < end_utc]
        expense_filters += [Expense.created_at >= start_utc, Expense.created_at < end_utc]
        sale_adj_filters += [SaleModel.timestamp >= start_utc, SaleModel.timestamp < end_utc]
        expense_adj_filters += [Expense.created_at >= start_utc, Expense.created_at < end_utc]

    sales_row = (
        db.query(
            func.coalesce(func.sum(SaleModel.total_price), 0.0),
            func.coalesce(func.sum(SaleModel.quantity), 0),
            # TRANSACTIONS, not sale lines — see checkout_key_expr(). A
            # 3-product checkout is one transaction; COUNT(SaleModel.id)
            # would have reported it as three.
            func.count(func.distinct(checkout_key_expr())),
            func.count(SaleModel.id),
        )
        .filter(*sales_filters)
        .first()
    )
    total_sales, total_units = float(sales_row[0] or 0), int(sales_row[1] or 0)
    transaction_count, sale_line_count = int(sales_row[2] or 0), int(sales_row[3] or 0)

    # SQL SUM ignores NULL, giving an exact subtotal of RECORDED snapshots.
    # Unknown rows are counted separately; Product is deliberately absent.
    cost_expr = SaleModel.unit_cost_at_sale
    total_known_cogs = float(
        db.query(func.coalesce(func.sum(SaleModel.quantity * cost_expr), 0.0))
        .filter(*sales_filters)
        .scalar() or 0
    )
    unknown_sale_row = db.query(
        func.count(SaleModel.id), func.coalesce(func.sum(SaleModel.quantity), 0),
        func.coalesce(func.sum(SaleModel.total_price), 0.0),
    ).filter(*sales_filters, SaleModel.unit_cost_at_sale.is_(None)).one()
    unknown_sale_lines = int(unknown_sale_row[0] or 0)
    unknown_sale_units = int(unknown_sale_row[1] or 0)
    unknown_sale_revenue = float(unknown_sale_row[2] or 0.0)

    # Sale corrections: amount_delta adjusts revenue directly; quantity_delta
    # (e.g. a partial return) adjusts COGS using the ORIGINAL sale's own cost
    # snapshot, never a current price — the correction is a delta on top of
    # a specific historical sale, so it must cost exactly what that sale cost.
    sale_adj_row = (
        db.query(
            func.coalesce(func.sum(SaleAdjustment.amount_delta), 0.0),
            func.coalesce(func.sum(SaleAdjustment.quantity_delta * cost_expr), 0.0),
            func.coalesce(func.sum(SaleAdjustment.quantity_delta), 0),
        )
        .join(SaleModel, SaleModel.id == SaleAdjustment.sale_id)
        .filter(*sale_adj_filters)
        .first()
    )
    sale_adj_amount, sale_adj_cogs, sale_adj_units = float(sale_adj_row[0] or 0), float(sale_adj_row[1] or 0), int(sale_adj_row[2] or 0)
    unknown_adjustment_count = int(
        db.query(func.count(SaleAdjustment.id))
        .join(SaleModel, SaleModel.id == SaleAdjustment.sale_id)
        .filter(*sale_adj_filters, SaleModel.unit_cost_at_sale.is_(None), SaleAdjustment.quantity_delta != 0)
        .scalar() or 0
    )
    total_sales += sale_adj_amount
    total_known_cogs += sale_adj_cogs
    total_units += sale_adj_units

    expense_row = (
        db.query(func.coalesce(func.sum(Expense.amount), 0.0), func.count(Expense.id))
        .filter(*expense_filters)
        .first()
    )
    total_expenses, expense_count = float(expense_row[0] or 0), int(expense_row[1] or 0)

    expense_adj_total = float(
        db.query(func.coalesce(func.sum(ExpenseAdjustment.amount_delta), 0.0))
        .join(Expense, Expense.id == ExpenseAdjustment.expense_id)
        .filter(*expense_adj_filters)
        .scalar() or 0
    )
    total_expenses += expense_adj_total

    # Refunds — filtered by the REFUND's own created_at (see docstring
    # above), never the original sale's timestamp. gross_* below preserves
    # the pre-refund figures for reporting (section 15/27 of the refund
    # spec: Gross Units Sold vs Refunded Units vs Net Units Sold must all
    # remain independently visible, never silently collapsed into one).
    gross_sales, known_gross_cogs, gross_units = total_sales, total_known_cogs, total_units
    refund_filters = [RefundLine.business_id == business_id]
    if start_utc is not None and end_utc is not None:
        refund_filters += [RefundLine.created_at >= start_utc, RefundLine.created_at < end_utc]
    refund_row = (
        db.query(
            func.coalesce(func.sum(RefundLine.refund_amount), 0.0),
            func.coalesce(func.sum(RefundLine.refund_cost), 0.0),
            func.coalesce(func.sum(RefundLine.quantity), 0),
        )
        .filter(*refund_filters)
        .first()
    )
    refund_amount, refund_cogs, refund_units = float(refund_row[0] or 0), float(refund_row[1] or 0), int(refund_row[2] or 0)
    unknown_refund_row = db.query(
        func.count(RefundLine.id), func.coalesce(func.sum(RefundLine.quantity), 0),
    ).filter(*refund_filters, RefundLine.refund_cost.is_(None)).one()
    unknown_refund_lines = int(unknown_refund_row[0] or 0)
    unknown_refund_units = int(unknown_refund_row[1] or 0)
    # Refund TRANSACTION count, kept strictly separate from sales
    # transaction_count above (section 16 — a refund must never inflate the
    # sales transaction count).
    refund_txn_filters = [RefundTransaction.business_id == business_id]
    if start_utc is not None and end_utc is not None:
        refund_txn_filters += [RefundTransaction.created_at >= start_utc, RefundTransaction.created_at < end_utc]
    refund_transaction_count = db.query(func.count(RefundTransaction.id)).filter(*refund_txn_filters).scalar() or 0

    total_sales -= refund_amount
    total_known_cogs -= refund_cogs
    total_units -= refund_units

    cogs_complete = unknown_sale_lines == 0 and unknown_adjustment_count == 0 and unknown_refund_lines == 0
    known_cogs = round(total_known_cogs, 2)
    gross_profit = round(total_sales - total_known_cogs, 2) if cogs_complete else None
    net_profit = round(gross_profit - total_expenses, 2) if gross_profit is not None else None
    margin = round((net_profit / total_sales) * 100, 1) if net_profit is not None and total_sales > 0 else None

    return {
        # NET figures (post-refund) under the original key names — every
        # existing caller (Dashboard, Profit page, exports) keeps reading
        # these same keys and now correctly gets refund-aware numbers
        # without needing to change anything.
        "sales": round(total_sales, 2), "cogs": known_cogs if cogs_complete else None, "gross_profit": gross_profit,
        "expenses": round(total_expenses, 2), "net_profit": net_profit, "profit_margin_percent": margin,
        "transaction_count": transaction_count, "units_sold": total_units, "expense_count": expense_count,
        # sale_line_count is the old COUNT(SaleModel.id) figure, kept as its
        # own clearly-named field: it answers "how many product lines?",
        # which is a genuinely different question from transaction_count's
        # "how many checkouts?" and must never be conflated with it again.
        "sale_line_count": sale_line_count,
        # Additive refund breakdown (section 27/28) — never required by
        # existing callers, available for anything that wants to show the
        # gross/refund split explicitly.
        "gross_sales": round(gross_sales, 2), "gross_cogs": round(known_gross_cogs, 2) if cogs_complete else None, "gross_units_sold": gross_units,
        "refund_amount": round(refund_amount, 2), "refunded_cogs": round(refund_cogs, 2) if unknown_refund_lines == 0 else None, "refunded_units": refund_units,
        "refund_transaction_count": int(refund_transaction_count),
        "cogs_complete": cogs_complete, "known_cogs": known_cogs,
        "known_gross_cogs": round(known_gross_cogs, 2), "known_refunded_cogs": round(refund_cogs, 2),
        "unknown_cogs_sale_lines": unknown_sale_lines, "unknown_cogs_sale_units": unknown_sale_units,
        "unknown_cogs_sale_revenue": round(unknown_sale_revenue, 2),
        "unknown_cogs_adjustment_count": unknown_adjustment_count,
        "unknown_cogs_refund_lines": unknown_refund_lines, "unknown_cogs_refund_units": unknown_refund_units,
        "period_start": to_utc_iso(start_utc), "period_end": to_utc_iso(end_utc),
    }

def financial_summary_for_period(db: Session, business: "BusinessProfile", period: str, custom_start: Optional[str] = None, custom_end: Optional[str] = None) -> Dict[str, Any]:
    start_utc, end_utc = resolve_financial_period(business, period, custom_start, custom_end)
    summary = compute_financial_summary(db, business.id, start_utc, end_utc)
    summary["period"] = period
    return summary

def get_business_by_code(db: Session, business_code: str) -> Optional[BusinessProfile]:
    cleaned = re.sub(r"[^A-Za-z0-9]", "", (business_code or "")).casefold()
    if not cleaned:
        return None
    businesses = db.query(BusinessProfile).all()
    return next((b for b in businesses if re.sub(r"[^A-Za-z0-9]", "", b.business_code).casefold() == cleaned), None)

def issue_token(user: User, db: Session) -> str:
    exp = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    jti = secrets.token_urlsafe(24)
    payload = {"sub": str(user.id), "business_id": user.business_id, "role": user.role, "jti": jti, "auth_version": int(user.auth_version or 1), "exp": exp}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def revoke_all_user_sessions(db: Session, user: User):
    now = datetime.utcnow()
    db.query(RefreshSession).filter(RefreshSession.user_id == user.id, RefreshSession.revoked_at.is_(None)).update({RefreshSession.revoked_at: now}, synchronize_session=False)
    user.auth_version = int(user.auth_version or 1) + 1

def create_refresh_session(db: Session, user: User) -> str:
    raw = secrets.token_urlsafe(48)
    row = RefreshSession(token_hash=hash_text(raw), user_id=user.id, business_id=user.business_id, expires_at=datetime.utcnow()+timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS))
    db.add(row); db.commit()
    return raw

def set_refresh_cookie(response: Response, raw: str):
    response.set_cookie(key=REFRESH_COOKIE_NAME, value=raw, httponly=True, secure=REFRESH_COOKIE_SECURE, samesite=REFRESH_COOKIE_SAMESITE if REFRESH_COOKIE_SAMESITE in {"lax","strict","none"} else "lax", max_age=REFRESH_TOKEN_EXPIRE_DAYS*86400, path="/")

def clear_refresh_cookie(response: Response):
    response.delete_cookie(REFRESH_COOKIE_NAME, path="/")

def token_from_payload(token: str) -> tuple[dict, str]:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        jti = payload.get("jti")
        if not jti: raise JWTError("Missing token id")
        return payload, jti
    except JWTError:
        raise HTTPException(status_code=401, detail="Your session could not be verified. Please sign in again.")

def require_active_user(user: User):
    if user.disabled:
        raise HTTPException(status_code=403, detail="This account is currently disabled.")
    return user

def get_authenticated_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> User:
    payload, jti = token_from_payload(token)
    if db.query(SessionRevocation).filter(SessionRevocation.jti == jti).first():
        raise HTTPException(status_code=401, detail="Your session is no longer valid. Please sign in again.")
    try:
        uid=int(payload.get("sub")); bid=int(payload.get("business_id")); av=int(payload.get("auth_version",0))
    except Exception:
        raise HTTPException(status_code=401, detail="Your session could not be verified. Please sign in again.")
    user=db.query(User).filter(User.id==uid, User.business_id==bid).first()
    if not user: raise HTTPException(status_code=401, detail="Your session could not be verified. Please sign in again.")
    require_active_user(user)
    if av and av != int(user.auth_version or 1):
        raise HTTPException(status_code=401, detail="Your session is no longer valid. Please sign in again.")
    return user

def get_current_user(user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)) -> User:
    if user.must_change_password:
        raise HTTPException(status_code=403, detail="You must change your temporary password before using the application.")
    require_subscription_access(db, user)
    return user


def enforce_offline_replay_identity(request: Request, user: User) -> None:
    """Reject an outbox replay that is not bound to the current identity.

    Ordinary online mutations do not carry the replay marker and remain
    unchanged. Every durable outbox replay carries the business, user, and
    account-auth generation captured when it was queued. This prevents a later
    user in the same browser/business from executing or being audited for the
    original user's operation.
    """
    if request.headers.get("x-cauldra-offline-replay") != "1":
        return
    try:
        queued_business_id = int(request.headers.get("x-cauldra-offline-business-id", ""))
        queued_user_id = int(request.headers.get("x-cauldra-offline-user-id", ""))
        queued_auth_version = int(request.headers.get("x-cauldra-offline-auth-version", ""))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Offline replay identity is incomplete.")
    if (
        queued_business_id != int(user.business_id)
        or queued_user_id != int(user.id)
        or queued_auth_version != int(user.auth_version or 1)
    ):
        raise HTTPException(
            status_code=409,
            detail="This offline change belongs to a different or expired signed-in identity and was not applied.",
        )

def require_ai_access(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> User:
    """Real security boundary for every AI-branded endpoint (Business Brain, AI
    chat, margin advisor, invoice OCR, predictive intelligence, etc). A locked
    frontend button is UX only — this dependency is what actually blocks a
    Core-plan business (included_ai_credits == 0) from calling AI endpoints
    directly, regardless of what the client sends or displays."""
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    plan = subscription_for(db, business)
    if not plan.get("included_ai_credits"):
        current_plan_id = (get_or_create_subscription(db, business).plan or "starter").lower()
        upgrade_label = PLAN_CONFIG.get(UPGRADE_PATH.get(current_plan_id, "starter"), PLAN_CONFIG["starter"])["label"]
        raise HTTPException(status_code=403, detail=f"AI features are not included in your {plan['label']} plan. Upgrade to {upgrade_label} to unlock AI-powered tools.")
    return user

# Action codes that describe presence/session state rather than a meaningful
# business/audit action. Activity History (list_audit_logs) excludes these at
# query time rather than deleting the underlying rows, so any such row already
# recorded historically is preserved but never re-surfaces going forward. New
# code should not add to AuditLog for presence/session events in the first
# place (see /presence/* and PresenceSession) — this set exists mainly to keep
# older rows written before that separation out of the business activity feed.
PRESENCE_SESSION_AUDIT_ACTIONS = {"LOGIN", "LOGOUT"}

def add_audit(db: Session, user: Optional[User], action: str, description: str, target: Optional[User] = None, business_id: Optional[int] = None, business_day_id: Optional[int] = None, metadata: Optional[dict] = None):
    """The single immutable audit trail for the whole app. business_day_id and
    metadata are additive/optional — every pre-existing call site is
    unaffected. metadata is stored as JSON text (never as arbitrary client
    input — every caller builds it server-side from values it already
    computed) so structured details (closing snapshots, correction old/new
    values) survive without needing a new table per event type."""
    actor_role = user.role if user else None
    actor_username = user.username if user else None
    bid = business_id or (user.business_id if user else None)
    if not bid:
        return
    db.add(AuditLog(
        business_id=bid,
        actor_user_id=user.id if user else None,
        actor_username=actor_username,
        actor_role=actor_role,
        action=action,
        target_user_id=target.id if target else None,
        target_username=target.username if target else None,
        description=description,
        business_day_id=business_day_id,
        metadata_json=json.dumps(metadata) if metadata is not None else None,
    ))

# -----------------------------------------------------------------------------
# NOTIFICATION ENGINE
#
# create_notification() is the ONE function anything in this app is allowed
# to call to notify a user — every business event in the file (subscription,
# AI credits, inventory, purchase orders, suppliers, security, team) routes
# through it. It owns: who receives it (role-aware recipients), whether it
# has already fired for this exact condition (dedup_key/stage), whether the
# recipient has opted out of it (preferences — mandatory categories can't
# be opted out), and whether it goes out as an external push in addition to
# living in the notification center. A push is never a second notification —
# it's a delivery channel on the same row (see Notification's docstring).
#
#   Business Event -> create_notification() -> recipients resolved ->
#   dedup/preference checked -> Notification row(s) written ->
#   push-eligible rows handed to deliver_push_notification() -> device
#
# This is deliberately NOT tied to Web Push specifically in its business-
# logic half (recipients/severity/dedup/category) — only
# deliver_push_notification() at the bottom knows anything about VAPID/Web
# Push, so a native provider (FCM/APNs) could replace just that one function
# later without touching a single call site above it.
# -----------------------------------------------------------------------------

# Categories a user is never allowed to silence — either because they're
# security-critical (the user needs to know regardless of preference) or
# because they're billing-continuity-critical (losing access is worse than
# one unwanted notification). NotificationPreference rows for these
# categories are never written or consulted; see user_wants_category().
NOTIFICATION_MANDATORY_CATEGORIES = {"security", "subscription"}

# Categories a user MAY opt out of via /notifications/preferences (section
# 16). Kept as an explicit allow-list — a category not listed here can never
# be toggled even if a future bug tried to write a preference row for it.
NOTIFICATION_OPTIONAL_CATEGORIES = {"inventory", "purchase_order", "supplier", "ai_usage", "financial", "team", "system"}

# Which roles are eligible to receive each category by default (section 11 —
# "Notification Audience"). "admin" is Cauldra's Owner role (see the
# one-time owner->admin role migration near startup). A category not listed
# here defaults to admin-only, the safest default for anything financial or
# account-sensitive.
NOTIFICATION_CATEGORY_ROLES: Dict[str, Tuple[str, ...]] = {
    "subscription": ("admin",),
    "ai_usage": ("admin",),
    "financial": ("admin",),
    "team": ("admin",),
    "inventory": ("admin", "manager"),
    "purchase_order": ("admin", "manager"),
    "supplier": ("admin", "manager"),
    "system": ("admin", "manager"),
    "security": ("admin", "manager", "staff"),  # narrowed further per-call via recipient_user_ids (affected user + admins)
}

def notification_recipients(db: Session, business_id: int, category: str, extra_user_ids: Optional[set] = None) -> set:
    """Resolves WHO is eligible to receive a notification in this category for
    this business, honoring Cauldra's existing role system (never a role of
    the caller's choosing) — this is the enforcement point for section 11's
    audience rules and for "employees must not receive sensitive billing/
    financial/subscription/owner-level notifications unless their role
    explicitly permits access." extra_user_ids adds specific individuals
    (e.g. the affected user in a security event) regardless of role, since a
    person is always entitled to know about their own security event."""
    roles = NOTIFICATION_CATEGORY_ROLES.get(category, ("admin",))
    ids = {u.id for u in db.query(User.id).filter(User.business_id == business_id, User.disabled == False, User.role.in_(roles)).all()}
    if extra_user_ids:
        ids.update(extra_user_ids)
    return ids

def user_wants_category(db: Session, user_id: int, category: str) -> bool:
    """Mandatory categories can never be silenced. Optional categories default
    to ON (a new user hears about everything their role permits until they
    explicitly turn a category off) and are only ever checked against the
    user's OWN preference row — a preference can narrow what this specific
    user receives, never widen it past what notification_recipients() (role)
    already allowed."""
    if category not in NOTIFICATION_OPTIONAL_CATEGORIES:
        return True
    pref = db.query(NotificationPreference).filter(NotificationPreference.user_id == user_id, NotificationPreference.category == category).first()
    return pref.enabled if pref else True

def create_notification(
    db: Session, *, business_id: int, category: str, severity: str, type: str, title: str, message: str,
    recipient_user_ids: Optional[set] = None, related_entity_type: Optional[str] = None,
    related_entity_id: Optional[int] = None, deep_link: Optional[str] = None,
    dedup_key: Optional[str] = None, stage: Optional[str] = None,
) -> List["Notification"]:
    """THE single entry point for every notification in Cauldra (see the
    section banner above). Returns the Notification rows actually created
    (an empty list means every eligible recipient already had this exact
    condition open, or opted out — not an error).

    severity is the explicit, deliberate channel decision this engine makes
    (section 19/20) — "critical" is not a visual style here, it is the one
    fact that makes a row push-eligible. Every call site in this file chose
    that severity by walking section 20's checklist (is this time-sensitive,
    would waiting create a real problem, etc.) at the moment it decided to
    call this function; "important" and "info" are never pushed, no matter
    how they're styled in the UI.

    dedup_key is what prevents the same condition from renotifying on every
    request/poll/background sweep (section 17): if an UNRESOLVED
    Notification with this exact dedup_key already exists for a recipient,
    no new row is created for them. Callers encode the "stage" that should
    reset the key into the key itself — e.g. an AI-credit threshold key
    includes the billing period string, so a new period is automatically a
    fresh key space; a subscription-countdown key includes the countdown
    day, so 7/3/1/expired are each their own key and all fire independently.
    stage is stored alongside purely for the notification center UI/API to
    show progression; it plays no part in the dedup decision itself (the key
    already encodes it)."""
    business = db.query(BusinessProfile).filter(BusinessProfile.id == business_id).first()
    if not business:
        return []
    recipients = recipient_user_ids if recipient_user_ids is not None else notification_recipients(db, business_id, category)
    created: List[Notification] = []
    for uid in recipients:
        if dedup_key:
            existing = db.query(Notification).filter(
                Notification.recipient_user_id == uid,
                Notification.dedup_key == dedup_key,
                Notification.resolved_at.is_(None),
            ).first()
            if existing:
                continue
        if not user_wants_category(db, uid, category):
            continue
        n = Notification(
            business_id=business_id, recipient_user_id=uid, type=type, category=category, severity=severity,
            title=title, message=message, related_entity_type=related_entity_type, related_entity_id=related_entity_id,
            deep_link=deep_link, in_app=True, push=(severity == "critical"), dedup_key=dedup_key, stage=stage,
        )
        db.add(n)
        created.append(n)
    db.flush()
    for n in created:
        if n.push:
            deliver_push_notification(db, n)
    return created

def resolve_notifications(db: Session, dedup_key: str, business_id: Optional[int] = None) -> int:
    """Marks every still-unresolved Notification for this dedup_key as
    resolved, WITHOUT deleting them (they remain in the recipient's history
    exactly as they were seen) — and frees the key so a future recurrence of
    the same condition (stock runs out again, a later payment fails again)
    can fire a fresh notification instead of being silently swallowed by a
    stale dedup match. Called when the underlying condition genuinely
    clears (stock replenished, payment succeeds, subscription reactivated).
    Returns the number of rows resolved."""
    q = db.query(Notification).filter(Notification.dedup_key == dedup_key, Notification.resolved_at.is_(None))
    if business_id is not None:
        q = q.filter(Notification.business_id == business_id)
    rows = q.all()
    now = datetime.utcnow()
    for row in rows:
        row.resolved_at = now
    return len(rows)

def deliver_push_notification(db: Session, notification: "Notification") -> None:
    """Best-effort Web Push delivery for one already-created, push-eligible
    Notification row. Never raises — a push provider outage or an expired
    subscription must never break the request that triggered the
    notification; the row already exists in the notification center
    regardless of whether this succeeds. A subscription the push service
    reports as permanently gone (404/410) is disabled so it is never retried
    again; any other failure is left alone for the next event to retry
    naturally (there is no separate retry queue)."""
    if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
        return  # push not configured — in-app notification still stands on its own
    subs = db.query(PushSubscription).filter(PushSubscription.user_id == notification.recipient_user_id, PushSubscription.disabled_at.is_(None)).all()
    if not subs:
        return
    payload = json.dumps({
        "notification_id": notification.id, "title": notification.title, "body": notification.message,
        "severity": notification.severity, "category": notification.category, "deep_link": notification.deep_link,
    })
    sent_any = False
    for sub in subs:
        try:
            webpush(
                subscription_info={"endpoint": sub.endpoint, "keys": {"p256dh": sub.p256dh, "auth": sub.auth}},
                data=payload, vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={"sub": f"mailto:{VAPID_CONTACT_EMAIL}"}, ttl=86400,
            )
            sub.last_seen_at = datetime.utcnow()
            sent_any = True
        except WebPushException as exc:
            status_code = getattr(exc.response, "status_code", None)
            if status_code in (404, 410):
                sub.disabled_at = datetime.utcnow()
            # Any other failure (5xx, timeout, offline device): leave the
            # subscription active and move on — no retry queue exists, and a
            # future event will naturally attempt delivery again.
        except Exception:
            pass
    if sent_any:
        notification.push_sent_at = datetime.utcnow()

def serialize_notification(n: "Notification") -> dict:
    return {
        "id": n.id, "type": n.type, "category": n.category, "severity": n.severity,
        "title": n.title, "message": n.message, "related_entity_type": n.related_entity_type,
        "related_entity_id": n.related_entity_id, "deep_link": n.deep_link,
        "in_app": n.in_app, "push": n.push, "is_read": n.is_read,
        "created_at": to_utc_iso(n.created_at), "read_at": to_utc_iso(n.read_at) if n.read_at else None,
        "push_sent_at": to_utc_iso(n.push_sent_at) if n.push_sent_at else None,
        "resolved": n.resolved_at is not None,
    }

def check_login_lockout_notification(db: Session, user: "User", scope: str, key: str) -> None:
    """Section 8 — repeated suspicious access attempts against a REAL
    account become a Critical/push security event, sent to the affected
    account and the business's Owner/Admin, at the exact moment Cauldra's
    existing rate-limit lockout (check_rate_limit/record_failure) actually
    triggers — never on every attempt while it stays locked, and never for
    a nonexistent username. Deliberately reuses the lockout Cauldra already
    computes rather than a second, parallel suspicious-activity detector."""
    row = db.query(AuthFailure).filter(AuthFailure.scope == scope, AuthFailure.key_hash == fail_key(scope, key)).first()
    if not row or row.failures != RATE_LIMIT_MAX_FAILURES:
        return
    admin_ids = {u.id for u in db.query(User.id).filter(User.business_id == user.business_id, User.role == "admin", User.disabled == False).all()}
    create_notification(
        db, business_id=user.business_id, category="security", severity="critical", type="REPEATED_LOGIN_FAILURES",
        title="Repeated failed sign-in attempts",
        message=f"There have been repeated failed sign-in attempts on the account '{user.username}'. If this wasn't you, consider changing your password.",
        recipient_user_ids={user.id} | admin_ids, related_entity_type="user", related_entity_id=user.id,
        deep_link="account_security", dedup_key=f"login_lockout:{user.id}:{row.window_started_at.isoformat()}",
    )
    # This runs on the FAILED-login path, which raises an HTTPException
    # immediately after returning here and never itself commits — without
    # this, the notification create_notification() just flushed would be
    # silently rolled back when the request's session closes. Self-committing
    # matches record_failure()/clear_failures()'s own established pattern of
    # owning their own commit rather than depending on the caller.
    db.commit()

def fail_key(scope: str, key: str) -> str:
    return hash_text(f"{scope}:{key}")

def check_rate_limit(db: Session, scope: str, key: str):
    k = fail_key(scope, key)
    row = db.query(AuthFailure).filter(AuthFailure.scope == scope, AuthFailure.key_hash == k).first()
    now = datetime.utcnow()
    if not row:
        return
    if row.locked_until and row.locked_until > now:
        seconds = int((row.locked_until - now).total_seconds())
        raise HTTPException(status_code=429, detail=f"Too many attempts. Please wait {max(1, seconds)} seconds and try again.")
    if (now - row.window_started_at).total_seconds() > RATE_LIMIT_WINDOW_SECONDS:
        row.failures = 0
        row.window_started_at = now
        row.locked_until = None
        db.commit()

def record_failure(db: Session, scope: str, key: str):
    k = fail_key(scope, key)
    now = datetime.utcnow()
    row = db.query(AuthFailure).filter(AuthFailure.scope == scope, AuthFailure.key_hash == k).first()
    if not row:
        row = AuthFailure(scope=scope, key_hash=k, failures=0, window_started_at=now)
        db.add(row)
    if (now - row.window_started_at).total_seconds() > RATE_LIMIT_WINDOW_SECONDS:
        row.failures = 0
        row.window_started_at = now
    row.failures += 1
    if row.failures >= RATE_LIMIT_MAX_FAILURES:
        row.locked_until = now + timedelta(seconds=RATE_LIMIT_WINDOW_SECONDS)
    db.commit()

def clear_failures(db: Session, scope: str, key: str):
    k = fail_key(scope, key)
    row = db.query(AuthFailure).filter(AuthFailure.scope == scope, AuthFailure.key_hash == k).first()
    if row:
        db.delete(row)
        db.commit()

def parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None

def serialize_user(u: User) -> dict:
    return {
        "id": u.id, "username": u.username, "role": u.role, "email": u.email, "phone": u.phone,
        "firstname": u.firstname, "lastname": u.lastname, "position": u.position,
        "must_change_password": u.must_change_password, "disabled": u.disabled,
        "auth_version": int(u.auth_version or 1),
    }

def serialize_business(b: BusinessProfile) -> dict:
    return {
        # Keep the business identifier explicit when this object is combined with
        # a user object in an authentication response. Both models have an id.
        "id": b.id, "business_id": b.id, "business_code": b.business_code, "company_name": b.company_name,
        "email": b.email, "phone": b.phone, "address": b.address, "currency": b.currency,
        "tax_id": b.tax_id, "country": b.country, "country_code": b.country_code,
        "locale": b.locale, "language": b.language, "timezone": b.timezone,
        "phone_country_code": b.phone_country_code,
        "subscription_plan": (b.subscription_plan or "starter").lower(), "billing_interval": b.billing_interval or "monthly",
    }

def auto_upsert_general_catalog(db: Session, product: Product):
    """Every business's own product create/edit quietly contributes its safe
    IDENTITY (name/barcode/size) to the shared catalog, so a later business
    scanning the same barcode gets an instant Cauldra-catalog hit instead of
    needing UPCitemdb. Deliberately never writes `category` (see
    GeneralCatalog docstring — a business's category choice is not shared
    identity) and never touches price/quantity/warehouse/supplier/business_id
    — none of that is part of this model at all."""
    barcode=normalize_barcode(product.barcode)
    key=f"barcode:{barcode}" if barcode else f"name:{(product.name or '').strip().casefold()}|category:{(product.category or 'General').strip().casefold()}"
    item=db.query(GeneralCatalog).filter(GeneralCatalog.catalog_key==key).first()
    if item:
        item.product_name=product.name or item.product_name; item.size=product.size or item.size; item.barcode=barcode or item.barcode; item.updated_at=datetime.utcnow(); return
    if barcode:
        item=db.query(GeneralCatalog).filter(GeneralCatalog.barcode==barcode).first()
        if item:
            item.catalog_key=f"barcode:{barcode}"; item.product_name=product.name or item.product_name; item.size=product.size or item.size; item.updated_at=datetime.utcnow(); return
    db.add(GeneralCatalog(barcode=barcode,catalog_key=key,product_name=product.name,size=product.size,source="business_submission"))

def lookup_general_catalog(db: Session, barcode: str) -> Optional["GeneralCatalog"]:
    """Exact-barcode General Catalog read — the one query every barcode chain
    (Add Product's /catalog/barcode-lookup) checks BEFORE ever considering
    UPCitemdb, so a barcode already known to Cauldra (from any business's
    past submission, or a past UPCitemdb hit) never triggers a new external
    request. `barcode` must already be normalize_barcode()-normalized."""
    return db.query(GeneralCatalog).filter(GeneralCatalog.barcode == barcode).first()

def upsert_general_catalog_identity(db: Session, barcode: str, product_name: str, brand: Optional[str], size: Optional[str], source: str) -> "GeneralCatalog":
    """Cache a SAFE identity result (product_name/brand/size only — see
    GeneralCatalog docstring) under an exact barcode key, called after a
    successful UPCitemdb lookup so the same barcode never needs a second
    external request from any business. Idempotent: a repeat call for the
    same barcode updates the existing row instead of creating a duplicate
    (the column's own unique index would reject a duplicate anyway)."""
    key = f"barcode:{barcode}"
    item = db.query(GeneralCatalog).filter(GeneralCatalog.catalog_key == key).first()
    if not item:
        item = db.query(GeneralCatalog).filter(GeneralCatalog.barcode == barcode).first()
    if item:
        item.product_name = product_name or item.product_name
        item.brand = brand if brand is not None else item.brand
        item.size = size if size is not None else item.size
        item.updated_at = datetime.utcnow()
        return item
    item = GeneralCatalog(barcode=barcode, catalog_key=key, product_name=product_name, brand=brand, size=size, source=source)
    db.add(item)
    db.flush()
    return item

def business_local_today(db: Session, business_id: int) -> str:
    business = db.query(BusinessProfile).filter(BusinessProfile.id == business_id).first()
    tz = business_local_zoneinfo(business) if business else timezone.utc
    return datetime.utcnow().replace(tzinfo=timezone.utc).astimezone(tz).date().isoformat()

BUSINESS_DAY_ALREADY_ACTIVE_MSG = "Another Business Day is currently open. Close it before opening or reopening another Business Day."

def get_active_business_day(db: Session, business_id: int) -> Optional[BusinessDay]:
    """The one source of truth for "what Business Day session is currently
    active" — the row where is_open is True, if any. NEVER date-based: a
    business may have many BusinessDay sessions on the same business-local
    date (see the BusinessDay model docstring), so this looks at lifecycle
    state only, not `date`. At most one such row can exist at a time,
    enforced by a partial unique index (see startup migrations); if a reopen
    ever raced its way past that somehow, ordering by opened_at desc at
    least returns the most-recently-opened one rather than an arbitrary row."""
    return (
        db.query(BusinessDay)
        .filter(BusinessDay.business_id == business_id, BusinessDay.is_open == True)
        .order_by(BusinessDay.opened_at.desc())
        .first()
    )

def _create_business_day_session(db: Session, business_id: int, opener: Optional[User], auto: bool = False, commit: bool = True) -> BusinessDay:
    """Unconditionally creates and audits a brand-new Business Day
    session row. Callers MUST have already confirmed no other session is
    currently active (see get_active_business_day) — this never looks up or
    reuses any prior row for today, closed or otherwise, because multiple
    sessions per calendar date are allowed (see the BusinessDay model
    docstring). `date` is business-local (see resolve_financial_period() for
    the same UTC<->local pattern) and is purely descriptive here — it plays
    no role in uniqueness.

    auto distinguishes the two ways a session can come into existence in the
    audit trail: BUSINESS_DAY_STARTED (someone deliberately pressed Open
    Business Day) vs BUSINESS_DAY_AUTO_OPENED (a sale/expense needed an owning
    day and none was active). Both are real, both are audited — but they mean
    different things when reconstructing what happened, so they must not be
    logged under the same action code."""
    today = business_local_today(db, business_id)
    day = BusinessDay(
        business_id=business_id, date=today, is_open=True, status="OPEN",
        opened_by_id=opener.id if opener else None,
        opened_by_name=opener.username if opener else None,
        opened_by_role=opener.role if opener else None,
    )
    db.add(day)
    if commit:
        db.commit()
        db.refresh(day)
    else:
        db.flush()
    add_audit(
        db, opener,
        "BUSINESS_DAY_AUTO_OPENED" if auto else "BUSINESS_DAY_STARTED",
        f"Business day {today} {'auto-opened for an operational action' if auto else 'started'}.",
        business_id=business_id, business_day_id=day.id,
        metadata={"date": today, "opened_by": day.opened_by_name, "role": day.opened_by_role, "auto": auto},
    )
    if commit:
        db.commit()
    else:
        db.flush()
    return day

def start_business_day(db: Session, business_id: int, opener: Optional[User]) -> BusinessDay:
    """The explicit "Open Business Day" action (see /sales/start-business-
    day). ALWAYS creates a brand-new session — a closed session from earlier
    today (or any other date) is history, never something this reuses or
    "continues" (that would be a reopen, a completely separate, more
    privileged, audited workflow that only exists in Sales History). The
    only thing that can block this is another Business Day currently being
    active — raises the exact required conflict error in that case, never a
    stale "already closed today" message.

    Race-safe via the partial unique index on (business_id) WHERE is_open
    (see startup migrations): if two concurrent opens both pass the
    active-day check, only one insert can succeed and the loser gets the
    same conflict error a strictly-sequential second call would have gotten."""
    if get_active_business_day(db, business_id):
        raise HTTPException(status_code=409, detail=BUSINESS_DAY_ALREADY_ACTIVE_MSG)
    try:
        return _create_business_day_session(db, business_id, opener)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail=BUSINESS_DAY_ALREADY_ACTIVE_MSG)

def ensure_open_business_day(db: Session, business_id: int, opener: Optional[User] = None, commit: bool = True) -> BusinessDay:
    """Fallback used only by write operations that need an active session to
    attribute themselves to (currently: sales checkout) — never by a
    read-only endpoint. Reuses whichever Business Day is currently active,
    exactly as new sales/expenses must (see get_active_business_day); only
    creates a brand-new session if none is active at all — this is the one
    path allowed to auto-open a session without the user pressing "Open
    Business Day" first. Race-safe: if two concurrent sales both find no
    active session and both try to create one, the partial unique index lets
    only one succeed — the loser simply re-reads and reuses the winner's new
    row instead of erroring."""
    active = get_active_business_day(db, business_id)
    if active:
        return active
    if not commit:
        # Checkout must keep auto-open, stock, sale lines, and both audits in
        # one transaction. Lock this tenant only while creating the missing
        # day, then re-check after acquiring the lock so two checkouts for
        # different products cannot race the partial unique active-day index.
        db.query(BusinessProfile).filter(BusinessProfile.id == business_id).with_for_update().one()
        active = get_active_business_day(db, business_id)
        if active:
            return active
        return _create_business_day_session(db, business_id, opener, auto=True, commit=False)
    try:
        return _create_business_day_session(db, business_id, opener, auto=True)
    except IntegrityError:
        db.rollback()
        active = get_active_business_day(db, business_id)
        if active:
            return active
        raise

def business_day_sales_query(db: Session, day: BusinessDay):
    """Sales belonging to this specific Business Day. Prefers the direct
    business_day_id relationship; falls back to the original timestamp-window
    match (this day's [opened_at, closed_at-or-now) span) only for rows
    recorded before that column existed and that the startup backfill
    couldn't already resolve — never a regression from previous behavior."""
    upper = day.closed_at or datetime.utcnow()
    return db.query(SaleModel).filter(
        SaleModel.business_id == day.business_id,
        or_(
            SaleModel.business_day_id == day.id,
            and_(SaleModel.business_day_id.is_(None), SaleModel.timestamp >= day.opened_at, SaleModel.timestamp <= upper),
        ),
    )

def business_day_expenses_query(db: Session, day: BusinessDay):
    upper = day.closed_at or datetime.utcnow()
    return db.query(Expense).filter(
        Expense.business_id == day.business_id,
        or_(
            Expense.business_day_id == day.id,
            and_(Expense.business_day_id.is_(None), Expense.created_at >= day.opened_at, Expense.created_at <= upper),
        ),
    )

def _business_day_refund_totals(db: Session, day: BusinessDay) -> Tuple[float, int, int]:
    """Refunds belong to the day they were PERFORMED on (RefundLine.
    business_day_id), never the original sale's day — see RefundTransaction
    model docstring. Returns (refund_total, refunded_units, refund_transaction_count)."""
    refund_total, refunded_units = (
        db.query(func.coalesce(func.sum(RefundLine.refund_amount), 0.0), func.coalesce(func.sum(RefundLine.quantity), 0))
        .filter(RefundLine.business_day_id == day.id, RefundLine.business_id == day.business_id)
        .one()
    )
    refund_transaction_count = (
        db.query(func.count(RefundTransaction.id))
        .filter(RefundTransaction.business_day_id == day.id, RefundTransaction.business_id == day.business_id)
        .scalar() or 0
    )
    return float(refund_total or 0), int(refunded_units or 0), int(refund_transaction_count)

def _business_day_financials(db: Session, day: BusinessDay) -> Dict[str, Any]:
    """The canonical per-Business-Day P&L — the ONE place a single session's
    sales/COGS/profit are computed, shared by the live current-day endpoint
    and the close snapshot so those can never disagree.

    Ownership is by business_day_id (business_day_sales_query /
    business_day_expenses_query), not by a timestamp window — a session that
    spans midnight, or two sessions on one calendar date, still attribute
    every row to exactly the session that owns it.

    COGS uses only each sale's own unit_cost_at_sale snapshot. Legacy NULL
    snapshots remain explicitly unknown, so a later catalog cost change can
    never move a closed day's reported profit."""
    sales_q = business_day_sales_query(db, day)
    gross_sales, transactions, units_sold = sales_q.with_entities(
        func.coalesce(func.sum(SaleModel.total_price), 0.0),
        func.count(func.distinct(checkout_key_expr())),
        func.coalesce(func.sum(SaleModel.quantity), 0),
    ).one()
    cost_expr = SaleModel.unit_cost_at_sale
    known_gross_cogs = sales_q.with_entities(
        func.coalesce(func.sum(SaleModel.quantity * cost_expr), 0.0),
    ).scalar() or 0.0
    unknown_sale_row = sales_q.filter(SaleModel.unit_cost_at_sale.is_(None)).with_entities(
        func.count(SaleModel.id), func.coalesce(func.sum(SaleModel.quantity), 0),
    ).one()
    expenses_total, expense_count = business_day_expenses_query(db, day).with_entities(
        func.coalesce(func.sum(Expense.amount), 0.0), func.count(Expense.id),
    ).one()
    refund_total, refunded_units, _ = _business_day_refund_totals(db, day)
    refund_cogs = float(
        db.query(func.coalesce(func.sum(RefundLine.refund_cost), 0.0))
        .filter(RefundLine.business_day_id == day.id, RefundLine.business_id == day.business_id)
        .scalar() or 0.0
    )
    unknown_refund_row = db.query(
        func.count(RefundLine.id), func.coalesce(func.sum(RefundLine.quantity), 0),
    ).filter(
        RefundLine.business_day_id == day.id, RefundLine.business_id == day.business_id,
        RefundLine.refund_cost.is_(None),
    ).one()
    net_sales = float(gross_sales or 0) - refund_total
    known_net_cogs = float(known_gross_cogs or 0) - refund_cogs
    unknown_sale_lines, unknown_sale_units = int(unknown_sale_row[0] or 0), int(unknown_sale_row[1] or 0)
    unknown_refund_lines, unknown_refund_units = int(unknown_refund_row[0] or 0), int(unknown_refund_row[1] or 0)
    cogs_complete = unknown_sale_lines == 0 and unknown_refund_lines == 0
    gross_profit = round(net_sales - known_net_cogs, 2) if cogs_complete else None
    net_profit = round(gross_profit - float(expenses_total or 0), 2) if gross_profit is not None else None
    return {
        "sales": round(net_sales, 2), "cogs": round(known_net_cogs, 2) if cogs_complete else None, "gross_profit": gross_profit,
        "expenses": round(float(expenses_total or 0), 2), "net_profit": net_profit,
        "profit_margin_percent": round((net_profit / net_sales) * 100, 1) if net_profit is not None and net_sales > 0 else None,
        "transactions": int(transactions or 0), "units_sold": int(units_sold or 0) - refunded_units,
        "expense_count": int(expense_count or 0),
        "gross_sales": round(float(gross_sales or 0), 2), "gross_units_sold": int(units_sold or 0),
        "refund_total": round(refund_total, 2), "refunded_units": refunded_units,
        "cogs_complete": cogs_complete, "known_cogs": round(known_net_cogs, 2),
        "unknown_cogs_sale_lines": unknown_sale_lines, "unknown_cogs_sale_units": unknown_sale_units,
        "unknown_cogs_refund_lines": unknown_refund_lines, "unknown_cogs_refund_units": unknown_refund_units,
    }

def serialize_business_day(day: Optional[BusinessDay], db: Session) -> Optional[dict]:
    if not day:
        return None
    # SQL aggregates instead of loading every Sale/Expense row for the day
    # just to sum them in Python (see performance refactor, section 22) —
    # same filter condition business_day_sales_query()/business_day_expenses_
    # query() use, just wrapped in SUM/COUNT so the database does the work.
    gross_sales, transactions, items_sold = business_day_sales_query(db, day).with_entities(
        func.coalesce(func.sum(SaleModel.total_price), 0.0), func.count(func.distinct(checkout_key_expr())), func.coalesce(func.sum(SaleModel.quantity), 0),
    ).one()
    expenses_total = business_day_expenses_query(db, day).with_entities(
        func.coalesce(func.sum(Expense.amount), 0.0),
    ).scalar()
    refund_total, refunded_units, refund_transaction_count = _business_day_refund_totals(db, day)
    return {
        "id": day.id, "date": day.date, "status": day.status, "is_open": day.is_open,
        "opened_at": to_utc_iso(day.opened_at), "opened_by_name": day.opened_by_name, "opened_by_role": day.opened_by_role,
        "closed_at": to_utc_iso(day.closed_at), "closed_by_name": day.closed_by_name, "closed_by_role": day.closed_by_role,
        "reopen_count": day.reopen_count,
        # net_sales keeps its established name/meaning to existing callers
        # (now correctly net of same-day refunds) — gross_sales/refund_total
        # are additive so a day's refund activity is never hidden.
        "net_sales": round(gross_sales - refund_total, 2), "gross_sales": round(gross_sales, 2),
        "refund_total": round(refund_total, 2), "refunded_units": refunded_units, "refund_transaction_count": refund_transaction_count,
        "transactions": transactions, "items_sold": items_sold,
        "expenses_total": expenses_total,
    }

def is_high_priority_product(db: Session, product: "Product") -> bool:
    """Heuristic for 'high-selling/high-value/business-critical product'
    (section 4 — Cauldra should not treat every product equally when
    deciding whether a stockout is push-worthy). There is no explicit
    product-importance flag in Cauldra's schema, so this ranks a product by
    its own recent sales velocity against the rest of the business's
    catalog — computed on demand, only at the rare moment a product actually
    reaches zero stock, never on every request. A product is high-priority
    when its units sold in the last 30 days are at or above double the
    business's own average UNITS-SOLD-PER-CATALOG-PRODUCT: simple,
    explainable, and self-calibrating per business rather than a fixed
    global number that would be meaningless across very different catalogs
    and sizes. The average is deliberately taken over every product in the
    catalog (not just the ones with any sales) — dividing only by products
    that sold something would make a business's single top seller
    mathematically indistinguishable from "average" whenever nothing else
    has sold recently, which defeats the whole comparison."""
    since = datetime.utcnow() - timedelta(days=30)
    per_product = dict(
        db.query(SaleModel.product_id, func.sum(SaleModel.quantity))
        .filter(SaleModel.business_id == product.business_id, SaleModel.timestamp >= since)
        .group_by(SaleModel.product_id).all()
    )
    this_product_units = per_product.get(product.id, 0)
    if not this_product_units:
        return False  # a product that hasn't sold can't be "high-selling", regardless of stockout
    catalog_size = db.query(func.count(Product.id)).filter(Product.business_id == product.business_id).scalar() or 1
    average = sum(per_product.values()) / catalog_size
    return this_product_units >= average * 2

def check_inventory_notifications(db: Session, business_id: int, product: "Product") -> None:
    """Fires/resolves the two quantity-driven inventory notification types
    for one product — called after every write that changes
    Product.quantity (sale checkout, stock adjustment/restock, transfer).
    Low stock is always Important/in-app only (section 4 — never externally
    pushed, expected day-to-day inventory activity). A stockout (quantity
    reaches zero) escalates to Critical/push, but only for a product this
    business's own sales data marks high-priority — see
    is_high_priority_product(). Either condition resolves itself the moment
    it's no longer true, so a restock frees the dedup key for a future
    recurrence instead of leaving it permanently silenced."""
    low_key, stockout_key = f"low_stock:{product.id}", f"stockout:{product.id}"
    if product.quantity <= 0:
        resolve_notifications(db, low_key, business_id)  # superseded by the more severe condition below
        if is_high_priority_product(db, product):
            create_notification(
                db, business_id=business_id, category="inventory", severity="critical", type="CRITICAL_STOCKOUT",
                title="Critical stockout", message=f"{product.name} is now out of stock.",
                related_entity_type="product", related_entity_id=product.id,
                deep_link=f"inventory:{product.id}", dedup_key=stockout_key, stage="out_of_stock",
            )
    elif product.quantity <= product.min_stock_level:
        resolve_notifications(db, stockout_key, business_id)
        create_notification(
            db, business_id=business_id, category="inventory", severity="important", type="LOW_STOCK",
            title="Low stock", message=f"{product.name} has fallen below its reorder level ({product.quantity} remaining, minimum {product.min_stock_level}).",
            related_entity_type="product", related_entity_id=product.id,
            deep_link=f"inventory:{product.id}", dedup_key=low_key, stage="low",
        )
    else:
        resolve_notifications(db, low_key, business_id)
        resolve_notifications(db, stockout_key, business_id)

def check_expiry_notifications(db: Session, business_id: int) -> None:
    """Product-expiry warnings (Important/in-app — never pushed; a slower-
    moving condition than a stockout). Lazily evaluated once per notification
    -center read, same trigger point the old build_alerts() used, rather
    than a dedicated write hook (expiry isn't driven by any single write)."""
    now = datetime.utcnow()
    for p in db.query(Product).filter(Product.business_id == business_id, Product.expiry_date.isnot(None)).all():
        days = (p.expiry_date.date() - now.date()).days
        if 0 <= days <= 60:
            create_notification(
                db, business_id=business_id, category="inventory", severity="important", type="PRODUCT_EXPIRING",
                title="Product approaching expiry", message=f"{p.name} expires in {days} day(s). Please review this stock.",
                related_entity_type="product", related_entity_id=p.id, deep_link=f"inventory:{p.id}",
                dedup_key=f"expiry:{p.id}", stage=str(days),
            )

# Severity threshold for section 6's "unusually significant" supplier price
# increase — kept as one named constant rather than a magic number so it can
# be tuned without hunting through the notification logic.
SUPPLIER_PRICE_SEVERE_INCREASE_PCT = 20.0

def check_price_change_notification(db: Session, source: "PriceMonitorSource", old_price: Optional[float], new_price: Optional[float]) -> None:
    """Section 6 — a normal supplier price change is Important/in-app only.
    It escalates to Critical/push ONLY when the increase exceeds
    SUPPLIER_PRICE_SEVERE_INCREASE_PCT *and* affects a product this
    business's own sales data marks high-priority (is_high_priority_product())
    — never on the size of the price move alone, and never on a decrease."""
    if not old_price or old_price <= 0 or new_price is None:
        return  # no prior price to compare against — nothing to notify about yet
    change_pct = ((new_price - old_price) / old_price) * 100
    if abs(change_pct) < 0.5:
        return  # rounding noise, not a real change
    product = db.query(Product).filter(Product.id == source.product_id).first() if source.product_id else None
    supplier = db.query(Supplier).filter(Supplier.id == source.supplier_id).first() if source.supplier_id else None
    product_name = product.name if product else "a monitored item"
    severe = change_pct >= SUPPLIER_PRICE_SEVERE_INCREASE_PCT and product is not None and is_high_priority_product(db, product)
    direction = "increased" if change_pct > 0 else "decreased"
    create_notification(
        db, business_id=source.business_id, category="supplier",
        severity="critical" if severe else "important",
        type="SUPPLIER_PRICE_INCREASE_SEVERE" if severe else "SUPPLIER_PRICE_CHANGE",
        title="Major supplier price increase" if severe else "Supplier price changed",
        message=(f"Supplier cost for {product_name} increased significantly ({change_pct:+.0f}%). Review the impact on your margin."
                 if severe else f"{supplier.name if supplier else 'A supplier'}'s price for {product_name} {direction} by {abs(change_pct):.0f}%."),
        related_entity_type="supplier", related_entity_id=source.supplier_id, deep_link=f"supplier:{source.supplier_id}",
        dedup_key=f"price_change:{source.id}:{datetime.utcnow().date().isoformat()}",
    )

# Section 3 — checked lowest-first so a usage jump that skips straight past
# multiple stages (a large batch of AI calls in one go) only ever fires the
# single most-severe stage actually reached, never every stage it passed
# through on the way down.
AI_CREDIT_THRESHOLDS = (0, 10, 20)

def check_ai_credit_notifications(db: Session, business: "BusinessProfile") -> None:
    """Section 3 — staged AI-credit threshold notifications (20% -> 10% ->
    0%), each firing exactly once per billing period (section 17's "must
    only fire once per billing cycle"). billing_period_start is folded into
    dedup_key, so a new billing period automatically gets a fresh key space
    with no explicit reset step required."""
    summary = usage_summary(db, business)
    included = summary["included_ai_credits"]
    if not included:
        return  # this plan has no AI-credit allowance to threshold against
    remaining_pct = (summary["remaining_ai_credits"] / included) * 100
    period = summary["billing_period_start"]
    for threshold in AI_CREDIT_THRESHOLDS:
        if remaining_pct <= threshold:
            if threshold == 0:
                title, message = "AI credits exhausted", "You've used all available AI credits for this billing period."
            elif threshold == 10:
                title, message = "AI credits critically low", "Only 10% of your AI credits remain. Review your usage to avoid interruption or unexpected overage."
            else:
                title, message = "AI credits running low", "You have 20% of your AI credits remaining for this billing period."
            create_notification(
                db, business_id=business.id, category="ai_usage", severity="critical", type=f"AI_CREDITS_{threshold}",
                title=title, message=message, deep_link="ai_center",
                dedup_key=f"ai_credits:{business.id}:{period}:{threshold}", stage=str(threshold),
            )
            break

# Section 2 — checked most-urgent-first, same "only the one crossed stage
# fires" rule as AI credits: a subscription that somehow jumps straight from
# 10 days to 2 days remaining (server downtime, clock skew) only ever gets
# the 1-day warning, not a backlog of 7-day/3-day ones for a moment already past.
SUBSCRIPTION_COUNTDOWN_STAGES = (1, 3, 7)

def check_subscription_countdown_notifications(db: Session, business: "BusinessProfile") -> None:
    """Section 2 — pre-expiration countdown warnings (7/3/1 days remaining).
    Actual expiration itself is handled at the authoritative point
    refresh_subscription_status() transitions status to expired/past_due,
    not here — this only ever fires for a subscription that is STILL
    trialing/active but approaching its boundary, which is why it needs a
    periodic background check (see notification_sweep_loop()) rather than a
    request-time hook: nothing else in the app naturally "happens" 3 days
    before a renewal date, so without this loop a business that doesn't
    open Cauldra in that window would never be warned before losing access."""
    sub = get_or_create_subscription(db, business, commit=False)
    if sub.status not in ("trialing", "active"):
        return
    period_end = sub.trial_end_at if sub.status == "trialing" else sub.current_period_end
    if not period_end:
        return
    days_remaining = (period_end.date() - datetime.utcnow().date()).days
    if days_remaining < 0:
        return  # refresh_subscription_status() owns the transition past this point
    period_anchor = period_end.date().isoformat()
    for stage in SUBSCRIPTION_COUNTDOWN_STAGES:
        if days_remaining <= stage:
            create_notification(
                db, business_id=business.id, category="subscription", severity=("important" if stage == 7 else "critical"),
                type=f"SUBSCRIPTION_EXPIRY_{stage}D", title="Subscription ending soon",
                message=f"Your Cauldra subscription expires in {stage} day{'s' if stage != 1 else ''}. Renew your plan to avoid interruption.",
                deep_link="subscription", dedup_key=f"sub_expiry_warning:{business.id}:{period_anchor}:{stage}", stage=str(stage),
            )
            break

NOTIFICATION_SWEEP_INTERVAL_SECONDS = int(os.getenv("SUPPLY_AI_NOTIFICATION_SWEEP_SECONDS", str(30 * 60)))

async def notification_sweep_loop():
    """Periodic background sweep for the one notification category that is
    genuinely time-based rather than tied to any single write: subscription
    day-countdown warnings (see check_subscription_countdown_notifications()
    — the whole reason this needs a background trigger is so a business
    gets warned even while nobody has opened Cauldra). Runs in-process for
    the app's own lifetime; every check inside is dedup-key-protected via
    create_notification(), so two overlapping sweeps (a slow tick plus the
    next one firing, or multiple worker processes each running their own
    loop) can only ever produce redundant no-op work, never a duplicate
    notification. Deliberately has no external scheduler dependency (no
    APScheduler/Celery/cron) — a single asyncio task the app owns."""
    while True:
        try:
            await asyncio.sleep(NOTIFICATION_SWEEP_INTERVAL_SECONDS)
            db = SessionLocal()
            try:
                for business in db.query(BusinessProfile).all():
                    try:
                        check_subscription_countdown_notifications(db, business)
                        db.commit()
                    except Exception:
                        db.rollback()
            finally:
                db.close()
        except Exception:
            pass  # a sweep-loop error must never kill the background task permanently

@app.on_event("startup")
async def _start_notification_sweep():
    asyncio.create_task(notification_sweep_loop())

def gemini_text_response(system_prompt: str, user_prompt: str) -> str:
    if not gemini_client:
        raise HTTPException(status_code=503, detail="Gemini AI is not configured. Add GEMINI_API_KEY to the server environment.")
    try:
        response = gemini_client.models.generate_content(model=GEMINI_MODEL, contents=f"{system_prompt}\n\n{user_prompt}")
        text = getattr(response, "text", None) or ""
        if not text.strip(): raise ValueError("Empty Gemini response")
        return text.strip()
    except Exception as exc:
        raise HTTPException(status_code=502, detail="Gemini could not complete that AI operation right now. Please try again.") from exc

def openai_json_response(prompt: str, image_data: Optional[str] = None) -> dict:
    if not openai_client:
        raise HTTPException(status_code=503, detail="OpenAI integration is not configured. Add OPENAI_API_KEY to the server environment.")
    content = [{"type": "input_text", "text": prompt}]
    if image_data:
        content.append({"type": "input_image", "image_url": image_data, "detail": "high"})
    schema = {
        "type": "object",
        "properties": {
            "supplier_name": {"type": "string"},
            "invoice_number": {"type": "string"},
            "invoice_date": {"type": "string"},
            "items": {"type": "array", "items": {"type": "object", "properties": {
                "description": {"type": "string"}, "quantity": {"type": "number"}, "unit_price": {"type": "number"}
            }, "required": ["description", "quantity", "unit_price"], "additionalProperties": False}},
            "subtotal": {"type": "number"}, "total": {"type": "number"}
        },
        "required": ["supplier_name", "invoice_number", "invoice_date", "items", "subtotal", "total"],
        "additionalProperties": False,
    }
    resp = openai_client.responses.create(
        model=OPENAI_MODEL,
        input=[{"role": "user", "content": content}],
        text={"format": {"type": "json_schema", "name": "invoice_extraction", "schema": schema, "strict": True}},
    )
    text = getattr(resp, "output_text", "")
    try:
        return json.loads(text)
    except Exception as exc:
        raise HTTPException(status_code=502, detail="The invoice could not be interpreted reliably. Please review it manually.") from exc

# -----------------------------------------------------------------------------
# SCHEMAS
# -----------------------------------------------------------------------------
class RegisterBusinessRequest(BaseModel):
    company_name: str
    email: EmailStr
    phone: str
    address: Optional[str] = None
    currency: str = "USD ($)"
    tax_id: Optional[str] = None
    firstname: str
    lastname: str
    owner_email: EmailStr
    owner_phone: str
    username: str
    password: str
    position: str = "Admin"
    country: Optional[str] = None
    country_code: Optional[str] = None
    locale: Optional[str] = None
    language: Optional[str] = "en"
    timezone: Optional[str] = "UTC"
    phone_country_code: Optional[str] = None
    plan: Optional[str] = None
    billing_interval: Optional[str] = None
    payment_reference: str

class BusinessVerifyRequest(BaseModel):
    business_id: str

class AdminLoginRequest(BaseModel):
    business_id: str
    username: str
    password: str

class EmployeeLoginRequest(BaseModel):
    business_id: str
    username: str
    password: str
    selected_role: str

class CreateUserRequest(BaseModel):
    username: str
    password: str
    role: str
    firstname: str
    lastname: str
    email: EmailStr
    phone: str
    position: str

class BusinessProfileSchema(BaseModel):
    company_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    currency: str = "USD ($)"
    tax_id: Optional[str] = None
    country: Optional[str] = None
    country_code: Optional[str] = None
    locale: Optional[str] = "en-US"
    language: Optional[str] = "en"
    timezone: Optional[str] = "UTC"
    phone_country_code: Optional[str] = None
    # Read-only identity fields. Sent back on every GET/POST response so the
    # frontend's in-memory businessProfile object stays complete (id,
    # business_id, business_code) no matter which endpoint last populated it —
    # previously omitted here, these were silently stripped from the response
    # by response_model filtering even though serialize_business() computes
    # them, which broke anything reading businessProfile.business_code after
    # visiting Team Management or saving Business Profile settings (e.g. the
    # "copy business code" button). Ignored on incoming POST bodies; the
    # update handler never reads them.
    id: Optional[int] = None
    business_id: Optional[int] = None
    business_code: Optional[str] = None

class WarehouseCreate(BaseModel):
    name: str

class WarehouseUpdate(BaseModel):
    name: str
    is_active: Optional[bool] = None

class ProductCreate(BaseModel):
    barcode: Optional[str] = None
    sku: Optional[str] = None
    name: str
    category: str
    size: Optional[str] = None
    quantity: int
    min_stock_level: int
    cost_price: float
    wholesale_price: Optional[float] = 0.0
    retail_price: float
    warehouse: Optional[str] = "Main Central Warehouse"
    expiry_date: Optional[datetime] = None
    # Optional: set by the offline sync client so a retried submission of the
    # same locally-created product (e.g. after a dropped connection) is
    # recognized instead of creating a duplicate product.
    client_ref: Optional[str] = None
    # V25 smart duplicate detection: the id of the EXISTING product the user
    # explicitly chose to override via "Add as different product". Never a
    # generic force=true - the server verifies this id belongs to the same
    # business AND matches a real duplicate candidate for this exact form
    # state before letting the create/edit proceed.
    duplicate_override_candidate_id: Optional[int] = None

class ProductUpdate(BaseModel):
    barcode: Optional[str] = None
    name: Optional[str] = None
    sku: Optional[str] = None
    category: Optional[str] = None
    size: Optional[str] = None
    quantity: Optional[int] = Field(default=None, ge=0)
    min_stock_level: Optional[int] = Field(default=None, ge=0)
    cost_price: Optional[float] = Field(default=None, ge=0)
    wholesale_price: Optional[float] = Field(default=None, ge=0)
    retail_price: Optional[float] = Field(default=None, ge=0)
    warehouse: Optional[str] = None
    expiry_date: Optional[datetime] = None
    client_ref: Optional[str] = None
    duplicate_override_candidate_id: Optional[int] = None

class StockUpdate(BaseModel):
    quantity_change: int
    action_type: str = "MANUAL"

class StockTransfer(BaseModel):
    from_warehouse: str
    to_warehouse: str
    quantity: int = Field(ge=1)

class SupplierCreate(BaseModel):
    name: str
    contact_email: Optional[str] = None
    phone: str
    lead_time_days: Optional[int] = 3

class ExpenseCreate(BaseModel):
    category: str
    amount: float
    payment_source: Optional[str] = None
    note: Optional[str] = None
    # Optional: only meaningful for a FUTURE offline-sync client, which does not
    # exist yet. Accepted now (and made unique per business at the DB layer via
    # the sync-retry lookup in create_expense()) purely so that feature can be
    # added later without a request-schema change; today every caller omits it.
    client_ref: Optional[str] = None

class PODraftUpdate(BaseModel):
    details: Optional[str] = None
    items_summary: Optional[str] = None
    supplier_id: Optional[int] = None

class NotificationPreferenceItem(BaseModel):
    category: str
    enabled: bool

class NotificationPreferenceUpdate(BaseModel):
    preferences: List[NotificationPreferenceItem]

class PushSubscriptionKeys(BaseModel):
    p256dh: str
    auth: str

class PushSubscribeRequest(BaseModel):
    endpoint: str
    keys: PushSubscriptionKeys
    user_agent: Optional[str] = None

class PushUnsubscribeRequest(BaseModel):
    endpoint: str

class PasswordChangeRequest(BaseModel):
    current_password: str
    new_password: str

class AdminPasswordResetRequest(BaseModel):
    new_password: str

class MarginRequest(BaseModel):
    name: str
    category: str
    cost_price: float

class ChatRequest(BaseModel):
    message: str

class CatalogBarcodeLookupRequest(BaseModel):
    barcode: str

class InvoiceScanRequest(BaseModel):
    image_data: str
    file_name: Optional[str] = None

class AccountActionRequestCreate(BaseModel):
    target_user_id: int
    action: str

class BusinessDayReopenRequestCreate(BaseModel):
    reason: str

class BusinessDayReopenResolution(BaseModel):
    note: Optional[str] = None

class RecordAdjustmentRequest(BaseModel):
    reason: str
    amount_delta: float = 0.0
    quantity_delta: Optional[int] = 0

class XlsxColumnSpec(BaseModel):
    key: str
    label: str
    type: str = "text"  # text | number | decimal | currency | percent | date | datetime | bool
    width: Optional[float] = None

class XlsxSheetSpec(BaseModel):
    name: str
    columns: List[XlsxColumnSpec]
    rows: List[List[Any]]
    title: Optional[str] = None
    metadata: Optional[List[List[str]]] = None

class XlsxExportRequest(BaseModel):
    filename: str
    report_title: Optional[str] = None
    sheets: List[XlsxSheetSpec]

class PriceSourceCreate(BaseModel):
    supplier_id: int
    product_id: int
    source_type: str = "manual"
    source_url: Optional[str] = None
    initial_price: Optional[float] = None

class ManualPriceUpdate(BaseModel):
    price: float

class SalesCheckoutItem(BaseModel):
    product_id: int
    quantity: int
    # Catalog modes are authoritative server choices: the browser names the
    # mode, then checkout reads the price from the locked Product row. A
    # submitted unit_price is considered only for the explicit negotiated
    # mode, whose permission, bounds, and reason are enforced in checkout.
    price_mode: Literal["retail", "wholesale", "negotiated"] = "retail"
    unit_price: Optional[float] = Field(default=None, gt=0, allow_inf_nan=False)
    negotiated_reason: Optional[str] = Field(default=None, max_length=200)

    @field_validator("unit_price", mode="before")
    @classmethod
    def _reject_non_finite_unit_price(cls, v):
        # A literal NaN/Infinity survives JSON parsing as a real Python float
        # (the stdlib json module accepts those tokens) and would otherwise
        # reach the gt/allow_inf_nan checks above as that literal value —
        # which correctly fails validation, but FastAPI's own 422 response
        # then crashes trying to JSON-encode the rejected value itself
        # (Starlette's JSONResponse forbids NaN/Infinity in its output).
        # Normalizing to an ordinary value that still fails gt=0 keeps the
        # rejection a clean 422 instead of an unhandled 500.
        if isinstance(v, float) and not math.isfinite(v):
            return -1.0
        return v

class SalesCheckoutRequest(BaseModel):
    items: List[SalesCheckoutItem]
    # Optional: one id for the whole cart, set by the offline sync client.
    # Retrying the same checkout (same client_ref) returns the original result
    # instead of decrementing inventory and creating sale rows a second time.
    client_ref: Optional[str] = None

class RefundLineRequest(BaseModel):
    sale_id: int
    quantity: int = Field(gt=0)
    # Independent per line — a multi-item refund may return some goods to
    # stock (customer return) and not others (damaged/consumed) at once.
    restock: bool = True

class RefundRequest(BaseModel):
    lines: List[RefundLineRequest]
    reason: Optional[str] = None
    note: Optional[str] = None
    # Idempotency key for this refund submission — a rapid double-click on
    # "Confirm Refund" that races two requests to the server with the same
    # value can only ever create one RefundTransaction (see create_refund).
    client_ref: Optional[str] = None

class BusinessBrainRecommendationAction(BaseModel):
    action: str

class PresenceHeartbeatRequest(BaseModel):
    session_id: Optional[str] = None

class PriceListUploadRequest(BaseModel):
    supplier_id: int
    product_id: Optional[int] = None
    file_name: str
    file_data: str

class ForgotPasswordRequest(BaseModel):
    business_id: str
    username: str
    channel: str = "email"
    email: Optional[EmailStr] = None
    phone: Optional[str] = None

class PasswordResetRequest(BaseModel):
    recovery_id: str
    code: str
    new_password: str

class AuditLogSchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    actor_username: Optional[str]
    actor_role: Optional[str]
    action: str
    target_username: Optional[str]
    description: str
    created_at: datetime

# -----------------------------------------------------------------------------
# AUTH
# -----------------------------------------------------------------------------
@app.post("/auth/register-business")
def register_business(data: RegisterBusinessRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    validate_password_strength(data.password)
    normalized = normalize_username(data.username)
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "register-business-ip", client_ip)

    # SECURITY / SPEC: the new-business flow is pay-before-register. A business
    # is never created without an already-verified Paystack card authorization
    # to point to. The plan and billing interval are taken ONLY from that
    # server-side verified record — never from data.plan/data.billing_interval
    # — so the frontend can never smuggle a different (e.g. cheaper) plan in
    # at registration time than the one the card was actually verified for.
    payment_reference = str(data.payment_reference or "").strip()
    if not payment_reference:
        raise HTTPException(status_code=400, detail="Please verify a payment method before registering your business.")
    auth_row = db.query(OnboardingAuthorization).filter(OnboardingAuthorization.paystack_reference == payment_reference).first()
    if not auth_row:
        raise HTTPException(status_code=404, detail="We couldn't find that payment verification. Please choose your plan and verify a card again.")
    if auth_row.status == "consumed":
        raise HTTPException(status_code=409, detail="This payment verification has already been used to register a business.")
    if auth_row.status != "verified":
        raise HTTPException(status_code=400, detail="Please complete payment verification before registering your business.")
    if datetime.utcnow() > auth_row.verified_at + timedelta(hours=ONBOARDING_AUTHORIZATION_CONSUME_HOURS):
        raise HTTPException(status_code=410, detail="Your payment verification has expired. Please choose your plan and verify a card again.")

    plan = str(auth_row.plan or "").strip().lower()
    if plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a subscription plan before registering your business.")
    billing_interval = str(auth_row.billing_interval or "monthly").strip().lower()
    if billing_interval not in ("monthly", "annual"):
        billing_interval = "monthly"

    country_values = canonical_business_country_values(data.country, data.country_code, data.language)
    language = str(data.language or "en").strip().lower() or "en"
    canonical_owner_phone = canonical_phone_for_country(data.owner_phone, country_values["phone_country_code"])

    # Risk signal only (never a hard block): flag when this owner identity has
    # started a trial before, for later review. A shared email/phone across
    # legitimate multi-business owners must not be auto-punished.
    prior_trial_admin = db.query(User).join(BusinessSubscription, BusinessSubscription.business_id == User.business_id).filter(
        User.role == "admin", BusinessSubscription.trial_start_at.isnot(None),
        or_(User.email == str(data.owner_email), User.phone == canonical_owner_phone)
    ).first()

    # Registration is one database transaction: consuming the verified card
    # authorization, creating the tenant/admin/default warehouse/trial, writing
    # the core audit trail and creating the first refresh session either all
    # commit together or all roll back together. The compare-and-swap UPDATE
    # also locks the authorization row on PostgreSQL until this transaction
    # finishes, so concurrent requests cannot create two businesses from one
    # payment reference.
    now = datetime.utcnow()
    trial_end = now + timedelta(days=PLAN_CONFIG[plan]["trial_days"])
    hashed_password = hash_password(data.password)
    refresh_raw = secrets.token_urlsafe(48)
    try:
        rows_updated = (
            db.query(OnboardingAuthorization)
            .filter(OnboardingAuthorization.id == auth_row.id, OnboardingAuthorization.status == "verified")
            .update({"status": "consumed", "consumed_at": now}, synchronize_session=False)
        )
        if rows_updated == 0:
            raise HTTPException(status_code=409, detail="This payment verification has already been used. Please choose your plan and verify a card again.")

        # Uniqueness is business-scoped, and the initial business has no users yet.
        new_biz = BusinessProfile(
            business_code=generate_dynamic_business_code(db, data.company_name),
            company_name=data.company_name,
            email=str(data.email),
            phone=canonical_phone_for_country(data.phone, country_values["phone_country_code"]),
            address=data.address,
            currency=country_values["currency"],
            tax_id=data.tax_id,
            country=country_values["country"],
            country_code=country_values["country_code"],
            locale=country_values["locale"],
            language=language,
            timezone=country_values["timezone"],
            phone_country_code=country_values["phone_country_code"],
            subscription_plan=plan,
            billing_interval=billing_interval,
        )
        db.add(new_biz)
        db.flush()

        admin = User(
            username=data.username.strip(), password=hashed_password, role="admin",
            firstname=data.firstname, lastname=data.lastname, position=data.position or "Admin",
            email=str(data.owner_email), phone=canonical_owner_phone, business_id=new_biz.id,
            must_change_password=False, disabled=False, auth_version=1
        )
        db.add(admin)
        db.flush()
        db.add(Warehouse(business_id=new_biz.id, name="Main Central Warehouse", is_active=True))

        # The already-verified card starts the selected trial in the same
        # transaction as the business. No partial tenant can exist without its
        # owner and subscription, and a rolled-back tenant does not consume the
        # payment authorization.
        new_sub = BusinessSubscription(
            business_id=new_biz.id, plan=plan, billing_interval=billing_interval, status="trialing",
            card_verified=True, trial_start_at=now, trial_end_at=trial_end,
            current_period_start=now, current_period_end=trial_end, next_billing_at=trial_end,
            trial_consent_at=now, paystack_customer_code=auth_row.paystack_customer_code,
            paystack_authorization_code=auth_row.paystack_authorization_code,
            card_last4=auth_row.card_last4, card_type=auth_row.card_type,
            card_exp_month=auth_row.card_exp_month, card_exp_year=auth_row.card_exp_year,
        )
        db.add(new_sub)
        db.flush()

        add_audit(db, admin, "BUSINESS_REGISTERED", f"Business and owner Admin account registered on the {PLAN_CONFIG[plan]['label']} plan ({billing_interval}), with payment method already verified.", admin)
        add_audit(db, admin, "TRIAL_STARTED", f"14-day free trial started on {PLAN_CONFIG[plan]['label']} ({billing_interval}) immediately at registration. No charge during trial.", admin)
        if prior_trial_admin:
            add_audit(db, admin, "TRIAL_RISK_SIGNAL", f"Owner identity (email/phone) previously started a trial for another business (ip={client_ip}). Flagged for review only; trial not restricted.", admin)
        db.add(RefreshSession(
            token_hash=hash_text(refresh_raw), user_id=admin.id, business_id=new_biz.id,
            expires_at=now + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS),
        ))
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise

    # Schedule the automatic post-trial charge with Paystack itself (same
    # pattern as /subscription/trial/confirm) — Cauldra does not invent its own
    # card-charging mechanism for the trial-to-paid conversion.
    plan_code = PLAN_CONFIG[plan].get("paystack_annual_plan_code" if billing_interval == "annual" else "paystack_monthly_plan_code")
    try:
        customer_code = auth_row.paystack_customer_code or paystack_get_or_create_customer(str(data.owner_email))
        if plan_code and auth_row.paystack_authorization_code:
            created = paystack_create_subscription(customer_code, plan_code, auth_row.paystack_authorization_code, start_date=trial_end)
            new_sub.paystack_subscription_code = created.get("subscription_code")
        else:
            add_audit(db, admin, "TRIAL_STARTED_WITHOUT_PAYSTACK_PLAN_CODE", f"Trial started for {PLAN_CONFIG[plan]['label']} ({billing_interval}) but no PAYSTACK_*_PLAN_CODE is configured for this plan/interval, so automatic conversion cannot be scheduled with Paystack yet. The trial will still expire correctly and require manual checkout to convert.", business_id=new_biz.id)
        new_sub.paystack_customer_code = customer_code
        new_sub.paystack_plan_code = plan_code or None
        db.commit()
    except Exception:
        db.rollback()
        try:
            add_audit(db, admin, "TRIAL_PAYSTACK_SUBSCRIPTION_CREATE_FAILED", "Business registered and trial started, but scheduling the recurring Paystack subscription failed. Automatic conversion at trial end may not occur; manual checkout will be needed.", business_id=new_biz.id)
            db.commit()
        except Exception:
            # The durable registration is already complete. A provider or
            # best-effort follow-up audit failure must not turn the successful
            # registration response into a retry that can only receive 409.
            db.rollback()

    access = issue_token(admin, db)
    set_refresh_cookie(response, refresh_raw)
    return {"access_token": access, "token_type": "bearer", "business_code": new_biz.business_code,
            "trial_end_at": to_utc_iso(trial_end), "card_last4": new_sub.card_last4, "card_type": new_sub.card_type,
            **serialize_business(new_biz), **serialize_user(admin)}

@app.post("/auth/verify-business")
def verify_business(data: BusinessVerifyRequest, request: Request, db: Session = Depends(get_db)):
    key = re.sub(r"\W", "", data.business_id or "").casefold()
    client_ip = request.client.host if request.client else "unknown"
    check_rate_limit(db, "business-verify-ip", client_ip)
    check_rate_limit(db, "business-verify-code", key or "blank")
    biz = get_business_by_code(db, data.business_id)
    if not biz:
        record_failure(db, "business-verify-ip", client_ip)
        record_failure(db, "business-verify-code", key or "blank")
        raise HTTPException(status_code=404, detail="Business not found. Please check the Business ID and try again.")
    clear_failures(db, "business-verify-ip", client_ip)
    return serialize_business(biz)

def authenticate_user_for_business(db: Session, business_code: str, username: str, password: str, role: Optional[str] = None, scope: str = "login", client_ip: Optional[str] = None) -> User:
    # --- TEMPORARY DEV DIAGNOSTIC (auth investigation) -----------------------
    diag_id = secrets.token_hex(4)
    username_fingerprint = hashlib.sha256(normalize_username(username).encode()).hexdigest()[:12]
    print(f"[auth-diag {diag_id}] LOGIN ATTEMPT scope={scope} pid={os.getpid()} db_target={engine.url.render_as_string(hide_password=True)} "
          f"business_code_received={business_code!r} username_fingerprint={username_fingerprint}")
    # --- END TEMPORARY DEV DIAGNOSTIC (continues below) ----------------------

    biz = get_business_by_code(db, business_code)

    print(f"[auth-diag {diag_id}] BUSINESS LOOKUP found={bool(biz)} business_db_id={getattr(biz, 'id', None)}")

    if not biz:
        print(f"[auth-diag {diag_id}] RESULT status=404 reason=business_not_found")
        raise HTTPException(status_code=404, detail="Business not found. Please check the Business ID and try again.")
    key = f"{biz.id}:{normalize_username(username)}"
    try:
        check_rate_limit(db, scope, key)
        if client_ip:
            check_rate_limit(db, scope + "-ip", client_ip)
    except HTTPException as rl_exc:
        print(f"[auth-diag {diag_id}] RATE LIMIT blocked=True status={rl_exc.status_code}")
        raise
    print(f"[auth-diag {diag_id}] RATE LIMIT blocked=False")

    user = next((u for u in db.query(User).filter(User.business_id == biz.id).all() if normalize_username(u.username) == normalize_username(username)), None)

    print(f"[auth-diag {diag_id}] USER LOOKUP found={bool(user)} user_db_id={getattr(user, 'id', None)} "
          f"user_business_id={getattr(user, 'business_id', None)} user_role={getattr(user, 'role', None)} "
          f"user_disabled={getattr(user, 'disabled', None)} password_hash_present={bool(getattr(user, 'password', None))}")

    password_ok = verify_password(password, user.password) if user else False
    print(f"[auth-diag {diag_id}] PASSWORD VERIFICATION result={password_ok}")

    if not user or not password_ok or user.disabled:
        record_failure(db, scope, key)
        if client_ip: record_failure(db, scope + "-ip", client_ip)
        if user:
            # Never notify for a nonexistent username — that reveals nothing
            # about any real account. Only a REAL account under repeated
            # attack fires this, and only at the exact moment Cauldra's
            # existing lockout actually triggers (see the helper's own
            # dedup/threshold check) — never on every subsequent locked-out
            # attempt.
            check_login_lockout_notification(db, user, scope, key)
        reason = "user_not_found" if not user else ("password_mismatch" if not password_ok else "user_disabled")
        print(f"[auth-diag {diag_id}] RESULT status=401 reason={reason}")
        raise HTTPException(status_code=401, detail="Incorrect username or password.")
    clear_failures(db, scope, key)
    if client_ip: clear_failures(db, scope + "-ip", client_ip)
    if role and user.role != role:
        print(f"[auth-diag {diag_id}] RESULT status=403 reason=role_mismatch expected={role} actual={user.role}")
        raise HTTPException(status_code=403, detail="This account does not have the selected role.")
    print(f"[auth-diag {diag_id}] RESULT status=200 reason=success")
    return user

@app.post("/auth/admin-login")
def admin_login(data: AdminLoginRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    print(f"[auth-diag] /auth/admin-login REQUEST RECEIVED pid={os.getpid()} port={request.url.port} db_target={engine.url.render_as_string(hide_password=True)}")  # TEMPORARY DEV DIAGNOSTIC
    user = authenticate_user_for_business(db, data.business_id, data.username, data.password, role="admin", scope="admin-login", client_ip=request.client.host if request.client else "unknown")
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    access = issue_token(user, db)
    set_refresh_cookie(response, create_refresh_session(db, user))
    return {"access_token": access, "token_type": "bearer", **serialize_business(biz), **serialize_user(user)}

@app.post("/auth/employee-login")
def employee_login(data: EmployeeLoginRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    if data.selected_role not in {"manager", "staff"}:
        raise HTTPException(status_code=400, detail="Select Manager or Staff for employee login.")
    user = authenticate_user_for_business(db, data.business_id, data.username, data.password, role=data.selected_role, scope="employee-login", client_ip=request.client.host if request.client else "unknown")
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    access = issue_token(user, db)
    set_refresh_cookie(response, create_refresh_session(db, user))
    return {"access_token": access, "token_type": "bearer", **serialize_business(biz), **serialize_user(user)}

@app.post("/token")
def login(form_data: OAuth2PasswordRequestForm = Depends(), request: Request = None, response: Response = None, db: Session = Depends(get_db)):
    # OAuth2 compatibility route: if a business code is supplied in 'scope', honor it.
    business_code = form_data.scopes[0] if form_data.scopes else ""
    if not business_code:
        raise HTTPException(status_code=400, detail="Business ID is required for token login.")
    user = authenticate_user_for_business(db, business_code, form_data.username, form_data.password, scope="token-login", client_ip=request.client.host if request and request.client else "unknown")
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    access = issue_token(user, db)
    if response is not None: set_refresh_cookie(response, create_refresh_session(db, user))
    return {"access_token": access, "token_type": "bearer", **serialize_business(biz), **serialize_user(user)}

@app.get("/auth/me")
def auth_me(user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    return {**serialize_business(biz), **serialize_user(user)}

@app.post("/auth/logout")
def auth_logout(response: Response, token: str = Depends(oauth2_scheme), refresh_token: Optional[str] = Cookie(default=None, alias=REFRESH_COOKIE_NAME), user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    payload, jti = token_from_payload(token)
    exp_ts = payload.get("exp")
    exp_dt = datetime.utcfromtimestamp(exp_ts) if exp_ts else datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    if not db.query(SessionRevocation).filter(SessionRevocation.jti == jti).first():
        db.add(SessionRevocation(jti=jti, user_id=user.id, expires_at=exp_dt)); db.commit()
    if refresh_token:
        rs = db.query(RefreshSession).filter(RefreshSession.token_hash==hash_text(refresh_token), RefreshSession.user_id==user.id, RefreshSession.revoked_at.is_(None)).first()
        if rs: rs.revoked_at = datetime.utcnow()
    # Sign-out is a presence/session event, not a business/audit event — it is
    # recorded on the user's PresenceSession row (signed_out_at) via the
    # separate /presence/logout call the frontend makes alongside this one,
    # not written into Activity History.
    db.commit(); clear_refresh_cookie(response)
    return {"message": "Signed out successfully."}

def _reject_refresh(response: Response) -> Response:
    clear_refresh_cookie(response)
    response.status_code = status.HTTP_204_NO_CONTENT
    return response

def _issue_refresh_success(db: Session, user: User) -> dict:
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    return {"access_token": issue_token(user, db), "token_type": "bearer", **serialize_business(biz), **serialize_user(user)}

@app.post("/auth/refresh")
def auth_refresh(response: Response, refresh_token: Optional[str] = Cookie(default=None, alias=REFRESH_COOKIE_NAME), db: Session = Depends(get_db)):
    # --- TEMPORARY DEV DIAGNOSTIC (auth investigation) -----------------------
    # Never logs the raw cookie/token value, its hash, or any other secret —
    # only which branch of this function executed and the non-sensitive facts
    # that decided it. Remove once the reload-logout investigation is closed.
    diag_id = secrets.token_hex(4)
    print(f"[auth-diag {diag_id}] /auth/refresh REQUEST RECEIVED pid={os.getpid()} cookie_present={bool(refresh_token)}")
    # --- END TEMPORARY DEV DIAGNOSTIC (continues below) ----------------------

    # A page can load without a session (new visitor, signed-out user, or an
    # expired browser cookie). That is a normal state, not an authentication
    # error, so return an explicit empty success response without a console 401.
    if not refresh_token:
        print(f"[auth-diag {diag_id}] RESULT status=204 branch=no_cookie_presented")
        return _reject_refresh(response)
    now = datetime.utcnow()
    presented_hash = hash_text(refresh_token)
    row = db.query(RefreshSession).filter(RefreshSession.token_hash == presented_hash).first()
    print(f"[auth-diag {diag_id}] SESSION ROW LOOKUP found={bool(row)}"
          + (f" revoked={bool(row.revoked_at)} has_replacement={bool(row.replaced_by_hash)} expired={row.expires_at <= now} row_id={row.id}" if row else ""))
    if not row:
        print(f"[auth-diag {diag_id}] RESULT status=204 branch=no_matching_session_row")
        return _reject_refresh(response)

    # Checked before attempting rotation, same as the prior implementation —
    # a disabled/deleted user's session must never be rotated *or* recovered
    # via the grace path below, so this short-circuits before either.
    user = db.query(User).filter(User.id == row.user_id, User.business_id == row.business_id).first()
    print(f"[auth-diag {diag_id}] USER LOOKUP found={bool(user)} disabled={getattr(user, 'disabled', None)}")
    if not user or user.disabled:
        db.query(RefreshSession).filter(RefreshSession.id == row.id, RefreshSession.revoked_at.is_(None)) \
            .update({RefreshSession.revoked_at: now}, synchronize_session=False)
        db.commit()
        print(f"[auth-diag {diag_id}] RESULT status=204 branch=user_missing_or_disabled")
        return _reject_refresh(response)

    # Atomic compare-and-swap: this UPDATE's WHERE clause is re-evaluated by
    # the database itself at the instant it runs, not against whatever `row`
    # looked like a moment ago — so if two requests present this same token
    # concurrently (e.g. two browser tabs refreshing around the same time),
    # the database guarantees only ONE of them can ever match and perform
    # the actual rotation below. The loser sees affected == 0 and falls
    # through to the grace-recovery check, never a corrupted/duplicated
    # rotation, via PostgreSQL's ordinary row-level locking. No application-
    # level lock is needed, so unrelated users/requests are never blocked by
    # this.)
    #
    # Revoking the old row and creating its replacement are committed
    # TOGETHER, in one transaction — not as two separate commits. Splitting
    # them would open a window where a concurrent loser could see "old row
    # is revoked, replaced_by_hash is set" before the replacement row it
    # points to actually exists yet, causing a spurious rejection instead of
    # the intended recovery (caught via reproduction while testing this).
    new_raw = secrets.token_urlsafe(48)
    new_hash = hash_text(new_raw)
    affected = db.query(RefreshSession).filter(
        RefreshSession.id == row.id,
        RefreshSession.revoked_at.is_(None),
        RefreshSession.expires_at > now,
    ).update({RefreshSession.revoked_at: now, RefreshSession.replaced_by_hash: new_hash}, synchronize_session=False)
    print(f"[auth-diag {diag_id}] ATOMIC ROTATION UPDATE affected_rows={affected}")

    if affected == 1:
        db.add(RefreshSession(token_hash=new_hash, user_id=user.id, business_id=user.business_id, expires_at=now + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)))
        db.commit()  # old-row revoke + replaced_by_hash + new row, all at once
        set_refresh_cookie(response, new_raw)
        print(f"[auth-diag {diag_id}] RESULT status=200 branch=normal_rotation sets_new_cookie=True")
        return _issue_refresh_success(db, user)

    # We lost the race, or this token was already invalid for some other
    # reason (expired, already revoked earlier, unknown). Roll back our own
    # no-op attempt cleanly, then re-read the row's current, committed state
    # and check — narrowly — whether this is the "another tab already
    # rotated this exact token moments ago" case.
    db.rollback()
    db.refresh(row)
    print(f"[auth-diag {diag_id}] ENTERED GRACE-CHECK BRANCH revoked={bool(row.revoked_at)} has_replacement={bool(row.replaced_by_hash)}")
    recovered_user = None
    if row.revoked_at and row.replaced_by_hash:
        elapsed_seconds = (now - row.revoked_at).total_seconds()
        print(f"[auth-diag {diag_id}] GRACE WINDOW elapsed_seconds={elapsed_seconds:.3f} limit={REFRESH_ROTATION_GRACE_SECONDS} within_window={elapsed_seconds <= REFRESH_ROTATION_GRACE_SECONDS}")
        if elapsed_seconds <= REFRESH_ROTATION_GRACE_SECONDS:
            # Exactly one hop: the immediate successor this row's own rotation
            # produced — never followed further, so a token from two or more
            # generations back can never recover through a longer chain.
            replacement = db.query(RefreshSession).filter(RefreshSession.token_hash == row.replaced_by_hash).first()
            print(f"[auth-diag {diag_id}] REPLACEMENT LOOKUP found={bool(replacement)}"
                  + (f" revoked={bool(replacement.revoked_at)} expired={replacement.expires_at <= now} same_user={replacement.user_id == row.user_id and replacement.business_id == row.business_id}" if replacement else ""))
            if (replacement and replacement.user_id == row.user_id and replacement.business_id == row.business_id
                    and not replacement.revoked_at and replacement.expires_at > now):
                candidate = db.query(User).filter(User.id == replacement.user_id, User.business_id == replacement.business_id).first()
                if candidate and not candidate.disabled:
                    recovered_user = candidate

    if not recovered_user:
        print(f"[auth-diag {diag_id}] RESULT status=204 branch=grace_recovery_failed")
        return _reject_refresh(response)
    print(f"[auth-diag {diag_id}] RESULT status=200 branch=grace_recovery_succeeded sets_new_cookie=False")

    # Grace recovery: hand this tab a fresh access token for the same user
    # and business, bound to the session the winning request already
    # created. Deliberately does NOT touch the cookie at all — not clearing
    # it (that would log the user out for no reason) and not setting a new
    # one either (the winning request's response already carries the
    # correct current refresh cookie; since we only ever stored a one-way
    # hash of that token, we have no way to reissue its actual raw value
    # here, and must not overwrite the browser's cookie with anything else).
    # No new RefreshSession row is created and no expiry is extended — this
    # path only ever reuses the row the other request already made.
    return _issue_refresh_success(db, recovered_user)

@app.post("/auth/change-password")
def change_password(data: PasswordChangeRequest, response: Response, user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    validate_password_strength(data.new_password)
    if not verify_password(data.current_password, user.password):
        raise HTTPException(status_code=400, detail="Incorrect current or temporary password.")
    if user.previous_password_hash and verify_password(data.new_password, user.previous_password_hash):
        raise HTTPException(status_code=400, detail="Last created password cannot be used.")
    user.previous_password_hash = user.password
    user.password = hash_password(data.new_password)
    user.must_change_password = False
    # Invalidate both refresh sessions and already issued access tokens.
    revoke_all_user_sessions(db, user)
    add_audit(db, user, "PASSWORD_CHANGED", "User changed their password.")
    db.commit()
    set_refresh_cookie(response, create_refresh_session(db, user))
    return {"message": "Password updated successfully.", "access_token": issue_token(user, db), "token_type": "bearer"}

@app.delete("/auth/account")
def delete_own_account(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "admin":
        raise HTTPException(status_code=403, detail="Business owner accounts cannot self-delete. Use the business recovery/closure process.")
    add_audit(db, user, "ACCOUNT_DELETED", "User deleted their own account.", user)
    user.disabled = True
    db.commit()
    return {"message": "Account disabled successfully."}

# -----------------------------------------------------------------------------
# PASSWORD RECOVERY / EMAIL
# -----------------------------------------------------------------------------
def send_recovery_email(to_email: str, username: str, code: str):
    api_key = os.getenv("RESEND_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="Email recovery is not configured. Add RESEND_API_KEY to the server environment.")
    import requests
    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "from": RESEND_FROM,
                "to": [to_email],
                "subject": f"{APP_NAME} password recovery code",
                "html": f"<p>Hello {username},</p><p>Your password recovery code is <strong>{code}</strong>.</p><p>This code expires in 10 minutes.</p>",
            },
            timeout=15,
        )
        if not r.ok:
            raise RuntimeError(r.text)
    except Exception as exc:
        raise HTTPException(status_code=502, detail="We could not send the recovery email right now.") from exc

def send_recovery_sms(phone: str, code: str):
    termii_key = os.getenv("TERMII_API_KEY", "").strip()
    twilio_sid = os.getenv("TWILIO_ACCOUNT_SID", "").strip()
    twilio_token = os.getenv("TWILIO_AUTH_TOKEN", "").strip()
    twilio_from = os.getenv("TWILIO_FROM", "").strip()
    # Termii-first is intentional. If Termii isn't configured, fall back to Twilio if configured.
    if termii_key:
        import requests
        try:
            r = requests.post(f"{TERMII_BASE_URL}/api/sms/send", json={
                "to": phone, "from": os.getenv("TERMII_SENDER_ID", "Cauldra"),
                "sms": f"Cauldra password recovery code: {code}. Expires in 10 minutes.",
                "type": "plain", "channel": "generic", "api_key": termii_key
            }, timeout=15)
            if r.ok:
                return
        except Exception:
            pass
    if twilio_sid and twilio_token and twilio_from:
        import requests
        try:
            auth = base64.b64encode(f"{twilio_sid}:{twilio_token}".encode()).decode()
            r = requests.post(
                f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json",
                data={"To": phone, "From": twilio_from, "Body": f"Cauldra password recovery code: {code}. Expires in 10 minutes."},
                headers={"Authorization": f"Basic {auth}"}, timeout=15,
            )
            if r.ok:
                return
        except Exception:
            pass
    raise HTTPException(status_code=503, detail="SMS recovery is not configured or the SMS provider could not deliver the code.")

@app.post("/auth/forgot-password")
def forgot_password(payload: ForgotPasswordRequest, request: Request, db: Session = Depends(get_db)):
    business_id = payload.business_id.strip()
    username = payload.username.strip()
    channel = payload.channel.strip().lower()
    if channel not in {"email", "sms"}:
        raise HTTPException(status_code=400, detail="Choose email or SMS for password recovery.")
    if channel == "email":
        if not payload.email:
            raise HTTPException(status_code=400, detail="Please provide your account email.")
        email = str(payload.email).strip()
    else:
        if not payload.phone or not normalize_phone(payload.phone):
            raise HTTPException(status_code=400, detail="Please provide your phone number.")
        phone = payload.phone.strip()
    biz = get_business_by_code(db, business_id)
    if not biz:
        raise HTTPException(status_code=404, detail="Business not found. Please check the Business ID and try again.")
    key = f"{biz.id}:{normalize_username(username)}:{channel}"
    client_ip = request.client.host if request.client else "unknown"
    check_rate_limit(db, "forgot-password", key)
    check_rate_limit(db, "forgot-password-ip", client_ip)
    user = next((u for u in db.query(User).filter(User.business_id == biz.id).all() if normalize_username(u.username) == normalize_username(username)), None)
    # Channel-specific identity verification only — email recovery never checks
    # phone, SMS recovery never checks email. The submitted value is used only
    # to verify the account; delivery always goes to the value already stored
    # on the User record (see send_recovery_email/send_recovery_sms below),
    # never to whatever the client supplied.
    identity_verified = False
    if user and not user.disabled:
        if channel == "email":
            identity_verified = bool(user.email) and user.email.strip().casefold() == email.casefold()
        else:
            identity_verified = bool(user.phone) and phones_match(phone, user.phone)
    if not identity_verified:
        record_failure(db, "forgot-password", key)
        record_failure(db, "forgot-password-ip", client_ip)
        raise HTTPException(status_code=400, detail="We could not verify those recovery details.")
    now = datetime.utcnow()
    active_recovery = (db.query(PasswordRecovery).filter(
        PasswordRecovery.user_id == user.id, PasswordRecovery.used == False,
        PasswordRecovery.expires_at > now
    ).order_by(PasswordRecovery.created_at.desc()).first())
    if active_recovery and active_recovery.resend_after > now:
        seconds = max(1, int((active_recovery.resend_after - now).total_seconds()))
        raise HTTPException(status_code=429, detail=f"Please wait {seconds} seconds before requesting another recovery code.")
    # A replacement code invalidates every existing usable recovery code.
    db.query(PasswordRecovery).filter(
        PasswordRecovery.user_id == user.id, PasswordRecovery.used == False,
        PasswordRecovery.expires_at > now,
    ).update({PasswordRecovery.used: True}, synchronize_session=False)
    code = f"{secrets.randbelow(1000000):06d}"
    recovery_id = secrets.token_urlsafe(18)
    row = PasswordRecovery(
        recovery_id=recovery_id, user_id=user.id, channel=channel,
        code_hash=hash_text(code), expires_at=now + timedelta(seconds=FORGOT_CODE_TTL_SECONDS),
        resend_after=now + timedelta(seconds=FORGOT_RESEND_SECONDS)
    )
    # Persist the recovery record only after delivery is accepted by the provider.
    if channel == "email":
        send_recovery_email(user.email, user.username, code)
    else:
        send_recovery_sms(normalize_phone(user.phone), code)
    db.add(row); db.commit()
    clear_failures(db, "forgot-password", key)
    clear_failures(db, "forgot-password-ip", client_ip)
    return {"recovery_id": recovery_id, "channel": channel, "expires_in_seconds": FORGOT_CODE_TTL_SECONDS, "resend_after_seconds": FORGOT_RESEND_SECONDS}

@app.post("/auth/reset-password")
def reset_password(payload: PasswordResetRequest, db: Session = Depends(get_db)):
    recovery_id = payload.recovery_id
    code = payload.code
    new_password = payload.new_password
    validate_password_strength(new_password)
    row = db.query(PasswordRecovery).filter(PasswordRecovery.recovery_id == recovery_id, PasswordRecovery.used == False).first()
    if not row or row.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="This recovery code has expired. Please request a new code.")
    row.attempts += 1
    if row.attempts > 5:
        row.used = True; db.commit()
        raise HTTPException(status_code=400, detail="Too many recovery attempts. Please request a new code.")
    if hash_text(code) != row.code_hash:
        db.commit(); raise HTTPException(status_code=400, detail="The recovery code is incorrect.")
    user = db.query(User).filter(User.id == row.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="The account could not be recovered.")
    if user.disabled:
        raise HTTPException(status_code=403, detail="This account is disabled. Please contact your business Admin.")
    if user.previous_password_hash and verify_password(new_password, user.previous_password_hash):
        raise HTTPException(status_code=400, detail="Last created password cannot be used.")
    user.previous_password_hash = user.password
    user.password = hash_password(new_password)
    user.must_change_password = False
    revoke_all_user_sessions(db, user)
    row.used = True
    add_audit(db, user, "PASSWORD_RECOVERED", "Password recovered through the account recovery flow.")
    db.commit()
    return {"message": "Password reset successfully. You can now sign in with your new password."}

# -----------------------------------------------------------------------------
# BUSINESS PROFILE
# -----------------------------------------------------------------------------
@app.get("/business-profile/", response_model=BusinessProfileSchema)
def get_business_profile(user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not biz: raise HTTPException(status_code=404, detail="Business profile is unavailable.")
    return BusinessProfileSchema(**serialize_business(biz))

@app.post("/business-profile/", response_model=BusinessProfileSchema)
def update_business_profile(profile_data: BusinessProfileSchema, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can modify Business Profile settings.")
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not biz: raise HTTPException(status_code=404, detail="Business profile is unavailable.")
    country_values = canonical_business_country_values(profile_data.country, profile_data.country_code, profile_data.language)
    biz.company_name = profile_data.company_name
    biz.email = profile_data.email
    biz.phone = canonical_phone_for_country(profile_data.phone, country_values["phone_country_code"])
    biz.address = profile_data.address
    biz.tax_id = profile_data.tax_id
    biz.country = country_values["country"]
    biz.country_code = country_values["country_code"]
    biz.currency = country_values["currency"]
    biz.phone_country_code = country_values["phone_country_code"]
    biz.language = str(profile_data.language or "en").strip().lower() or "en"
    biz.locale = f"{biz.language}-{biz.country_code}"
    biz.timezone = country_values["timezone"]
    add_audit(db, user, "BUSINESS_PROFILE_UPDATED", "Business profile settings were updated.")
    db.commit(); db.refresh(biz)
    return BusinessProfileSchema(**serialize_business(biz))

# -----------------------------------------------------------------------------
# USERS / ACCOUNT MANAGEMENT
# -----------------------------------------------------------------------------
@app.delete("/business-profile/")
def delete_business_profile(response: Response, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Deletes the CALLER's own business and everything it owns.

    Authorization: admin-only, and the tenant is ALWAYS user.business_id —
    never a client-supplied id — so one business can never delete another's
    (see get_current_user's own tenant resolution; there is no business_id
    parameter on this endpoint at all for that reason).

    Deletion strategy: the database's own FK graph is the deletion engine,
    not a hand-maintained list of models. Every tenant-owned table's
    business_id column is a real `ON DELETE CASCADE` foreign key to
    business_profile.id in the live schema — verified directly against
    PostgreSQL's information_schema (not assumed from the ORM) for every
    single table that has a business_id column, with zero exceptions found.
    Tables one step removed from business_profile (alert_reads via alerts,
    password_recoveries/session_revocations via users, price_history via
    price_monitor_sources) are ALSO real CASCADE, so they're covered
    transitively by the exact same one statement.

    This is deliberately NOT a hardcoded per-model DELETE list — that
    approach is exactly how the previous version of this endpoint went
    stale: it was written before RefundTransaction, RefundLine,
    SaleTransaction, and Expense's Business Day ownership existed, so it
    silently missed all of them. A schema-verified cascade cannot go stale
    the same way; a future tenant-owned table only needs its own
    `ForeignKey("business_profile.id", ondelete="CASCADE")` (the existing,
    established convention for every table in this app) to be correctly
    swept up here with no code change.

    The ONE thing a database cascade cannot do is delete files on disk —
    StoredUpload rows disappear with everything else, but the physical
    files under UPLOAD_STORAGE_DIR do not, so those paths are captured
    before deletion and removed after commit succeeds."""
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only the business Admin can delete the business profile.")
    biz = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not biz:
        raise HTTPException(status_code=404, detail="Business profile not found.")

    business_id, business_code = biz.id, biz.business_code
    upload_rows = db.query(StoredUpload).filter(StoredUpload.business_id == business_id).all()
    upload_paths = [(UPLOAD_STORAGE_DIR / row.storage_key).resolve() for row in upload_rows]

    # No audit_logs entry is written for this action: audit_logs.business_id
    # is itself part of the cascade (by design — the whole business's audit
    # trail is tenant-owned data, exactly like everything else), so a row
    # written here would be deleted in the same transaction anyway. This
    # print is a lightweight, external (outside the database, so it cannot
    # be deleted by the cascade it's recording) operational record — not a
    # new logging system, just visibility for whoever operates the server.
    print(f"[business-deleted] id={business_id} code={business_code} by user_id={user.id} username={user.username}")

    try:
        db.delete(biz)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="The business could not be deleted. Please try again.")

    for path in upload_paths:
        if UPLOAD_STORAGE_DIR in path.parents:
            path.unlink(missing_ok=True)
    clear_refresh_cookie(response)
    return {"message": "Business profile deleted successfully."}

@app.post("/auth/clear-client-session")
def clear_client_session(response: Response):
    """Clear only the browser's Cauldra refresh cookie. No business or session row is deleted."""
    clear_refresh_cookie(response)
    return {"message": "Client session cookie cleared."}

@app.get("/users")
def list_users(limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    q = db.query(User).filter(User.business_id == user.business_id)
    if user.role == "manager": q = q.filter(User.role == "staff")
    users = q.order_by(User.id.asc()).offset(offset).limit(limit).all()
    return [serialize_user(u) for u in users]

@app.post("/users")
def create_user(data: CreateUserRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Your account is not allowed to create employee accounts.")
    role = data.role.casefold()
    if role not in {"admin","manager","staff"}: raise HTTPException(status_code=400, detail="Unsupported account role.")
    if user.role == "manager" and role != "staff": raise HTTPException(status_code=403, detail="Managers can only create Staff accounts.")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    current = db.query(User).filter(User.business_id == user.business_id, User.role == role).count()
    check_plan_limit(db, business, role, current)
    validate_password_strength(data.password)
    if normalize_username(data.username) == normalize_username(user.username):
        raise HTTPException(status_code=409, detail="That username is already in use in this business.")
    exists = next((u for u in db.query(User).filter(User.business_id == user.business_id).all() if normalize_username(u.username) == normalize_username(data.username)), None)
    if exists: raise HTTPException(status_code=409, detail="That username is already in use in this business.")
    new_user = User(
        username=data.username.strip(), password=hash_password(data.password), role=role,
        firstname=data.firstname, lastname=data.lastname, email=str(data.email), phone=data.phone, position=data.position,
        business_id=user.business_id, must_change_password=True, disabled=False, auth_version=1,
    )
    db.add(new_user); db.flush()
    add_audit(db, user, "USER_CREATED", f"Created {role} account.", new_user)
    db.commit(); db.refresh(new_user)
    # The creator already supplied the temporary password; never echo it back
    # from the API where it could be retained in logs or browser tooling.
    return serialize_user(new_user)

@app.patch("/users/{user_id}/disable")
def disable_user(user_id: int, actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    target = db.query(User).filter(User.id == user_id, User.business_id == actor.business_id).first()
    if not target: raise HTTPException(status_code=404, detail="Account is unavailable.")
    if actor.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can disable accounts directly.")
    if target.id == actor.id and target.role == "admin": raise HTTPException(status_code=400, detail="The active business owner account cannot be disabled here.")
    if target.role == "admin" and len(db.query(User).filter(User.business_id == actor.business_id, User.role == "admin", User.disabled == False).all()) <= 1:
        raise HTTPException(status_code=400, detail="The last active Admin cannot be disabled.")
    target.disabled = True
    revoke_all_user_sessions(db, target)
    add_audit(db, actor, "USER_DISABLED", "Account disabled.", target)
    db.commit()
    return {"message": "Account disabled successfully."}

@app.patch("/users/{user_id}/enable")
def enable_user(user_id: int, actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    target = db.query(User).filter(User.id == user_id, User.business_id == actor.business_id).first()
    if not target: raise HTTPException(status_code=404, detail="Account is unavailable.")
    if actor.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can enable accounts directly.")
    target.disabled = False
    target.auth_version = int(target.auth_version or 1) + 1
    add_audit(db, actor, "USER_ENABLED", "Account enabled.", target)
    db.commit()
    return {"message": "Account enabled successfully."}

@app.patch("/users/{user_id}/reset-password")
def reset_user_password(user_id: int, data: AdminPasswordResetRequest, actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if actor.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied.")
    target = db.query(User).filter(User.id == user_id, User.business_id == actor.business_id).first()
    if not target: raise HTTPException(status_code=404, detail="Account is unavailable.")
    if actor.role == "manager" and target.role != "staff": raise HTTPException(status_code=403, detail="Managers can only reset Staff accounts.")
    validate_password_strength(data.new_password)
    if target.previous_password_hash and verify_password(data.new_password, target.previous_password_hash):
        raise HTTPException(status_code=400, detail="Last created password cannot be used.")
    target.previous_password_hash = target.password
    target.password = hash_password(data.new_password)
    target.must_change_password = True
    revoke_all_user_sessions(db, target)
    target.disabled = False
    add_audit(db, actor, "PASSWORD_RESET", "Temporary password reset by authorized staff.", target)
    db.commit()
    return {"message": "Temporary password reset. The employee must change it at next sign-in."}

@app.delete("/users/{user_id}")
def delete_user(user_id: int, actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    target = db.query(User).filter(User.id == user_id, User.business_id == actor.business_id).first()
    if not target: raise HTTPException(status_code=404, detail="Account is unavailable.")
    if actor.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can delete accounts directly.")
    if target.role == "admin":
        admins = db.query(User).filter(User.business_id == actor.business_id, User.role == "admin", User.disabled == False).count()
        if target.id == actor.id or admins <= 1: raise HTTPException(status_code=400, detail="The last active Admin cannot be deleted.")
    add_audit(db, actor, "USER_DELETED", "Account deleted. Business-owned records were retained.", target)
    db.delete(target); db.commit()
    return {"message": "Account deleted. Business data was preserved."}

# -----------------------------------------------------------------------------
# ACCOUNT ACTION REQUESTS + AUDIT LOGS
# -----------------------------------------------------------------------------
@app.post("/account-action-requests")
def create_account_action_request(data: AccountActionRequestCreate, actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if actor.role != "manager": raise HTTPException(status_code=403, detail="Only Managers submit these approval requests.")
    if data.action not in {"delete", "disable"}: raise HTTPException(status_code=400, detail="Unsupported account action.")
    target = db.query(User).filter(User.id == data.target_user_id, User.business_id == actor.business_id).first()
    if not target or target.role != "staff": raise HTTPException(status_code=403, detail="Managers may request actions only for Staff accounts.")
    row = AccountActionRequest(
        business_id=actor.business_id, target_user_id=target.id, target_username=target.username, target_position=target.position,
        action=data.action, requested_by_id=actor.id, requested_by_name=actor.username, requested_by_position=actor.position,
    )
    db.add(row); add_audit(db, actor, "ACCOUNT_ACTION_REQUESTED", f"Requested Admin approval to {data.action} Staff account.", target); db.commit(); db.refresh(row)
    return {"id": row.id, "message": "Admin approval request sent."}

def serialize_account_action_request(x: AccountActionRequest, include_resolution: bool = False) -> dict:
    data = {
        "id": x.id, "target_username": x.target_username, "target_position": x.target_position,
        "action": x.action, "status": x.status,
        "requested_by_name": x.requested_by_name, "requested_by_position": x.requested_by_position,
        "created_at": to_utc_iso(x.created_at),
    }
    if include_resolution:
        data["resolved_by_name"] = x.resolved_by_name
        data["resolved_at"] = to_utc_iso(x.resolved_at) if x.resolved_at else None
    return data

@app.get("/account-action-requests")
def list_account_action_requests(actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(AccountActionRequest).filter(AccountActionRequest.business_id == actor.business_id, AccountActionRequest.status == "PENDING")
    if actor.role == "manager": q = q.filter(AccountActionRequest.requested_by_id == actor.id)
    elif actor.role != "admin": raise HTTPException(status_code=403, detail="Access denied")
    items = q.order_by(AccountActionRequest.created_at.desc()).all()
    return [serialize_account_action_request(x) for x in items]

@app.get("/account-action-requests/history")
def list_account_action_request_history(limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0), actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(AccountActionRequest).filter(AccountActionRequest.business_id == actor.business_id, AccountActionRequest.status != "PENDING")
    if actor.role == "manager": q = q.filter(AccountActionRequest.requested_by_id == actor.id)
    elif actor.role != "admin": raise HTTPException(status_code=403, detail="Access denied")
    items = q.order_by(AccountActionRequest.resolved_at.desc()).offset(offset).limit(limit).all()
    return [serialize_account_action_request(x, include_resolution=True) for x in items]

@app.post("/account-action-requests/{request_id}/{resolution}")
def resolve_account_action_request(request_id: int, resolution: str, actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if actor.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can resolve approval requests.")
    row = db.query(AccountActionRequest).filter(AccountActionRequest.id == request_id, AccountActionRequest.business_id == actor.business_id, AccountActionRequest.status == "PENDING").first()
    if not row: raise HTTPException(status_code=404, detail="Approval request is no longer pending.")
    target = db.query(User).filter(User.id == row.target_user_id, User.business_id == actor.business_id).first() if row.target_user_id else None
    row.status = "APPROVED" if resolution == "approve" else "REJECTED"; row.resolved_by_id = actor.id; row.resolved_by_name = actor.username; row.resolved_at = datetime.utcnow()
    add_audit(db, actor, f"ACCOUNT_REQUEST_{row.status}", f"Admin {row.status.lower()} request to {row.action} account.", target)
    if resolution == "approve" and target:
        # The approval decision and the resulting execution are recorded as two
        # distinct Activity History events — using the same action codes as the
        # direct admin disable/delete paths — so the full story (request →
        # decision → completed action) is visible, not collapsed into one entry.
        if row.action == "disable":
            target.disabled = True
            revoke_all_user_sessions(db, target)
            add_audit(db, actor, "USER_DISABLED", f"Account disabled (approved request #{row.id}).", target)
        elif row.action == "delete":
            add_audit(db, actor, "USER_DELETED", f"Account deleted (approved request #{row.id}). Business-owned records were retained.", target)
            db.delete(target)
    db.commit()
    return {"message": f"Request {row.status.lower()}."}

def _build_audit_log_query(db: Session, actor: User, q, actor_username, action, date_from, date_to):
    """Shared filter-building for /audit-logs and /audit-logs/export — kept
    in one place so the CSV export can never show a different result set
    than what the same filters display on screen."""
    query = db.query(AuditLog).filter(AuditLog.business_id == actor.business_id, AuditLog.action.notin_(PRESENCE_SESSION_AUDIT_ACTIONS))
    if q and q.strip():
        like = f"%{q.strip()}%"
        query = query.filter(or_(
            AuditLog.description.ilike(like), AuditLog.action.ilike(like),
            AuditLog.actor_username.ilike(like), AuditLog.target_username.ilike(like),
        ))
    if actor_username and actor_username.strip():
        query = query.filter(AuditLog.actor_username.ilike(f"%{actor_username.strip()}%"))
    if action and action.strip():
        query = query.filter(AuditLog.action.ilike(f"%{action.strip()}%"))
    from_dt = parse_iso_datetime(date_from)
    if from_dt: query = query.filter(AuditLog.created_at >= from_dt)
    # The frontend sends a precise UTC instant (already the correct exclusive
    # upper bound, computed in the business's own timezone) whenever it can —
    # recognizable by containing "T". A bare "YYYY-MM-DD" (e.g. from a caller
    # that didn't do that conversion) is still treated as an inclusive whole
    # calendar day for backward compatibility.
    to_dt = parse_iso_datetime(date_to)
    if to_dt:
        upper_bound = to_dt if (date_to and "T" in date_to) else to_dt + timedelta(days=1)
        query = query.filter(AuditLog.created_at < upper_bound)
    return query

@app.get("/audit-logs")
def list_audit_logs(
    limit: int = Query(100, ge=1, le=300), offset: int = Query(0, ge=0),
    q: Optional[str] = Query(None, description="Keyword search across description/action/actor/target"),
    actor_username: Optional[str] = Query(None, description="Filter by the person who performed the action"),
    action: Optional[str] = Query(None, description="Filter by action code (partial match)"),
    date_from: Optional[str] = Query(None, description="Only entries on/after this date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Only entries on/before this date (YYYY-MM-DD)"),
    actor: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    if actor.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    query = _build_audit_log_query(db, actor, q, actor_username, action, date_from, date_to)
    total = query.count()
    rows = query.order_by(AuditLog.created_at.desc()).offset(offset).limit(limit).all()
    return {
        "items": [{"id": r.id, "actor_username": r.actor_username, "actor_role": r.actor_role, "action": r.action, "target_username": r.target_username, "description": r.description, "created_at": to_utc_iso(r.created_at)} for r in rows],
        "total": total,
    }

AUDIT_LOG_EXPORT_COLUMNS = [
    {"key": "id", "label": "ID", "type": "number"},
    {"key": "date", "label": "DATE", "type": "datetime"},
    {"key": "actor", "label": "ACTOR", "type": "text"},
    {"key": "role", "label": "ROLE", "type": "text"},
    {"key": "action", "label": "ACTION", "type": "text"},
    {"key": "target", "label": "TARGET", "type": "text"},
    {"key": "description", "label": "DESCRIPTION", "type": "text"},
]

def _audit_log_export_rows(db, actor, q, actor_username, action, date_from, date_to):
    """Shared by the CSV and Excel Activity History exports. Capped at the
    same 300-row ceiling the JSON endpoint already enforces (`le=300`);
    Activity History is a live, fast-growing table, so this export is a
    bounded, filtered slice, not a full historical dump."""
    query = _build_audit_log_query(db, actor, q, actor_username, action, date_from, date_to)
    rows = query.order_by(AuditLog.created_at.desc()).limit(300).all()
    return [[r.id, to_utc_iso(r.created_at), r.actor_username or "", r.actor_role or "", r.action or "", r.target_username or "", r.description or ""] for r in rows]

@app.get("/audit-logs/export")
def export_audit_logs_csv(
    q: Optional[str] = Query(None), actor_username: Optional[str] = Query(None), action: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    actor: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    """CSV export of Activity History. Same role check, same business
    scoping, and the exact same filter-building function as GET
    /audit-logs — this can never expose a row the JSON endpoint wouldn't
    also show for the same filters."""
    if actor.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    out_rows = _audit_log_export_rows(db, actor, q, actor_username, action, date_from, date_to)
    if not out_rows:
        return Response(status_code=204)
    header = [c["label"] for c in AUDIT_LOG_EXPORT_COLUMNS]
    return build_csv_response(f"cauldra_activity_history_{business_local_today(db, actor.business_id)}.csv", header, out_rows)

@app.get("/audit-logs/export/xlsx")
def export_audit_logs_xlsx(
    q: Optional[str] = Query(None), actor_username: Optional[str] = Query(None), action: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    actor: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    """Same authorization, tenant scoping, filters, and row data as the CSV
    export above — only the presentation differs."""
    if actor.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    out_rows = _audit_log_export_rows(db, actor, q, actor_username, action, date_from, date_to)
    if not out_rows:
        return Response(status_code=204)
    filters_desc = ", ".join(f"{k}={v}" for k, v in [("Search", q), ("Actor", actor_username), ("Action", action), ("From", date_from), ("To", date_to)] if v) or "None"
    metadata = [["Generated", datetime.utcnow().strftime("%d %B %Y %H:%M UTC")], ["Filters", filters_desc]]
    sheets = [{"name": "Activity History", "title": "CAULDRA ACTIVITY HISTORY", "metadata": metadata, "columns": AUDIT_LOG_EXPORT_COLUMNS, "rows": out_rows}]
    return build_xlsx_response(f"cauldra_activity_history_{business_local_today(db, actor.business_id)}.xlsx", "Cauldra Activity History", sheets)

@app.get("/audit-logs/actions")
def list_audit_log_actions(actor: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Distinct action codes and actor usernames actually present in this business's
    Activity History, so the frontend can offer filter suggestions without ever
    hardcoding a fixed list of activity types."""
    if actor.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    base = db.query(AuditLog).filter(AuditLog.business_id == actor.business_id, AuditLog.action.notin_(PRESENCE_SESSION_AUDIT_ACTIONS))
    actions = [r[0] for r in base.with_entities(AuditLog.action).distinct().order_by(AuditLog.action.asc()).all()]
    actors = [r[0] for r in base.filter(AuditLog.actor_username.isnot(None)).with_entities(AuditLog.actor_username).distinct().order_by(AuditLog.actor_username.asc()).all()]
    return {"actions": actions, "actors": actors}

def get_warehouse_for_business(db: Session, business_id: int, name: str, active_only: bool = True) -> Optional[Warehouse]:
    cleaned = (name or "").strip()
    if not cleaned:
        return None
    q = db.query(Warehouse).filter(Warehouse.business_id == business_id, func.lower(Warehouse.name) == cleaned.casefold())
    if active_only:
        q = q.filter(Warehouse.is_active == True)
    return q.first()

def serialize_warehouse(w: Warehouse, db: Session) -> dict:
    sku_count = db.query(WarehouseStock.product_id).filter(WarehouseStock.business_id == w.business_id, WarehouseStock.warehouse == w.name).distinct().count()
    return {"id": w.id, "name": w.name, "is_active": w.is_active, "sku_count": sku_count, "created_at": to_utc_iso(w.created_at)}

@app.get("/warehouses/")
def list_warehouses(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.query(Warehouse).filter(Warehouse.business_id == user.business_id, Warehouse.is_active == True).order_by(Warehouse.name.asc()).all()
    return [serialize_warehouse(w, db) for w in rows]

@app.post("/warehouses/")
def create_warehouse(data: WarehouseCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Only Admins and Managers can create warehouses.")
    name = data.name.strip()
    if len(name) < 2:
        raise HTTPException(status_code=400, detail="Warehouse name must contain at least 2 characters.")
    existing = db.query(Warehouse).filter(Warehouse.business_id == user.business_id, func.lower(Warehouse.name) == name.casefold()).first()
    if existing:
        if not existing.is_active:
            business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
            check_plan_limit(db, business, "warehouse", db.query(Warehouse).filter(Warehouse.business_id == user.business_id, Warehouse.is_active == True).count())
            existing.is_active = True
            existing.updated_at = datetime.utcnow()
            db.commit(); db.refresh(existing)
            return serialize_warehouse(existing, db)
        raise HTTPException(status_code=409, detail="That warehouse already exists in this business.")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    check_plan_limit(db, business, "warehouse", db.query(Warehouse).filter(Warehouse.business_id == user.business_id, Warehouse.is_active == True).count())
    row = Warehouse(business_id=user.business_id, name=name, is_active=True)
    db.add(row)
    add_audit(db, user, "WAREHOUSE_CREATED", f"Created warehouse {name}.")
    db.commit(); db.refresh(row)
    return serialize_warehouse(row, db)

@app.patch("/warehouses/{warehouse_id}")
def update_warehouse(warehouse_id: int, data: WarehouseUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Only Admins and Managers can update warehouses.")
    row = db.query(Warehouse).filter(Warehouse.id == warehouse_id, Warehouse.business_id == user.business_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Warehouse not found.")
    new_name = data.name.strip()
    if len(new_name) < 2:
        raise HTTPException(status_code=400, detail="Warehouse name must contain at least 2 characters.")
    duplicate = db.query(Warehouse).filter(Warehouse.business_id == user.business_id, Warehouse.id != row.id, func.lower(Warehouse.name) == new_name.casefold()).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="That warehouse already exists in this business.")
    old_name = row.name
    if old_name != new_name:
        db.query(Product).filter(Product.business_id == user.business_id, Product.warehouse == old_name).update({Product.warehouse: new_name}, synchronize_session=False)
        db.query(WarehouseStock).filter(WarehouseStock.business_id == user.business_id, WarehouseStock.warehouse == old_name).update({WarehouseStock.warehouse: new_name}, synchronize_session=False)
        row.name = new_name
    if data.is_active is not None:
        row.is_active = bool(data.is_active)
    row.updated_at = datetime.utcnow()
    add_audit(db, user, "WAREHOUSE_UPDATED", f"Updated warehouse {old_name} to {new_name}.")
    db.commit(); db.refresh(row)
    return serialize_warehouse(row, db)

@app.delete("/warehouses/{warehouse_id}")
def deactivate_warehouse(warehouse_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Only Admins and Managers can deactivate warehouses.")
    row = db.query(Warehouse).filter(Warehouse.id == warehouse_id, Warehouse.business_id == user.business_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Warehouse not found.")
    active_count = db.query(Warehouse).filter(Warehouse.business_id == user.business_id, Warehouse.is_active == True).count()
    if active_count <= 1 and row.is_active:
        raise HTTPException(status_code=400, detail="The last active warehouse cannot be deactivated.")
    row.is_active = False
    row.updated_at = datetime.utcnow()
    add_audit(db, user, "WAREHOUSE_DEACTIVATED", f"Deactivated warehouse {row.name}.")
    db.commit()
    return {"message": "Warehouse deactivated successfully."}

# -----------------------------------------------------------------------------
# PRODUCTS
# -----------------------------------------------------------------------------
def generate_unique_sku(db: Session, business_id: int) -> str:
    for _ in range(50):
        sku = f"SKU-{random.randint(100000, 999999)}"
        if not db.query(Product).filter(Product.business_id == business_id, Product.sku == sku).first():
            return sku
    raise HTTPException(status_code=500, detail="We could not generate a unique SKU. Please try again.")

@app.get("/products/")
def list_products(limit: int = Query(200, ge=1, le=500), offset: int = Query(0, ge=0), warehouse: Optional[str] = Query(None), stock_status: Optional[str] = Query(None, alias="status"), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    warehouse_name = None
    if warehouse and warehouse.upper() != "ALL":
        wh = get_warehouse_for_business(db, user.business_id, warehouse)
        if not wh:
            raise HTTPException(status_code=404, detail="The selected warehouse is unavailable.")
        warehouse_name = wh.name

    status_value = (stock_status or "").strip().lower()
    if status_value not in {"", "healthy", "low", "out"}:
        raise HTTPException(status_code=400, detail="Invalid stock status filter.")

    if warehouse_name:
        q = (db.query(Product, WarehouseStock.quantity.label("warehouse_quantity"))
             .join(WarehouseStock, and_(WarehouseStock.product_id == Product.id, WarehouseStock.business_id == user.business_id, WarehouseStock.warehouse == warehouse_name))
             .filter(Product.business_id == user.business_id))
        effective_qty = WarehouseStock.quantity
    else:
        q = db.query(Product).filter(Product.business_id == user.business_id)
        effective_qty = Product.quantity

    if status_value == "healthy":
        q = q.filter(effective_qty > Product.min_stock_level)
    elif status_value == "low":
        q = q.filter(effective_qty > 0, effective_qty <= Product.min_stock_level)
    elif status_value == "out":
        q = q.filter(effective_qty == 0)

    q = q.order_by(Product.id.desc()).offset(offset).limit(limit).all()
    rows = []
    if warehouse_name:
        for p, warehouse_quantity in q:
            rows.append({
                "id": p.id, "sku": p.sku, "barcode": p.barcode, "name": p.name, "category": p.category, "size": p.size, "quantity": int(warehouse_quantity or 0),
                "total_quantity": p.quantity, "min_stock_level": p.min_stock_level, "cost_price": p.cost_price, "wholesale_price": p.wholesale_price,
                "retail_price": p.retail_price, "warehouse": warehouse_name, "created_at": to_utc_iso(p.created_at),
                "expiry_date": p.expiry_date.isoformat() if p.expiry_date else None
            })
    else:
        for p in q:
            rows.append({
                "id": p.id, "sku": p.sku, "barcode": p.barcode, "name": p.name, "category": p.category, "size": p.size, "quantity": p.quantity,
                "total_quantity": p.quantity, "min_stock_level": p.min_stock_level, "cost_price": p.cost_price, "wholesale_price": p.wholesale_price,
                "retail_price": p.retail_price, "warehouse": p.warehouse, "created_at": to_utc_iso(p.created_at),
                "expiry_date": p.expiry_date.isoformat() if p.expiry_date else None
            })
    return rows

@app.get("/products/inventory-summary")
def inventory_summary(warehouse: Optional[str] = Query(None), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    warehouse_name = None
    if warehouse and warehouse.upper() != "ALL":
        wh = get_warehouse_for_business(db, user.business_id, warehouse)
        if not wh:
            raise HTTPException(status_code=404, detail="The selected warehouse is unavailable.")
        warehouse_name = wh.name

    if warehouse_name:
        rows = (db.query(Product.min_stock_level, WarehouseStock.quantity)
                .join(WarehouseStock, and_(WarehouseStock.product_id == Product.id, WarehouseStock.business_id == user.business_id, WarehouseStock.warehouse == warehouse_name))
                .filter(Product.business_id == user.business_id).all())
        quantities = [(int(qty or 0), int(min_level or 0)) for min_level, qty in rows]
    else:
        rows = db.query(Product.quantity, Product.min_stock_level).filter(Product.business_id == user.business_id).all()
        quantities = [(int(qty or 0), int(min_level or 0)) for qty, min_level in rows]

    return {
        "total_products": len(quantities),
        "healthy": sum(1 for qty, minimum in quantities if qty > minimum),
        "low": sum(1 for qty, minimum in quantities if qty > 0 and qty <= minimum),
        "out": sum(1 for qty, minimum in quantities if qty == 0),
    }

# =============================================================================
# SMART PRODUCT DUPLICATE DETECTION (V25)
# One reusable, deterministic function. Business-scoped: it never looks at any
# other business's products and never consults the General Catalog / UPCitemdb
# (those are identity lookups, not "do I already stock this?"). No AI/LLM.
# Both the create and the edit product paths run through _dup_enforce_or_raise
# BEFORE writing a row, so the check cannot be bypassed by a stale frontend,
# a second client, or a direct API call.
#
# Signals & weights (barcode/SKU strongest, name/size strong, category/price
# supporting):
#   same_barcode            +100     same_size (normalized)     +30
#   same_sku                 +80     same_category              +10
#   name_similarity   up to  +35     similar_cost_price          +8
#                                    similar_wholesale_price     +8
#                                    similar_retail_price        +8
# Decision:  exact barcode OR exact SKU OR score >= 95  -> "definite"
#            score >= 55                                 -> "possible"
#            else                                        -> "low"/"none"
# =============================================================================
DUP_PRICE_PERCENT_TOLERANCE = 0.07      # 7%
DUP_PRICE_ABSOLUTE_ALLOWANCE = 1.0      # currency units - absorbs rounding
DUP_NAME_SIMILAR_MIN = 0.72            # below this, names contribute nothing
DUP_WEIGHT_BARCODE = 100
DUP_WEIGHT_SKU = 80
DUP_WEIGHT_NAME = 35
DUP_WEIGHT_SIZE = 30
DUP_WEIGHT_CATEGORY = 10
DUP_WEIGHT_PRICE_EACH = 8
DUP_HARD_THRESHOLD = 95               # >= this (or exact barcode/SKU) -> definite
DUP_POSSIBLE_THRESHOLD = 55           # >= this -> possible
DUP_HINT_THRESHOLD = 25              # >= this -> worth a non-blocking "similar" hint

_DUP_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_DUP_WS_RE = re.compile(r"\s+")


def _dup_norm_name(value: Optional[str]) -> str:
    """lowercase, unify hyphen/underscore/slash to space, drop other
    punctuation, collapse whitespace, trim. 'Cream-Crackers' / 'CREAM   CRACKERS'
    / 'cream crackers' all normalize to the same string."""
    s = (value or "").strip().casefold()
    s = s.replace("-", " ").replace("_", " ").replace("/", " ")
    s = _DUP_PUNCT_RE.sub(" ", s)
    return _DUP_WS_RE.sub(" ", s).strip()


def _dup_norm_size(value: Optional[str]) -> str:
    """'200g' == '200 g' == '200G'; '200g' != '500g'. A few obvious spelling
    variants are unified but NO unit conversion is attempted (the project has
    no safe size-parsing utility, and 0.2kg vs 200g should stay 'different')."""
    s = (value or "").strip().casefold()
    s = _DUP_WS_RE.sub("", s).replace(".", "").replace(",", "")
    s = s.replace("litres", "l").replace("litre", "l").replace("liters", "l").replace("liter", "l")
    s = s.replace("grams", "g").replace("gram", "g")
    s = s.replace("kilograms", "kg").replace("kilogram", "kg").replace("kgs", "kg")
    s = s.replace("millilitres", "ml").replace("millilitre", "ml").replace("milliliters", "ml").replace("milliliter", "ml")
    return s


def _dup_name_similarity(a: Optional[str], b: Optional[str]) -> float:
    """0.0-1.0. Fuzzy, not exact-string-only. Combines a character-level
    SequenceMatcher ratio with a token Jaccard so word-order and small
    spelling/plural differences ('Cream Crackers' vs 'Cream Cracker') score
    high, while an added meaningful word ('Blue Cream Crackers') scores
    'similar but not identical' rather than 1.0."""
    na, nb = _dup_norm_name(a), _dup_norm_name(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    from difflib import SequenceMatcher
    ratio = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    union = ta | tb
    jaccard = (len(ta & tb) / len(union)) if union else 0.0
    return max(ratio, jaccard)


def _dup_price_similar(candidate_price, existing_price) -> bool:
    try:
        c = float(candidate_price or 0)
        e_ = float(existing_price or 0)
    except (TypeError, ValueError):
        return False
    if c <= 0 or e_ <= 0:
        return False   # a missing/zero price is not evidence of anything
    return abs(c - e_) <= max(DUP_PRICE_ABSOLUTE_ALLOWANCE, e_ * DUP_PRICE_PERCENT_TOLERANCE)


def _dup_score_pair(cand: dict, existing: "Product") -> "Tuple[int, list]":
    """cand = normalized candidate fields. Returns (score, matched_signals)."""
    score = 0
    signals: list = []

    cb = (cand.get("barcode") or "").strip()
    if cb and (existing.barcode or "").strip() == cb:
        score += DUP_WEIGHT_BARCODE
        signals.append("same_barcode")

    cs = (cand.get("sku") or "").strip().casefold()
    if cs and (existing.sku or "").strip().casefold() == cs:
        score += DUP_WEIGHT_SKU
        signals.append("same_sku")

    name_sim = _dup_name_similarity(cand.get("name"), existing.name)
    if name_sim >= DUP_NAME_SIMILAR_MIN:
        score += int(round(DUP_WEIGHT_NAME * name_sim))
        signals.append("name_similarity")

    csize, esize = _dup_norm_size(cand.get("size")), _dup_norm_size(existing.size)
    if csize and esize and csize == esize:
        score += DUP_WEIGHT_SIZE
        signals.append("same_size")

    ccat = _dup_norm_name(cand.get("category"))
    if ccat and _dup_norm_name(existing.category) == ccat:
        score += DUP_WEIGHT_CATEGORY
        signals.append("same_category")

    if _dup_price_similar(cand.get("cost_price"), existing.cost_price):
        score += DUP_WEIGHT_PRICE_EACH
        signals.append("similar_cost_price")
    if _dup_price_similar(cand.get("wholesale_price"), existing.wholesale_price):
        score += DUP_WEIGHT_PRICE_EACH
        signals.append("similar_wholesale_price")
    if _dup_price_similar(cand.get("retail_price"), existing.retail_price):
        score += DUP_WEIGHT_PRICE_EACH
        signals.append("similar_retail_price")

    return score, signals


def _dup_level_for(score: int, signals: list) -> str:
    if "same_barcode" in signals or "same_sku" in signals or score >= DUP_HARD_THRESHOLD:
        return "definite"
    if score >= DUP_POSSIBLE_THRESHOLD:
        return "possible"
    return "low"


def _dup_candidate_payload(p: "Product", score: int, signals: list, level: str) -> dict:
    return {
        "id": p.id, "name": p.name, "sku": p.sku, "barcode": p.barcode,
        "size": p.size, "category": p.category,
        "cost_price": p.cost_price, "wholesale_price": p.wholesale_price,
        "retail_price": p.retail_price,
        "score": score, "matched_signals": signals, "duplicate_level": level,
    }


def assess_product_duplicate(db: Session, business_id: int, *, name: str,
                             size: Optional[str] = None, category: Optional[str] = None,
                             barcode: Optional[str] = None, sku: Optional[str] = None,
                             cost_price=None, wholesale_price=None, retail_price=None,
                             exclude_id: Optional[int] = None) -> dict:
    """The single source of truth for 'is this a duplicate of a product THIS
    business already has?'. Deterministic, business-scoped, no external calls.

    Returns:
      {
        "duplicate_detected": bool,            # top candidate is definite/possible
        "duplicate_level": "definite"|"possible"|"low"|"none",
        "score": int, "matched_signals": [...],   # of the top candidate
        "candidate":  {...} | None,               # top flagged candidate
        "candidates": [ {...}, ... up to 3 ],      # flagged, ranked best-first
        "hint_candidate": {id,name,size,category} | None,  # best match for the live hint
      }
    """
    cand = {
        "name": name, "size": size, "category": category,
        "barcode": normalize_barcode(barcode) if barcode else None,
        "sku": sku, "cost_price": cost_price,
        "wholesale_price": wholesale_price, "retail_price": retail_price,
    }
    q = db.query(Product).filter(Product.business_id == business_id)
    if exclude_id is not None:
        q = q.filter(Product.id != exclude_id)

    scored = []
    for p in q.all():
        s, sig = _dup_score_pair(cand, p)
        if s <= 0:
            continue
        scored.append((s, _dup_level_for(s, sig), p, sig))

    _rank = {"definite": 0, "possible": 1, "low": 2}
    scored.sort(key=lambda x: (_rank[x[1]], -x[0], x[2].id))

    empty = {"duplicate_detected": False, "duplicate_level": "none", "score": 0,
             "matched_signals": [], "candidate": None, "candidates": [], "hint_candidate": None}
    if not scored:
        return empty

    flagged = [x for x in scored if x[1] in ("definite", "possible")][:3]
    best = scored[0]
    hint = None
    if best[0] >= DUP_HINT_THRESHOLD:
        hp = best[2]
        hint = {"id": hp.id, "name": hp.name, "size": hp.size, "category": hp.category}

    if not flagged:
        return {**empty, "duplicate_level": "low", "score": best[0],
                "matched_signals": best[3], "hint_candidate": hint}

    candidates = [_dup_candidate_payload(p, s, sig, lv) for (s, lv, p, sig) in flagged]
    return {
        "duplicate_detected": True,
        "duplicate_level": flagged[0][1],
        "score": flagged[0][0],
        "matched_signals": flagged[0][3],
        "candidate": candidates[0],
        "candidates": candidates,
        "hint_candidate": hint,
    }


def _dup_enforce_or_raise(db: Session, business_id: int, *, name, size, category,
                          barcode, sku, cost_price, wholesale_price, retail_price,
                          override_id: Optional[int], exclude_id: Optional[int]) -> None:
    """Run the assessment and, unless the caller supplied a valid explicit
    override for this exact candidate, raise 409 with the structured duplicate
    body the frontend renders in its blocking dialog. An exact barcode or SKU
    match ('hard' signal) can NEVER be overridden - it is a real collision, the
    user must use the existing product or change the identifier."""
    assessment = assess_product_duplicate(
        db, business_id, name=name, size=size, category=category,
        barcode=barcode, sku=sku, cost_price=cost_price,
        wholesale_price=wholesale_price, retail_price=retail_price,
        exclude_id=exclude_id,
    )
    if not assessment["duplicate_detected"]:
        return

    signals = assessment.get("matched_signals") or []
    hard = ("same_barcode" in signals) or ("same_sku" in signals)

    if override_id is not None and not hard:
        allowed = {c["id"] for c in (assessment.get("candidates") or [])}
        if assessment.get("candidate"):
            allowed.add(assessment["candidate"]["id"])
        if int(override_id) in allowed:
            return   # explicit, verified, same-business override -> allow

    raise HTTPException(status_code=409, detail={
        "duplicate_detected": True,
        "duplicate_level": assessment["duplicate_level"],
        "score": assessment["score"],
        "matched_signals": signals,
        "candidate": assessment["candidate"],
        "candidates": assessment["candidates"],
        "message": ("This product already exists in your inventory."
                    if assessment["duplicate_level"] == "definite"
                    else "This looks like a product you may already have."),
    })


class ProductDuplicateCheckRequest(BaseModel):
    name: str
    size: Optional[str] = None
    category: Optional[str] = None
    barcode: Optional[str] = None
    sku: Optional[str] = None
    cost_price: Optional[float] = None
    wholesale_price: Optional[float] = None
    retail_price: Optional[float] = None
    exclude_id: Optional[int] = None


@app.post("/products/duplicate-check")
def products_duplicate_check(req: ProductDuplicateCheckRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Read-only, business-scoped. Powers the live non-blocking 'similar
    product already in inventory' hint and the pre-save check. The blocking
    guarantee still lives in the create/edit endpoints (see
    _dup_enforce_or_raise) - this endpoint only informs the UI."""
    if user.role == "staff":
        raise HTTPException(status_code=403, detail="Staff accounts cannot add products.")
    if not (req.name or "").strip():
        return {"duplicate_detected": False, "duplicate_level": "none", "score": 0,
                "matched_signals": [], "candidate": None, "candidates": [], "hint_candidate": None}
    return assess_product_duplicate(
        db, user.business_id, name=req.name, size=req.size, category=req.category,
        barcode=req.barcode, sku=req.sku, cost_price=req.cost_price,
        wholesale_price=req.wholesale_price, retail_price=req.retail_price,
        exclude_id=req.exclude_id,
    )


@app.post("/products/")
def create_product(data: ProductCreate, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    enforce_offline_replay_identity(request, user)
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot add products.")
    client_ref = (data.client_ref or "").strip()[:100] or None
    claim, replay = claim_idempotent_mutation(
        db, user.business_id, "product_create", client_ref,
        data.model_dump(mode="json", exclude={"client_ref", "duplicate_override_candidate_id"}),
    )
    if replay:
        return replay
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    check_plan_limit(db, business, "product", db.query(Product).filter(Product.business_id == user.business_id).count())
    sku=(data.sku or "").strip()
    if not sku:
        for _ in range(50):
            candidate=f"SKU-{random.randint(100000,999999)}"
            if not db.query(Product).filter(Product.business_id==user.business_id, Product.sku==candidate).first(): sku=candidate; break
        if not sku: raise HTTPException(status_code=500, detail="We could not create a unique product code. Please try again.")
    elif db.query(Product).filter(Product.business_id==user.business_id, Product.sku==sku).first():
        raise HTTPException(status_code=409, detail="That SKU is already used in this business.")
    warehouse_name = (data.warehouse or "Main Central Warehouse").strip()
    if not get_warehouse_for_business(db, user.business_id, warehouse_name):
        raise HTTPException(status_code=400, detail="The selected warehouse does not exist in this business.")
    barcode = normalize_barcode(data.barcode) if data.barcode else None
    # V25: business-scoped smart duplicate detection. Runs before the row is
    # created so a stale/bypassed frontend, a 2nd client or a direct API call
    # still cannot slip an accidental duplicate through. An exact barcode match
    # in this business is a "definite" duplicate and is not overridable.
    _dup_enforce_or_raise(
        db, user.business_id,
        name=data.name, size=data.size, category=data.category,
        barcode=barcode, sku=sku, cost_price=data.cost_price,
        wholesale_price=(data.wholesale_price or data.retail_price * 0.85),
        retail_price=data.retail_price,
        override_id=data.duplicate_override_candidate_id, exclude_id=None,
    )
    p=Product(sku=sku, barcode=barcode, name=data.name, category=data.category, size=data.size, quantity=data.quantity, min_stock_level=data.min_stock_level, cost_price=data.cost_price, wholesale_price=data.wholesale_price or data.retail_price*0.85, retail_price=data.retail_price, warehouse=warehouse_name, initial_stock=data.quantity, expiry_date=data.expiry_date, business_id=user.business_id, owner_id=user.id, client_ref=client_ref, synced_at=(datetime.utcnow() if client_ref else None))
    db.add(p); db.flush(); db.add(WarehouseStock(business_id=user.business_id,product_id=p.id,warehouse=p.warehouse,quantity=p.quantity)); auto_upsert_general_catalog(db,p); add_audit(db,user,"PRODUCT_CREATED",f"Added product {p.name}."); mark_business_brain_dirty(db, user.business_id)
    response = {"id":p.id,"sku":p.sku,"barcode":p.barcode,"name":p.name}
    complete_idempotent_mutation(claim, response)
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    db.refresh(p)
    return response

@app.patch("/products/{product_id}")
def update_product(product_id: int, data: ProductUpdate, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    enforce_offline_replay_identity(request, user)
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot edit products.")
    changes = data.model_dump(exclude_unset=True)
    client_ref = str(changes.pop("client_ref", "") or "").strip()[:100] or None
    _dup_override_id = changes.pop("duplicate_override_candidate_id", None)
    claim, replay = claim_idempotent_mutation(
        db, user.business_id, f"product_update:{product_id}", client_ref,
        {key: (value.isoformat() if isinstance(value, datetime) else value) for key, value in changes.items()},
    )
    if replay:
        return replay
    p = db.query(Product).filter(Product.id == product_id, Product.business_id == user.business_id).first()
    if not p: raise HTTPException(status_code=404, detail="The product could not be found in this inventory.")
    if "sku" in changes and changes["sku"] and changes["sku"] != p.sku:
        if db.query(Product).filter(Product.business_id == user.business_id, Product.sku == changes["sku"], Product.id != p.id).first():
            raise HTTPException(status_code=409, detail="That SKU is already in use in this business.")
    if "barcode" in changes:
        normalized_barcode = normalize_barcode(changes["barcode"]) if changes["barcode"] else None
        if normalized_barcode and db.query(Product).filter(Product.business_id == user.business_id, Product.barcode == normalized_barcode, Product.id != p.id).first():
            raise HTTPException(status_code=409, detail="That barcode is already used by another product in this business.")
        changes["barcode"] = normalized_barcode
    # V25: business-scoped smart duplicate detection on edits. Excludes this
    # product's own id; still compares against every other product in this
    # business. Uses the post-edit effective value of each field.
    _dup_enforce_or_raise(
        db, user.business_id,
        name=changes.get("name", p.name), size=changes.get("size", p.size),
        category=changes.get("category", p.category),
        barcode=(changes["barcode"] if "barcode" in changes else p.barcode),
        sku=changes.get("sku", p.sku),
        cost_price=changes.get("cost_price", p.cost_price),
        wholesale_price=changes.get("wholesale_price", p.wholesale_price),
        retail_price=changes.get("retail_price", p.retail_price),
        override_id=_dup_override_id, exclude_id=p.id,
    )
    if "warehouse" in changes:
        requested_warehouse = (changes["warehouse"] or "").strip()
        warehouse = get_warehouse_for_business(db, user.business_id, requested_warehouse)
        if not warehouse:
            raise HTTPException(status_code=400, detail="The selected warehouse does not exist in this business.")
        changes["warehouse"] = warehouse.name

    # Product.quantity is the total of its warehouse ledger. Keep direct edits
    # compatible with the existing form while applying the difference to the
    # selected warehouse row, so inventory cannot silently drift out of sync.
    if "quantity" in changes:
        target_warehouse = changes.get("warehouse") or p.warehouse or "Main Central Warehouse"
        stock = db.query(WarehouseStock).filter(
            WarehouseStock.business_id == user.business_id,
            WarehouseStock.product_id == p.id,
            WarehouseStock.warehouse == target_warehouse,
        ).first()
        if not stock:
            stock = WarehouseStock(
                business_id=user.business_id, product_id=p.id,
                warehouse=target_warehouse, quantity=0,
            )
            db.add(stock)
            db.flush()
        other_total = sum(
            int(row.quantity or 0) for row in db.query(WarehouseStock).filter(
                WarehouseStock.business_id == user.business_id,
                WarehouseStock.product_id == p.id,
                WarehouseStock.warehouse != target_warehouse,
            ).all()
        )
        new_stock_quantity = changes["quantity"] - other_total
        if new_stock_quantity < 0:
            raise HTTPException(
                status_code=400,
                detail="Total quantity cannot be lower than stock held in the other warehouses.",
            )
        stock.quantity = new_stock_quantity
    for k, v in changes.items():
        setattr(p, k, v)
    if "quantity" in changes:
        p.quantity = sum(
            int(row.quantity or 0) for row in db.query(WarehouseStock).filter(
                WarehouseStock.business_id == user.business_id,
                WarehouseStock.product_id == p.id,
            ).all()
        )
    auto_upsert_general_catalog(db, p)
    add_audit(db, user, "PRODUCT_UPDATED", f"Updated product {p.name} ({p.sku}).")
    # Quantity/min-stock-level changes affect Business Brain's low-stock and
    # velocity recommendations — mark dirty rather than recomputing inline.
    if "quantity" in changes or "min_stock_level" in changes:
        mark_business_brain_dirty(db, user.business_id)
    # Returns the authoritative updated fields so the frontend can patch its
    # local product state directly instead of reloading the whole inventory
    # (see performance refactor, section 13).
    response = {
        "message": "Product updated successfully.",
        "product": {
            "id": p.id, "sku": p.sku, "barcode": p.barcode, "name": p.name, "category": p.category,
            "size": p.size, "quantity": p.quantity, "min_stock_level": p.min_stock_level,
            "cost_price": p.cost_price, "wholesale_price": p.wholesale_price, "retail_price": p.retail_price,
            "warehouse": p.warehouse, "expiry_date": p.expiry_date,
        },
    }
    complete_idempotent_mutation(claim, response)
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return response

@app.delete("/products/{product_id}")
def delete_product(product_id: int, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    enforce_offline_replay_identity(request, user)
    p = db.query(Product).filter(Product.id == product_id, Product.business_id == user.business_id).first()
    if not p: raise HTTPException(status_code=404, detail="The product could not be found in this inventory.")
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot delete products.")
    if user.role == "manager":
        row = ProductDeletionRequest(business_id=user.business_id, product_id=p.id, product_name=p.name, requested_by_id=user.id, requested_by_name=user.username)
        db.add(row); add_audit(db, user, "PRODUCT_DELETE_REQUESTED", f"Requested Admin approval to delete product {p.name}.")
        db.commit(); return {"message": "Admin approval request sent."}
    add_audit(db, user, "PRODUCT_DELETED", f"Deleted product {p.name} ({p.sku}).")
    mark_business_brain_dirty(db, user.business_id)
    db.delete(p); db.commit(); return {"message": "Product deleted successfully."}

@app.get("/product-deletion-requests")
def list_product_deletion_requests(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Access denied")
    rows = db.query(ProductDeletionRequest).filter(ProductDeletionRequest.business_id == user.business_id, ProductDeletionRequest.status == "PENDING").all()
    return [{"id": r.id, "product_name": r.product_name, "sku": (db.query(Product).filter(Product.id == r.product_id).first().sku if r.product_id and db.query(Product).filter(Product.id == r.product_id).first() else ""), "requested_by_name": r.requested_by_name} for r in rows]

@app.post("/product-deletion-requests/{request_id}/{resolution}")
def resolve_product_deletion(request_id: int, resolution: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can resolve deletion requests.")
    row = db.query(ProductDeletionRequest).filter(ProductDeletionRequest.id == request_id, ProductDeletionRequest.business_id == user.business_id, ProductDeletionRequest.status == "PENDING").first()
    if not row: raise HTTPException(status_code=404, detail="Deletion request is no longer pending.")
    product = db.query(Product).filter(Product.id == row.product_id, Product.business_id == user.business_id).first() if row.product_id else None
    row.status = "APPROVED" if resolution == "approve" else "REJECTED"; row.resolved_by_id = user.id; row.resolved_by_name = user.username; row.resolved_at = datetime.utcnow()
    if resolution == "approve" and product:
        add_audit(db, user, "PRODUCT_DELETED", f"Approved and deleted product {product.name}.")
        mark_business_brain_dirty(db, user.business_id)
        db.delete(product)
    else:
        add_audit(db, user, "PRODUCT_DELETE_REQUEST_REJECTED", f"Rejected product deletion request for {row.product_name}.")
    db.commit(); return {"message": f"Request {row.status.lower()}."}

@app.patch("/products/{product_id}/stock")
def update_stock(product_id: int, data: StockUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.id == product_id, Product.business_id == user.business_id).first()
    if not p: raise HTTPException(status_code=404, detail="The product could not be found in this inventory.")
    warehouse_name = p.warehouse or "Main Central Warehouse"
    stock = db.query(WarehouseStock).filter(
        WarehouseStock.business_id == user.business_id,
        WarehouseStock.product_id == p.id,
        WarehouseStock.warehouse == warehouse_name,
    ).first()
    if not stock:
        stock = WarehouseStock(business_id=user.business_id, product_id=p.id, warehouse=warehouse_name, quantity=0)
        db.add(stock)
        db.flush()
    if stock.quantity + data.quantity_change < 0:
        raise HTTPException(status_code=400, detail=f"Stock in {warehouse_name} cannot become negative.")
    stock.quantity += data.quantity_change
    p.quantity = sum(
        int(row.quantity or 0) for row in db.query(WarehouseStock).filter(
            WarehouseStock.business_id == user.business_id, WarehouseStock.product_id == p.id,
        ).all()
    )
    mark_business_brain_dirty(db, user.business_id)
    check_inventory_notifications(db, user.business_id, p)
    db.commit(); return {"message": "Stock updated successfully.", "quantity": p.quantity}

@app.patch("/products/{product_id}/transfer")
def transfer_stock(product_id: int, data: StockTransfer, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot transfer stock between warehouses.")
    if data.quantity<1 or data.from_warehouse==data.to_warehouse: raise HTTPException(status_code=400, detail="Choose different warehouses and enter a positive quantity.")
    p=db.query(Product).filter(Product.id==product_id,Product.business_id==user.business_id).first()
    if not p: raise HTTPException(status_code=404, detail="The selected product is unavailable.")
    if not get_warehouse_for_business(db, user.business_id, data.from_warehouse) or not get_warehouse_for_business(db, user.business_id, data.to_warehouse):
        raise HTTPException(status_code=400, detail="Both selected warehouses must be active warehouses in this business.")
    source=db.query(WarehouseStock).filter(WarehouseStock.product_id==p.id,WarehouseStock.warehouse==data.from_warehouse).first()
    if not source:
        if p.warehouse==data.from_warehouse:
            source=WarehouseStock(business_id=user.business_id,product_id=p.id,warehouse=data.from_warehouse,quantity=p.quantity); db.add(source); db.flush()
        else: raise HTTPException(status_code=400, detail="There is no stock recorded in the selected source warehouse.")
    if source.quantity<data.quantity: raise HTTPException(status_code=400, detail=f"Only {source.quantity} units are recorded in {data.from_warehouse}.")
    target=db.query(WarehouseStock).filter(WarehouseStock.product_id==p.id,WarehouseStock.warehouse==data.to_warehouse).first()
    if not target: target=WarehouseStock(business_id=user.business_id,product_id=p.id,warehouse=data.to_warehouse,quantity=0); db.add(target); db.flush()
    source.quantity-=data.quantity; target.quantity+=data.quantity
    p.quantity=sum(w.quantity for w in db.query(WarehouseStock).filter(WarehouseStock.product_id==p.id).all())
    add_audit(db,user,"STOCK_TRANSFER",f"Transferred {data.quantity} units of {p.name} from {data.from_warehouse} to {data.to_warehouse}.")
    db.commit(); return {"message":"Stock transfer completed successfully.","quantity_transferred":data.quantity,"from_warehouse":data.from_warehouse,"to_warehouse":data.to_warehouse,"total_quantity":p.quantity}

@app.get("/products/{product_id}/warehouse-stocks")
def product_warehouse_stocks(product_id:int,user:User=Depends(get_current_user),db:Session=Depends(get_db)):
    p=db.query(Product).filter(Product.id==product_id,Product.business_id==user.business_id).first()
    if not p: raise HTTPException(status_code=404, detail="The selected product is unavailable.")
    rows=db.query(WarehouseStock).filter(WarehouseStock.product_id==p.id,WarehouseStock.business_id==user.business_id).order_by(WarehouseStock.warehouse.asc()).all()
    if not rows:
        row=WarehouseStock(business_id=user.business_id,product_id=p.id,warehouse=p.warehouse or "Main Central Warehouse",quantity=p.quantity); db.add(row); db.commit(); rows=[row]
    return [{"warehouse":r.warehouse,"quantity":r.quantity} for r in rows]

@app.get("/products/predictive-forecast")
def predictive_forecast(user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    sales_count = db.query(SaleModel).filter(SaleModel.business_id == user.business_id).count()
    # Distinct calendar dates with at least one closed session — not a raw
    # row count. A business can close and reopen multiple sessions on the
    # same date (see the BusinessDay model docstring); counting rows would
    # inflate "history days" and understate daily velocity.
    history_days = db.query(func.count(func.distinct(BusinessDay.date))).filter(BusinessDay.business_id == user.business_id, BusinessDay.is_open == False).scalar() or 0
    products = db.query(Product).filter(Product.business_id == user.business_id).all()
    forecast = []
    for p in products:
        product_sales = db.query(func.sum(SaleModel.quantity)).filter(SaleModel.business_id == user.business_id, SaleModel.product_id == p.id).scalar() or 0
        daily_velocity = (product_sales / max(history_days, 1)) if history_days else 0
        days_to_stockout = round(p.quantity / daily_velocity, 1) if daily_velocity > 0 else None
        risk = "unknown"
        if days_to_stockout is not None:
            risk = "critical" if days_to_stockout <= 7 else "moderate" if days_to_stockout <= 21 else "low"
        forecast.append({"name": p.name, "sku": p.sku, "quantity": p.quantity, "daily_velocity": round(daily_velocity, 2), "days_to_stockout": days_to_stockout, "risk": risk})
    return {"report_ready": history_days >= 6, "history_days": history_days, "sales_count": sales_count, "forecast": forecast}

# -----------------------------------------------------------------------------
# INVENTORY FINANCIAL INTELLIGENCE
#
# Extends the predictive-forecast calculations above into the four money
# questions: what's at risk of a missed sale, what capital is tied up and not
# moving, where margin is under pressure, and what might be recoverable.
# Every number here is computed from this business's actual products/sales —
# nothing is invented, and any category without enough underlying data reports
# that honestly instead of guessing. Reuses the same sales-velocity/burn-rate
# formula as predictive_forecast() above rather than a second calculation, and
# the business's own configured BusinessProfile.currency (never assumes NGN).
# -----------------------------------------------------------------------------
@app.get("/products/financial-intelligence")
def financial_intelligence(user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    # Distinct calendar dates with a closed session — see predictive_forecast().
    history_days = db.query(func.count(func.distinct(BusinessDay.date))).filter(BusinessDay.business_id == user.business_id, BusinessDay.is_open == False).scalar() or 0
    currency = business.currency if business and business.currency else "USD ($)"

    if history_days < FIN_INTEL_MIN_HISTORY_DAYS:
        return {
            "report_ready": False, "history_days": history_days, "currency": currency,
            "insufficient_data_message": "Not enough sales history to calculate a reliable prediction.",
            "money_at_risk": {"total": 0, "products": []}, "money_tied_up": {"total": 0, "products": []},
            "margin_pressure": {"total": 0, "products": []}, "potentially_recoverable": {"total": 0, "products": []},
            "products": [],
        }

    now = datetime.utcnow()
    slow_moving_cutoff = now - timedelta(days=FIN_INTEL_SLOW_MOVING_LOOKBACK_DAYS)
    margin_window_start = now - timedelta(days=FIN_INTEL_MARGIN_WINDOW_DAYS)
    products = db.query(Product).filter(Product.business_id == user.business_id).all()

    # Everything below used to run 3-4 separate queries PER PRODUCT inside
    # this loop (total sold, last sale, recent-window sales, multi-warehouse
    # check) — classic N+1 behavior that scaled with catalog size. Replaced
    # with a fixed, small number of grouped aggregate queries up front, then
    # merged into per-product lookups in Python. Query count now stays
    # constant regardless of how many products exist.
    total_sold_map = dict(
        db.query(SaleModel.product_id, func.sum(SaleModel.quantity))
        .filter(SaleModel.business_id == user.business_id)
        .group_by(SaleModel.product_id).all()
    )
    last_sale_map = dict(
        db.query(SaleModel.product_id, func.max(SaleModel.timestamp))
        .filter(SaleModel.business_id == user.business_id)
        .group_by(SaleModel.product_id).all()
    )
    recent_sales_map = {
        pid: (units or 0, revenue or 0.0)
        for pid, units, revenue in (
            db.query(SaleModel.product_id, func.sum(SaleModel.quantity), func.sum(SaleModel.total_price))
            .filter(SaleModel.business_id == user.business_id, SaleModel.timestamp >= margin_window_start)
            .group_by(SaleModel.product_id).all()
        )
    }
    multi_warehouse_map = dict(
        db.query(WarehouseStock.product_id, func.count(func.distinct(WarehouseStock.warehouse)))
        .filter(WarehouseStock.business_id == user.business_id, WarehouseStock.quantity > 0)
        .group_by(WarehouseStock.product_id).all()
    )

    at_risk_rows, tied_up_rows, margin_rows, recoverable_rows, products_out = [], [], [], [], []
    at_risk_total = tied_up_total = margin_total = recoverable_total = 0.0
    products_missing_cost = products_missing_price = 0

    for p in products:
        # Same all-time-sold / completed-business-days velocity used by
        # /products/predictive-forecast — deliberately not duplicated with a
        # different formula.
        total_sold = total_sold_map.get(p.id, 0)
        daily_velocity = (total_sold / history_days) if history_days else 0.0
        days_to_stockout = round(p.quantity / daily_velocity, 1) if daily_velocity > 0 else None

        last_sale_at = last_sale_map.get(p.id)
        days_since_last_sale = (now - last_sale_at).days if last_sale_at else None

        why_flags = []
        product_at_risk = product_tied_up = product_margin_impact = 0.0
        is_slow_moving = False

        # 1) MONEY AT RISK FROM STOCKOUTS — expected demand during the risk
        # window that current stock can't cover, priced at the actual selling
        # price. NOT current-stock-value; only the shortfall.
        if daily_velocity > 0 and days_to_stockout is not None and days_to_stockout <= FIN_INTEL_STOCKOUT_RISK_DAYS:
            expected_demand_in_window = daily_velocity * FIN_INTEL_STOCKOUT_RISK_DAYS
            units_at_risk = max(0.0, expected_demand_in_window - p.quantity)
            if units_at_risk > 0 and p.retail_price > 0:
                product_at_risk = units_at_risk * p.retail_price
                at_risk_total += product_at_risk
                at_risk_rows.append({
                    "product_id": p.id, "name": p.name, "sku": p.sku, "quantity": p.quantity,
                    "days_to_stockout": days_to_stockout, "potential_sales_at_risk": round(product_at_risk, 2),
                })
                why_flags.append(f"Selling about {daily_velocity:.1f} units/day — at this pace stock runs out in ~{days_to_stockout} days, inside the {FIN_INTEL_STOCKOUT_RISK_DAYS}-day risk window.")
            elif units_at_risk > 0 and p.retail_price <= 0:
                products_missing_price += 1

        # 2) MONEY TIED UP IN SLOW-MOVING INVENTORY — actual last-sale history,
        # not "any product with a gap".
        if p.quantity > 0 and (last_sale_at is None or last_sale_at < slow_moving_cutoff):
            is_slow_moving = True
            if p.cost_price > 0:
                product_tied_up = p.quantity * p.cost_price
                tied_up_total += product_tied_up
                tied_up_rows.append({
                    "product_id": p.id, "name": p.name, "sku": p.sku, "quantity": p.quantity,
                    "cost_per_unit": p.cost_price, "capital_tied_up": round(product_tied_up, 2),
                    "last_sale": to_utc_iso(last_sale_at),
                })
                why_flags.append(f"No sale recorded {'in the last ' + str(days_since_last_sale) + ' days' if days_since_last_sale is not None else 'yet'}, with {p.quantity} units still in stock.")
            else:
                products_missing_cost += 1

        # 3) MARGIN PRESSURE — compares the NOMINAL margin (listed retail minus
        # cost) against the REALIZED margin actually achieved on recent sales
        # (recent revenue / recent units sold, minus cost). A gap means recent
        # sales are effectively going out below the listed margin (discounting,
        # promotions, price erosion) — this is the only "previous vs current
        # margin" comparison the available data can honestly support, since
        # individual historical sale prices aren't tracked separately.
        recent_units, recent_revenue = recent_sales_map.get(p.id, (0, 0.0))
        nominal_margin = current_margin = None
        if recent_units > 0 and p.retail_price > 0 and p.cost_price >= 0:
            realized_avg_price = recent_revenue / recent_units
            nominal_margin_per_unit = p.retail_price - p.cost_price
            realized_margin_per_unit = realized_avg_price - p.cost_price
            nominal_margin = round((nominal_margin_per_unit / p.retail_price) * 100, 1) if p.retail_price else None
            current_margin = round((realized_margin_per_unit / realized_avg_price) * 100, 1) if realized_avg_price else None
            if nominal_margin_per_unit > 0 and realized_margin_per_unit < nominal_margin_per_unit:
                gap_per_unit = nominal_margin_per_unit - realized_margin_per_unit
                product_margin_impact = gap_per_unit * recent_units  # window IS the monthlyized period
                margin_total += product_margin_impact
                margin_rows.append({
                    "product_id": p.id, "name": p.name, "sku": p.sku,
                    "previous_margin_pct": nominal_margin, "current_margin_pct": current_margin,
                    "estimated_monthly_impact": round(product_margin_impact, 2),
                })
                why_flags.append(f"Recent average selling price is below list — realized margin ~{current_margin}% vs. listed ~{nominal_margin}%.")
        elif recent_units == 0:
            pass  # not enough recent sales history for this product's margin trend — omitted honestly, not guessed

        # 4) POTENTIALLY RECOVERABLE — the slow-moving subset, with a
        # deterministic (not AI-guessed) recommended action.
        if is_slow_moving and product_tied_up > 0:
            multi_warehouse = multi_warehouse_map.get(p.id, 0) > 1
            if multi_warehouse:
                action = "Transfer to a branch/warehouse where this product moves faster"
            elif daily_velocity == 0:
                action = "Run a promotion or discount to move stock"
            elif days_since_last_sale is not None and days_since_last_sale > FIN_INTEL_SEVERELY_STALE_DAYS:
                action = "Bundle with faster-moving products or discount"
            else:
                action = "Reduce future purchasing of this product"
            recoverable_total += product_tied_up
            recoverable_rows.append({
                "product_id": p.id, "name": p.name, "sku": p.sku, "quantity": p.quantity,
                "cost_value": round(product_tied_up, 2), "last_sale": to_utc_iso(last_sale_at),
                "recommended_action": action, "potential_value": round(product_tied_up, 2),
            })

        products_out.append({
            "product_id": p.id, "name": p.name, "sku": p.sku, "quantity": p.quantity,
            "daily_velocity": round(daily_velocity, 2), "days_to_stockout": days_to_stockout,
            "potential_sales_at_risk": round(product_at_risk, 2) if product_at_risk else 0,
            "capital_tied_up": round(product_tied_up, 2) if product_tied_up else 0,
            "is_slow_moving": is_slow_moving,
            "previous_margin_pct": nominal_margin, "current_margin_pct": current_margin,
            "estimated_monthly_margin_impact": round(product_margin_impact, 2) if product_margin_impact else 0,
            "last_sale": to_utc_iso(last_sale_at),
            "why_flagged": why_flags,
        })

    notes = []
    if products_missing_price:
        notes.append("Cost and selling price information are required to calculate margin pressure for some products, so they were excluded from that estimate.")
    if products_missing_cost:
        notes.append("Cost information is required to calculate capital tied up for some slow-moving products, so they were excluded from that total.")

    return {
        "report_ready": True, "history_days": history_days, "currency": currency, "notes": notes,
        "money_at_risk": {"total": round(at_risk_total, 2), "products": sorted(at_risk_rows, key=lambda r: -r["potential_sales_at_risk"])},
        "money_tied_up": {"total": round(tied_up_total, 2), "products": sorted(tied_up_rows, key=lambda r: -r["capital_tied_up"])},
        "margin_pressure": {"total": round(margin_total, 2), "products": sorted(margin_rows, key=lambda r: -r["estimated_monthly_impact"])},
        "potentially_recoverable": {"total": round(recoverable_total, 2), "products": sorted(recoverable_rows, key=lambda r: -r["potential_value"])},
        "products": products_out,
    }


# -----------------------------------------------------------------------------
# BUSINESS BRAIN
# -----------------------------------------------------------------------------
BUSINESS_BRAIN_HISTORY_DAYS = 6
BUSINESS_BRAIN_FORECAST_HORIZON_DAYS = 7

# --- Prediction self-calibration (additive) ---------------------------------
# Learns, per product and per prediction kind, whether Cauldra's own past
# forecasts ran high or low — then nudges future forecast MAGNITUDE toward
# reality. This is separate from the confidence score above: confidence says
# how sure Cauldra is, calibration corrects for a known, measured bias.
CALIBRATION_MIN_SAMPLES = 3     # don't correct anything until this many evaluated predictions exist
CALIBRATION_LOOKBACK = 12       # weight recent performance, not the entire lifetime of the product
CALIBRATION_MIN_FACTOR = 0.5    # never adjust a forecast by more than 2x in either direction —
CALIBRATION_MAX_FACTOR = 1.75   # a single noisy stretch of evaluations shouldn't be over-trusted

def _prediction_bias_factor(db: Session, business_id: int, product_id: int, kind: str) -> Optional[float]:
    """Average of (actual/predicted) across this product's recent evaluated
    predictions of this kind. None until there's enough evaluated history to
    trust a correction — an untested bias guess would be exactly the kind of
    fabricated confidence this system exists to avoid."""
    rows = (
        db.query(BusinessBrainPrediction)
        .filter(
            BusinessBrainPrediction.business_id == business_id,
            BusinessBrainPrediction.product_id == product_id,
            BusinessBrainPrediction.kind == kind,
            BusinessBrainPrediction.actual_units.isnot(None),
        )
        .order_by(BusinessBrainPrediction.evaluated_at.desc())
        .limit(CALIBRATION_LOOKBACK)
        .all()
    )
    if len(rows) < CALIBRATION_MIN_SAMPLES:
        return None
    ratios = [row.actual_units / row.predicted_units for row in rows if row.predicted_units and row.predicted_units > 0]
    if len(ratios) < CALIBRATION_MIN_SAMPLES:
        return None
    factor = sum(ratios) / len(ratios)
    return round(max(CALIBRATION_MIN_FACTOR, min(CALIBRATION_MAX_FACTOR, factor)), 3)

# --- Seasonal pattern detection (additive) ----------------------------------
# Learns each product's own recurring week-of-year, not a generic calendar.
# Every threshold below exists to keep Cauldra silent when evidence is weak
# rather than to force a claim: see SEASONAL_MIN_ALERT_CONFIDENCE.
SEASONAL_MIN_TOTAL_WEEKS = 10          # don't even look for a pattern before this much trading history exists
SEASONAL_MIN_WEEK_UNITS = 3            # ignore buckets too small in volume to mean anything
SEASONAL_MIN_LIFT_RATIO = 1.3          # bucket must clearly exceed the product's own baseline
SEASONAL_LOOKAHEAD_DAYS = 21           # how far ahead Cauldra watches for an approaching known period
SEASONAL_MIN_ALERT_CONFIDENCE = 0.30   # below this: stay silent (prefer no alert over a low-quality one)
SEASONAL_RECOMPUTE_INTERVAL_HOURS = 12 # throttle so this never runs on every page load

def _iso_year_week(ts: datetime) -> tuple[int, int]:
    iso = ts.isocalendar()
    return int(iso[0]), int(iso[1])

def _weekly_sales_totals(db: Session, business_id: int, product_id: int) -> Dict[tuple, float]:
    """This product's own units sold, grouped by (iso_year, iso_week). Shared
    by seasonal-pattern detection and cross-product relationship detection so
    both learn from the exact same weekly ledger."""
    sales = db.query(SaleModel.timestamp, SaleModel.quantity).filter(
        SaleModel.business_id == business_id, SaleModel.product_id == product_id
    ).all()
    weekly_totals: Dict[tuple, float] = {}
    for ts, qty in sales:
        if ts is None: continue
        key = _iso_year_week(ts)
        weekly_totals[key] = weekly_totals.get(key, 0.0) + float(qty)
    return weekly_totals

def _seasonal_confidence(cycles_observed: int, consistency: float, lift_ratio: float) -> float:
    cycles_component = min(0.35, cycles_observed * 0.15)
    consistency_component = consistency * 0.25
    lift_component = min(0.15, max(0.0, (lift_ratio - 1.0) * 0.3))
    return round(min(0.90, 0.15 + cycles_component + consistency_component + lift_component), 2)

def _seasonal_confidence_label(confidence: float) -> str:
    if confidence >= 0.70: return "Strong seasonal pattern detected."
    if confidence >= 0.45: return "Demand is likely to increase soon."
    return "Possible increase in demand."

def _next_occurrence_of_week(week_of_year: int, now: datetime) -> datetime:
    year = now.isocalendar()[0]
    try:
        candidate = datetime.fromisocalendar(year, week_of_year, 1)
    except ValueError:
        candidate = datetime.fromisocalendar(year, 52, 1)
    if candidate.date() < now.date():
        try:
            candidate = datetime.fromisocalendar(year + 1, week_of_year, 1)
        except ValueError:
            candidate = datetime.fromisocalendar(year + 1, 52, 1)
    return candidate

def _recompute_seasonal_pattern(db: Session, business_id: int, product: Product, now: datetime) -> None:
    """Evidence-gated: only ever stores the single best-supported recurring
    week for this product, learned entirely from this product's own sales —
    never a generic seasonal assumption. Withdraws a stored pattern the
    moment the evidence behind it no longer clears the confidence floor."""
    weekly_totals = _weekly_sales_totals(db, business_id, product.id)
    existing = db.query(BusinessBrainSeasonalPattern).filter(
        BusinessBrainSeasonalPattern.business_id == business_id,
        BusinessBrainSeasonalPattern.product_id == product.id,
    ).first()

    distinct_weeks = len(weekly_totals)
    if distinct_weeks < SEASONAL_MIN_TOTAL_WEEKS:
        _withdraw_seasonal_pattern(db, business_id, product.id, existing)
        return
    baseline_avg = sum(weekly_totals.values()) / distinct_weeks
    if baseline_avg <= 0:
        _withdraw_seasonal_pattern(db, business_id, product.id, existing)
        return

    by_bucket: Dict[int, List[float]] = {}
    for (_yr, wk), total in weekly_totals.items():
        by_bucket.setdefault(wk, []).append(total)

    best = None  # (week, cycles_observed, avg_in_week, lift_ratio, consistency, confidence)
    for wk, values in by_bucket.items():
        cycles_observed = len(values)
        avg_in_week = sum(values) / cycles_observed
        if avg_in_week < SEASONAL_MIN_WEEK_UNITS:
            continue
        lift_ratio = avg_in_week / baseline_avg
        if lift_ratio < SEASONAL_MIN_LIFT_RATIO:
            continue
        if cycles_observed >= 2:
            variance = sum((v - avg_in_week) ** 2 for v in values) / cycles_observed
            consistency = max(0.0, min(1.0, 1 - ((variance ** 0.5) / avg_in_week)))
        else:
            consistency = 0.0  # only one prior occurrence: nothing to compare it against yet
        confidence = _seasonal_confidence(cycles_observed, consistency, lift_ratio)
        if confidence < SEASONAL_MIN_ALERT_CONFIDENCE:
            continue
        if best is None or confidence > best[5]:
            best = (wk, cycles_observed, avg_in_week, lift_ratio, consistency, confidence)

    if best is None:
        _withdraw_seasonal_pattern(db, business_id, product.id, existing)
        return
    wk, cycles_observed, avg_in_week, lift_ratio, consistency, confidence = best
    if existing is None:
        existing = BusinessBrainSeasonalPattern(business_id=business_id, product_id=product.id)
        db.add(existing)
    existing.week_of_year = wk
    existing.cycles_observed = cycles_observed
    existing.avg_units_in_week = round(avg_in_week, 2)
    existing.baseline_avg_units = round(baseline_avg, 2)
    existing.lift_ratio = round(lift_ratio, 3)
    existing.consistency = round(consistency, 3)
    existing.confidence = confidence
    existing.last_computed_at = now

def _withdraw_seasonal_pattern(db: Session, business_id: int, product_id: int, existing) -> None:
    """One place to retire a product's seasonal pattern: drop the stored row and
    its plain-language Memory together, so a Memory statement can never outlive
    the evidence that produced it."""
    if existing is not None:
        db.delete(existing)
    _delete_brain_memory(db, business_id, f"seasonal:{product_id}")

def _apply_seasonal_pattern(db: Session, business_id: int, product: Product, now: datetime) -> set:
    """Recomputes (throttled) then, if a pattern is on file, records the plain-
    language memory and — only when the known period is actually approaching
    and current stock looks short — a timely, actionable recommendation. Returns
    the set of recommendation fingerprints this pass still supports."""
    if product.seasonal_checked_at is None or (now - product.seasonal_checked_at) > timedelta(hours=SEASONAL_RECOMPUTE_INTERVAL_HOURS):
        _recompute_seasonal_pattern(db, business_id, product, now)
        product.seasonal_checked_at = now

    pattern = db.query(BusinessBrainSeasonalPattern).filter(
        BusinessBrainSeasonalPattern.business_id == business_id,
        BusinessBrainSeasonalPattern.product_id == product.id,
    ).first()
    if pattern is None:
        return set()

    upcoming = _next_occurrence_of_week(pattern.week_of_year, now)
    days_until = (upcoming.date() - now.date()).days
    calibration_factor = _prediction_bias_factor(db, business_id, product.id, "seasonal")
    calibrated_units = round(pattern.avg_units_in_week * (calibration_factor or 1.0), 2)
    evidence = {
        "week_of_year": pattern.week_of_year,
        "cycles_observed": pattern.cycles_observed,
        "average_units_in_period": pattern.avg_units_in_week,
        "calibrated_units": calibrated_units,
        "calibration_factor": calibration_factor,
        "baseline_average_units": pattern.baseline_avg_units,
        "lift_ratio": pattern.lift_ratio,
        "days_until_period": days_until,
    }
    times = "time" if pattern.cycles_observed == 1 else "times"
    _upsert_brain_memory(
        db, business_id, product.id, f"seasonal:{product.id}",
        f"{product.name} tends to sell more around the week of {upcoming.strftime('%b %d')}, based on {pattern.cycles_observed} previous {times} in this business's own sales history.",
        evidence, pattern.confidence, now,
    )

    # Track this cycle's forecast so its accuracy can be measured once the
    # window passes (_evaluate_due_brain_predictions, already generic, does
    # the comparison) and future occurrences of this product's pattern can be
    # calibrated against it — the self-correction loop from section 9/10.
    target_at = upcoming + timedelta(days=7)
    open_prediction = db.query(BusinessBrainPrediction).filter(
        BusinessBrainPrediction.business_id == business_id,
        BusinessBrainPrediction.product_id == product.id,
        BusinessBrainPrediction.kind == "seasonal",
        BusinessBrainPrediction.actual_units.is_(None),
    ).first()
    if open_prediction is None or open_prediction.target_at.date() != target_at.date():
        db.add(BusinessBrainPrediction(
            business_id=business_id, product_id=product.id, kind="seasonal",
            forecast_at=now, target_at=target_at, horizon_days=7,
            predicted_units=calibrated_units, confidence=pattern.confidence,
            evidence_json=json.dumps(evidence),
        ))
    else:
        open_prediction.predicted_units = calibrated_units
        open_prediction.confidence = pattern.confidence
        open_prediction.evidence_json = json.dumps(evidence)

    if -3 <= days_until <= SEASONAL_LOOKAHEAD_DAYS and product.quantity < calibrated_units:
        shortfall_ratio = (product.quantity / calibrated_units) if calibrated_units else 0
        priority = "critical" if shortfall_ratio < 0.5 else "important"
        when_text = "around now" if days_until <= 0 else f"in about {days_until} day{'s' if days_until != 1 else ''}"
        _upsert_brain_recommendation(
            db, business_id, product.id, f"seasonal-lead:{product.id}", "seasonal_demand", priority,
            f"{product.name} demand may rise soon",
            f"{_seasonal_confidence_label(pattern.confidence)} {product.name} usually picks up {when_text}. "
            f"Current stock is {product.quantity}, which may not cover the expected demand of about {calibrated_units:g} units. "
            f"Consider replenishing before the increase.",
            {**evidence, "current_stock": product.quantity},
        )
        return {f"seasonal-lead:{product.id}"}
    return set()

# --- Cross-product relationship detection (additive) ------------------------
# Learns whether two products' weekly sales move together, purely from this
# business's own numbers — see section 12: only used when evidence supports
# it, never a guessed or hardcoded relationship.
RELATIONSHIP_MAX_PRODUCTS = 30            # bound the pairwise comparison to the business's most active products
RELATIONSHIP_MIN_OVERLAP_WEEKS = 8        # need real shared trading history before trusting any correlation
RELATIONSHIP_MIN_ABS_CORRELATION = 0.6    # only a clearly strong co-movement counts as a relationship
RELATIONSHIP_MAX_STORED = 10              # keep only the business's most confident relationships, not a full matrix
RELATIONSHIP_RECOMPUTE_INTERVAL_HOURS = 24

def _pearson_correlation(a: List[float], b: List[float]) -> Optional[float]:
    n = len(a)
    if n < 2: return None
    mean_a, mean_b = sum(a) / n, sum(b) / n
    cov = sum((a[i] - mean_a) * (b[i] - mean_b) for i in range(n))
    var_a = sum((v - mean_a) ** 2 for v in a)
    var_b = sum((v - mean_b) ** 2 for v in b)
    if var_a <= 0 or var_b <= 0: return None  # a constant series can't be correlated with anything
    return cov / ((var_a ** 0.5) * (var_b ** 0.5))

def _relationship_confidence(overlapping_weeks: int, correlation: float) -> float:
    weeks_component = min(0.40, (overlapping_weeks / 40) * 0.40)
    strength_component = min(0.35, max(0.0, (abs(correlation) - RELATIONSHIP_MIN_ABS_CORRELATION) / (1 - RELATIONSHIP_MIN_ABS_CORRELATION) * 0.35))
    return round(min(0.85, 0.15 + weeks_component + strength_component), 2)

def _recompute_business_relationships(db: Session, business_id: int, now: datetime) -> None:
    """O(n^2) over a bounded, most-active product set, throttled to run at
    most once a day per business — see section 26 on not letting a richer
    intelligence layer make the app slow."""
    active_products = (
        db.query(Product.id)
        .join(SaleModel, SaleModel.product_id == Product.id)
        .filter(Product.business_id == business_id)
        .group_by(Product.id)
        .order_by(func.sum(SaleModel.quantity).desc())
        .limit(RELATIONSHIP_MAX_PRODUCTS)
        .all()
    )
    product_ids = [row[0] for row in active_products]
    series_by_product = {pid: _weekly_sales_totals(db, business_id, pid) for pid in product_ids}

    found: List[tuple] = []  # (product_a, product_b, overlapping_weeks, correlation, confidence)
    for i in range(len(product_ids)):
        for j in range(i + 1, len(product_ids)):
            pid_a, pid_b = product_ids[i], product_ids[j]
            weeks_a, weeks_b = series_by_product[pid_a], series_by_product[pid_b]
            shared_weeks = sorted(set(weeks_a.keys()) & set(weeks_b.keys()))
            if len(shared_weeks) < RELATIONSHIP_MIN_OVERLAP_WEEKS:
                continue
            series_a = [weeks_a[wk] for wk in shared_weeks]
            series_b = [weeks_b[wk] for wk in shared_weeks]
            correlation = _pearson_correlation(series_a, series_b)
            if correlation is None or abs(correlation) < RELATIONSHIP_MIN_ABS_CORRELATION:
                continue
            confidence = _relationship_confidence(len(shared_weeks), correlation)
            found.append((min(pid_a, pid_b), max(pid_a, pid_b), len(shared_weeks), round(correlation, 3), confidence))

    found.sort(key=lambda row: row[4], reverse=True)
    keep = found[:RELATIONSHIP_MAX_STORED]
    keep_pairs = {(a, b) for a, b, *_ in keep}

    existing_rows = db.query(BusinessBrainRelationship).filter(BusinessBrainRelationship.business_id == business_id).all()
    for row in existing_rows:
        if (row.product_a_id, row.product_b_id) not in keep_pairs:
            _delete_brain_memory(db, business_id, f"relationship:{row.product_a_id}:{row.product_b_id}")
            db.delete(row)  # evidence no longer supports this pairing; withdraw it

    for product_a, product_b, overlapping_weeks, correlation, confidence in keep:
        row = db.query(BusinessBrainRelationship).filter(
            BusinessBrainRelationship.business_id == business_id,
            BusinessBrainRelationship.product_a_id == product_a,
            BusinessBrainRelationship.product_b_id == product_b,
        ).first()
        if row is None:
            row = BusinessBrainRelationship(business_id=business_id, product_a_id=product_a, product_b_id=product_b)
            db.add(row)
        row.overlapping_weeks = overlapping_weeks
        row.correlation = correlation
        row.confidence = confidence
        row.last_computed_at = now

def _apply_business_relationships(db: Session, business_id: int, now: datetime) -> None:
    """Recomputes (throttled) then records each surviving relationship as a
    plain-language memory — reusing the same memory surface everything else
    in the Business Brain already writes to, so no new UI is needed."""
    business = db.query(BusinessProfile).filter(BusinessProfile.id == business_id).first()
    if business is None:
        return
    if business.brain_relationships_checked_at is None or (now - business.brain_relationships_checked_at) > timedelta(hours=RELATIONSHIP_RECOMPUTE_INTERVAL_HOURS):
        _recompute_business_relationships(db, business_id, now)
        business.brain_relationships_checked_at = now

    relationships = db.query(BusinessBrainRelationship).filter(BusinessBrainRelationship.business_id == business_id).all()
    for rel in relationships:
        product_a = db.query(Product).filter(Product.id == rel.product_a_id).first()
        product_b = db.query(Product).filter(Product.id == rel.product_b_id).first()
        if not product_a or not product_b:
            continue
        direction = "tend to rise and fall together" if rel.correlation > 0 else "tend to move in opposite directions"
        evidence = {
            "product_a": product_a.name, "product_b": product_b.name,
            "overlapping_weeks": rel.overlapping_weeks, "correlation": rel.correlation,
        }
        _upsert_brain_memory(
            db, business_id, product_a.id, f"relationship:{rel.product_a_id}:{rel.product_b_id}",
            f"{product_a.name} and {product_b.name} {direction} in weekly sales, based on {rel.overlapping_weeks} shared weeks of history.",
            evidence, rel.confidence, now,
        )

def _brain_confidence(history_days: int, prior_accuracy: Optional[float]) -> float:
    """A deterministic confidence score, calibrated by usable history and results."""
    history_component = min(0.55, history_days / 90 * 0.55)
    accuracy_component = 0.15 if prior_accuracy is None else max(0.0, min(0.30, prior_accuracy * 0.30))
    return round(min(0.90, 0.20 + history_component + accuracy_component), 2)

def _brain_confidence_label(confidence: float) -> str:
    return "High confidence" if confidence >= 0.70 else "Moderate confidence" if confidence >= 0.45 else "Limited confidence"

def _evaluate_due_brain_predictions(db: Session, business_id: int, now: datetime) -> None:
    due = db.query(BusinessBrainPrediction).filter(BusinessBrainPrediction.business_id == business_id, BusinessBrainPrediction.actual_units.is_(None), BusinessBrainPrediction.target_at <= now).all()
    for prediction in due:
        actual = db.query(func.sum(SaleModel.quantity)).filter(SaleModel.business_id == business_id, SaleModel.product_id == prediction.product_id, SaleModel.timestamp >= prediction.forecast_at, SaleModel.timestamp <= prediction.target_at).scalar() or 0
        prediction.actual_units = float(actual)
        prediction.accuracy_score = round(max(0.0, 1 - abs(float(actual) - prediction.predicted_units) / max(prediction.predicted_units, 1.0)), 3)
        prediction.evaluated_at = now

# --- Business Brain lifecycle (data organisation, not new intelligence) ------
# Recommendations, predictions and memories are STATE, not an append-only feed.
# The engine already dedupes by fingerprint (one row per condition, updated in
# place). These helpers add the rest of the lifecycle: an active item that the
# engine no longer re-affirms leaves the active brief for History; an item it
# had auto-resolved comes back if its condition returns; a Memory is withdrawn
# when the evidence behind it is withdrawn. No item is deleted purely for age.
RECOMMENDATION_ACTIVE_STATUSES = ("new", "opened")
RECOMMENDATION_RESOLVED_AUTO = ("resolved", "expired")
RECOMMENDATION_HISTORY_STATUSES = ("acted", "dismissed", "resolved", "expired")
CONDITION_RECOMMENDATION_KINDS = ("stock_review", "forecast_stockout", "seasonal_demand")

def _delete_brain_memory(db: Session, business_id: int, fingerprint: str) -> None:
    db.query(BusinessBrainMemory).filter(
        BusinessBrainMemory.business_id == business_id,
        BusinessBrainMemory.fingerprint == fingerprint,
    ).delete(synchronize_session=False)

def _upsert_brain_recommendation(db: Session, business_id: int, product_id: Optional[int], fingerprint: str, kind: str, priority: str, title: str, summary: str, evidence: Dict[str, Any]):
    row = db.query(BusinessBrainRecommendation).filter(BusinessBrainRecommendation.business_id == business_id, BusinessBrainRecommendation.fingerprint == fingerprint).first()
    if row is None:
        row = BusinessBrainRecommendation(business_id=business_id, product_id=product_id, fingerprint=fingerprint, kind=kind, priority=priority, title=title, summary=summary, evidence_json=json.dumps(evidence))
        db.add(row)
    else:
        row.priority, row.title, row.summary, row.evidence_json = priority, title, summary, json.dumps(evidence)
        # A condition Cauldra had auto-resolved has returned: this is a genuinely
        # new occurrence, so re-open it. A user's explicit acted/dismissed is
        # respected and never auto-revived.
        if row.status in RECOMMENDATION_RESOLVED_AUTO:
            row.status = "new"
            row.opened_at = row.acted_at = row.dismissed_at = None
    return row

def _retire_stale_brain_recommendations(db: Session, business_id: int, now: datetime, affirmed: set, history_days: int, product_ids: set) -> None:
    """Condition-based recommendations a full refresh did NOT re-affirm this pass
    (stock replenished, demand risk gone, seasonal window passed, product
    removed) leave the active brief and move to History. A kind is only eligible
    when this pass actually re-checked its condition."""
    active = db.query(BusinessBrainRecommendation).filter(
        BusinessBrainRecommendation.business_id == business_id,
        BusinessBrainRecommendation.status.in_(RECOMMENDATION_ACTIVE_STATUSES),
        BusinessBrainRecommendation.kind.in_(CONDITION_RECOMMENDATION_KINDS),
    ).all()
    for rec in active:
        if rec.fingerprint in affirmed:
            continue
        product_gone = rec.product_id is not None and rec.product_id not in product_ids
        if not product_gone and rec.kind in ("forecast_stockout", "seasonal_demand") and history_days < BUSINESS_BRAIN_HISTORY_DAYS:
            # The velocity/seasonal loop is skipped below this history threshold,
            # so a missing re-affirmation is not evidence the condition cleared.
            continue
        rec.status = "resolved"  # updated_at auto-bumps and is the resolution time History shows

def _upsert_brain_memory(db: Session, business_id: int, product_id: int, fingerprint: str, statement: str, evidence: Dict[str, Any], confidence: float, observed_at: datetime):
    row = db.query(BusinessBrainMemory).filter(BusinessBrainMemory.business_id == business_id, BusinessBrainMemory.fingerprint == fingerprint).first()
    if row is None:
        row = BusinessBrainMemory(business_id=business_id, product_id=product_id, fingerprint=fingerprint, statement=statement, evidence_json=json.dumps(evidence), confidence=confidence, last_observed_at=observed_at)
        db.add(row)
    else:
        row.statement, row.evidence_json, row.confidence, row.last_observed_at = statement, json.dumps(evidence), confidence, observed_at
    return row

def _brain_accuracy_trend(evaluated_predictions: List["BusinessBrainPrediction"]) -> Optional[str]:
    """Compares the accuracy of Cauldra's more recent evaluated predictions
    against its earlier ones for this business. Requires enough evaluated
    predictions on both sides of the split to be a meaningful comparison —
    otherwise returns None rather than guessing at a trend."""
    if len(evaluated_predictions) < 6:
        return None
    ordered = sorted(evaluated_predictions, key=lambda p: p.evaluated_at or datetime.min)
    midpoint = len(ordered) // 2
    earlier, recent = ordered[:midpoint], ordered[midpoint:]
    if len(earlier) < 3 or len(recent) < 3:
        return None
    earlier_avg = sum(p.accuracy_score for p in earlier) / len(earlier)
    recent_avg = sum(p.accuracy_score for p in recent) / len(recent)
    delta = recent_avg - earlier_avg
    if delta > 0.03: return "improving"
    if delta < -0.03: return "declining"
    return "stable"

def _brain_revenue_at_risk(db: Session, business_id: int) -> Optional[float]:
    """Sums the revenue exposed by currently open (non-dismissed) demand-
    shortfall recommendations, using only the expected-demand figure and
    live product price already computed elsewhere — never an invented or
    historical 'savings' number. When a product has both a 7-day forecast
    and a seasonal forecast open at once, only the larger shortfall counts,
    so the same units are never priced twice."""
    rows = db.query(BusinessBrainRecommendation).filter(
        BusinessBrainRecommendation.business_id == business_id,
        BusinessBrainRecommendation.kind.in_(["forecast_stockout", "seasonal_demand"]),
        BusinessBrainRecommendation.status.in_(RECOMMENDATION_ACTIVE_STATUSES),
    ).all()
    if not rows:
        return None
    per_product: Dict[int, float] = {}
    for row in rows:
        if row.product_id is None:
            continue
        evidence = json.loads(row.evidence_json or "{}")
        expected = evidence.get("calibrated_units") if row.kind == "seasonal_demand" else evidence.get("expected_units")
        current_stock = evidence.get("current_stock")
        if expected is None or current_stock is None:
            continue
        shortfall = max(0.0, float(expected) - float(current_stock))
        if shortfall <= 0:
            continue
        product = db.query(Product).filter(Product.id == row.product_id, Product.business_id == business_id).first()
        if not product or not product.retail_price:
            continue
        risk_value = shortfall * product.retail_price
        per_product[row.product_id] = max(per_product.get(row.product_id, 0.0), risk_value)
    if not per_product:
        return None
    return round(sum(per_product.values()), 2)

# --- Business Brain invalidation (performance) ------------------------------
# GET /business-brain used to call refresh_business_brain() unconditionally
# on every single read — a per-product loop with several queries each,
# meaning opening the dashboard or the Business Brain modal repeatedly (or
# from multiple components at once) repeatedly re-ran the entire analytical
# pass. Replaced with a dirty-flag: mutations that could change Business
# Brain's output mark the business dirty; GET only recomputes when dirty,
# otherwise it's a cheap read of already-stored recommendations/predictions/
# memory. See BUSINESS_BRAIN_DIRTY_TRIGGERS call sites (sale checkout,
# Business Day close/reopen, product mutations) for what marks it dirty.
def mark_business_brain_dirty(db: Session, business_id: int) -> None:
    """Never commits on its own — the caller's own mutation commit carries
    this flag, so the flag can never persist without the mutation that
    justified it (or vice versa)."""
    db.query(BusinessProfile).filter(BusinessProfile.id == business_id).update(
        {BusinessProfile.business_brain_dirty: True}, synchronize_session=False,
    )

def _try_claim_business_brain_refresh(db: Session, business_id: int) -> bool:
    """Atomic compare-and-swap, same pattern as the Business Day refresh-
    token rotation: the WHERE clause is re-checked by the database at the
    instant this runs, so if several concurrent GETs see the business as
    dirty at once, only ONE of them ever flips it to false and performs the
    actual (expensive) refresh — the rest just read already-stored state.
    Commits immediately so the claim is visible/durable before the caller
    starts the potentially-slow refresh work."""
    affected = db.query(BusinessProfile).filter(
        BusinessProfile.id == business_id, BusinessProfile.business_brain_dirty == True,
    ).update({BusinessProfile.business_brain_dirty: False}, synchronize_session=False)
    db.commit()
    return affected == 1

def _business_brain_meta(db: Session, business_id: int) -> Dict[str, Any]:
    """Cheap, side-effect-free summary — safe to call on every GET regardless
    of dirty state (bounded by completed-prediction count, not product
    count). Never touches per-product state; that only happens inside
    refresh_business_brain() when a dirty refresh actually runs."""
    # Distinct calendar dates with a closed session — see predictive_forecast().
    history_days = db.query(func.count(func.distinct(BusinessDay.date))).filter(BusinessDay.business_id == business_id, BusinessDay.is_open == False).scalar() or 0
    completed = db.query(BusinessBrainPrediction).filter(BusinessBrainPrediction.business_id == business_id, BusinessBrainPrediction.accuracy_score.isnot(None)).all()
    prior_accuracy = sum(row.accuracy_score for row in completed) / len(completed) if completed else None
    confidence = _brain_confidence(history_days, prior_accuracy)
    return {"history_days": history_days, "confidence": confidence, "prior_accuracy": prior_accuracy, "evaluated_predictions": len(completed), "accuracy_trend": _brain_accuracy_trend(completed)}

def refresh_business_brain(db: Session, business_id: int) -> Dict[str, Any]:
    """Learn from one business only; never infer seasons or product facts
    generically. Only ever invoked from a claimed dirty refresh (see
    _try_claim_business_brain_refresh) — never directly from GET anymore."""
    now = datetime.utcnow(); _evaluate_due_brain_predictions(db, business_id, now)
    meta = _business_brain_meta(db, business_id)
    history_days, confidence, prior_accuracy = meta["history_days"], meta["confidence"], meta["prior_accuracy"]
    products = db.query(Product).filter(Product.business_id == business_id).all()
    product_ids = {p.id for p in products}
    affirmed: set = set()  # recommendation fingerprints this pass still supports
    # Grouped once instead of one SUM query per product (see the identical
    # fix in /products/financial-intelligence) — query count no longer
    # scales with catalog size.
    total_sold_map = dict(
        db.query(SaleModel.product_id, func.sum(SaleModel.quantity))
        .filter(SaleModel.business_id == business_id).group_by(SaleModel.product_id).all()
    )
    for product in products:
        if product.quantity <= product.min_stock_level:
            _upsert_brain_recommendation(db, business_id, product.id, f"low-stock:{product.id}", "stock_review", "critical", f"Review stock for {product.name}", f"Current stock is {product.quantity} units, at or below this product's minimum level of {product.min_stock_level}.", {"current_stock": product.quantity, "minimum_stock": product.min_stock_level, "source": "current_inventory"})
            affirmed.add(f"low-stock:{product.id}")
        if history_days < BUSINESS_BRAIN_HISTORY_DAYS: continue
        total_sold = total_sold_map.get(product.id, 0)
        daily_velocity = float(total_sold) / history_days
        if daily_velocity <= 0:
            # No longer selling: an "averages X units/day" Memory is stale, so
            # withdraw it rather than keep presenting an out-of-date figure.
            _delete_brain_memory(db, business_id, f"velocity-28:{product.id}")
            continue
        raw_expected_units = round(daily_velocity * BUSINESS_BRAIN_FORECAST_HORIZON_DAYS, 2)
        calibration_factor = _prediction_bias_factor(db, business_id, product.id, "velocity")
        expected_units = round(raw_expected_units * (calibration_factor or 1.0), 2)
        today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        evidence = {"completed_business_days": history_days, "units_sold": total_sold, "average_daily_units": round(daily_velocity, 2), "horizon_days": BUSINESS_BRAIN_FORECAST_HORIZON_DAYS, "prior_prediction_accuracy": round(prior_accuracy, 3) if prior_accuracy is not None else None, "raw_expected_units": raw_expected_units, "calibration_factor": calibration_factor}
        prediction = db.query(BusinessBrainPrediction).filter(BusinessBrainPrediction.business_id == business_id, BusinessBrainPrediction.product_id == product.id, BusinessBrainPrediction.kind == "velocity", BusinessBrainPrediction.forecast_at >= today, BusinessBrainPrediction.actual_units.is_(None)).first()
        if prediction is None:
            db.add(BusinessBrainPrediction(business_id=business_id, product_id=product.id, kind="velocity", forecast_at=now, target_at=now + timedelta(days=BUSINESS_BRAIN_FORECAST_HORIZON_DAYS), predicted_units=expected_units, confidence=confidence, evidence_json=json.dumps(evidence)))
        else:
            prediction.predicted_units, prediction.confidence, prediction.evidence_json = expected_units, confidence, json.dumps(evidence)
        if product.quantity < expected_units:
            days_to_stockout = round(product.quantity / daily_velocity, 1)
            _upsert_brain_recommendation(db, business_id, product.id, f"forecast-stockout:{product.id}", "forecast_stockout", "critical" if days_to_stockout <= 3 else "important", f"Prepare for demand on {product.name}", f"Based on {history_days} completed business days, expected demand for the next 7 days is about {expected_units:g} units while current stock is {product.quantity}.", {**evidence, "current_stock": product.quantity, "days_to_stockout": days_to_stockout, "expected_units": expected_units})
            affirmed.add(f"forecast-stockout:{product.id}")
        if history_days >= 28:
            _upsert_brain_memory(db, business_id, product.id, f"velocity-28:{product.id}", f"{product.name} has averaged about {daily_velocity:.1f} units per completed business day across {history_days} recorded days.", evidence, confidence, now)
        affirmed |= _apply_seasonal_pattern(db, business_id, product, now)
    _apply_business_relationships(db, business_id, now)
    _retire_stale_brain_recommendations(db, business_id, now, affirmed, history_days, product_ids)
    db.commit()
    # Re-read: _evaluate_due_brain_predictions()/the loop above may have
    # changed accuracy_score/prediction rows since meta was first computed.
    return _business_brain_meta(db, business_id)

@app.get("/business-brain")
def business_brain(user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if business and (business.business_brain_dirty or business.business_brain_refreshed_at is None):
        if _try_claim_business_brain_refresh(db, user.business_id):
            try:
                refresh_business_brain(db, user.business_id)
                db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).update(
                    {BusinessProfile.business_brain_refreshed_at: datetime.utcnow()}, synchronize_session=False,
                )
                db.commit()
            except Exception:
                # Refresh failed partway — leave the business dirty so the
                # next read retries, rather than silently serving a state
                # that was never actually completed.
                db.rollback()
                mark_business_brain_dirty(db, user.business_id)
                db.commit()
                raise
        # else: another concurrent request already claimed and is
        # performing (or just performed) the refresh — read whatever is
        # currently stored rather than double-computing.
    meta = _business_brain_meta(db, user.business_id)
    now = datetime.utcnow()
    # Quiet business: no dirty trigger for a while, but a forecast window has
    # closed. Evaluate only genuinely-due rows so their Outcome becomes known
    # and they leave "Coming Up". Skipped when we just refreshed above.
    if business and not business.business_brain_dirty:
        due_exists = db.query(BusinessBrainPrediction.id).filter(
            BusinessBrainPrediction.business_id == user.business_id,
            BusinessBrainPrediction.actual_units.is_(None),
            BusinessBrainPrediction.target_at <= now,
        ).first()
        if due_exists is not None:
            _evaluate_due_brain_predictions(db, user.business_id, now)
            db.commit()
            meta = _business_brain_meta(db, user.business_id)
    # ACTIVE brief only — current, unresolved items. Acted / dismissed / auto-
    # resolved items are never here; they live in History, which is paginated
    # via GET /business-brain/history and never shipped whole to the dashboard.
    rows = db.query(BusinessBrainRecommendation).filter(BusinessBrainRecommendation.business_id == user.business_id, BusinessBrainRecommendation.status.in_(RECOMMENDATION_ACTIVE_STATUSES)).order_by(BusinessBrainRecommendation.updated_at.desc()).limit(30).all()
    rows.sort(key=lambda row: {"critical": 0, "important": 1, "opportunity": 2}.get(row.priority, 3))
    if user.role == "staff": rows = [r for r in rows if r.kind == "stock_review"]
    def rec_out(r): return {"id": r.id, "kind": r.kind, "priority": r.priority, "title": r.title, "summary": r.summary, "evidence": json.loads(r.evidence_json or "{}"), "status": r.status, "updated_at": to_utc_iso(r.updated_at)}
    output = {"learning": meta["history_days"] < BUSINESS_BRAIN_HISTORY_DAYS, "history_days": meta["history_days"], "currency": business.currency if business else "USD ($)", "attention": [rec_out(r) for r in rows], "recommendations": [rec_out(r) for r in rows], "history_available": True, "learning_message": "Cauldra is still learning this part of your business. Continue recording sales and closing business days to build reliable predictions." if meta["history_days"] < BUSINESS_BRAIN_HISTORY_DAYS else None}
    if user.role != "staff":
        memories = db.query(BusinessBrainMemory).filter(BusinessBrainMemory.business_id == user.business_id).order_by(BusinessBrainMemory.last_observed_at.desc()).limit(20).all()
        # COMING UP = what is likely to matter next: only forecasts whose window
        # is still open, and only the newest one per product per kind so a
        # standing condition never stacks near-identical rows.
        open_predictions = db.query(BusinessBrainPrediction, Product.name).join(Product, Product.id == BusinessBrainPrediction.product_id).filter(BusinessBrainPrediction.business_id == user.business_id, Product.business_id == user.business_id, BusinessBrainPrediction.actual_units.is_(None), BusinessBrainPrediction.target_at >= now).order_by(BusinessBrainPrediction.forecast_at.desc()).all()
        seen_pred: set = set(); coming = []
        for p, name in open_predictions:
            key = (p.product_id, p.kind)
            if key in seen_pred: continue
            seen_pred.add(key)
            coming.append({"product_name": name, "predicted_units": p.predicted_units, "target_at": to_utc_iso(p.target_at), "confidence": _brain_confidence_label(p.confidence), "evidence": json.loads(p.evidence_json or "{}")})
            if len(coming) >= 8: break
        actions_taken = db.query(BusinessBrainRecommendation).filter(BusinessBrainRecommendation.business_id == user.business_id, BusinessBrainRecommendation.status == "acted").count()
        revenue_at_risk = _brain_revenue_at_risk(db, user.business_id)
        output.update({"memory": [{"statement": m.statement, "evidence": json.loads(m.evidence_json or "{}"), "confidence": _brain_confidence_label(m.confidence), "last_observed_at": to_utc_iso(m.last_observed_at), "first_seen": to_utc_iso(m.created_at), "reinforced": bool(m.last_observed_at and m.created_at and (m.last_observed_at - m.created_at) > timedelta(days=1))} for m in memories], "coming": coming, "outcomes": {"evaluated_predictions": meta["evaluated_predictions"], "accuracy": round(meta["prior_accuracy"] * 100, 1) if meta["prior_accuracy"] is not None else None, "accuracy_trend": meta["accuracy_trend"], "actions_taken": actions_taken, "revenue_at_risk": revenue_at_risk, "message": "Prediction outcomes will appear after forecast periods complete." if not meta["evaluated_predictions"] else None}})
    return output

@app.post("/business-brain/recommendations/{recommendation_id}/action")
def action_business_brain_recommendation(recommendation_id: int, payload: BusinessBrainRecommendationAction, user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Only Admins and Managers can update recommendation actions.")
    row = db.query(BusinessBrainRecommendation).filter(BusinessBrainRecommendation.id == recommendation_id, BusinessBrainRecommendation.business_id == user.business_id).first()
    if not row: raise HTTPException(status_code=404, detail="Recommendation not found.")
    action = payload.action.strip().lower()
    if action not in {"opened", "acted", "dismissed"}: raise HTTPException(status_code=400, detail="Unsupported recommendation action.")
    now = datetime.utcnow(); row.status = action
    if action == "opened": row.opened_at = now
    elif action == "acted": row.acted_at = now
    else: row.dismissed_at = now
    add_audit(db, user, "BUSINESS_BRAIN_RECOMMENDATION_" + action.upper(), f"Updated Business Brain recommendation #{row.id} to {action}."); db.commit()
    return {"id": row.id, "status": row.status}

# --- Business Brain History (long-term archive, paginated at the database) ---
BRAIN_HISTORY_RANGE_DAYS = {"week": 7, "month": 31, "quarter": 92, "3months": 92, "year": 366}

def _brain_history_window(range_key: str, date_from: Optional[str], date_to: Optional[str], now: datetime):
    """Resolve a range keyword (or an explicit custom YYYY-MM-DD span) into a
    (start, end) datetime pair for the History filter. 'all'/unknown => no bound."""
    if range_key == "custom":
        try:
            start = datetime.fromisoformat(date_from) if date_from else None
            end = (datetime.fromisoformat(date_to) + timedelta(days=1)) if date_to else None
        except ValueError:
            raise HTTPException(status_code=400, detail="Custom History dates must be YYYY-MM-DD.")
        return start, end
    days = BRAIN_HISTORY_RANGE_DAYS.get(range_key)
    return (now - timedelta(days=days), None) if days else (None, None)

def _brain_prediction_outcome(accuracy_score: Optional[float]) -> Optional[str]:
    """A plain outcome label derived only from the already-measured accuracy of
    a completed forecast — never a fabricated confidence."""
    if accuracy_score is None:
        return None
    if accuracy_score >= 0.7:
        return "Confirmed"
    if accuracy_score >= 0.4:
        return "Partly confirmed"
    return "Missed"

@app.get("/business-brain/history")
def business_brain_history(
    type: str = Query("all"),
    range: str = Query("all"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=50),
    user: User = Depends(require_ai_access), db: Session = Depends(get_db),
):
    """Paginated long-term Business Brain archive. Filtering and pagination
    happen in the database — the browser never downloads a whole multi-year
    history to filter it locally. Every row is a record this engine already
    stores (resolved/acted/dismissed recommendations, evaluated forecasts);
    no separate history ledger is written."""
    now = datetime.utcnow()
    staff = user.role == "staff"
    want = (type or "all").strip().lower()
    range_key = (range or "all").strip().lower()
    start, end = _brain_history_window(range_key, date_from, date_to, now)
    want_recs = want in ("all", "attention", "recommendations")
    want_preds = (want in ("all", "predictions", "outcomes")) and not staff
    # Over-fetch bound per source: enough to serve this page after the merge,
    # never the whole table. Each source is independently ordered newest-first,
    # so the merged slice [offset:offset+limit] is exact once each source has at
    # least offset+limit+1 rows (or is fully exhausted).
    take = offset + limit + 1
    events: List[Dict[str, Any]] = []

    if want_recs:
        rq = db.query(BusinessBrainRecommendation).filter(
            BusinessBrainRecommendation.business_id == user.business_id,
            BusinessBrainRecommendation.status.in_(RECOMMENDATION_HISTORY_STATUSES),
        )
        if staff or want == "attention":
            rq = rq.filter(BusinessBrainRecommendation.kind == "stock_review")
        elif want == "recommendations":
            rq = rq.filter(BusinessBrainRecommendation.kind != "stock_review")
        if start is not None:
            rq = rq.filter(BusinessBrainRecommendation.updated_at >= start)
        if end is not None:
            rq = rq.filter(BusinessBrainRecommendation.updated_at < end)
        for row in rq.order_by(BusinessBrainRecommendation.updated_at.desc()).limit(take).all():
            occurred_at = row.acted_at if row.status == "acted" else (row.dismissed_at if row.status == "dismissed" else None)
            occurred_at = occurred_at or row.updated_at
            category = ("Completed action" if row.status == "acted"
                        else "Dismissed" if row.status == "dismissed"
                        else "Resolved warning" if row.kind == "stock_review"
                        else "Resolved automatically")
            events.append({"type": "recommendation", "category": category, "title": row.title,
                           "summary": row.summary, "status": row.status, "priority": row.priority,
                           "occurred_at": to_utc_iso(occurred_at), "_sort": occurred_at})

    if want_preds:
        pq = (db.query(BusinessBrainPrediction, Product.name)
              .join(Product, Product.id == BusinessBrainPrediction.product_id)
              .filter(BusinessBrainPrediction.business_id == user.business_id,
                      Product.business_id == user.business_id,
                      BusinessBrainPrediction.actual_units.isnot(None)))
        if start is not None:
            pq = pq.filter(func.coalesce(BusinessBrainPrediction.evaluated_at, BusinessBrainPrediction.target_at) >= start)
        if end is not None:
            pq = pq.filter(func.coalesce(BusinessBrainPrediction.evaluated_at, BusinessBrainPrediction.target_at) < end)
        for prediction, product_name in pq.order_by(BusinessBrainPrediction.evaluated_at.desc()).limit(take).all():
            occurred_at = prediction.evaluated_at or prediction.target_at
            outcome = _brain_prediction_outcome(prediction.accuracy_score)
            accuracy = round(prediction.accuracy_score * 100, 1) if prediction.accuracy_score is not None else None
            acc_text = f" Accuracy {accuracy:g}%." if accuracy is not None else ""
            events.append({"type": "forecast",
                           "category": "Seasonal forecast" if prediction.kind == "seasonal" else "Demand forecast",
                           "title": f"{product_name} forecast evaluated",
                           "summary": f"Forecast {prediction.predicted_units:g} units; actual {prediction.actual_units:g} units.{acc_text}",
                           "status": "evaluated", "outcome": outcome,
                           "occurred_at": to_utc_iso(occurred_at), "_sort": occurred_at})

    events.sort(key=lambda e: e["_sort"] or datetime.min, reverse=True)
    window = events[offset:offset + limit + 1]
    has_more = len(window) > limit
    window = window[:limit]
    for e in window:
        e.pop("_sort", None)
    return {"events": window, "offset": offset, "limit": limit, "has_more": has_more,
            "type": want, "range": range_key}

# -----------------------------------------------------------------------------
# EXPENSES
#
# Foundation-only per explicit scope: no profit/loss, no revenue calculation,
# no payment-source model (none exists in this codebase — payment_source is a
# free-text field until a real account/payment-source system is built), no
# offline sync (none exists in this codebase either — client_ref exists in the
# schema for a future sync client but nothing populates it today). Every
# financial fact the frontend could lie about — business, creator, timestamp —
# is set here from the authenticated session, never from the request body.
# -----------------------------------------------------------------------------
@app.get("/expenses/categories")
def list_expense_categories(user: User = Depends(get_current_user)):
    return {"categories": EXPENSE_CATEGORIES}

@app.post("/expenses/")
def create_expense(data: ExpenseCreate, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    enforce_offline_replay_identity(request, user)
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    category = (data.category or "").strip()
    if not category: raise HTTPException(status_code=400, detail="Please choose or enter an expense category.")
    if len(category) > 100: raise HTTPException(status_code=400, detail="Expense category is too long.")
    if data.amount is None or data.amount <= 0: raise HTTPException(status_code=400, detail="Expense amount must be greater than zero.")
    payment_source = (data.payment_source or "").strip()[:100] or None
    note = (data.note or "").strip()[:500] or None
    client_ref = (data.client_ref or "").strip()[:100] or None

    claim, replay = claim_idempotent_mutation(
        db, user.business_id, "expense_create", client_ref,
        {"category": category, "amount": float(data.amount), "payment_source": payment_source, "note": note},
    )
    if replay:
        return replay

    # Recording an expense is an operational financial action, exactly like
    # completing a sale — so it OWNS a Business Day the same way: it attaches
    # to the active session, or auto-opens one if none is active. Previously
    # this only attached to an already-open day and otherwise silently wrote
    # business_day_id = NULL, leaving the expense outside every business-day
    # report it belongs in. Read-only endpoints still never auto-open a day.
    day = ensure_open_business_day(db, user.business_id, opener=user, commit=False)
    business_day_id = day.id

    expense = Expense(
        business_id=user.business_id, category=category, amount=float(data.amount),
        payment_source=payment_source, note=note, owner_id=user.id,
        created_at=datetime.utcnow(), client_ref=client_ref,
        synced_at=(datetime.utcnow() if client_ref else None),
        business_day_id=business_day_id,
    )
    db.add(expense); db.flush()
    add_audit(db, user, "EXPENSE_RECORDED", f"Recorded a {category} expense of {expense.amount:,.2f}.", business_day_id=business_day_id)
    response = {"id": expense.id, "message": "Expense recorded successfully."}
    complete_idempotent_mutation(claim, response)
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    db.refresh(expense)
    return response

def _build_expenses_query(db: Session, user: User, business: "BusinessProfile", category, payment_source, user_id, search, date_from, date_to):
    """Shared filter-building for /expenses/ and /expenses/export — the two
    endpoints must never be able to disagree about which rows a given set of
    filters matches, so both call this exact same function rather than each
    maintaining their own copy of the filter logic."""
    tz = business_local_zoneinfo(business)
    q = db.query(Expense).filter(Expense.business_id == user.business_id)  # tenant scope — never trusts the request
    if category: q = q.filter(Expense.category == category)
    if payment_source: q = q.filter(Expense.payment_source == payment_source)
    if user_id: q = q.filter(Expense.owner_id == user_id)
    if search:
        like = f"%{search.strip()}%"
        q = q.filter(or_(Expense.category.ilike(like), Expense.note.ilike(like)))
    # date_from/date_to are business-local calendar dates (as picked in the
    # UI's date-range filter), inclusive on both ends. Expense.created_at is
    # stored as naive UTC, so the boundaries must be converted from
    # business-local midnight into UTC before filtering — comparing the
    # local calendar string directly against a UTC timestamp would shift
    # results by the business's UTC offset (the same bug the Business Day
    # "Opened" time had). This mirrors resolve_financial_period()'s "custom"
    # period branch, which every other date-range filter in the app uses.
    if date_from:
        try:
            start_local = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=tz)
        except ValueError:
            raise HTTPException(status_code=400, detail="date_from must be in YYYY-MM-DD format.")
        q = q.filter(Expense.created_at >= start_local.astimezone(timezone.utc).replace(tzinfo=None))
    if date_to:
        try:
            end_local = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=tz) + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="date_to must be in YYYY-MM-DD format.")
        q = q.filter(Expense.created_at < end_local.astimezone(timezone.utc).replace(tzinfo=None))
    return q

def _owner_name_lookup(db: Session, rows):
    owner_ids = {r.owner_id for r in rows if r.owner_id}
    owners = {u.id: u for u in db.query(User).filter(User.id.in_(owner_ids)).all()} if owner_ids else {}

    def owner_name(uid):
        o = owners.get(uid)
        if not o: return "Unknown"
        return (f"{o.firstname} {o.lastname}".strip() if (o.firstname or o.lastname) else o.username)
    return owner_name

@app.get("/expenses/")
def list_expenses(
    category: Optional[str] = Query(None), payment_source: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None), search: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500), offset: int = Query(0, ge=0),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    q = _build_expenses_query(db, user, business, category, payment_source, user_id, search, date_from, date_to)
    total = q.count()
    rows = q.order_by(Expense.created_at.desc()).offset(offset).limit(limit).all()
    owner_name = _owner_name_lookup(db, rows)

    return {
        "total": total, "limit": limit, "offset": offset,
        "expenses": [{
            "id": r.id, "category": r.category, "amount": r.amount, "payment_source": r.payment_source,
            "note": r.note, "recorded_by": owner_name(r.owner_id), "owner_id": r.owner_id,
            "created_at": to_utc_iso(r.created_at),
        } for r in rows],
    }

EXPENSES_EXPORT_COLUMNS = [
    {"key": "id", "label": "EXPENSE ID", "type": "number"},
    {"key": "date", "label": "DATE", "type": "datetime"},
    {"key": "category", "label": "CATEGORY", "type": "text"},
    {"key": "description", "label": "DESCRIPTION", "type": "text"},
    {"key": "amount", "label": "AMOUNT", "type": "currency"},
    {"key": "payment_method", "label": "PAYMENT METHOD", "type": "text"},
    {"key": "recorded_by", "label": "RECORDED BY", "type": "text"},
]

def _expenses_export_rows(db, user, business, category, payment_source, user_id, search, date_from, date_to):
    """Shared by the CSV and Excel Expense History exports — one query, one
    row-building pass, so the two formats can never show different data.
    Capped at 2000 rows (well above the 500-row page-fetch cap used
    everywhere else in the app) so "export everything matching my filters"
    actually works in practice rather than being limited to one page, while
    still bounding the query."""
    q = _build_expenses_query(db, user, business, category, payment_source, user_id, search, date_from, date_to)
    rows = q.order_by(Expense.created_at.desc()).limit(2000).all()
    owner_name = _owner_name_lookup(db, rows)
    return [[r.id, to_utc_iso(r.created_at), r.category, r.note or "", r.amount, r.payment_source or "", owner_name(r.owner_id)] for r in rows]

@app.get("/expenses/export")
def export_expenses_csv(
    category: Optional[str] = Query(None), payment_source: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None), search: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    """CSV export of Expense History. Same role check and business scoping
    as GET /expenses/ — the business_id always comes from the authenticated
    session, never from the request."""
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    out_rows = _expenses_export_rows(db, user, business, category, payment_source, user_id, search, date_from, date_to)
    if not out_rows:
        return Response(status_code=204)
    header = [c["label"] for c in EXPENSES_EXPORT_COLUMNS]
    return build_csv_response(f"cauldra_expenses_{business_local_today(db, user.business_id)}.csv", header, out_rows)

@app.get("/expenses/export/xlsx")
def export_expenses_xlsx(
    category: Optional[str] = Query(None), payment_source: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None), search: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    """Same authorization, tenant scoping, filters, and row data as the CSV
    export above (both call _expenses_export_rows) — this only changes how
    the rows are presented (styled, typed, filterable workbook)."""
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    out_rows = _expenses_export_rows(db, user, business, category, payment_source, user_id, search, date_from, date_to)
    if not out_rows:
        return Response(status_code=204)
    filters_desc = ", ".join(f"{k}={v}" for k, v in [("Category", category), ("Payment Source", payment_source), ("Search", search), ("From", date_from), ("To", date_to)] if v) or "None"
    metadata = [["Generated", datetime.utcnow().strftime("%d %B %Y %H:%M UTC")], ["Filters", filters_desc]]
    sheets = [{"name": "Expenses", "title": "CAULDRA EXPENSE HISTORY", "metadata": metadata, "columns": EXPENSES_EXPORT_COLUMNS, "rows": out_rows}]
    return build_xlsx_response(f"cauldra_expenses_{business_local_today(db, user.business_id)}.xlsx", "Cauldra Expense History", sheets)

# -----------------------------------------------------------------------------
# GENERIC EXCEL EXPORT — for every feature whose data is already fully
# available client-side (Inventory, Suppliers, Purchase Orders, Warehouses,
# Employees, Team Presence, Price Monitor, Predictive, Profit). The frontend
# builds the exact same {columns, rows} spec it already uses for its CSV
# export of that same screen (see index.html's buildXExportSpec()-style
# functions) and posts it here for styled-workbook rendering.
#
# This performs NO new data access — it only formats rows the authenticated
# caller already legitimately holds (fetched moments earlier through that
# feature's own tenant-scoped, role-checked endpoint). Excel export can
# therefore never expose more than the corresponding CSV/JSON view already
# did (section 17): there is no query here for this endpoint to leak through.
# Authentication is still required so this can't be used as an open,
# unauthenticated rendering/DoS oracle, and the size guards below bound the
# work this endpoint will ever do for one request.
# -----------------------------------------------------------------------------
@app.post("/export/xlsx")
def export_generic_xlsx(payload: XlsxExportRequest, user: User = Depends(get_current_user)):
    if not payload.sheets:
        raise HTTPException(status_code=400, detail="At least one sheet is required.")
    if len(payload.sheets) > XLSX_MAX_SHEETS:
        raise HTTPException(status_code=400, detail=f"Too many sheets (max {XLSX_MAX_SHEETS}).")
    total_rows = 0
    sheets_dicts = []
    for sheet in payload.sheets:
        if len(sheet.columns) > XLSX_MAX_COLUMNS:
            raise HTTPException(status_code=400, detail=f"Too many columns in sheet '{sheet.name}' (max {XLSX_MAX_COLUMNS}).")
        if len(sheet.rows) > XLSX_MAX_ROWS_PER_SHEET:
            raise HTTPException(status_code=400, detail=f"Too many rows in sheet '{sheet.name}' (max {XLSX_MAX_ROWS_PER_SHEET}).")
        total_rows += len(sheet.rows)
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, str) and len(cell) > XLSX_MAX_CELL_TEXT:
                    raise HTTPException(status_code=400, detail="A cell value exceeds the maximum allowed length.")
        sheets_dicts.append({
            "name": sheet.name,
            "title": sheet.title,
            "metadata": sheet.metadata,
            "columns": [c.model_dump() for c in sheet.columns],
            "rows": sheet.rows,
        })
    if total_rows == 0:
        return Response(status_code=204)
    return build_xlsx_response(payload.filename, payload.report_title, sheets_dicts)

# -----------------------------------------------------------------------------
# SUPPLIERS
# -----------------------------------------------------------------------------
@app.get("/suppliers/")
def list_suppliers(limit: int = Query(200, ge=1, le=500), offset: int = Query(0, ge=0), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.query(Supplier).filter(Supplier.business_id == user.business_id).order_by(Supplier.id.desc()).offset(offset).limit(limit).all()
    return [{"id": s.id, "name": s.name, "contact_email": s.contact_email, "phone": s.phone, "lead_time_days": s.lead_time_days} for s in rows]

@app.post("/suppliers/")
def create_supplier(data: SupplierCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot add suppliers.")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    check_plan_limit(db, business, "supplier", db.query(Supplier).filter(Supplier.business_id == user.business_id).count())
    s = Supplier(name=data.name, contact_email=data.contact_email, phone=data.phone, lead_time_days=data.lead_time_days or 3, business_id=user.business_id)
    db.add(s); db.flush(); add_audit(db, user, "SUPPLIER_CREATED", f"Added supplier {s.name}."); db.commit(); db.refresh(s)
    return {"id": s.id, "message": "Supplier added successfully."}

@app.delete("/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot delete suppliers.")
    s = db.query(Supplier).filter(Supplier.id == supplier_id, Supplier.business_id == user.business_id).first()
    if not s: raise HTTPException(status_code=404, detail="The supplier could not be found.")
    add_audit(db, user, "SUPPLIER_DELETED", f"Deleted supplier {s.name}.")
    db.delete(s); db.commit()
    return {"message": "Supplier deleted successfully."}

# -----------------------------------------------------------------------------
# PURCHASE ORDERS
# -----------------------------------------------------------------------------
@app.get("/purchase-orders/")
def get_purchase_orders(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    rows = db.query(PurchaseOrder).filter(PurchaseOrder.business_id == user.business_id).order_by(PurchaseOrder.id.desc()).all()
    owner_name = _owner_name_lookup(db, rows)
    out=[]
    for po in rows:
        supplier = db.query(Supplier).filter(Supplier.id == po.supplier_id).first() if po.supplier_id else None
        # created_at/created_by were always on the model but never serialized
        # here — additive fields, existing consumers of this response are
        # unaffected (the frontend already references po.created_at, which
        # was silently undefined until now).
        out.append({"id": po.id, "supplier_id": po.supplier_id, "vendor_name": supplier.name if supplier else "General Vendor", "status": po.status, "total_estimated_cost": po.total_estimated_cost, "items_summary": po.email_draft or "", "email_draft": po.email_draft or "", "created_at": to_utc_iso(po.created_at), "created_by": owner_name(po.owner_id)})
    return out

@app.post("/purchase-orders/generate")
def generate_po(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    period_start, period_end, _ = billing_period_for(db, business)
    current = db.query(PurchaseOrder).filter(PurchaseOrder.business_id == user.business_id, PurchaseOrder.created_at >= period_start, PurchaseOrder.created_at < period_end).count()
    check_plan_limit(db, business, "purchase_order", current)
    low_stock = db.query(Product).filter(Product.business_id == user.business_id, Product.quantity <= Product.min_stock_level).all()
    if not low_stock: raise HTTPException(status_code=400, detail="No low stock items requiring restock.")
    total_cost = sum(p.cost_price * max(p.min_stock_level * 2, 1) for p in low_stock)
    items = ", ".join([f"{p.name} ({max(p.min_stock_level * 2, 1)} units)" for p in low_stock])
    draft = f"Please confirm availability for: {items}."
    po = PurchaseOrder(status="DRAFT", total_estimated_cost=total_cost, email_draft=draft, business_id=user.business_id, owner_id=None)
    db.add(po); db.flush(); add_audit(db, user, "PURCHASE_ORDER_CREATED", f"Generated purchase order #{po.id}.")
    create_notification(
        db, business_id=user.business_id, category="purchase_order", severity="info", type="PO_GENERATED",
        title="Purchase order generated", message=f"Purchase order #{po.id} was generated for {len(low_stock)} low-stock item(s).",
        related_entity_type="purchase_order", related_entity_id=po.id, deep_link=f"purchase_order:{po.id}",
    )
    db.commit(); db.refresh(po)
    return {"message": "Purchase Order generated successfully.", "id": po.id, "po_id": po.id}

@app.put("/purchase-orders/{po_id}")
def update_po_draft(po_id: int, update: PODraftUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id, PurchaseOrder.business_id == user.business_id).first()
    if not po: raise HTTPException(status_code=404, detail="Purchase order not found.")
    details = update.details if update.details is not None else update.items_summary
    if details is not None: po.email_draft = details
    if update.supplier_id is not None: po.supplier_id = update.supplier_id
    db.commit(); add_audit(db, user, "PURCHASE_ORDER_UPDATED", f"Updated purchase order #{po.id}."); db.commit()
    return {"message": "PO draft updated."}

@app.delete("/purchase-orders/{po_id}")
def delete_po(po_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id, PurchaseOrder.business_id == user.business_id).first()
    if not po: raise HTTPException(status_code=404, detail="Purchase order not found.")
    add_audit(db, user, "PURCHASE_ORDER_DELETED", f"Deleted purchase order #{po.id}.")
    db.delete(po); db.commit(); return {"message": "Purchase order deleted."}

@app.delete("/purchase-orders")
def delete_all_pos(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can delete all purchase orders.")
    db.query(PurchaseOrder).filter(PurchaseOrder.business_id == user.business_id).delete(synchronize_session=False)
    add_audit(db, user, "PURCHASE_ORDERS_CLEARED", "Cleared all purchase orders."); db.commit(); return {"message": "All purchase orders deleted."}

@app.post("/purchase-orders/{po_id}/dispatch")
def dispatch_po(po_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id, PurchaseOrder.business_id == user.business_id).first()
    if not po or not po.supplier_id: raise HTTPException(status_code=400, detail="Purchase order or assigned supplier is unavailable.")
    supplier = db.query(Supplier).filter(Supplier.id == po.supplier_id, Supplier.business_id == user.business_id).first()
    if not supplier: raise HTTPException(status_code=404, detail="Supplier is unavailable.")
    phone = normalize_phone(supplier.phone)
    encoded = urllib.parse.quote(po.email_draft or "")
    po.status = "SENT"; add_audit(db, user, "PURCHASE_ORDER_DISPATCHED", f"Dispatched purchase order #{po.id} via WhatsApp.")
    create_notification(
        db, business_id=user.business_id, category="purchase_order", severity="important", type="PO_SUBMITTED",
        title="Purchase order submitted", message=f"Purchase order #{po.id} was sent to {supplier.name}.",
        related_entity_type="purchase_order", related_entity_id=po.id, deep_link=f"purchase_order:{po.id}",
    )
    db.commit()
    return {"message": f"Purchase Order dispatched to {supplier.name}.", "whatsapp_url": f"https://wa.me/{phone}?text={encoded}"}

@app.post("/purchase-orders/{po_id}/dispatch-email")
def dispatch_po_email(po_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id, PurchaseOrder.business_id == user.business_id).first()
    if not po or not po.supplier_id: raise HTTPException(status_code=400, detail="Purchase order or assigned supplier is unavailable.")
    supplier = db.query(Supplier).filter(Supplier.id == po.supplier_id, Supplier.business_id == user.business_id).first()
    if not supplier: raise HTTPException(status_code=404, detail="Supplier is unavailable.")
    if not supplier.contact_email: raise HTTPException(status_code=400, detail="This supplier has no contact email on file.")
    api_key = os.getenv("RESEND_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="Email dispatch is not configured. Add RESEND_API_KEY to the server environment.")
    import requests
    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "from": RESEND_FROM,
                "to": [supplier.contact_email],
                "subject": f"Purchase Order #{po.id}",
                "html": f"<p>Hello {supplier.name},</p><p>{(po.email_draft or '').replace(chr(10), '<br>')}</p>",
            },
            timeout=15,
        )
        if not r.ok:
            raise RuntimeError(r.text)
    except Exception as exc:
        raise HTTPException(status_code=502, detail="We could not email this purchase order right now.") from exc
    po.status = "SENT"; db.commit(); add_audit(db, user, "PURCHASE_ORDER_DISPATCHED", f"Emailed purchase order #{po.id} to {supplier.name}."); db.commit()
    return {"message": f"Purchase Order emailed to {supplier.name}."}

# -----------------------------------------------------------------------------
# SALES / BUSINESS DAYS
# -----------------------------------------------------------------------------
@app.post("/sales/checkout")
def sales_checkout(payload: SalesCheckoutRequest, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    enforce_offline_replay_identity(request, user)
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    client_ref = (payload.client_ref or "").strip()[:100] or None

    def _existing_result(ref: str) -> Optional[dict]:
        """The already-recorded result for this client_ref, if any — so a
        duplicate submission is answered with the original sale instead of
        an error or a second sale."""
        rows = db.query(SaleModel).filter(SaleModel.business_id == user.business_id, SaleModel.client_ref == ref).all()
        if not rows:
            return None
        return {
            "message": "Sale completed successfully.",
            "daily_total": sum(r.total_price for r in rows),
            "business_day_id": rows[0].business_day_id,
            "transaction_id": ref, "updated_products": [], "duplicate": True,
        }

    # Fast path for the common, already-settled duplicate (an offline replay,
    # or a retry long after the original committed). This read alone is NOT
    # the actual guarantee — two genuinely concurrent submissions can both
    # pass it before either commits. The real guard is the SaleTransaction
    # header insert below, which the database itself serializes.
    if client_ref:
        settled = _existing_result(client_ref)
        if settled:
            return settled

    items = payload.items
    if not items:
        raise HTTPException(status_code=400, detail="Cart is empty.")

    # Every checkout submission gets a stable transaction identifier stored
    # on every line item it creates — the client's own client_ref for an
    # offline-synced batch, or one generated here for a normal online
    # checkout that never sent one. It is both the grouping key (see
    # checkout_key_expr / _sales_for_transaction_key) and the idempotency
    # key. It does NOT change synced_at's meaning below, which must stay
    # true only for a genuine offline-sync replay.
    transaction_ref = client_ref or secrets.token_urlsafe(12)

    # Claim idempotency BEFORE stock validation. A concurrent retry with the
    # same client_ref now waits on this unique insert, then returns the winner's
    # committed result even if that winner consumed the last available unit.
    txn = SaleTransaction(
        business_id=user.business_id, business_day_id=None, client_ref=transaction_ref,
        created_by_id=user.id, created_by_name=user.username, created_by_role=user.role,
    )
    db.add(txn)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        settled = _existing_result(transaction_ref)
        if settled:
            return settled
        raise HTTPException(status_code=409, detail="This sale is already being processed.")

    # A real sale is a genuine operational action — recording it legitimately
    # counts as starting a brand-new Business Day session if none is active
    # yet, and is correctly attributed to whichever authenticated user is
    # doing the selling (see ensure_open_business_day). This is the one
    # exception to "no automatic open sessions" from a WRITE path; every
    # read-only endpoint below (current-day, history, analytics) never
    # creates one. ensure_open_business_day always returns an ACTIVE session
    # (either the existing one or a freshly created one) — it can never
    # return a closed day, so there is nothing to reject here; a day closed
    # earlier today never blocks new sales, a new session simply attaches to
    # whatever is active right now (see the BusinessDay model docstring).
    # --- PHASE 1: lock and validate EVERY line before mutating anything -------
    # Atomicity (section 7): a cart where item A is fine but item B is short
    # on stock must change nothing at all — not A's stock, not a partial sale
    # row. The idempotency header above is deliberately flushed first, but it
    # remains uncommitted and is rolled back with the request on any validation
    # or write failure. Inventory and sale-line mutation starts only after the
    # complete cart is locked and known-good.
    #
    # requested_by_product accumulates across lines so the SAME product
    # appearing on two cart lines is validated against its combined quantity,
    # not twice independently against full stock.
    requested_by_product: Dict[int, int] = {}
    for item in items:
        qty = item.quantity
        if qty < 1:
            raise HTTPException(status_code=409, detail="Invalid sale quantity.")
        requested_by_product[item.product_id] = requested_by_product.get(item.product_id, 0) + qty

    # Deterministic lock order prevents deadlocks for overlapping multi-line
    # carts. PostgreSQL makes the second checkout wait, then it validates the
    # winner's committed quantity rather than the stale pre-sale quantity.
    product_ids = sorted(requested_by_product)
    locked_products = (
        db.query(Product)
        .filter(Product.business_id == user.business_id, Product.id.in_(product_ids))
        .order_by(Product.id.asc())
        .with_for_update()
        .all()
    )
    products_by_id = {product.id: product for product in locked_products}
    if len(products_by_id) != len(product_ids):
        raise HTTPException(status_code=409, detail="A selected item is no longer in this inventory.")

    for product_id, requested in requested_by_product.items():
        product = products_by_id[product_id]
        if requested > product.quantity:
            raise HTTPException(status_code=409, detail=f"Not enough stock for {product.name}.")

    locked_stocks = (
        db.query(WarehouseStock)
        .filter(WarehouseStock.business_id == user.business_id, WarehouseStock.product_id.in_(product_ids))
        .order_by(WarehouseStock.product_id.asc(), WarehouseStock.warehouse.asc(), WarehouseStock.id.asc())
        .with_for_update()
        .all()
    )
    stock_by_source = {(stock.product_id, stock.warehouse): stock for stock in locked_stocks}
    source_stock_by_product = {}
    for product_id, requested in requested_by_product.items():
        product = products_by_id[product_id]
        source_name = product.warehouse or "Main Central Warehouse"
        stock = stock_by_source.get((product_id, source_name))
        if stock and requested > stock.quantity:
            raise HTTPException(status_code=409, detail=f"Not enough stock for {product.name} in {source_name}.")
        source_stock_by_product[product_id] = stock

    validated = []
    negotiated_lines = []
    for item in items:
        p = products_by_id[item.product_id]
        qty = item.quantity
        retail_price = float(p.retail_price or 0.0)
        wholesale_price = float(p.wholesale_price or retail_price)
        if retail_price <= 0 or not math.isfinite(retail_price):
            raise HTTPException(status_code=409, detail=f"{p.name} has no valid retail price.")
        if wholesale_price <= 0 or not math.isfinite(wholesale_price):
            wholesale_price = retail_price

        if item.price_mode == "retail":
            price = retail_price
        elif item.price_mode == "wholesale":
            price = wholesale_price
        else:
            if user.role not in {"admin", "manager"}:
                raise HTTPException(status_code=403, detail="Only Admins and Managers can negotiate a sale price.")
            reason = (item.negotiated_reason or "").strip()
            if len(reason) < 5:
                raise HTTPException(status_code=400, detail="A negotiated-price reason of at least 5 characters is required.")
            if item.unit_price is None:
                raise HTTPException(status_code=400, detail="A negotiated selling price is required.")
            price = float(item.unit_price)
            minimum_price = max(0.01, float(p.cost_price or 0.0))
            maximum_price = max(retail_price, wholesale_price)
            if price < minimum_price or price > maximum_price:
                raise HTTPException(
                    status_code=400,
                    detail=f"Negotiated price for {p.name} must be between {minimum_price:.2f} and {maximum_price:.2f}.",
                )
            negotiated_lines.append({
                "product_id": p.id, "product_name": p.name,
                "catalog_retail_price": retail_price,
                "catalog_wholesale_price": wholesale_price,
                "unit_cost": float(p.cost_price or 0.0),
                "negotiated_price": price, "reason": reason,
            })
        validated.append((p, qty, price, item.price_mode))

    # --- PHASE 2: the whole cart is valid — now open the day and write -------
    # Only reached once every line passed, so a rejected cart never leaves a
    # spuriously auto-opened Business Day behind either.
    day = ensure_open_business_day(db, user.business_id, opener=user, commit=False)
    txn.business_day_id = day.id

    daily_total = 0.0
    synced_at = datetime.utcnow() if client_ref else None
    updated_products = []  # authoritative post-sale quantities — lets the frontend patch its local state instead of reloading everything (see performance refactor, section 13)
    for product_id in product_ids:
        p = products_by_id[product_id]
        qty = requested_by_product[product_id]
        p.quantity -= qty
        stock = source_stock_by_product[product_id]
        if stock:
            stock.quantity -= qty
        updated_products.append({"id": p.id, "quantity": p.quantity})

    for p, qty, price, price_mode in validated:
        total = qty * price
        # Every sale-time fact this line will ever need is captured here (see
        # the SaleModel snapshot comments) — nothing about this row's
        # financial meaning depends on the Product row afterwards.
        db.add(SaleModel(
            business_id=user.business_id, product_id=p.id, quantity=qty, total_price=total,
            unit_price=price, unit_cost_at_sale=p.cost_price, pricing_type=price_mode, product_name_snapshot=p.name,
            timestamp=datetime.utcnow(), client_ref=transaction_ref, synced_at=synced_at,
            business_day_id=day.id,
        ))
        daily_total += total
    # A sale changes stock levels and sales history — both are Business
    # Brain inputs (low-stock/forecast recommendations, velocity), so mark it
    # dirty here rather than recomputing it inline on this write path; the
    # next GET /business-brain performs one controlled refresh (see the
    # Business Brain invalidation strategy above).
    txn.total_price = daily_total
    mark_business_brain_dirty(db, user.business_id)
    add_audit(
        db, user, "SALE_COMPLETED", f"Completed a sale worth {daily_total:.2f}.", business_day_id=day.id,
        metadata={"transaction_id": transaction_ref, "total": round(daily_total, 2),
                  "lines": len(validated), "units": sum(q for _, q, _, _ in validated),
                  "pricing_policy": "server_catalog_or_authorized_negotiation",
                  "negotiated_lines": negotiated_lines},
    )
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    # Best-effort: a sale is already fully recorded above regardless of
    # whether this succeeds — never let a notification failure turn a
    # completed sale into an error response.
    try:
        for product_id in product_ids:
            check_inventory_notifications(db, user.business_id, products_by_id[product_id])
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "Sale completed successfully.", "daily_total": daily_total, "business_day_id": day.id,
            "transaction_id": transaction_ref, "updated_products": updated_products}

# -----------------------------------------------------------------------------
# REFUNDS
#
# A refund always references an original completed sale line (SaleModel row)
# — it never edits or deletes that row, and never exists as a bare negative
# SaleModel row. "Transaction" here means every SaleModel row created from
# one /sales/checkout submission, grouped by SaleModel.client_ref (now
# generated for every checkout, not just offline ones — see sales_checkout).
# A sale recorded before that grouping existed has no client_ref, so it is
# its own one-line "transaction" — addressed as "S{sale_id}" — see
# _sales_for_transaction_key.
# -----------------------------------------------------------------------------
def _sales_for_transaction_key(db: Session, business_id: int, transaction_key: str) -> List[SaleModel]:
    """Resolves a transaction_key (as handed out by list_sale_transactions/
    get_sale_transaction) back to its SaleModel rows. "S{id}" only ever
    matches a real single sale that genuinely has no client_ref — guards
    against a real client_ref string that happens to start with "S" being
    misread as the synthetic legacy form."""
    if transaction_key.startswith("S") and transaction_key[1:].isdigit():
        row = db.query(SaleModel).filter(
            SaleModel.id == int(transaction_key[1:]), SaleModel.business_id == business_id, SaleModel.client_ref.is_(None),
        ).first()
        if row: return [row]
    return (
        db.query(SaleModel)
        .filter(SaleModel.business_id == business_id, SaleModel.client_ref == transaction_key)
        .order_by(SaleModel.timestamp.asc()).all()
    )

def _serialize_sale_transactions(sales: List[SaleModel], refunded_map: Dict[int, int], products_map: Dict[int, "Product"]) -> List[dict]:
    """Groups a flat list of SaleModel rows into transactions by client_ref
    (see module note above), enriching each line with how much of it has
    already been refunded (refunded_map: sale_id -> cumulative refunded
    quantity, from SUM(RefundLine.quantity) — see the callers) so the
    remaining refundable quantity is always `sold - already_refunded`,
    never independently tracked or guessed."""
    groups: Dict[str, List[SaleModel]] = {}
    order: List[str] = []
    for s in sales:
        key = s.client_ref if s.client_ref else f"S{s.id}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(s)

    out = []
    for key in order:
        rows = groups[key]
        items = []
        for s in rows:
            refunded_qty = int(refunded_map.get(s.id, 0))
            # Prefer the sale's OWN price snapshot; derive from total_price
            # only for rows recorded before unit_price existed.
            unit_price = s.unit_price if s.unit_price is not None else (round(s.total_price / s.quantity, 2) if s.quantity else 0.0)
            product = products_map.get(s.product_id) if s.product_id else None
            # Name resolution order: sale-time snapshot -> live product ->
            # explicit placeholder. A deleted or renamed product therefore
            # never erases what history says was actually sold.
            product_name = s.product_name_snapshot or (product.name if product else "Deleted product")
            items.append({
                "sale_id": s.id, "product_id": s.product_id,
                "product_name": product_name,
                "quantity": s.quantity, "unit_price": unit_price,
                "unit_cost_known": s.unit_cost_at_sale is not None,
                "refunded_quantity": refunded_qty, "available_quantity": max(0, s.quantity - refunded_qty),
                "product_exists": product is not None,
            })
        original_total = round(sum(s.total_price for s in rows), 2)
        refunded_total = round(sum(i["refunded_quantity"] * i["unit_price"] for i in items), 2)
        total_original_qty = sum(i["quantity"] for i in items)
        total_refundable = sum(i["available_quantity"] for i in items)
        if total_refundable <= 0: status = "fully_refunded"
        elif total_refundable == total_original_qty: status = "not_refunded"
        else: status = "partially_refunded"
        out.append({
            "transaction_key": key, "timestamp": to_utc_iso(min(s.timestamp for s in rows)),
            "business_day_id": rows[0].business_day_id,
            "original_total": original_total, "refunded_total": refunded_total,
            "net_total": round(original_total - refunded_total, 2), "status": status, "items": items,
        })
    return out

def _refunded_quantity_map(db: Session, business_id: int, sale_ids: List[int]) -> Dict[int, int]:
    if not sale_ids: return {}
    rows = (
        db.query(RefundLine.original_sale_id, func.coalesce(func.sum(RefundLine.quantity), 0))
        .filter(RefundLine.business_id == business_id, RefundLine.original_sale_id.in_(sale_ids))
        .group_by(RefundLine.original_sale_id).all()
    )
    return {sid: int(qty) for sid, qty in rows}

@app.get("/sales/transactions")
def list_sale_transactions(
    business_day_id: Optional[int] = Query(None), period: str = Query("today"),
    custom_start: Optional[str] = Query(None), custom_end: Optional[str] = Query(None),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    """Individual completed transactions (grouped sale rows), the entry
    point for locating a sale to refund from — never used to auto-open a
    Business Day (read-only, see section 6 of the refund spec). Staff can
    read this (unlike the day-level /sales/history) because staff are
    explicitly authorized to create refunds and need a way to find what
    they're refunding."""
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    if business_day_id is not None:
        day = db.query(BusinessDay).filter(BusinessDay.id == business_day_id, BusinessDay.business_id == user.business_id).first()
        if not day: raise HTTPException(status_code=404, detail="Business Day not found.")
        query = business_day_sales_query(db, day)
    else:
        business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
        if not business: raise HTTPException(status_code=404, detail="Business not found.")
        start_utc, end_utc = resolve_financial_period(business, period, custom_start, custom_end)
        query = db.query(SaleModel).filter(SaleModel.business_id == user.business_id)
        if start_utc is not None and end_utc is not None:
            query = query.filter(SaleModel.timestamp >= start_utc, SaleModel.timestamp < end_utc)
    # Bounded like every other list/export in this app (see _sales_export_rows's
    # 5000-row cap) — a UI list needs far fewer, so this stays cheap.
    sales = query.order_by(SaleModel.timestamp.desc()).limit(500).all()
    if not sales: return []
    sale_ids = [s.id for s in sales]
    refunded_map = _refunded_quantity_map(db, user.business_id, sale_ids)
    product_ids = list({s.product_id for s in sales if s.product_id})
    products_map = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    return _serialize_sale_transactions(sales, refunded_map, products_map)

@app.get("/sales/transactions/{transaction_key}")
def get_sale_transaction(transaction_key: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    sales = _sales_for_transaction_key(db, user.business_id, transaction_key)
    if not sales: raise HTTPException(status_code=404, detail="Transaction not found.")
    sale_ids = [s.id for s in sales]
    refunded_map = _refunded_quantity_map(db, user.business_id, sale_ids)
    product_ids = list({s.product_id for s in sales if s.product_id})
    products_map = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    return _serialize_sale_transactions(sales, refunded_map, products_map)[0]

REFUND_REASONS = {"Customer return", "Wrong item", "Damaged product", "Pricing error", "Duplicate charge", "Other"}

@app.post("/sales/transactions/{transaction_key}/refund")
def create_refund(transaction_key: str, payload: RefundRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Atomic: every line is validated (ownership, remaining refundable
    quantity, transaction membership) BEFORE anything is written; only once
    every line passes does this create the RefundTransaction/RefundLine rows
    and restore stock, all inside this one request/commit. Any failure
    raises before the first db.add, so there is nothing to roll back — and
    if a rare DB-level failure happens after that point, FastAPI's
    exception propagation plus get_db()'s uncommitted session.close()
    discards every uncommitted change together (see resolve_business_day_
    reopen_request for the same pattern elsewhere in this file)."""
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    if not payload.lines: raise HTTPException(status_code=400, detail="Select at least one item to refund.")

    # Idempotency: a rapid double-submit with the same client_ref returns the
    # original result instead of refunding twice (section 22).
    client_ref = (payload.client_ref or "").strip()[:100] or None
    if client_ref:
        existing = db.query(RefundTransaction).filter(RefundTransaction.business_id == user.business_id, RefundTransaction.client_ref == client_ref).first()
        if existing:
            return {"message": "Refund already recorded.", "refund_transaction_id": existing.id, "refund_total": existing.refund_total, "duplicate": True}

    transaction_sales = _sales_for_transaction_key(db, user.business_id, transaction_key)
    if not transaction_sales: raise HTTPException(status_code=404, detail="Transaction not found.")
    transaction_sale_ids = {s.id for s in transaction_sales}
    sales_by_id = {s.id: s for s in transaction_sales}

    # REFUND_REASONS is a UI suggestion list, not an enforced enum — "Other"
    # plus a free-text note (below) covers anything not on the short list,
    # per section 18's "do not overcomplicate the UI".
    reason = (payload.reason or "").strip()[:100] or None

    # --- Validate EVERY line before touching anything -----------------------
    resolved = []
    seen_sale_ids = set()
    for line in payload.lines:
        if line.sale_id in seen_sale_ids:
            raise HTTPException(status_code=400, detail="Each item can only appear once per refund submission.")
        seen_sale_ids.add(line.sale_id)
        if line.sale_id not in transaction_sale_ids:
            raise HTTPException(status_code=400, detail="One or more items do not belong to this transaction.")
        sale = sales_by_id[line.sale_id]
        already_refunded = db.query(func.coalesce(func.sum(RefundLine.quantity), 0)).filter(
            RefundLine.original_sale_id == sale.id, RefundLine.business_id == user.business_id,
        ).scalar() or 0
        remaining = sale.quantity - already_refunded
        if line.quantity > remaining:
            raise HTTPException(status_code=409, detail=f"Cannot refund {line.quantity} unit(s) — only {remaining} remaining refundable for this item.")
        product = db.query(Product).filter(Product.id == sale.product_id, Product.business_id == user.business_id).first() if sale.product_id else None
        unit_price = (sale.total_price / sale.quantity) if sale.quantity else 0.0
        # A legacy NULL snapshot is unknown forever. Never replace it with
        # today's Product cost (or zero after deletion).
        unit_cost = sale.unit_cost_at_sale
        resolved.append({
            "sale": sale, "product": product, "quantity": line.quantity,
            "unit_price": unit_price, "unit_cost": unit_cost,
            "restock_requested": line.restock,
        })

    # Refunds are a genuine operational action, same as a sale — they open a
    # new Business Day session if none is active, never by merely viewing
    # history (section 6). Only reached after every line above validated.
    day = ensure_open_business_day(db, user.business_id, opener=user)

    refund_total = round(sum(r["quantity"] * r["unit_price"] for r in resolved), 2)
    refund_cost_total = (
        round(sum(r["quantity"] * r["unit_cost"] for r in resolved), 2)
        if all(r["unit_cost"] is not None for r in resolved) else None
    )

    rt = RefundTransaction(
        business_id=user.business_id, business_day_id=day.id,
        # transaction_sales all share the same client_ref by construction
        # (see _sales_for_transaction_key) — None for the synthetic "S{id}"
        # legacy single-sale form, the real shared ref otherwise.
        original_client_ref=transaction_sales[0].client_ref,
        reason=reason, note=(payload.note or "").strip()[:500] or None,
        refund_total=refund_total, refund_cost_total=refund_cost_total,
        created_by_id=user.id, created_by_name=user.username, created_by_role=user.role,
        client_ref=client_ref,
    )
    db.add(rt)
    try:
        db.flush()
    except IntegrityError:
        # Race: two near-simultaneous submissions with the same client_ref
        # both passed the check above before either committed — the partial
        # unique index (business_id, client_ref) lets only one insert
        # through. The loser rolls back and returns the winner's result
        # instead of erroring, exactly like start_business_day's own
        # IntegrityError handling.
        db.rollback()
        if client_ref:
            existing = db.query(RefundTransaction).filter(RefundTransaction.business_id == user.business_id, RefundTransaction.client_ref == client_ref).first()
            if existing:
                return {"message": "Refund already recorded.", "refund_transaction_id": existing.id, "refund_total": existing.refund_total, "duplicate": True}
        raise

    updated_products = []
    restock_skipped = []
    for r in resolved:
        sale, product, qty = r["sale"], r["product"], r["quantity"]
        restock = bool(r["restock_requested"] and product is not None)
        if r["restock_requested"] and product is None:
            # section 24: never fabricate/recreate a deleted product — the
            # financial refund still proceeds, restocking simply cannot.
            restock_skipped.append(sale.product_id)
        line_amount = round(qty * r["unit_price"], 2)
        line_cost = round(qty * r["unit_cost"], 2) if r["unit_cost"] is not None else None
        db.add(RefundLine(
            business_id=user.business_id, business_day_id=day.id, refund_transaction_id=rt.id,
            original_sale_id=sale.id, product_id=product.id if product else None,
            product_name_snapshot=(sale.product_name_snapshot or (product.name if product else f"Deleted product #{sale.product_id}" if sale.product_id else "Deleted product")),
            quantity=qty, unit_price=r["unit_price"], unit_cost=r["unit_cost"],
            refund_amount=line_amount, refund_cost=line_cost, restocked=restock,
        ))
        if restock:
            # Exactly the refunded quantity, on both Product and its
            # matching WarehouseStock row — the same two writes checkout
            # itself makes in reverse, so the two never drift apart
            # (sections 7/9). Never more, never less, never twice.
            product.quantity += qty
            stock = db.query(WarehouseStock).filter(WarehouseStock.product_id == product.id, WarehouseStock.warehouse == (product.warehouse or "Main Central Warehouse")).first()
            if stock: stock.quantity += qty
            updated_products.append({"id": product.id, "quantity": product.quantity})

    mark_business_brain_dirty(db, user.business_id)
    db.commit(); db.refresh(rt)

    add_audit(
        db, user, "REFUND_COMPLETED", f"Refunded {refund_total:.2f} across {len(resolved)} item(s).",
        business_day_id=day.id,
        metadata={
            "refund_transaction_id": rt.id, "original_transaction": transaction_key,
            "refund_total": refund_total, "refund_cost_total": refund_cost_total,
            "cogs_complete": refund_cost_total is not None,
            "quantity": sum(r["quantity"] for r in resolved),
            "restocked": any(r["restock_requested"] and r["product"] is not None for r in resolved),
            "restock_skipped_products": restock_skipped or None,
        },
    )
    db.commit()

    return {
        "message": "Refund completed.", "refund_transaction_id": rt.id,
        "refund_total": refund_total, "refund_cost_total": refund_cost_total,
        "business_day_id": day.id, "updated_products": updated_products,
        "restock_skipped_products": restock_skipped or None,
    }

@app.get("/sales/current-day")
def current_sales_day(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Read-only: looks up the currently ACTIVE Business Day session, if any
    # — never date-based, never creates one. A session closed earlier today
    # (or any prior day) is history, not "the current day" — once nothing is
    # active, this deliberately looks exactly like NOT_STARTED, business_day
    # null, so the dashboard/Today's Sales view has no way to render a
    # leftover "Reopen" affordance from a closed session's last state.
    # Opening this view must never, by itself, start a session.
    day = get_active_business_day(db, user.business_id)
    if not day:
        return {"open": False, "status": "NOT_STARTED", "business_day": None}
    return {"open": True, "status": day.status, "business_day": serialize_business_day(day, db)}

@app.post("/sales/start-business-day")
def start_business_day_endpoint(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Business Day open/close is available to every operational role — Admin,
    # Manager, and Staff — since Staff are often the ones physically running
    # the counter (see the Business Day role-permission model). Reopening a
    # CLOSED day is a separate, more privileged workflow (direct-reopen for
    # Admin, reopen-request for Manager) that Staff are never part of, and
    # that lives exclusively in Sales History — never here.
    #
    # Open ALWAYS creates a brand-new session (see start_business_day) — a
    # session closed earlier today is never reused and never treated as
    # something to "continue". The only thing that can block this is
    # another session currently being active, in which case
    # start_business_day itself raises the required conflict message.
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    day = start_business_day(db, user.business_id, opener=user)
    return {"message": "Business day started.", "business_day": serialize_business_day(day, db)}

@app.post("/sales/open-business-day")
def open_business_day_endpoint(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Idempotent "Open Business Day": returns the currently active session
    if one already exists rather than erroring, otherwise opens a fresh one.
    Deliberately distinct from /sales/start-business-day above, which is
    strict (409 if a session is already active) and is what the Dashboard
    control calls — the two exist because "make sure a day is open" and
    "start a new day, and tell me if that's not possible" are different
    intents and must not silently collapse into one behaviour."""
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    existing = get_active_business_day(db, user.business_id)
    if existing:
        return {"message": "This Business Day is already open.", "business_day": serialize_business_day(existing, db), "already_open": True}
    day = start_business_day(db, user.business_id, opener=user)
    return {"message": "Business day started.", "business_day": serialize_business_day(day, db), "already_open": False}

@app.get("/business-days/current-summary")
def current_business_day_summary(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """The live financial state of the CURRENTLY ACTIVE Business Day session
    — the operational "what has happened since we opened" view, which is a
    different question from the calendar-period "today" that
    /financial-summary?period=today answers (a day opened yesterday evening
    and still open now belongs to this session, not to today's calendar
    date). Both are legitimate; they must not be conflated.

    Strictly read-only: viewing the dashboard must never auto-open a
    session, so with nothing active this reports open=false and zeroes
    rather than creating a day.

    Every figure comes from the same canonical per-business-day aggregates
    the close snapshot and Sales History use — never a second, independently
    written formula."""
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    day = get_active_business_day(db, user.business_id)
    if not day:
        return {
            "open": False, "business_day": None,
            "sales": 0.0, "cogs": 0.0, "gross_profit": 0.0, "expenses": 0.0, "net_profit": 0.0,
            "profit_margin_percent": None, "transactions": 0, "units_sold": 0, "expense_count": 0,
            "refund_total": 0.0, "refunded_units": 0,
            "cogs_complete": True, "known_cogs": 0.0,
            "unknown_cogs_sale_lines": 0, "unknown_cogs_sale_units": 0,
            "unknown_cogs_refund_lines": 0, "unknown_cogs_refund_units": 0,
        }
    return {"open": True, "business_day": serialize_business_day(day, db), **_business_day_financials(db, day)}

@app.get("/financial-summary")
def financial_summary(period: str = Query("today"), custom_start: Optional[str] = Query(None), custom_end: Optional[str] = Query(None), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """The one endpoint behind every profit/revenue/expense figure shown
    anywhere in Cauldra — Dashboard, Sales History period totals, Expense
    History period totals, and the Profit page all call this exact endpoint
    rather than calculating anything independently, so they can never
    disagree with each other."""
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    return financial_summary_for_period(db, business, period, custom_start, custom_end)

@app.get("/financial-summary/breakdown")
def financial_summary_breakdown(period: str = Query("today"), custom_start: Optional[str] = Query(None), custom_end: Optional[str] = Query(None), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Expense-by-category and top-selling-products for the same period
    financial_summary() uses — deliberately calls resolve_financial_period()
    directly (not a separately re-derived boundary) so a Profit page's
    breakdown rows always add up to the same headline totals shown above
    them, never a second, independently-computed version."""
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    start_utc, end_utc = resolve_financial_period(business, period, custom_start, custom_end)

    expense_filters = [Expense.business_id == business.id]
    sales_filters = [SaleModel.business_id == business.id]
    if start_utc is not None and end_utc is not None:
        expense_filters += [Expense.created_at >= start_utc, Expense.created_at < end_utc]
        sales_filters += [SaleModel.timestamp >= start_utc, SaleModel.timestamp < end_utc]

    expense_rows = (
        db.query(Expense.category, func.coalesce(func.sum(Expense.amount), 0.0))
        .filter(*expense_filters).group_by(Expense.category).order_by(func.sum(Expense.amount).desc()).all()
    )
    # Grouped by the sale's own product_id with the sale-time NAME snapshot —
    # an OUTER join so a sale whose product was later deleted still reports
    # its units/revenue instead of silently dropping out of the breakdown
    # (and so the totals here keep adding up to the headline figures above).
    # Product is many-to-one from SaleModel, so this join cannot fan out and
    # multiply quantities — verified against the generated SQL.
    name_expr = func.coalesce(SaleModel.product_name_snapshot, Product.name, literal("Deleted product"))
    product_rows = (
        db.query(name_expr, func.coalesce(func.sum(SaleModel.quantity), 0), func.coalesce(func.sum(SaleModel.total_price), 0.0))
        .outerjoin(Product, Product.id == SaleModel.product_id)
        .filter(*sales_filters).group_by(SaleModel.product_id, name_expr).order_by(func.sum(SaleModel.total_price).desc()).limit(5).all()
    )
    return {
        "period": period,
        "expense_breakdown": [{"category": c, "amount": round(float(a or 0), 2)} for c, a in expense_rows],
        "top_products": [{"name": n, "units_sold": int(u or 0), "revenue": round(float(r or 0), 2)} for n, u, r in product_rows],
    }

@app.get("/sales/history")
def sales_history(period: str = Query("all"), custom_start: Optional[str] = Query(None), custom_end: Optional[str] = Query(None), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    # Same shared period resolver every other date-range filter in the app
    # uses (Profit, Expense History's totals) — business-timezone-aware, and
    # period="all" (the default, preserving this endpoint's original
    # behavior when called with no params) returns (None, None): no filter.
    start_utc, end_utc = resolve_financial_period(business, period, custom_start, custom_end)
    # A day that has been closed at least once belongs in Sales History even
    # while it is currently REOPENED (is_open == True) awaiting re-close —
    # otherwise a reopened day would vanish from history entirely instead of
    # showing its "Close Business Day" action (see business-day-permission
    # spec section 7). A day that has never been closed even once (the
    # ordinary still-open "today") is Today's Sales, not history yet.
    days_query = db.query(BusinessDay).filter(BusinessDay.business_id == user.business_id, BusinessDay.closed_at.isnot(None))
    if start_utc is not None and end_utc is not None:
        days_query = days_query.filter(BusinessDay.opened_at >= start_utc, BusinessDay.opened_at < end_utc)
    # Ordered by opened_at, not date — a business can have multiple sessions
    # on the same calendar date (see the BusinessDay model docstring), and
    # `date` alone can't distinguish or order them, only opened_at can.
    days = days_query.order_by(BusinessDay.opened_at.desc()).all()

    # Was one Sale query PER Business Day (business_day_sales_query(db, d)
    # inside this loop) — N+1 behavior that scaled with historical day
    # count. Replaced with one grouped aggregate for every day that already
    # has business_day_id set, plus a single bounded fallback query for any
    # legacy pre-migration sales still missing it (expected to be rare/zero
    # once the startup backfill has run — see _backfill_business_day_ids()).
    totals_map: Dict[int, Dict[str, float]] = {}
    day_ids = [d.id for d in days]
    if day_ids:
        grouped = (
            db.query(SaleModel.business_day_id, func.sum(SaleModel.total_price), func.sum(SaleModel.quantity), func.count(func.distinct(checkout_key_expr())))
            .filter(SaleModel.business_id == user.business_id, SaleModel.business_day_id.in_(day_ids))
            .group_by(SaleModel.business_day_id).all()
        )
        for bd_id, total, qty, cnt in grouped:
            totals_map[bd_id] = {"gross_sales": total or 0.0, "items_sold": qty or 0, "transactions": cnt or 0}

        orphan_sales = (
            db.query(SaleModel.timestamp, SaleModel.total_price, SaleModel.quantity)
            .filter(
                SaleModel.business_id == user.business_id, SaleModel.business_day_id.is_(None),
                SaleModel.timestamp >= min(d.opened_at for d in days),
                SaleModel.timestamp <= max(d.closed_at or datetime.utcnow() for d in days),
            ).all()
        )
        for ts, total, qty in orphan_sales:
            for d in days:
                if d.opened_at <= ts <= (d.closed_at or datetime.utcnow()):
                    bucket = totals_map.setdefault(d.id, {"gross_sales": 0.0, "items_sold": 0, "transactions": 0})
                    bucket["gross_sales"] += total or 0.0
                    bucket["items_sold"] += qty or 0
                    bucket["transactions"] += 1
                    break

        # Refund activity for these days — grouped by RefundLine's OWN
        # business_day_id (the day the refund was PERFORMED on), which can
        # differ from the original sale's day (section 5/27 of the refund
        # spec) — never mixed into the sales grouped-aggregate above.
        refund_grouped = (
            db.query(RefundLine.business_day_id, func.sum(RefundLine.refund_amount), func.sum(RefundLine.quantity))
            .filter(RefundLine.business_id == user.business_id, RefundLine.business_day_id.in_(day_ids))
            .group_by(RefundLine.business_day_id).all()
        )
        refund_totals_map = {bd_id: {"refund_total": total or 0.0, "refunded_units": qty or 0} for bd_id, total, qty in refund_grouped}
        refund_txn_grouped = (
            db.query(RefundTransaction.business_day_id, func.count(RefundTransaction.id))
            .filter(RefundTransaction.business_id == user.business_id, RefundTransaction.business_day_id.in_(day_ids))
            .group_by(RefundTransaction.business_day_id).all()
        )
        refund_txn_map = {bd_id: cnt for bd_id, cnt in refund_txn_grouped}
    else:
        refund_totals_map, refund_txn_map = {}, {}

    out = []
    for d in days:
        t = totals_map.get(d.id, {"gross_sales": 0.0, "items_sold": 0, "transactions": 0})
        r = refund_totals_map.get(d.id, {"refund_total": 0.0, "refunded_units": 0})
        out.append({
            "date": d.date, "transactions": t["transactions"], "items_sold": t["items_sold"],
            # net_sales keeps its established name (now net of same-day
            # refunds); gross_sales/refund_total are additive, never
            # required by existing callers.
            "net_sales": round(t["gross_sales"] - r["refund_total"], 2), "gross_sales": round(t["gross_sales"], 2),
            "refund_total": round(r["refund_total"], 2), "refunded_units": r["refunded_units"],
            "refund_transaction_count": refund_txn_map.get(d.id, 0),
            "business_day_id": d.id, "status": d.status,
            "opened_at": to_utc_iso(d.opened_at), "opened_by_name": d.opened_by_name, "opened_by_role": d.opened_by_role,
            "closed_at": to_utc_iso(d.closed_at), "closed_by_name": d.closed_by_name, "closed_by_role": d.closed_by_role,
            "reopen_count": d.reopen_count,
        })
    return out

SALES_EXPORT_COLUMNS = [
    {"key": "id", "label": "SALE ID", "type": "number"},
    {"key": "date", "label": "DATE", "type": "date"},
    {"key": "time", "label": "TIME", "type": "time"},
    {"key": "product", "label": "PRODUCT", "type": "text"},
    {"key": "sku", "label": "SKU", "type": "text"},
    {"key": "warehouse", "label": "WAREHOUSE", "type": "text"},
    {"key": "quantity", "label": "QUANTITY", "type": "number"},
    {"key": "unit_price", "label": "UNIT PRICE", "type": "currency"},
    {"key": "total", "label": "TOTAL", "type": "currency"},
]

def _sales_export_rows(db, user, business, period, custom_start, custom_end):
    """Shared by the CSV and Excel line-item Sales exports — the detailed
    view /sales/history itself cannot provide, since that endpoint only
    returns per-Business-Day aggregates. Only columns that exist on
    SaleModel/Product are included — this app has no per-sale customer or
    payment-method field and no per-sale staff attribution, so those are
    intentionally omitted rather than invented. Capped at 5000 rows, well
    above what a single filtered period is expected to contain."""
    start_utc, end_utc = resolve_financial_period(business, period, custom_start, custom_end)
    q = db.query(SaleModel, Product).outerjoin(Product, Product.id == SaleModel.product_id).filter(SaleModel.business_id == user.business_id)
    if start_utc is not None and end_utc is not None:
        q = q.filter(SaleModel.timestamp >= start_utc, SaleModel.timestamp < end_utc)
    rows = q.order_by(SaleModel.timestamp.asc()).limit(5000).all()
    tz = business_local_zoneinfo(business)
    out_rows = []
    for sale, product in rows:
        local_dt = sale.timestamp.replace(tzinfo=timezone.utc).astimezone(tz)
        unit_price = round(sale.total_price / sale.quantity, 2) if sale.quantity else sale.total_price
        out_rows.append([
            sale.id, local_dt.strftime("%Y-%m-%d"), local_dt.strftime("%H:%M:%S"),
            product.name if product else "Deleted product", product.sku if product else "", (product.warehouse if product else "") or "",
            sale.quantity, unit_price, sale.total_price,
        ])
    return out_rows

@app.get("/sales/export")
def export_sales_csv(period: str = Query("all"), custom_start: Optional[str] = Query(None), custom_end: Optional[str] = Query(None), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Line-item CSV export of individual Sale rows. Same role restriction
    and period resolver as /sales/history (staff sees Today's Sales but not
    this historical export, matching the existing restriction there)."""
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    out_rows = _sales_export_rows(db, user, business, period, custom_start, custom_end)
    if not out_rows:
        return Response(status_code=204)
    header = [c["label"] for c in SALES_EXPORT_COLUMNS]
    return build_csv_response(f"cauldra_sales_{business_local_today(db, user.business_id)}.csv", header, out_rows)

@app.get("/sales/export/xlsx")
def export_sales_xlsx(period: str = Query("all"), custom_start: Optional[str] = Query(None), custom_end: Optional[str] = Query(None), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Same authorization, tenant scoping, filters, and row data as the CSV
    export above — only the presentation differs."""
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    if not business: raise HTTPException(status_code=404, detail="Business not found.")
    out_rows = _sales_export_rows(db, user, business, period, custom_start, custom_end)
    if not out_rows:
        return Response(status_code=204)
    metadata = [["Generated", datetime.utcnow().strftime("%d %B %Y %H:%M UTC")], ["Period", period if period != "custom" else f"{custom_start or '?'} - {custom_end or '?'}"]]
    sheets = [{"name": "Sales", "title": "CAULDRA SALES REPORT", "metadata": metadata, "columns": SALES_EXPORT_COLUMNS, "rows": out_rows}]
    return build_xlsx_response(f"cauldra_sales_{business_local_today(db, user.business_id)}.xlsx", "Cauldra Sales Report", sheets)

@app.get("/sales/analytics")
def sales_analytics(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    days = db.query(BusinessDay).filter(BusinessDay.business_id == user.business_id, BusinessDay.is_open == False).order_by(BusinessDay.date.asc()).all()
    rows=[]
    for d in days:
        sales = business_day_sales_query(db, d).all()
        rows.append({"date": d.date, "sales": sum(s.total_price for s in sales)})
    vals = [r["sales"] for r in rows]
    avg = sum(vals)/len(vals) if vals else 0
    change = 0
    if len(vals) >= 2 and vals[-2]: change = round(((vals[-1]-vals[-2])/vals[-2])*100,1)
    trend = "Growing" if change > 0 else "Declining" if change < 0 else "Stable"
    return {"days": rows, "average_daily_sales": avg, "change_percent": change, "trend": trend}

def _close_business_day(db: Session, day: BusinessDay, user: User) -> dict:
    """The one place a BusinessDay row is ever closed — used by both
    /sales/end-business-day (whichever session is currently active) and
    /business-days/{id}/close (any specific session by id, including a
    historical one reopened from Sales History). Caller is responsible for
    the role check and for confirming `day` is currently open; this only
    ever mutates the row and writes the audit trail.

    is_reclose distinguishes a day's first-ever close (BUSINESS_DAY_CLOSED)
    from a later close after a reopen (BUSINESS_DAY_CLOSED_AGAIN) — the
    original close's snapshot is never overwritten: it stays forever in its
    own immutable AuditLog row (see business_day_timeline / section 8's
    original-vs-corrected requirement), this just records a new one."""
    # The closing snapshot is the SAME canonical per-day P&L the live
    # current-day endpoint serves (_business_day_financials) — one formula,
    # so a day's numbers cannot shift the moment it is closed. It is written
    # into the immutable AuditLog purely as a record of what was reported at
    # close time; the figures stay independently recomputable forever from
    # the attached sale/expense/refund snapshots, so this is a receipt, not a
    # second source of truth that could drift.
    snapshot = _business_day_financials(db, day)
    # Legacy key aliases. The closing snapshot is written into the immutable
    # AuditLog, so rows recorded before this refactor use the older names —
    # and the Sales History original-vs-corrected timeline reads
    # meta.sales_total to render them. Emitting both shapes keeps every
    # historical audit row and its reader working unchanged, while new rows
    # also carry the full canonical P&L above.
    snapshot.update({
        "sales_total": snapshot["gross_sales"],
        "items_sold": snapshot["gross_units_sold"],
        "expenses_total": snapshot["expenses"],
        "net_sales_total": snapshot["sales"],
    })
    is_reclose = day.closed_at is not None  # a prior close_at means this day was reopened since
    day.is_open = False
    day.status = "CLOSED"
    day.closed_at = datetime.utcnow()
    day.closed_by_id = user.id
    day.closed_by_name = user.username
    day.closed_by_role = user.role
    action = "BUSINESS_DAY_CLOSED_AGAIN" if is_reclose else "BUSINESS_DAY_CLOSED"
    add_audit(
        db, user, action, f"Closed business day {day.date}.",
        business_day_id=day.id, metadata=snapshot,
    )
    # Closing finalizes a day's sales/expense totals — a Business Brain input
    # (history_days, velocity). Mark dirty rather than recomputing inline.
    mark_business_brain_dirty(db, day.business_id)
    db.commit()
    return {"message": "Business day closed and recorded.", "business_day": serialize_business_day(day, db), "closing_snapshot": snapshot}

@app.post("/sales/end-business-day")
def end_business_day(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Business Day open/close is available to every operational role — Admin,
    # Manager, and Staff (see the Business Day role-permission model).
    # Reopening a CLOSED day remains a separate, more privileged workflow
    # that Staff are never part of — enforced in the dedicated reopen/
    # reopen-request endpoints, not here.
    #
    # Close ALWAYS ends whatever session is currently ACTIVE — never
    # "today's" row by date. If a previously-closed session from earlier
    # today (or a reopened older one) isn't the active one, this correctly
    # has nothing to do with it.
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    day = get_active_business_day(db, user.business_id)
    if not day: raise HTTPException(status_code=409, detail="No Business Day is currently open to close.")
    return _close_business_day(db, day, user)

@app.post("/business-days/{business_day_id}/close")
def close_business_day_by_id(business_day_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Closes a SPECIFIC Business Day session by id — the counterpart
    /sales/end-business-day can't provide, since that always targets
    whichever session is currently active. Needed so a historical session
    reopened from Sales History can be closed again correctly
    even if the real calendar day has since moved on (e.g. a day from three
    days ago is reopened for a correction and must be re-closed as ITSELF,
    not conflated with today's own Business Day). Open/close is available to
    Admin, Manager, and Staff alike, same as the current-day endpoints above.
    """
    if user.role not in {"admin", "manager", "staff"}: raise HTTPException(status_code=403, detail="Access denied")
    day = db.query(BusinessDay).filter(BusinessDay.id == business_day_id, BusinessDay.business_id == user.business_id).first()
    if not day: raise HTTPException(status_code=404, detail="Business Day not found.")
    if not day.is_open: raise HTTPException(status_code=409, detail="This Business Day is already closed.")
    return _close_business_day(db, day, user)

@app.get("/business-days/{business_day_id}/timeline")
def business_day_timeline(business_day_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """The Business Day activity/timeline — a filtered read of the existing
    immutable AuditLog, not a separate log. Every lifecycle event (started,
    closed, reopen requested/approved/rejected, reopened, closed again) that
    ever touched this day is here, in order, forever."""
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    day = db.query(BusinessDay).filter(BusinessDay.id == business_day_id, BusinessDay.business_id == user.business_id).first()
    if not day: raise HTTPException(status_code=404, detail="Business Day not found.")
    rows = db.query(AuditLog).filter(AuditLog.business_day_id == business_day_id, AuditLog.business_id == user.business_id).order_by(AuditLog.created_at.asc()).all()
    events = []
    for r in rows:
        meta = None
        if r.metadata_json:
            try: meta = json.loads(r.metadata_json)
            except Exception: meta = None
        events.append({
            "id": r.id, "action": r.action, "description": r.description,
            "actor_username": r.actor_username, "actor_role": r.actor_role,
            "created_at": to_utc_iso(r.created_at), "metadata": meta,
        })
    return {"business_day": serialize_business_day(day, db), "events": events}

@app.post("/business-days/{business_day_id}/reopen-request")
def request_business_day_reopen(business_day_id: int, data: BusinessDayReopenRequestCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Staff cannot request reopening at all; Manager must; Admin uses direct-reopen instead.
    if user.role != "manager": raise HTTPException(status_code=403, detail="Only Managers can request that a Business Day be reopened.")
    day = db.query(BusinessDay).filter(BusinessDay.id == business_day_id, BusinessDay.business_id == user.business_id).first()
    if not day: raise HTTPException(status_code=404, detail="Business Day not found.")
    if day.is_open: raise HTTPException(status_code=409, detail="This Business Day is not closed.")
    reason = (data.reason or "").strip()
    if not reason: raise HTTPException(status_code=400, detail="A reason is required to request reopening.")
    existing_pending = db.query(BusinessDayReopenRequest).filter(BusinessDayReopenRequest.business_day_id == business_day_id, BusinessDayReopenRequest.status == "PENDING").first()
    if existing_pending: raise HTTPException(status_code=409, detail="A reopening request for this Business Day is already pending.")
    row = BusinessDayReopenRequest(
        business_id=user.business_id, business_day_id=business_day_id, reason=reason,
        requested_by_id=user.id, requested_by_name=user.username, requested_by_role=user.role,
    )
    db.add(row); db.flush()
    add_audit(
        db, user, "BUSINESS_DAY_REOPEN_REQUESTED", f"Requested reopening of business day {day.date}: {reason}",
        business_day_id=business_day_id, metadata={"reason": reason, "request_id": row.id},
    )
    db.commit(); db.refresh(row)
    return {"id": row.id, "message": "Reopening request submitted for Admin approval."}

@app.get("/business-days/reopen-requests")
def list_business_day_reopen_requests(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Admin sees every pending request (the approval queue). A Manager cannot
    # approve anything here, but is allowed to see the status of requests
    # *they themselves* filed — otherwise there is no honest way for the
    # frontend to show "reopening requested, awaiting Admin approval" instead
    # of silently pretending the day is just closed.
    if user.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    filters = [BusinessDayReopenRequest.business_id == user.business_id, BusinessDayReopenRequest.status == "PENDING"]
    if user.role == "manager":
        filters.append(BusinessDayReopenRequest.requested_by_id == user.id)
    rows = db.query(BusinessDayReopenRequest).filter(*filters).order_by(BusinessDayReopenRequest.created_at.desc()).all()
    return [{
        "id": r.id, "business_day_id": r.business_day_id, "reason": r.reason,
        "requested_by_name": r.requested_by_name, "requested_by_role": r.requested_by_role,
        "created_at": to_utc_iso(r.created_at), "status": r.status,
    } for r in rows]

@app.get("/business-days/reopen-requests/history")
def list_business_day_reopen_request_history(limit: int = Query(50, ge=1, le=200), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Access denied")
    rows = db.query(BusinessDayReopenRequest).filter(BusinessDayReopenRequest.business_id == user.business_id, BusinessDayReopenRequest.status != "PENDING").order_by(BusinessDayReopenRequest.resolved_at.desc()).limit(limit).all()
    return [{
        "id": r.id, "business_day_id": r.business_day_id, "reason": r.reason,
        "requested_by_name": r.requested_by_name, "requested_by_role": r.requested_by_role,
        "created_at": to_utc_iso(r.created_at), "status": r.status,
        "resolved_by_name": r.resolved_by_name, "resolution_note": r.resolution_note,
        "resolved_at": to_utc_iso(r.resolved_at),
    } for r in rows]

@app.post("/business-days/reopen-requests/{request_id}/{resolution}")
def resolve_business_day_reopen_request(request_id: int, resolution: str, data: BusinessDayReopenResolution = BusinessDayReopenResolution(), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can approve or reject reopening requests.")
    if resolution not in {"approve", "reject"}: raise HTTPException(status_code=400, detail="Unsupported resolution.")
    row = db.query(BusinessDayReopenRequest).filter(BusinessDayReopenRequest.id == request_id, BusinessDayReopenRequest.business_id == user.business_id, BusinessDayReopenRequest.status == "PENDING").first()
    if not row: raise HTTPException(status_code=404, detail="Reopening request is no longer pending.")
    day = db.query(BusinessDay).filter(BusinessDay.id == row.business_day_id, BusinessDay.business_id == user.business_id).first()
    if not day: raise HTTPException(status_code=404, detail="Business Day not found.")

    # Reopening actually reactivates this session (is_open True) — must
    # never collide with another Business Day that's already active (see
    # the BusinessDay model docstring: at most one active session at a
    # time). Checked BEFORE any mutation, so a blocked approval leaves the
    # request PENDING for a later retry instead of silently consuming it.
    if resolution == "approve" and not day.is_open:
        active = get_active_business_day(db, user.business_id)
        if active and active.id != day.id:
            raise HTTPException(status_code=409, detail=BUSINESS_DAY_ALREADY_ACTIVE_MSG)

    row.status = "APPROVED" if resolution == "approve" else "REJECTED"
    row.resolved_by_id = user.id
    row.resolved_by_name = user.username
    row.resolution_note = (data.note or "").strip() or None
    row.resolved_at = datetime.utcnow()

    if resolution == "approve":
        if not day.is_open:  # still closed at this exact moment — reopen it
            day.is_open = True
            day.status = "REOPENED"
            day.reopen_count = (day.reopen_count or 0) + 1
        add_audit(
            db, user, "BUSINESS_DAY_REOPEN_APPROVED", f"Approved reopening of business day {day.date}.",
            business_day_id=day.id, metadata={"request_id": row.id, "requested_by": row.requested_by_name, "reason": row.reason},
        )
        # The reopening itself is its own distinct, separately reportable event
        # from the approval decision — the original closure this replaces stays
        # permanently visible in this same timeline, never overwritten.
        add_audit(
            db, user, "BUSINESS_DAY_REOPENED", f"Business day {day.date} reopened.",
            business_day_id=day.id, metadata={"request_id": row.id, "reopen_count": day.reopen_count},
        )
        message = "Business day reopened."
    else:
        add_audit(
            db, user, "BUSINESS_DAY_REOPEN_REJECTED", f"Rejected reopening of business day {day.date}.",
            business_day_id=day.id, metadata={"request_id": row.id, "requested_by": row.requested_by_name, "reason": row.reason, "resolution_note": row.resolution_note},
        )
        message = "Business day reopening request rejected."
    db.commit()
    return {"message": message, "business_day": serialize_business_day(day, db)}

@app.post("/business-days/{business_day_id}/direct-reopen")
def direct_reopen_business_day(business_day_id: int, data: BusinessDayReopenRequestCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Admin full authority still means every action is audited — an Admin
    reopening without a prior Manager request is allowed (section 23), but
    it is never silent: a reason is required and the event is recorded
    exactly like an approved request would be."""
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can directly reopen a Business Day.")
    day = db.query(BusinessDay).filter(BusinessDay.id == business_day_id, BusinessDay.business_id == user.business_id).first()
    if not day: raise HTTPException(status_code=404, detail="Business Day not found.")
    if day.is_open: raise HTTPException(status_code=409, detail="This Business Day is not closed.")
    # Reactivating this session must never collide with another Business Day
    # that's already active (see the BusinessDay model docstring).
    active = get_active_business_day(db, user.business_id)
    if active and active.id != day.id:
        raise HTTPException(status_code=409, detail=BUSINESS_DAY_ALREADY_ACTIVE_MSG)
    reason = (data.reason or "").strip()
    if not reason: raise HTTPException(status_code=400, detail="A reason is required to reopen a Business Day.")
    day.is_open = True
    day.status = "REOPENED"
    day.reopen_count = (day.reopen_count or 0) + 1
    add_audit(
        db, user, "BUSINESS_DAY_REOPENED", f"Business day {day.date} reopened directly by Admin: {reason}",
        business_day_id=day.id, metadata={"reason": reason, "direct": True, "reopen_count": day.reopen_count},
    )
    db.commit()
    return {"message": "Business day reopened.", "business_day": serialize_business_day(day, db)}

@app.post("/sales/{sale_id}/adjustments")
def create_sale_adjustment(sale_id: int, data: RecordAdjustmentRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """The general correction mechanism (section 15): the original Sale row
    is never touched. This only ever ADDS a new, permanently-linked
    adjustment record — reporting sums original + adjustments, it never
    rewrites the original."""
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can record a correction.")
    sale = db.query(SaleModel).filter(SaleModel.id == sale_id, SaleModel.business_id == user.business_id).first()
    if not sale: raise HTTPException(status_code=404, detail="Sale not found.")
    reason = (data.reason or "").strip()
    if not reason: raise HTTPException(status_code=400, detail="A reason is required to record a correction.")
    if sale.business_day_id:
        day = db.query(BusinessDay).filter(BusinessDay.id == sale.business_day_id).first()
        if day and day.is_open is False and day.status == "CLOSED":
            raise HTTPException(status_code=409, detail="This Business Day is closed. Reopen it before recording a correction to its records.")
    adj = SaleAdjustment(
        business_id=user.business_id, sale_id=sale_id,
        quantity_delta=data.quantity_delta or 0, amount_delta=data.amount_delta or 0.0, reason=reason,
        created_by_id=user.id, created_by_name=user.username, created_by_role=user.role,
    )
    db.add(adj); db.flush()
    add_audit(
        db, user, "CORRECTION_CREATED", f"Recorded a correction to sale #{sale_id}: {reason}",
        business_day_id=sale.business_day_id,
        metadata={"record_type": "sale", "record_id": sale_id, "quantity_delta": adj.quantity_delta, "amount_delta": adj.amount_delta, "reason": reason, "adjustment_id": adj.id},
    )
    db.commit(); db.refresh(adj)
    return {"id": adj.id, "message": "Correction recorded."}

@app.get("/sales/{sale_id}/adjustments")
def list_sale_adjustments(sale_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    sale = db.query(SaleModel).filter(SaleModel.id == sale_id, SaleModel.business_id == user.business_id).first()
    if not sale: raise HTTPException(status_code=404, detail="Sale not found.")
    rows = db.query(SaleAdjustment).filter(SaleAdjustment.sale_id == sale_id, SaleAdjustment.business_id == user.business_id).order_by(SaleAdjustment.created_at.asc()).all()
    original_total = sale.total_price
    final_total = original_total + sum(r.amount_delta for r in rows)
    return {
        "original": {"quantity": sale.quantity, "total_price": original_total},
        "adjustments": [{"id": r.id, "quantity_delta": r.quantity_delta, "amount_delta": r.amount_delta, "reason": r.reason, "created_by_name": r.created_by_name, "created_by_role": r.created_by_role, "created_at": to_utc_iso(r.created_at)} for r in rows],
        "final_total": final_total,
    }

@app.post("/expenses/{expense_id}/adjustments")
def create_expense_adjustment(expense_id: int, data: RecordAdjustmentRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin": raise HTTPException(status_code=403, detail="Only Admins can record a correction.")
    expense = db.query(Expense).filter(Expense.id == expense_id, Expense.business_id == user.business_id).first()
    if not expense: raise HTTPException(status_code=404, detail="Expense not found.")
    reason = (data.reason or "").strip()
    if not reason: raise HTTPException(status_code=400, detail="A reason is required to record a correction.")
    if expense.business_day_id:
        day = db.query(BusinessDay).filter(BusinessDay.id == expense.business_day_id).first()
        if day and day.is_open is False and day.status == "CLOSED":
            raise HTTPException(status_code=409, detail="This Business Day is closed. Reopen it before recording a correction to its records.")
    adj = ExpenseAdjustment(
        business_id=user.business_id, expense_id=expense_id,
        amount_delta=data.amount_delta or 0.0, reason=reason,
        created_by_id=user.id, created_by_name=user.username, created_by_role=user.role,
    )
    db.add(adj); db.flush()
    add_audit(
        db, user, "CORRECTION_CREATED", f"Recorded a correction to expense #{expense_id}: {reason}",
        business_day_id=expense.business_day_id,
        metadata={"record_type": "expense", "record_id": expense_id, "amount_delta": adj.amount_delta, "reason": reason, "adjustment_id": adj.id},
    )
    db.commit(); db.refresh(adj)
    return {"id": adj.id, "message": "Correction recorded."}

@app.get("/expenses/{expense_id}/adjustments")
def list_expense_adjustments(expense_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Access denied")
    expense = db.query(Expense).filter(Expense.id == expense_id, Expense.business_id == user.business_id).first()
    if not expense: raise HTTPException(status_code=404, detail="Expense not found.")
    rows = db.query(ExpenseAdjustment).filter(ExpenseAdjustment.expense_id == expense_id, ExpenseAdjustment.business_id == user.business_id).order_by(ExpenseAdjustment.created_at.asc()).all()
    original_total = expense.amount
    final_total = original_total + sum(r.amount_delta for r in rows)
    return {
        "original": {"amount": original_total},
        "adjustments": [{"id": r.id, "amount_delta": r.amount_delta, "reason": r.reason, "created_by_name": r.created_by_name, "created_by_role": r.created_by_role, "created_at": to_utc_iso(r.created_at)} for r in rows],
        "final_total": final_total,
    }

# -----------------------------------------------------------------------------
# PRESENCE
# -----------------------------------------------------------------------------
@app.post("/presence/heartbeat")
def presence_heartbeat(payload: PresenceHeartbeatRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    session_id = payload.session_id or secrets.token_urlsafe(18)
    row = db.query(PresenceSession).filter(PresenceSession.session_id == session_id, PresenceSession.user_id == user.id).first()
    if not row: row = PresenceSession(session_id=session_id, business_id=user.business_id, user_id=user.id); db.add(row)
    row.last_seen_at = datetime.utcnow(); row.signed_out_at = None; db.commit(); return {"session_id": session_id}

@app.post("/presence/logout")
def presence_logout(payload: PresenceHeartbeatRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    session_id = payload.session_id
    if session_id:
        row = db.query(PresenceSession).filter(PresenceSession.session_id == session_id, PresenceSession.user_id == user.id).first()
        if row: row.signed_out_at = datetime.utcnow(); db.commit()
    return {"message": "Presence session ended."}

@app.get("/presence/team")
def team_presence(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Full employee roster (same directory as /users) merged with each
    employee's current/last known presence state. An employee's profile
    fields never disappear when they go offline — only the presence portion
    (online/online_since/last_seen_at) changes, and only when the backend
    actually has a presence record for them."""
    if user.role not in {"admin", "manager"}: raise HTTPException(status_code=403, detail="Access denied")
    cutoff = datetime.utcnow() - timedelta(minutes=2)
    q = db.query(User).filter(User.business_id == user.business_id, User.id != user.id)
    if user.role == "manager": q = q.filter(User.role == "staff")
    employees = q.order_by(User.id.asc()).all()

    # One most-recent PresenceSession per user (a new session_id is created per
    # browser session, so multiple rows can exist per user over time — only the
    # latest by last_seen_at reflects their current/last known presence).
    latest_by_user: dict = {}
    for row in db.query(PresenceSession).filter(PresenceSession.business_id == user.business_id).all():
        current = latest_by_user.get(row.user_id)
        if not current or row.last_seen_at > current.last_seen_at:
            latest_by_user[row.user_id] = row

    result = []
    for emp in employees:
        row = latest_by_user.get(emp.id)
        is_online = bool(row and row.signed_out_at is None and row.last_seen_at >= cutoff)
        result.append({
            **serialize_user(emp),
            "online": is_online,
            "online_since": to_utc_iso(row.signed_in_at) if is_online and row else None,
            "last_seen_at": to_utc_iso(row.last_seen_at) if row else None,
            # Set only when the session ended via an explicit sign-out; left null for
            # a session that simply went stale (timed out / tab closed), so the
            # frontend never has to guess or invent which of the two occurred.
            "signed_out_at": to_utc_iso(row.signed_out_at) if row and row.signed_out_at else None,
        })
    return {"employees": result}

# -----------------------------------------------------------------------------
# NOTIFICATIONS (notification center — see NOTIFICATION ENGINE above for the
# create/dedup/recipient logic; these endpoints only ever READ/mark rows this
# user is the recipient_user_id of, or write the two per-user side tables
# (preferences, push subscriptions) — never another user's rows.)
# -----------------------------------------------------------------------------
def sync_condition_driven_notifications(db: Session, business_id: int) -> None:
    """Refreshes the notification types that are conditions on current data
    rather than discrete write events (low stock / stockout / expiry) —
    called once per notification-center read, the same lazy-refresh trigger
    point the old alerts system used. Write-time hooks (checkout, stock
    adjustment) already fire these instantly when they happen; this is the
    safety net for a condition that changed without going through one of
    those hooks (e.g. min_stock_level itself was edited)."""
    for p in db.query(Product).filter(Product.business_id == business_id).all():
        check_inventory_notifications(db, business_id, p)
    check_expiry_notifications(db, business_id)
    db.commit()

@app.get("/notifications")
def list_notifications(unread_only: bool = Query(False), limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    sync_condition_driven_notifications(db, user.business_id)
    q = db.query(Notification).filter(Notification.recipient_user_id == user.id)
    if unread_only:
        q = q.filter(Notification.is_read == False)
    rows = q.order_by(Notification.created_at.desc()).offset(offset).limit(limit).all()
    unread_count = db.query(Notification).filter(Notification.recipient_user_id == user.id, Notification.is_read == False).count()
    return {"notifications": [serialize_notification(n) for n in rows], "unread_count": unread_count}

@app.get("/notifications/unread-count")
def notifications_unread_count(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return {"unread_count": db.query(Notification).filter(Notification.recipient_user_id == user.id, Notification.is_read == False).count()}

@app.post("/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    n = db.query(Notification).filter(Notification.id == notification_id, Notification.recipient_user_id == user.id).first()
    if not n: raise HTTPException(status_code=404, detail="Notification is unavailable.")
    if not n.is_read:
        n.is_read = True; n.read_at = datetime.utcnow(); db.commit()
    return {"message": "Notification marked as read."}

@app.post("/notifications/mark-all-read")
def mark_all_notifications_read(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    now = datetime.utcnow()
    updated = db.query(Notification).filter(Notification.recipient_user_id == user.id, Notification.is_read == False) \
        .update({Notification.is_read: True, Notification.read_at: now}, synchronize_session=False)
    db.commit()
    return {"message": "All notifications marked as read.", "updated": updated}

@app.get("/notifications/preferences")
def get_notification_preferences(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Only categories this user's ROLE can even receive are offered as
    toggles (section 16 — "do not allow ordinary employees to enable
    notifications for information they do not have permission to access").
    Mandatory categories (security, subscription) are never listed here —
    there is nothing to toggle."""
    allowed_categories = [c for c in NOTIFICATION_OPTIONAL_CATEGORIES if user.role in NOTIFICATION_CATEGORY_ROLES.get(c, ("admin",))]
    existing = {p.category: p.enabled for p in db.query(NotificationPreference).filter(NotificationPreference.user_id == user.id).all()}
    return {"preferences": [{"category": c, "enabled": existing.get(c, True)} for c in sorted(allowed_categories)]}

@app.put("/notifications/preferences")
def update_notification_preferences(payload: NotificationPreferenceUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Silently ignores any category the caller's role isn't eligible for and
    any category outside the optional allow-list — preferences can only ever
    narrow what a role already receives, never grant access to a category
    Cauldra's own authorization rules withhold from that role (section 16's
    'must never bypass Cauldra's authorization system')."""
    allowed_categories = {c for c in NOTIFICATION_OPTIONAL_CATEGORIES if user.role in NOTIFICATION_CATEGORY_ROLES.get(c, ("admin",))}
    applied = []
    for item in payload.preferences:
        if item.category not in allowed_categories:
            continue
        row = db.query(NotificationPreference).filter(NotificationPreference.user_id == user.id, NotificationPreference.category == item.category).first()
        if row:
            row.enabled = item.enabled
        else:
            db.add(NotificationPreference(user_id=user.id, category=item.category, enabled=item.enabled))
        applied.append(item.category)
    db.commit()
    return {"message": "Notification preferences updated.", "updated": applied}

# -----------------------------------------------------------------------------
# PUSH SUBSCRIPTIONS (Web Push registration — see deliver_push_notification())
# -----------------------------------------------------------------------------
@app.get("/push/vapid-public-key")
def get_vapid_public_key():
    """Public by design — this is the applicationServerKey the browser's own
    pushManager.subscribe() call needs; it grants no access to anything and
    is useless without also possessing VAPID_PRIVATE_KEY (server-only)."""
    if not VAPID_PUBLIC_KEY:
        raise HTTPException(status_code=503, detail="Push notifications are not configured on this server.")
    return {"public_key": VAPID_PUBLIC_KEY}

@app.post("/push/subscribe")
def subscribe_to_push(payload: PushSubscribeRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    existing = db.query(PushSubscription).filter(PushSubscription.endpoint == payload.endpoint).first()
    if existing:
        # The SAME browser endpoint re-subscribing (e.g. after a permission
        # reset) always reassigns to whichever account just subscribed it —
        # never left silently pointing at a previous, possibly different
        # user's account.
        existing.user_id = user.id; existing.business_id = user.business_id
        existing.p256dh = payload.keys.p256dh; existing.auth = payload.keys.auth
        existing.user_agent = payload.user_agent; existing.last_seen_at = datetime.utcnow(); existing.disabled_at = None
    else:
        db.add(PushSubscription(
            business_id=user.business_id, user_id=user.id, endpoint=payload.endpoint,
            p256dh=payload.keys.p256dh, auth=payload.keys.auth, user_agent=payload.user_agent,
        ))
    db.commit()
    return {"message": "Push notifications enabled."}

@app.post("/push/unsubscribe")
def unsubscribe_from_push(payload: PushUnsubscribeRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.query(PushSubscription).filter(PushSubscription.endpoint == payload.endpoint, PushSubscription.user_id == user.id).first()
    if row:
        db.delete(row); db.commit()
    return {"message": "Push notifications disabled."}

# -----------------------------------------------------------------------------
# PRICE MONITOR
# -----------------------------------------------------------------------------
@app.get("/price-monitor")
def price_monitor(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.query(PriceMonitorSource).filter(PriceMonitorSource.business_id == user.business_id).order_by(PriceMonitorSource.id.desc()).all()
    sources=[]
    for s in rows:
        product = db.query(Product).filter(Product.id == s.product_id).first() if s.product_id else None
        supplier = db.query(Supplier).filter(Supplier.id == s.supplier_id).first() if s.supplier_id else None
        hist = db.query(PriceHistory).filter(PriceHistory.source_id == s.id).order_by(PriceHistory.recorded_at.asc()).all()
        change = None
        if len(hist) >= 2 and hist[-2].price:
            change = round(((hist[-1].price - hist[-2].price)/hist[-2].price)*100, 2)
        sources.append({"id": s.id, "product_name": product.name if product else "Unknown product", "sku": product.sku if product else "", "supplier_name": supplier.name if supplier else "General Vendor", "source_type": s.source_type, "last_price": s.last_price, "change_percent": change, "history": [{"price": h.price, "recorded_at": to_utc_iso(h.recorded_at)} for h in hist]})
    return {"sources": sources}

@app.post("/price-monitor/sources")
def create_price_source(payload: PriceSourceCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    supplier_id = payload.supplier_id; product_id = payload.product_id; source_type = payload.source_type
    if not db.query(Supplier).filter(Supplier.id == supplier_id, Supplier.business_id == user.business_id).first(): raise HTTPException(status_code=404, detail="Supplier is unavailable.")
    if not db.query(Product).filter(Product.id == product_id, Product.business_id == user.business_id).first(): raise HTTPException(status_code=404, detail="Product is unavailable.")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    check_plan_limit(db, business, "price_monitor", db.query(PriceMonitorSource).filter(PriceMonitorSource.business_id == user.business_id).count())
    s=PriceMonitorSource(business_id=user.business_id, supplier_id=supplier_id, product_id=product_id, source_type=source_type, source_url=payload.source_url, last_price=payload.initial_price)
    db.add(s); db.commit(); return {"id": s.id, "message": "Price source added."}

@app.post("/price-monitor/{source_id}/price")
def manual_price_update(source_id: int, payload: ManualPriceUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    s=db.query(PriceMonitorSource).filter(PriceMonitorSource.id == source_id, PriceMonitorSource.business_id == user.business_id).first()
    if not s: raise HTTPException(status_code=404, detail="Price source is unavailable.")
    old_price = s.last_price
    price=payload.price; s.last_price=price; s.last_checked_at=datetime.utcnow(); db.add(PriceHistory(source_id=s.id, price=price))
    check_price_change_notification(db, s, old_price, price)
    db.commit(); return {"message":"Supplier price recorded."}

@app.post("/price-monitor/{source_id}/check")
def check_price_source(source_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    s=db.query(PriceMonitorSource).filter(PriceMonitorSource.id == source_id, PriceMonitorSource.business_id == user.business_id).first()
    if not s: raise HTTPException(status_code=404, detail="Price source is unavailable.")
    if s.source_type != "website": raise HTTPException(status_code=400, detail="Only website monitoring sources can be checked automatically.")
    # Safe behavior until a real website scraper/parser is configured: retain the last price.
    if s.last_price is None: raise HTTPException(status_code=422, detail="No supplier price has been recorded for this source yet.")
    s.last_checked_at=datetime.utcnow(); db.add(PriceHistory(source_id=s.id, price=s.last_price)); db.commit(); return {"message":"Supplier price checked successfully."}

@app.post("/price-monitor/upload-price-list")
def upload_price_list(payload: PriceListUploadRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # The browser sends a data URL, which is retained privately after validation.
    supplier_id=payload.supplier_id; product_id=payload.product_id
    if not db.query(Supplier).filter(Supplier.id == supplier_id, Supplier.business_id == user.business_id).first():
        raise HTTPException(status_code=404, detail="Supplier is unavailable.")
    if product_id and not db.query(Product).filter(Product.id == product_id, Product.business_id == user.business_id).first():
        raise HTTPException(status_code=404, detail="Product is unavailable.")
    file_name = safe_upload_name(payload.file_name, "price-list.csv")
    if Path(file_name).suffix.lower() != ".csv":
        raise HTTPException(status_code=422, detail="Please upload a CSV price list.")
    raw_bytes, content_type = decode_base64_upload(payload.file_data, {"text/csv", "application/vnd.ms-excel"})
    try:
        raw = raw_bytes.decode("utf-8-sig")
    except Exception as exc:
        raise HTTPException(status_code=422, detail="The price list could not be read.") from exc
    count=0
    for line in raw.splitlines()[1:]:
        parts=[p.strip() for p in line.split(",")]
        if len(parts) < 2: continue
        try: price=float(parts[-1])
        except: continue
        pid=int(product_id) if product_id else (int(parts[0]) if parts[0].isdigit() else None)
        if not pid: continue
        s=db.query(PriceMonitorSource).filter(PriceMonitorSource.business_id == user.business_id, PriceMonitorSource.supplier_id == supplier_id, PriceMonitorSource.product_id == pid).first()
        if not s:
            s=PriceMonitorSource(business_id=user.business_id, supplier_id=supplier_id, product_id=pid, source_type="price_list"); db.add(s); db.flush()
        s.last_price=price; s.last_checked_at=datetime.utcnow(); db.add(PriceHistory(source_id=s.id, price=price)); count += 1
    if not count:
        raise HTTPException(status_code=422, detail="No valid price rows were found in this CSV file.")
    upload = persist_upload(db, user, "price_list", file_name, content_type, raw_bytes)
    add_audit(db, user, "PRICE_LIST_UPLOADED", f"Uploaded and processed price list {upload.original_name}.")
    db.commit(); return {"count": count, "upload_id": upload.id}

# -----------------------------------------------------------------------------
# GENERAL CATALOG + BARCODE
#
# This endpoint belongs to the ADD PRODUCT workflow only:
#   scanned barcode -> own inventory duplicate check -> General Catalog
#   -> UPCitemdb (only on a catalog miss) -> cache identity -> caller fills
#   the Add Product form.
# New Sale / POS barcode scanning is a COMPLETELY SEPARATE, isolated chain
# (see resolveScannedProduct() in index.html) that matches only against the
# authenticated business's own already-loaded Product rows and never calls
# this endpoint, GeneralCatalog, or UPCitemdb — an item General Catalog or
# UPCitemdb recognizes is still "not in your current inventory" at the
# register unless this business actually stocks it. Do not wire POS scanning
# to this endpoint.
# -----------------------------------------------------------------------------
@app.post("/catalog/barcode-lookup")
def catalog_barcode_lookup(req: CatalogBarcodeLookupRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "staff": raise HTTPException(status_code=403, detail="Staff accounts cannot add products.")
    barcode = normalize_barcode(req.barcode)
    if not barcode:
        raise HTTPException(status_code=400, detail="Please scan or enter a valid barcode.")
    print(f"[barcode-flow] request received: {barcode}")

    # 1. Duplicate check — this business's OWN inventory only. Never reveals
    # whether any other business stocks this barcode (see spec: business
    # isolation must hold even when the barcode is otherwise recognized).
    existing = db.query(Product).filter(Product.business_id == user.business_id, Product.barcode == barcode).first()
    if existing:
        print("[barcode-flow] own inventory: HIT")
        print("[barcode-flow] final response source: own_inventory")
        return {
            "found": True, "source": "own_inventory", "barcode": barcode,
            "duplicate": True, "product_id": existing.id, "product_name": existing.name,
        }
    print("[barcode-flow] own inventory: MISS")

    # 2. Cauldra General Catalog — shared product IDENTITY only (see
    # GeneralCatalog docstring). Checked before ever considering UPCitemdb,
    # so a barcode already known to Cauldra (from any business's past
    # submission, or a past UPCitemdb hit) never costs a new external call.
    cataloged = lookup_general_catalog(db, barcode)
    if cataloged:
        print("[barcode-flow] general catalog: HIT")
        print("[barcode-flow] final response source: cauldra_catalog")
        return {
            "found": True, "source": "cauldra_catalog", "barcode": barcode,
            "product_name": cataloged.product_name, "brand": cataloged.brand, "size": cataloged.size,
        }
    print("[barcode-flow] general catalog: MISS")

    # 3. UPCitemdb — reached on EVERY General Catalog miss. A miss here is not
    # the end of the lookup, and a provider error/rate-limit here is NOT a
    # "catalog did not contain it" outcome — the two are reported separately
    # (source "not_found" vs "upcitemdb_unavailable") so the frontend can show
    # an accurate message.
    print("[barcode-flow] ENTERING UPCITEMDB FALLBACK")
    from upcitemdb_provider import lookup_upcitemdb_detailed
    print("[barcode-flow] UPCitemdb request started")
    upc = lookup_upcitemdb_detailed(barcode)
    print(f"[barcode-flow] UPCitemdb raw outcome: {upc['outcome'].upper()} ({upc['detail']})")

    if upc["outcome"] == "hit" and upc["identity"]:
        identity = upc["identity"]
        cached = upsert_general_catalog_identity(
            db, barcode, identity["product_name"], identity.get("brand"), identity.get("size"), source="upcitemdb",
        )
        try:
            db.commit()
        except Exception:
            db.rollback()
            raise
        print("[barcode-flow] final response source: upcitemdb")
        return {
            "found": True, "source": "upcitemdb", "barcode": barcode,
            "product_name": cached.product_name, "brand": cached.brand, "size": cached.size,
        }

    if upc["outcome"] in ("error", "rate_limited"):
        # The barcode service could not answer — DO NOT claim the catalog
        # lacked it. Manual entry is still offered, with an honest message.
        print(f"[barcode-flow] final response source: upcitemdb_unavailable ({upc['outcome']})")
        return {
            "found": False, "source": "upcitemdb_unavailable", "barcode": barcode,
            "upcitemdb_outcome": upc["outcome"], "upcitemdb_detail": upc["detail"],
            "manual_entry": True,
        }

    print("[barcode-flow] final response source: not_found")
    return {"found": False, "source": "not_found", "barcode": barcode, "manual_entry": True}

@app.get("/general-catalog")
def search_general_catalog(barcode: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Exact-barcode identity lookup only — General Catalog exists to help
    identify a product, not to provide a downloadable shared product
    database, so this deliberately requires a specific barcode instead of
    returning an unrestricted browseable list (see GeneralCatalog docstring;
    `category` is never returned — deprecated)."""
    normalized = normalize_barcode(barcode)
    if not normalized:
        raise HTTPException(status_code=400, detail="Please provide a valid barcode.")
    item = lookup_general_catalog(db, normalized)
    if not item:
        return {"found": False, "barcode": normalized}
    return {"found": True, "barcode": item.barcode, "product_name": item.product_name, "brand": item.brand, "size": item.size}

# -----------------------------------------------------------------------------
# AI
# -----------------------------------------------------------------------------
@app.get("/subscription/usage")
def subscription_usage(user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    # Read-only status/usage view. Deliberately NOT gated by require_subscription_access:
    # a business with no active entitlement yet (status="pending_payment_method",
    # "expired", or "cancelled") must still be able to see its own subscription
    # state so the Settings > Subscription & Billing screen can render and offer
    # Start Trial / Subscribe. Entitlement is still enforced independently on
    # every actual resource-mutating endpoint via get_current_user.
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    summary = usage_summary(db, business)
    plan = subscription_for(db, business)
    period_start, period_end, _ = billing_period_for(db, business)
    storage_used_bytes = db.query(func.coalesce(func.sum(StoredUpload.size_bytes), 0)).filter(StoredUpload.business_id == business.id).scalar() or 0
    po_this_period = db.query(PurchaseOrder).filter(PurchaseOrder.business_id == business.id, PurchaseOrder.created_at >= period_start, PurchaseOrder.created_at < period_end).count()
    role_counts = {role: db.query(User).filter(User.business_id == business.id, User.role == role).count() for role in ("admin", "manager", "staff")}
    resources = {
        "products": db.query(Product).filter(Product.business_id == business.id).count(),
        "suppliers": db.query(Supplier).filter(Supplier.business_id == business.id).count(),
        "warehouses": db.query(Warehouse).filter(Warehouse.business_id == business.id, Warehouse.is_active == True).count(),
        "users": sum(role_counts.values()),
        "price_monitor_sources": db.query(PriceMonitorSource).filter(PriceMonitorSource.business_id == business.id).count(),
        "purchase_orders": po_this_period,
        "storage_bytes": storage_used_bytes,
    }
    total_user_limit = None if all(plan.get(role) is None for role in ("admin", "manager", "staff")) else sum(plan.get(role) or 0 for role in ("admin", "manager", "staff"))
    summary["resources"] = {name: {"used": used, "limit": (total_user_limit if key == "users" else plan.get(key))} for name, used, key in [
        ("products", resources["products"], "product"), ("suppliers", resources["suppliers"], "supplier"), ("warehouses", resources["warehouses"], "warehouse"),
        ("users", resources["users"], "users"), ("price_monitor_sources", resources["price_monitor_sources"], "price_monitor"),
        ("purchase_orders", resources["purchase_orders"], "purchase_order"),
    ]}
    summary["resources"]["storage"] = {"used": round(storage_used_bytes / (1024**3), 3), "limit": plan["storage_gb"], "unit": "GB"}
    # Role breakdown lives at the top level, deliberately outside "resources" —
    # it's a nested {role: {used, limit}} shape, not a simple {used, limit}
    # usage card like every entry in resources, and must never be iterated
    # over as if it were one.
    summary["users_by_role"] = {role: {"used": role_counts[role], "limit": plan.get(role)} for role in ("admin", "manager", "staff")}
    summary["unlimited_people_and_locations"] = plan.get("admin") is None
    return summary

class ChangePlanRequest(BaseModel):
    plan: str
    billing_interval: Optional[str] = None

def plan_amount_naira(plan: str, interval: str) -> int:
    cfg = PLAN_CONFIG[plan]
    return int(cfg["annual_price" if interval == "annual" else "monthly_price"])

@app.post("/subscription/change-plan")
def change_plan(data: ChangePlanRequest, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can change the subscription plan.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "change-plan", f"business:{user.business_id}")
    check_rate_limit(db, "change-plan-ip", client_ip)
    plan = str(data.plan or "").strip().lower()
    if plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    interval = str(data.billing_interval or "").strip().lower()
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    if interval not in ("monthly", "annual"):
        interval = sub.billing_interval or "monthly"

    # SECURITY: this endpoint must never be usable to grant paid entitlements for
    # free. While a business is still on its unpaid trial, switching plans just
    # changes what will be billed later, so it's safe to apply immediately (and
    # must NOT reset the trial clock — see below). Once a business is an actual
    # paying customer (active/past_due), a plan+interval combination that costs
    # MORE than what they are currently paying must go through the paid checkout
    # flow instead, where the backend computes the authoritative price and
    # Paystack actually verifies payment. Only lateral/downgrade moves — which
    # cannot be used to obtain higher entitlements without paying — are allowed
    # here for paying customers.
    current_plan_id = (sub.plan or "starter").strip().lower()
    if sub.status not in ("trialing",):
        current_amount = plan_amount_naira(current_plan_id, sub.billing_interval or "monthly")
        new_amount = plan_amount_naira(plan, interval)
        if new_amount > current_amount:
            add_audit(db, user, "SUBSCRIPTION_PLAN_CHANGE_REJECTED", f"Blocked attempt to change to {PLAN_CONFIG[plan]['label']} ({interval}) without payment — current plan costs less than requested plan.", business_id=business.id)
            db.commit()
            record_failure(db, "change-plan", f"business:{user.business_id}")
            raise HTTPException(status_code=402, detail="Upgrading plans requires payment. Please use Subscribe / Upgrade to start checkout for the new plan.")
        # SECURITY: a genuine plan-tier downgrade (as opposed to a same-plan
        # interval change) must never apply immediately for a paying customer —
        # it has to be scheduled for the end of their current, already-paid
        # billing period instead. This endpoint stays reserved for trial
        # switches and same-tier interval moves; use /subscription/downgrade.
        if PLAN_RANK.get(plan, -1) < PLAN_RANK.get(current_plan_id, -1):
            raise HTTPException(status_code=409, detail="Downgrading to a lower plan takes effect at the end of your current billing period. Please use the downgrade option to schedule it.")

    # A downgrade must not silently strip data the business already has — surface
    # it as a clear, actionable error instead of a partial/confusing state.
    new_limits = PLAN_CONFIG[plan]
    role_counts = {role: db.query(User).filter(User.business_id == business.id, User.role == role).count() for role in ("admin", "manager", "staff")}
    checks = [
        ("admin", role_counts["admin"]), ("manager", role_counts["manager"]), ("staff", role_counts["staff"]),
        ("product", db.query(Product).filter(Product.business_id == business.id).count()),
        ("supplier", db.query(Supplier).filter(Supplier.business_id == business.id).count()),
        ("warehouse", db.query(Warehouse).filter(Warehouse.business_id == business.id, Warehouse.is_active == True).count()),
        ("price_monitor", db.query(PriceMonitorSource).filter(PriceMonitorSource.business_id == business.id).count()),
    ]
    for resource, current in checks:
        maximum = new_limits.get(resource)
        if maximum is not None and current > maximum:
            raise HTTPException(status_code=409, detail=f"Your current usage ({current} {resource.replace('_',' ')}) exceeds the {new_limits['label']} limit of {maximum}. Reduce usage before switching to this plan.")

    sub.plan = plan; sub.billing_interval = interval
    business.subscription_plan = plan; business.billing_interval = interval
    clear_pending_downgrade(sub)
    add_audit(db, user, "SUBSCRIPTION_PLAN_CHANGED", f"Subscription changed to {new_limits['label']} ({interval}).")
    db.commit()
    return usage_summary(db, business)

class DowngradeRequest(BaseModel):
    plan: str
    billing_interval: Optional[str] = None

@app.post("/subscription/downgrade")
def schedule_downgrade(data: DowngradeRequest, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Schedule a genuine plan-tier downgrade for the end of the current,
    already-paid billing period. Never applied immediately — see the
    downgrade-aware branch of the recurring-charge Paystack webhook, which is
    the only place `sub.plan` actually changes as a result of this."""
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can schedule a subscription downgrade.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "downgrade", f"business:{user.business_id}")
    check_rate_limit(db, "downgrade-ip", client_ip)

    plan = str(data.plan or "").strip().lower()
    if plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    interval = str(data.billing_interval or "").strip().lower()
    if interval not in ("monthly", "annual"):
        interval = sub.billing_interval or "monthly"

    if sub.status not in ("active", "past_due"):
        raise HTTPException(status_code=409, detail="Your subscription must be active to schedule a downgrade. While on trial, use the plan picker to switch directly — no scheduling is needed.")
    if not sub.current_period_end:
        raise HTTPException(status_code=409, detail="Your current billing period isn't available yet. Please try again shortly or contact support.")

    current_plan_id = (sub.plan or "starter").strip().lower()
    if PLAN_RANK.get(plan, -1) >= PLAN_RANK.get(current_plan_id, -1):
        raise HTTPException(status_code=400, detail="That isn't a downgrade. Choose a plan below your current one, or use Upgrade for a higher plan.")

    now = datetime.utcnow()
    # The server alone determines the effective date — always the trusted
    # subscription's own current_period_end, never anything the client sends.
    effective_at = sub.current_period_end
    current_interval = sub.billing_interval or "monthly"

    # Best-effort Paystack handoff, mirroring the existing upgrade flow exactly:
    # stop the current plan's recurring charge (so it can't also bill at the
    # boundary) and pre-arrange the new lower-plan subscription to start right
    # at the preserved boundary. A missing plan code or a failed Paystack call
    # never blocks the customer's recorded intent — it only means the switch
    # will need manual verification at the boundary, same tolerance already
    # used elsewhere in this file when a plan has no configured Paystack code.
    new_plan_code = PLAN_CONFIG[plan].get("paystack_annual_plan_code" if interval == "annual" else "paystack_monthly_plan_code")
    pending_paystack_code = None
    try:
        if sub.paystack_subscription_code:
            try:
                fetched = paystack_fetch_subscription(sub.paystack_subscription_code)
                email_token = fetched.get("email_token")
                if email_token:
                    paystack_disable_subscription(sub.paystack_subscription_code, email_token)
            except Exception:
                add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_OLD_PAYSTACK_DISABLE_FAILED", "Downgrade scheduled, but disabling the current recurring Paystack subscription failed. Verify manually to ensure it does not also charge at the boundary.", business_id=business.id)
        if new_plan_code and sub.paystack_authorization_code and sub.paystack_customer_code:
            created = paystack_create_subscription(sub.paystack_customer_code, new_plan_code, sub.paystack_authorization_code, start_date=effective_at)
            pending_paystack_code = created.get("subscription_code")
        else:
            add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_WITHOUT_PAYSTACK_PLAN_CODE", f"Downgrade to {PLAN_CONFIG[plan]['label']} ({interval}) scheduled, but no PAYSTACK_*_PLAN_CODE is configured for it, so automatic conversion cannot be scheduled with Paystack yet. The switch will need manual verification at the boundary.", business_id=business.id)
    except Exception:
        add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_PAYSTACK_SCHEDULE_FAILED", "Downgrade scheduled locally, but arranging the new recurring Paystack subscription failed. Automatic conversion at the boundary may not occur; verify manually.", business_id=business.id)

    sub.pending_downgrade_plan = plan
    sub.pending_downgrade_billing_interval = interval
    sub.pending_downgrade_effective_at = effective_at
    sub.pending_downgrade_requested_at = now
    sub.pending_downgrade_requested_by_user_id = user.id
    sub.pending_downgrade_paystack_subscription_code = pending_paystack_code
    add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_SCHEDULED",
              f"Downgrade scheduled: {PLAN_CONFIG[current_plan_id]['label']} ({current_interval}) -> {PLAN_CONFIG[plan]['label']} ({interval}), effective {to_utc_iso(effective_at)}.",
              business_id=business.id)
    db.commit()
    return usage_summary(db, business)

@app.post("/subscription/downgrade/cancel")
def cancel_pending_downgrade(request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can cancel a scheduled downgrade.")
    check_rate_limit(db, "downgrade-cancel", f"business:{user.business_id}")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    if not sub.pending_downgrade_plan:
        raise HTTPException(status_code=409, detail="There is no scheduled downgrade to cancel.")

    old_pending_plan, old_pending_interval = sub.pending_downgrade_plan, sub.pending_downgrade_billing_interval
    current_plan_id = (sub.plan or "starter").strip().lower()
    current_interval = sub.billing_interval or "monthly"

    # Undo the Paystack handoff: disable the pre-arranged lower-plan
    # subscription and restore normal auto-renewal for the CURRENT plan at the
    # same preserved boundary — no refund, no new billing period, just a
    # restoration of the subscription state to what it was before scheduling.
    try:
        if sub.pending_downgrade_paystack_subscription_code:
            try:
                fetched = paystack_fetch_subscription(sub.pending_downgrade_paystack_subscription_code)
                email_token = fetched.get("email_token")
                if email_token:
                    paystack_disable_subscription(sub.pending_downgrade_paystack_subscription_code, email_token)
            except Exception:
                add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_CANCEL_PAYSTACK_DISABLE_FAILED", "Scheduled downgrade cancelled locally, but disabling its pre-arranged Paystack subscription failed. Verify manually to ensure it does not also charge at the boundary.", business_id=business.id)
        current_plan_code = PLAN_CONFIG[current_plan_id].get("paystack_annual_plan_code" if current_interval == "annual" else "paystack_monthly_plan_code")
        if current_plan_code and sub.paystack_authorization_code and sub.paystack_customer_code and sub.current_period_end:
            created = paystack_create_subscription(sub.paystack_customer_code, current_plan_code, sub.paystack_authorization_code, start_date=sub.current_period_end)
            sub.paystack_subscription_code = created.get("subscription_code")
            sub.paystack_plan_code = current_plan_code
    except Exception:
        add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_CANCEL_PAYSTACK_RESCHEDULE_FAILED", "Scheduled downgrade cancelled, but restoring normal recurring billing for the current plan failed. Verify manually.", business_id=business.id)

    clear_pending_downgrade(sub)
    add_audit(db, user, "SUBSCRIPTION_DOWNGRADE_CANCELLED",
              f"Cancelled scheduled downgrade to {PLAN_CONFIG.get(old_pending_plan, {}).get('label', old_pending_plan)} ({old_pending_interval}). Continuing on {PLAN_CONFIG[current_plan_id]['label']} ({current_interval}).",
              business_id=business.id)
    db.commit()
    return usage_summary(db, business)

# -----------------------------------------------------------------------------
# PAYSTACK HELPERS
#
# Thin wrappers around the current officially-supported Paystack REST API.
# PAYSTACK_SECRET_KEY never leaves the server: it is only ever used here, in
# server-side request headers, and is never returned in any API response.
# -----------------------------------------------------------------------------
class PaystackRequestError(RuntimeError):
    """A provider rejection is safe to retry; a transport failure is not.

    `definitive` is true only when Paystack returned a readable negative
    response. A timeout/disconnect can happen after Paystack accepted a write,
    so callers must reconcile rather than blindly create the mutation again.
    """
    def __init__(self, message: str, *, definitive: bool):
        super().__init__(message)
        self.definitive = definitive


def paystack_request(method: str, path: str, json_body: Optional[dict] = None, timeout: int = 15) -> dict:
    import requests
    resp = requests.request(
        method, f"https://api.paystack.co{path}",
        headers={"Authorization": f"Bearer {PAYSTACK_SECRET_KEY}", "Content-Type": "application/json"},
        json=json_body, timeout=timeout,
    )
    try:
        payload = resp.json()
    except Exception:
        raise PaystackRequestError("Paystack returned an unreadable response.", definitive=False)
    if not resp.ok or not payload.get("status"):
        raise PaystackRequestError(payload.get("message", "Paystack rejected the request."), definitive=True)
    return payload

def paystack_verify_transaction(reference: str) -> dict:
    """Backend-authoritative check. We never trust a frontend 'payment successful'
    callback — every payment/authorization is independently re-verified here
    against Paystack's own transaction record before we act on it."""
    payload = paystack_request("GET", f"/transaction/verify/{reference}")
    return payload.get("data") or {}

def parse_paystack_metadata(value) -> dict:
    """Normalize Paystack metadata without ever treating malformed data as trusted."""
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}

def first_subscription_verification_mismatches(
    transaction: dict, record: PaymentRecord, business: Optional[BusinessProfile]
) -> list[str]:
    """Compare Paystack's server response with the checkout facts we own."""
    expected = parse_paystack_metadata(record.transaction_metadata)
    actual = parse_paystack_metadata(transaction.get("metadata"))
    mismatches = []
    if transaction.get("status") != "success":
        mismatches.append("status")
    if str(transaction.get("reference") or "") != record.paystack_reference:
        mismatches.append("reference")
    if int(transaction.get("amount") or 0) != record.amount_kobo:
        mismatches.append("amount")
    if (transaction.get("currency") or "").upper() != (record.currency or "").upper():
        mismatches.append("currency")
    if not str(transaction.get("id") or "").strip():
        mismatches.append("transaction_id")
    if not business or business.id != record.business_id:
        mismatches.append("business")

    expected_email = str(expected.get("customer_email") or (business.email if business else "") or "").strip().casefold()
    actual_email = str((transaction.get("customer") or {}).get("email") or "").strip().casefold()
    if not expected_email or actual_email != expected_email:
        mismatches.append("customer_email")

    expected_fields = {
        "business_id": record.business_id,
        "plan": record.plan,
        "billing_interval": record.billing_interval,
        "purpose": "subscription",
    }
    for key, expected_value in expected_fields.items():
        if str(expected.get(key)) != str(expected_value) or str(actual.get(key)) != str(expected_value):
            mismatches.append(f"metadata_{key}")
    return mismatches


def reconcile_first_subscription_payment(
    db: Session,
    record: PaymentRecord,
    business: BusinessProfile,
    verified_transaction: dict,
    now: Optional[datetime] = None,
) -> dict:
    """Apply one server-verified first-subscription payment exactly once.

    Both the signed Paystack webhook and the authenticated browser callback use
    this function. The browser therefore cannot become a second payment
    authority: it supplies only our reference, while every financial fact comes
    from the locked PaymentRecord and Paystack's verification response.
    """
    now = now or datetime.utcnow()
    if record.status == "success":
        return {"status": "success", "already_processed": True}

    provider_status = str(verified_transaction.get("status") or "").strip().lower()
    metadata = parse_paystack_metadata(record.transaction_metadata)
    metadata.update({
        "paystack_transaction_id": str(verified_transaction.get("id") or ""),
        "server_verified_at": now.replace(microsecond=0).isoformat() + "Z",
    })

    if provider_status != "success":
        if provider_status in {"failed", "abandoned", "reversed"}:
            transitioned = record.status != "failed"
            record.status = "failed"
            metadata["verification_provider_status"] = provider_status
            record.transaction_metadata = json.dumps(metadata, sort_keys=True)
            if transitioned:
                add_audit(
                    db, None, "SUBSCRIPTION_PAYMENT_FAILED",
                    f"Paystack transaction for reference {record.paystack_reference} is {provider_status}; subscription not activated.",
                    business_id=record.business_id,
                )
            db.flush()
            return {"status": "failed", "already_processed": False}

        # Paystack may still be processing immediately after the browser
        # returns. This is retryable and must not be converted into either a
        # successful subscription or a durable mismatch.
        record.status = "pending"
        metadata["verification_provider_status"] = provider_status or "pending"
        record.transaction_metadata = json.dumps(metadata, sort_keys=True)
        db.flush()
        return {"status": "pending", "already_processed": False}

    mismatches = first_subscription_verification_mismatches(verified_transaction, record, business)
    if mismatches:
        transitioned = record.status != "flagged_verification_mismatch"
        record.status = "flagged_verification_mismatch"
        metadata["verification_mismatches"] = sorted(set(mismatches))
        record.transaction_metadata = json.dumps(metadata, sort_keys=True)
        if transitioned:
            add_audit(
                db, None, "SUBSCRIPTION_PAYMENT_VERIFICATION_MISMATCH",
                f"Paystack server verification did not match the server-owned checkout for reference {record.paystack_reference}. Subscription not activated; payment flagged for review.",
                business_id=record.business_id,
            )
        db.flush()
        return {"status": "flagged_verification_mismatch", "already_processed": False,
                "verification_mismatches": sorted(set(mismatches))}

    metadata["verification_source"] = "paystack_verify_transaction"
    record.transaction_metadata = json.dumps(metadata, sort_keys=True)
    record.paystack_transaction_id = str(verified_transaction.get("id") or "") or None
    record.status = "success"
    record.paid_at = now

    sub = get_or_create_subscription(db, business, commit=False)
    sub.plan = record.plan
    sub.billing_interval = record.billing_interval
    sub.status = "active"
    sub.payment_status = "paid"
    sub.paid_at = now
    sub.current_period_start = now
    sub.current_period_end = add_billing_interval(now, record.billing_interval)
    sub.next_billing_at = sub.current_period_end
    sub.latest_transaction_reference = record.paystack_reference
    sub.cancel_at_period_end = False
    sub.cancelled_at = None
    sub.grace_period_ends_at = None
    clear_pending_downgrade(sub)
    business.subscription_plan = record.plan
    business.billing_interval = record.billing_interval
    add_audit(
        db, None, "SUBSCRIPTION_ACTIVATED",
        f"Payment verified with Paystack; subscription active on {PLAN_CONFIG[record.plan]['label']} ({record.billing_interval}).",
        business_id=business.id,
    )
    resolve_notifications(db, f"sub_expired:{business.id}", business.id)
    resolve_notifications(db, f"payment_failed:{business.id}", business.id)
    db.flush()
    return {"status": "success", "already_processed": False}

def paystack_get_or_create_customer(email: str) -> str:
    payload = paystack_request("POST", "/customer", {"email": email})
    return (payload.get("data") or {}).get("customer_code", "")

def paystack_create_subscription(customer_code: str, plan_code: str, authorization_code: str, start_date: Optional[datetime] = None) -> dict:
    body = {"customer": customer_code, "plan": plan_code, "authorization": authorization_code}
    if start_date:
        body["start_date"] = start_date.replace(microsecond=0).isoformat() + "Z"
    payload = paystack_request("POST", "/subscription", body)
    return payload.get("data") or {}

def paystack_disable_subscription(subscription_code: str, email_token: str):
    paystack_request("POST", "/subscription/disable", {"code": subscription_code, "token": email_token})

def paystack_fetch_subscription(subscription_code: str) -> dict:
    payload = paystack_request("GET", f"/subscription/{subscription_code}")
    return payload.get("data") or {}

def _apply_paystack_refund_state(row, provider_refund: dict, now: Optional[datetime] = None) -> str:
    """Persist provider truth without treating initiation as completion."""
    now = now or datetime.utcnow()
    provider_status = str(provider_refund.get("status") or "pending").strip().lower()
    provider_id = str(provider_refund.get("id") or "").strip() or None
    if provider_id:
        row.refund_provider_id = provider_id
    row.refund_provider_status = provider_status
    row.refund_updated_at = now
    if provider_status == "processed":
        row.refund_status = "succeeded"
        row.refunded_at = now
        row.refund_last_error = None
    elif provider_status == "failed":
        row.refund_status = "failed"
        row.refunded_at = None
        row.refund_last_error = "Paystack reported that the refund failed."
    else:
        # Paystack's pending, processing, and needs-attention states are all
        # unfinished. Never tell a customer that any of them is a refund.
        row.refund_status = "pending"
        row.refunded_at = None
        row.refund_last_error = None
    return row.refund_status


def _paystack_refund_result(row) -> dict:
    return {
        "refund_status": row.refund_status or "not_requested",
        "refund_provider_status": row.refund_provider_status,
        "refund_pending": (row.refund_status or "not_requested") == "pending",
    }


def ensure_verification_refund(db: Session, row, transaction_reference: str, provider_transaction_id) -> dict:
    """Initiate or reconcile one verification-charge refund safely.

    A DB compare-and-swap allows only one caller to initiate. If the outbound
    result is ambiguous, state stays pending and later calls only query Paystack;
    they never issue a second POST. A definitive rejection becomes failed and
    may be retried, but reconciliation is attempted first.
    """
    provider_tx_id = str(provider_transaction_id or row.paystack_transaction_id or "").strip()
    if provider_tx_id:
        row.paystack_transaction_id = provider_tx_id

    # Reconcile known provider work first. Fetch-by-refund-id is exact; if the
    # create response was lost, list-by-transaction-id discovers the accepted
    # refund without creating another one.
    try:
        provider_refund = None
        if row.refund_provider_id:
            payload = paystack_request("GET", f"/refund/{row.refund_provider_id}")
            provider_refund = payload.get("data") or {}
        elif provider_tx_id.isdigit() and row.refund_status in {"pending", "failed"}:
            payload = paystack_request("GET", f"/refund?transaction={provider_tx_id}&perPage=50")
            candidates = payload.get("data") or []
            matching = [item for item in candidates if int(item.get("amount") or 0) == int(row.amount_kobo or 0)]
            if matching:
                provider_refund = matching[0]
        if provider_refund:
            _apply_paystack_refund_state(row, provider_refund)
            db.commit()
            return _paystack_refund_result(row)
    except Exception:
        # Reconciliation unavailability is not evidence of failure. Preserve
        # the last durable state and, critically, do not duplicate a pending
        # provider write.
        db.rollback()
        row = db.query(type(row)).filter(type(row).id == row.id).first()

    if row.refund_status == "succeeded" or row.refund_status == "pending":
        return _paystack_refund_result(row)

    now = datetime.utcnow()
    claimed = (
        db.query(type(row))
        .filter(type(row).id == row.id, type(row).refund_status.in_(["not_requested", "failed"]))
        .update({
            "refund_status": "pending", "refund_provider_status": "initiating",
            "refund_attempt_count": type(row).refund_attempt_count + 1,
            "refund_requested_at": now, "refund_updated_at": now,
            "refund_last_error": None, "paystack_transaction_id": provider_tx_id or None,
        }, synchronize_session=False)
    )
    db.commit()
    db.refresh(row)
    if claimed == 0:
        return _paystack_refund_result(row)

    try:
        payload = paystack_request("POST", "/refund", {
            "transaction": transaction_reference,
            "amount": int(row.amount_kobo),
            "currency": "NGN",
            "customer_note": "Refund of Cauldra card verification charge",
            "merchant_note": "Automated card verification refund",
        })
        _apply_paystack_refund_state(row, payload.get("data") or {})
    except PaystackRequestError as exc:
        row.refund_updated_at = datetime.utcnow()
        row.refunded_at = None
        if exc.definitive:
            row.refund_status = "failed"
            row.refund_provider_status = "failed"
            row.refund_last_error = "Paystack rejected the refund request."
        else:
            row.refund_status = "pending"
            row.refund_provider_status = "status_unknown"
            row.refund_last_error = "Refund outcome is awaiting reconciliation."
    except Exception:
        row.refund_status = "pending"
        row.refund_provider_status = "status_unknown"
        row.refund_updated_at = datetime.utcnow()
        row.refunded_at = None
        row.refund_last_error = "Refund outcome is awaiting reconciliation."
    db.commit()
    return _paystack_refund_result(row)


# -----------------------------------------------------------------------------
# EMAIL VERIFICATION BEFORE PAYSTACK (new-business onboarding)
#
# The person registering a business must prove they control the email address
# BEFORE Cauldra initialises any Paystack transaction for them. This reuses the
# project's already-configured Supabase project (SUPABASE_URL / SUPABASE_SECRET_KEY
# via supabase_client.py) and Supabase Auth's own magic-link email plus its
# user.email_confirmed_at state. No new table, no new column, no second email
# provider, no custom verification codes: Supabase sends the email and Supabase
# is the ONLY source of truth this code trusts. It sits strictly between the
# "enter email" step and /onboarding/payment/init — the rest of onboarding,
# Paystack, payment verification and business creation are untouched.
# -----------------------------------------------------------------------------
# Where Supabase should send the guest after they click the verification link.
# Must also be added to the Supabase project's Auth "Redirect URLs" allow-list.
SUPABASE_EMAIL_REDIRECT_URL = os.getenv("SUPABASE_EMAIL_REDIRECT_URL", "").strip().rstrip("/")
ONBOARDING_EMAIL_RESEND_SECONDS = int(os.getenv("ONBOARDING_EMAIL_RESEND_SECONDS", "60"))

def _supabase_auth_or_503():
    """The trusted server-side Supabase Auth client. 503 (never a silent pass)
    when Supabase is not configured — email verification is a hard requirement."""
    try:
        client = get_supabase_client(required=False)
    except SupabaseConfigurationError as exc:
        raise HTTPException(status_code=503, detail="Email verification is not available right now. Please contact support.") from exc
    if client is None:
        raise HTTPException(status_code=503, detail="Email verification is not available right now. Please contact support.")
    return client.auth

def _supabase_email_verify_redirect(request: Request, plan: str, interval: str) -> str:
    base = (SUPABASE_EMAIL_REDIRECT_URL or SUPPLY_AI_FRONTEND_URL
            or str(paystack_callback_url(request)).rstrip("/"))
    from urllib.parse import urlencode
    # plan/interval are public catalogue values, never sensitive data.
    return base.rstrip("/") + "/?" + urlencode({"cauldra_email_verify": "1", "ev_plan": plan, "ev_interval": interval})

def _supabase_email_confirmed(email: str) -> Optional[bool]:
    """True / False from Supabase's own user record for this email; None if
    Supabase has never seen it. Uses admin.generate_link purely as an O(1)
    lookup-by-email — it returns the user record and does NOT send an email."""
    email_l = (email or "").strip().lower()
    if not email_l:
        return None
    auth = _supabase_auth_or_503()
    try:
        resp = auth.admin.generate_link({"type": "magiclink", "email": email_l})
    except Exception as exc:
        msg = str(getattr(exc, "message", "") or exc).lower()
        if any(m in msg for m in ("not found", "user_not_found", "no user", "does not exist")):
            return None
        raise HTTPException(status_code=502, detail="We couldn't check your verification status right now. Please try again.") from exc
    user = getattr(resp, "user", None)
    if user is None:
        return None
    return bool(getattr(user, "email_confirmed_at", None) or getattr(user, "confirmed_at", None))

def _validated_onboarding_plan_interval(plan, interval):
    p = str(plan or "").strip().lower()
    if p not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    i = str(interval or "monthly").strip().lower()
    return p, (i if i in ("monthly", "annual") else "monthly")

class OnboardingEmailVerifyRequest(BaseModel):
    email: EmailStr
    plan: str
    billing_interval: str = "monthly"

@app.post("/onboarding/email/verify")
def onboarding_email_verify(data: OnboardingEmailVerifyRequest, request: Request, db: Session = Depends(get_db)):
    """Send (or resend) a Supabase magic-link verification email to the address
    the guest wants to register/pay with. This NEVER claims the email is
    verified — it only asks Supabase to send its own email. Calling it again is
    a resend; Supabase enforces its own per-email cooldown and that error is
    surfaced as a friendly 'please wait' message."""
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "onboarding-email-verify-ip", client_ip)
    email_l = str(data.email).strip().lower()
    check_rate_limit(db, "onboarding-email-verify-email", email_l)
    plan, interval = _validated_onboarding_plan_interval(data.plan, data.billing_interval)

    # Already verified in Supabase -> do not send another email; let them continue.
    if _supabase_email_confirmed(email_l) is True:
        return {"status": "verified", "email": email_l, "plan": plan, "billing_interval": interval}

    auth = _supabase_auth_or_503()
    redirect = _supabase_email_verify_redirect(request, plan, interval)
    try:
        auth.sign_in_with_otp({
            "email": email_l,
            "options": {"email_redirect_to": redirect, "should_create_user": True},
        })
    except Exception as exc:
        detail = str(getattr(exc, "message", "") or exc).lower()
        if any(m in detail for m in ("rate", "limit", "too many", "seconds", "429")):
            record_failure(db, "onboarding-email-verify-email", email_l)
            raise HTTPException(status_code=429, detail="Please wait a little before requesting another verification email.") from exc
        raise HTTPException(status_code=502, detail="We couldn't send the verification email right now. Please try again.") from exc
    return {"status": "sent", "email": email_l, "plan": plan, "billing_interval": interval,
            "resend_after_seconds": ONBOARDING_EMAIL_RESEND_SECONDS}

class OnboardingEmailConfirmRequest(BaseModel):
    access_token: str = ""
    email: Optional[EmailStr] = None

@app.post("/onboarding/email/verify/confirm")
def onboarding_email_verify_confirm(data: OnboardingEmailConfirmRequest, request: Request, db: Session = Depends(get_db)):
    """Called when the guest returns from the Supabase link. Verification is
    decided ONLY from trusted Supabase state:
      * if a Supabase session access token is supplied (from the link's URL
        fragment), it is validated WITH Supabase (auth.get_user) and the
        confirmed email is read from the returned user record;
      * the email is then re-checked against Supabase's stored user state.
    A URL parameter such as ?verified=true, or a frontend boolean, is never
    trusted. The frontend-supplied email is only ever a lookup key against
    Supabase, never proof on its own."""
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "onboarding-email-confirm-ip", client_ip)
    token = str(data.access_token or "").strip()
    auth = _supabase_auth_or_503()

    token_email = None
    if token:
        try:
            resp = auth.get_user(token)
            user = getattr(resp, "user", None)
        except Exception:
            user = None
        if user is not None and getattr(user, "email", None):
            confirmed = getattr(user, "email_confirmed_at", None) or getattr(user, "confirmed_at", None)
            if confirmed:
                token_email = str(user.email).strip().lower()

    check_email = (str(data.email).strip().lower() if data.email else "") or token_email
    if not check_email:
        raise HTTPException(status_code=400, detail="We couldn't confirm this verification link. Please request a new verification email.")

    if _supabase_email_confirmed(check_email) is not True:
        raise HTTPException(status_code=400, detail="We couldn't confirm this verification link. It may have expired. Please request a new verification email.")
    if token_email and token_email != check_email:
        raise HTTPException(status_code=400, detail="The verified email does not match the email you entered. Please verify the correct address.")
    return {"status": "verified", "email": check_email}

# -----------------------------------------------------------------------------
# NEW-BUSINESS ONBOARDING: PAY-BEFORE-REGISTER
#
# Flow: Choose Plan -> verify card with Paystack (no business/user exists yet)
# -> Register Your Business -> 14-day trial starts immediately using the
# already-verified card -> Enter Cauldra.
#
# This intentionally does NOT reuse /subscription/trial/init + /confirm as-is,
# because those require get_authenticated_user (an existing business/user) —
# there is no business yet at this point. It DOES reuse every other piece of
# the existing architecture: the same paystack_* helpers above, the same
# PLAN_CONFIG pricing/trial-day source of truth, the same independent
# backend-side Paystack re-verification pattern, and the same
# card_verified/authorization_code/card_last4/card_type fields already on
# BusinessSubscription. /subscription/trial/init + /confirm are left fully
# intact for any business still sitting in "pending_payment_method" (e.g.
# older registrations) — the two flows don't interfere with each other.
# -----------------------------------------------------------------------------
class OnboardingPaymentInitRequest(BaseModel):
    email: EmailStr
    plan: str
    billing_interval: str = "monthly"

@app.post("/onboarding/payment/init")
def onboarding_payment_init(data: OnboardingPaymentInitRequest, request: Request, db: Session = Depends(get_db)):
    """Step 1: begin card verification for a guest who has not registered a
    business yet. Only ever charges the same small, refundable verification
    amount used by the existing trial flow — never the plan's actual price."""
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "onboarding-payment-init-ip", client_ip)
    check_rate_limit(db, "onboarding-payment-init-email", str(data.email).strip().lower())
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured yet. Please contact support.")

    # Backend-authoritative validation — the frontend can never smuggle an
    # arbitrary plan/interval (or amount) into this flow.
    plan = str(data.plan or "").strip().lower()
    if plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    interval = str(data.billing_interval or "monthly").strip().lower()
    if interval not in ("monthly", "annual"):
        interval = "monthly"

    # GATE: the email must be verified through Supabase BEFORE any Paystack
    # transaction is initialised. Checked against Supabase's own user state on
    # every call, so this also blocks a direct API request that skips the UI.
    if _supabase_email_confirmed(str(data.email).strip().lower()) is not True:
        raise HTTPException(status_code=403, detail="Please verify your email address before continuing to payment.")

    amount_kobo = PAYSTACK_TRIAL_VERIFICATION_AMOUNT_KOBO
    reference = f"cauldra_onboard_{secrets.token_hex(10)}"
    now = datetime.utcnow()
    row = OnboardingAuthorization(
        paystack_reference=reference, email=str(data.email), plan=plan, billing_interval=interval,
        amount_kobo=amount_kobo, status="initialized",
        expires_at=now + timedelta(minutes=ONBOARDING_PAYMENT_SESSION_MINUTES),
    )
    db.add(row); db.commit()

    try:
        payload = paystack_request("POST", "/transaction/initialize", {
            "email": str(data.email), "amount": amount_kobo, "currency": "NGN", "reference": reference,
            "channels": ["card"], "callback_url": paystack_callback_url(request),
            "metadata": {"plan": plan, "billing_interval": interval, "purpose": "onboarding_card_verification"},
        })
    except Exception:
        row.status = "failed"; db.commit()
        record_failure(db, "onboarding-payment-init-ip", client_ip)
        raise HTTPException(status_code=502, detail="We couldn't start card verification right now. Please try again.")

    return {"authorization_url": payload["data"]["authorization_url"], "reference": reference,
            "verification_amount_kobo": amount_kobo, "plan": plan, "billing_interval": interval}


class OnboardingPaymentConfirmRequest(BaseModel):
    reference: str

@app.post("/onboarding/payment/confirm")
def onboarding_payment_confirm(data: OnboardingPaymentConfirmRequest, request: Request, db: Session = Depends(get_db)):
    """Step 2: the frontend can only report that Paystack's checkout finished —
    it can never decide verification succeeded. This independently re-verifies
    the transaction and the resulting authorization with Paystack itself before
    marking anything as verified, exactly like /subscription/trial/confirm."""
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "onboarding-payment-confirm-ip", client_ip)
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured.")

    reference = str(data.reference or "").strip()
    row = db.query(OnboardingAuthorization).filter(OnboardingAuthorization.paystack_reference == reference).first()
    if not row:
        # Never trust a reference the backend doesn't already own a record for.
        raise HTTPException(status_code=404, detail="We couldn't find that payment verification attempt.")

    # Idempotent: replaying confirm (double click, retried request, replayed
    # callback) after verification already succeeded must not re-run any of this.
    if row.status == "verified":
        refund = ensure_verification_refund(db, row, reference, row.paystack_transaction_id)
        return {"status": "verified", "plan": row.plan, "billing_interval": row.billing_interval,
                "card_last4": row.card_last4, "card_type": row.card_type,
                "already_processed": True, **refund}
    if row.status == "consumed":
        raise HTTPException(status_code=409, detail="This payment verification has already been used to register a business.")
    if datetime.utcnow() > row.expires_at:
        row.status = "failed"; db.commit()
        raise HTTPException(status_code=410, detail="This payment verification session has expired. Please select your plan again.")

    try:
        tx = paystack_verify_transaction(reference)
    except Exception:
        raise HTTPException(status_code=502, detail="We couldn't verify that payment method with Paystack right now. Please try again.")

    if (tx.get("status") != "success" or int(tx.get("amount") or 0) != row.amount_kobo
            or (tx.get("currency") or "").upper() != "NGN" or tx.get("reference") != reference):
        row.status = "failed"; db.commit()
        record_failure(db, "onboarding-payment-confirm-ip", client_ip)
        raise HTTPException(status_code=400, detail="Payment method verification did not succeed. Please try again with a valid card.")

    authorization = tx.get("authorization") or {}
    if not authorization.get("reusable") or authorization.get("channel") != "card" or not authorization.get("authorization_code"):
        # This specific authorization can't be relied on for future recurring
        # charges — never promise automatic billing the backend can't deliver.
        row.status = "failed"; db.commit()
        raise HTTPException(status_code=400, detail="This card can't be used for automatic recurring billing. Please try again with a different card.")

    # Atomic compare-and-swap: if two confirm requests race, only the first
    # UPDATE (matching status = 'initialized') affects a row; a second racing
    # request affects zero rows and is treated as already-processed.
    rows_updated = (
        db.query(OnboardingAuthorization)
        .filter(OnboardingAuthorization.id == row.id, OnboardingAuthorization.status == "initialized")
        .update({
            "status": "verified", "verified_at": datetime.utcnow(),
            "paystack_transaction_id": str(tx.get("id") or "") or None,
            "paystack_authorization_code": authorization.get("authorization_code"),
            "card_last4": authorization.get("last4"), "card_type": authorization.get("card_type"),
            "card_exp_month": str(authorization.get("exp_month") or ""), "card_exp_year": str(authorization.get("exp_year") or ""),
        }, synchronize_session=False)
    )
    db.commit()
    db.refresh(row)
    if rows_updated == 0:
        return {"status": row.status, "plan": row.plan, "billing_interval": row.billing_interval, "already_processed": True}

    try:
        row.paystack_customer_code = paystack_get_or_create_customer(row.email)
        db.commit()
    except Exception:
        # Non-fatal — registration re-attempts this lookup if it's still missing.
        pass

    refund = ensure_verification_refund(db, row, reference, tx.get("id"))

    # Card brand/last4 only — never the authorization_code, which stays
    # server-side and is only ever used in outbound Paystack API calls.
    return {"status": "verified", "plan": row.plan, "billing_interval": row.billing_interval,
            "card_last4": row.card_last4, "card_type": row.card_type, **refund}


class TrialInitRequest(BaseModel):
    plan: Optional[str] = None
    billing_interval: Optional[str] = None

@app.post("/subscription/trial/init")
def trial_init(data: TrialInitRequest, request: Request, user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    """Step 1 of the card-required trial. Charges a small, refundable card-
    tokenization amount so Paystack can authenticate the card via 2FA and return
    a reusable authorization — this is Paystack's own currently-supported route
    to a recurring-billing-ready authorization; Paystack does not guarantee a
    true zero-amount charge succeeds across all banks/card brands."""
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can start a subscription trial.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "trial-init", f"business:{user.business_id}")
    check_rate_limit(db, "trial-init-ip", client_ip)
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured yet. Please contact support to start your trial.")

    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)

    # Authoritative one-trial-per-business decision: based on the business's own
    # server-side subscription record, not on any client-supplied signal. Once a
    # business has ever been card_verified or has trial_start_at set, it is
    # permanently ineligible for another introductory trial regardless of how
    # many users, sessions, devices, or emails are involved.
    if sub.card_verified or sub.trial_start_at is not None:
        add_audit(db, user, "TRIAL_DENIED_REPEAT_ATTEMPT", "Trial start blocked: this business has already used its one introductory trial.", business_id=business.id)
        db.commit()
        record_failure(db, "trial-init", f"business:{user.business_id}")
        raise HTTPException(status_code=409, detail="This business has already used its introductory trial. Please subscribe directly to continue.")
    if sub.status not in ("pending_payment_method",):
        raise HTTPException(status_code=409, detail="A trial can't be started for this subscription's current status.")

    plan = str(data.plan or sub.plan or "starter").strip().lower()
    if plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    interval = str(data.billing_interval or sub.billing_interval or "monthly").strip().lower()
    if interval not in ("monthly", "annual"):
        interval = "monthly"

    amount_kobo = PAYSTACK_TRIAL_VERIFICATION_AMOUNT_KOBO
    reference = f"cauldra_trialcard_{business.business_code}_{secrets.token_hex(8)}"
    record = PaymentRecord(business_id=business.id, subscription_id=sub.id, plan=plan, billing_interval=interval,
                            amount_kobo=amount_kobo, currency="NGN", paystack_reference=reference, status="initialized",
                            purpose="card_verification",
                            transaction_metadata=json.dumps({"business_id": business.id, "plan": plan, "billing_interval": interval, "purpose": "trial_card_verification"}))
    db.add(record); db.commit()

    try:
        payload = paystack_request("POST", "/transaction/initialize", {
            "email": business.email or user.email, "amount": amount_kobo, "currency": "NGN", "reference": reference,
            "channels": ["card"], "callback_url": paystack_callback_url(request),
            "metadata": {"business_id": business.id, "plan": plan, "billing_interval": interval, "purpose": "trial_card_verification"},
        })
    except Exception:
        record.status = "failed"; db.commit()
        raise HTTPException(status_code=502, detail="We couldn't start card verification right now. Please try again.")

    add_audit(db, user, "TRIAL_CARD_VERIFICATION_INITIATED", f"Started card verification for a {PLAN_CONFIG[plan]['label']} ({interval}) trial.", business_id=business.id)
    db.commit()
    return {"authorization_url": payload["data"]["authorization_url"], "reference": reference, "verification_amount_kobo": amount_kobo, "plan": plan, "billing_interval": interval}


class TrialConfirmRequest(BaseModel):
    reference: str

@app.post("/subscription/trial/confirm")
def trial_confirm(data: TrialConfirmRequest, request: Request, user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    """Step 2 of the card-required trial. The frontend can only report that
    Paystack's popup finished — it can NEVER decide the trial has started. This
    endpoint independently re-verifies the transaction and authorization with
    Paystack before granting anything."""
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can confirm a subscription trial.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "trial-confirm", f"business:{user.business_id}")
    check_rate_limit(db, "trial-confirm-ip", client_ip)
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured.")

    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    reference = str(data.reference or "").strip()
    record = db.query(PaymentRecord).filter(PaymentRecord.paystack_reference == reference, PaymentRecord.business_id == business.id, PaymentRecord.purpose == "card_verification").first()
    if not record:
        # Never trust a client-supplied reference we don't already own for this business.
        raise HTTPException(status_code=404, detail="We couldn't find that verification attempt for your business.")

    # Idempotent: replaying confirm (double-click, retried request, replayed call)
    # after a trial already started must not re-run any of this.
    if sub.card_verified or sub.status == "trialing":
        refund = ensure_verification_refund(db, record, reference, record.paystack_transaction_id)
        return {"status": sub.status, "plan": sub.plan, "billing_interval": sub.billing_interval,
                "trial_end_at": to_utc_iso(sub.trial_end_at), "already_processed": True, **refund}
    if record.status == "success":
        return {"status": sub.status, "already_processed": True}

    try:
        tx = paystack_verify_transaction(reference)
    except Exception:
        raise HTTPException(status_code=502, detail="We couldn't verify that payment method with Paystack right now. Please try again.")

    expected_meta = json.loads(record.transaction_metadata or "{}")
    if (tx.get("status") != "success" or int(tx.get("amount") or 0) != record.amount_kobo
            or (tx.get("currency") or "").upper() != "NGN" or tx.get("reference") != reference):
        record.status = "failed"; db.commit()
        raise HTTPException(status_code=400, detail="Payment method verification did not succeed. Please try again with a valid card.")

    authorization = tx.get("authorization") or {}
    if not authorization.get("reusable") or authorization.get("channel") != "card" or not authorization.get("authorization_code"):
        # Paystack's own supported flow says this specific authorization can't be
        # relied on for future recurring charges — never promise automatic billing
        # the backend can't actually deliver.
        record.status = "failed"; db.commit()
        add_audit(db, user, "TRIAL_DENIED_UNUSABLE_AUTHORIZATION", "Card verification succeeded but the resulting authorization is not valid for recurring billing.", business_id=business.id)
        db.commit()
        raise HTTPException(status_code=400, detail="This card can't be used for automatic recurring billing. Please try again with a different card.")

    plan = str(expected_meta.get("plan") or sub.plan or "starter").strip().lower()
    if plan not in PLAN_CONFIG:
        plan = "starter"
    interval = str(expected_meta.get("billing_interval") or sub.billing_interval or "monthly").strip().lower()
    if interval not in ("monthly", "annual"):
        interval = "monthly"

    # Atomic compare-and-swap: if two confirm requests race, only the first
    # UPDATE (matching card_verified = 0) affects a row; the second affects zero
    # rows and is treated as already-processed rather than granting a 2nd trial.
    now = datetime.utcnow()
    trial_end = now + timedelta(days=PLAN_CONFIG[plan]["trial_days"])
    # Atomic compare-and-swap via a conditional UPDATE routed through the ORM's
    # own column types (so datetime/bool binding is handled correctly by
    # SQLAlchemy rather than the raw DBAPI): the WHERE clause only matches a row
    # that is still un-verified, so if two confirm requests race, only the first
    # UPDATE affects a row — the second affects zero rows and is treated as
    # already-processed rather than granting a second trial.
    rows_updated = (
        db.query(BusinessSubscription)
        .filter(BusinessSubscription.id == sub.id, BusinessSubscription.card_verified == False)
        .update({
            "card_verified": True, "status": "trialing", "plan": plan, "billing_interval": interval,
            "trial_start_at": now, "trial_end_at": trial_end, "current_period_start": now,
            "current_period_end": trial_end, "next_billing_at": trial_end, "trial_consent_at": now,
            "paystack_authorization_code": authorization.get("authorization_code"),
            "card_last4": authorization.get("last4"), "card_type": authorization.get("card_type"),
            "card_exp_month": str(authorization.get("exp_month") or ""), "card_exp_year": str(authorization.get("exp_year") or ""),
            "payment_status": None, "updated_at": now,
        }, synchronize_session=False)
    )
    db.commit()
    if rows_updated == 0:
        # Lost the race (or was already processed) — report current state, don't retry.
        db.refresh(sub)
        return {"status": sub.status, "already_processed": True}
    db.refresh(sub)

    record.status = "success"; record.paid_at = now
    record.paystack_transaction_id = str(tx.get("id") or "") or None
    business.subscription_plan = plan; business.billing_interval = interval

    customer_code = None
    subscription_code = None
    plan_code = PLAN_CONFIG[plan].get("paystack_annual_plan_code" if interval == "annual" else "paystack_monthly_plan_code")
    try:
        customer_code = paystack_get_or_create_customer(business.email or user.email)
        if plan_code:
            # Paystack's supported start_date parameter lets Paystack itself
            # schedule and execute the first real debit automatically at trial
            # end — Cauldra does not need to invent its own card-charging
            # mechanism for the conversion charge.
            created = paystack_create_subscription(customer_code, plan_code, authorization["authorization_code"], start_date=trial_end)
            subscription_code = created.get("subscription_code")
        else:
            add_audit(db, user, "TRIAL_STARTED_WITHOUT_PAYSTACK_PLAN_CODE", f"Trial started for {PLAN_CONFIG[plan]['label']} ({interval}) but no PAYSTACK_*_PLAN_CODE is configured for this plan/interval, so automatic conversion cannot be scheduled with Paystack yet. The trial will still expire correctly and require manual checkout to convert.", business_id=business.id)
    except Exception:
        add_audit(db, user, "TRIAL_PAYSTACK_SUBSCRIPTION_CREATE_FAILED", "Card verification succeeded and the trial started, but creating the recurring Paystack subscription failed. Automatic conversion at trial end may not occur; manual checkout will be needed.", business_id=business.id)

    sub.paystack_customer_code = customer_code
    sub.paystack_subscription_code = subscription_code
    sub.paystack_plan_code = plan_code or None
    db.commit()

    refund = ensure_verification_refund(db, record, reference, tx.get("id"))
    if refund["refund_status"] == "succeeded":
        refund_message = "Verification amount refund succeeded."
        add_audit(db, user, "TRIAL_VERIFICATION_REFUND_SUCCEEDED", refund_message, business_id=business.id)
    elif refund["refund_status"] == "failed":
        refund_message = "Verification amount refund failed and requires a safe retry."
        add_audit(db, user, "TRIAL_VERIFICATION_REFUND_FAILED", refund_message, business_id=business.id)
    else:
        refund_message = "Verification amount refund is pending provider processing."
        add_audit(db, user, "TRIAL_VERIFICATION_REFUND_PENDING", refund_message, business_id=business.id)
    add_audit(db, user, "TRIAL_STARTED", f"14-day free trial started on {PLAN_CONFIG[plan]['label']} ({interval}). {refund_message}", business_id=business.id)
    db.commit()

    return {"status": "trialing", "plan": plan, "billing_interval": interval,
            "trial_start_at": to_utc_iso(now), "trial_end_at": to_utc_iso(trial_end), **refund}


@app.post("/subscription/trial/cancel")
def trial_cancel(request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can cancel the trial.")
    check_rate_limit(db, "trial-cancel", f"business:{user.business_id}")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    if sub.status != "trialing":
        raise HTTPException(status_code=409, detail="There is no active trial to cancel.")
    if sub.paystack_subscription_code:
        try:
            fetched = paystack_fetch_subscription(sub.paystack_subscription_code)
            email_token = fetched.get("email_token")
            if email_token:
                paystack_disable_subscription(sub.paystack_subscription_code, email_token)
        except Exception:
            add_audit(db, user, "TRIAL_CANCEL_PAYSTACK_DISABLE_FAILED", "Trial cancelled locally, but disabling the linked Paystack subscription failed. Verify manually to ensure no future charge occurs.", business_id=business.id)
    sub.status = "cancelled"; sub.cancelled_at = datetime.utcnow()
    add_audit(db, user, "TRIAL_CANCELLED", "Trial cancelled before conversion. No paid subscription will begin.", business_id=business.id)
    db.commit()
    return {"status": sub.status}


@app.post("/subscription/cancel")
def subscription_cancel(request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can cancel the subscription.")
    check_rate_limit(db, "subscription-cancel", f"business:{user.business_id}")
    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    if sub.status not in ("active", "past_due"):
        raise HTTPException(status_code=409, detail="There is no active paid subscription to cancel.")
    if sub.paystack_subscription_code:
        try:
            fetched = paystack_fetch_subscription(sub.paystack_subscription_code)
            email_token = fetched.get("email_token")
            if email_token:
                paystack_disable_subscription(sub.paystack_subscription_code, email_token)
        except Exception:
            add_audit(db, user, "SUBSCRIPTION_CANCEL_PAYSTACK_DISABLE_FAILED", "Subscription cancelled locally, but disabling the linked Paystack subscription failed. Verify manually to ensure no future charge occurs.", business_id=business.id)
    sub.cancel_at_period_end = True
    sub.cancelled_at = datetime.utcnow()
    add_audit(db, user, "SUBSCRIPTION_CANCELLED", f"Subscription cancelled. Access continues until the current billing period ends ({to_utc_iso(sub.current_period_end) if sub.current_period_end else 'unknown'}); no further charge will occur.", business_id=business.id)
    db.commit()
    return {"status": sub.status, "cancel_at_period_end": True, "access_until": to_utc_iso(sub.current_period_end)}


@app.get("/subscription/payments")
def subscription_payments(limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0), user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    """Non-sensitive payment history for the authenticated user's own business
    only — business_id is always taken from the session, never from the client."""
    rows = (db.query(PaymentRecord)
            .filter(PaymentRecord.business_id == user.business_id, PaymentRecord.purpose != "card_verification")
            .order_by(PaymentRecord.created_at.desc()).offset(offset).limit(limit).all())
    results = []
    for r in rows:
        entry = {"plan": r.plan, "plan_label": PLAN_CONFIG.get(r.plan, {}).get("label", r.plan), "billing_interval": r.billing_interval,
                 "amount_naira": r.amount_kobo / 100, "currency": r.currency, "status": r.status, "purpose": r.purpose,
                 "reference": r.paystack_reference, "paid_at": to_utc_iso(r.paid_at),
                 "created_at": to_utc_iso(r.created_at)}
        if r.purpose == "subscription_upgrade":
            # Enough detail for a future UI to render this distinctly from a
            # normal renewal (e.g. "Business -> Enterprise Upgrade — ₦175,000"
            # instead of implying a full-price Enterprise charge occurred).
            try:
                meta = json.loads(r.transaction_metadata or "{}")
            except Exception:
                meta = {}
            entry["from_plan"] = meta.get("from_plan")
            entry["from_plan_label"] = PLAN_CONFIG.get(meta.get("from_plan"), {}).get("label")
            entry["unused_credit_naira"] = (meta.get("unused_credit_kobo") or 0) / 100
            entry["label"] = f"{entry['plan_label']} Upgrade"
        results.append(entry)
    return results


class CheckoutRequest(BaseModel):
    plan: str
    billing_interval: str = "monthly"

@app.post("/subscription/checkout")
def start_checkout(data: CheckoutRequest, request: Request, user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can start a subscription payment.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "checkout", f"business:{user.business_id}")
    check_rate_limit(db, "checkout-ip", client_ip)
    plan = str(data.plan or "").strip().lower()
    if plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    interval = str(data.billing_interval or "monthly").strip().lower()
    if interval not in ("monthly", "annual"):
        interval = "monthly"
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured yet. Please contact support to subscribe.")

    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    amount_naira = PLAN_CONFIG[plan]["annual_price" if interval == "annual" else "monthly_price"]
    amount_kobo = int(amount_naira) * 100
    reference = f"cauldra_{business.business_code}_{secrets.token_hex(8)}"

    sub = get_or_create_subscription(db, business)
    checkout_email = str(business.email or user.email or "").strip().casefold()
    checkout_metadata = {
        "business_id": business.id, "plan": plan, "billing_interval": interval,
        "purpose": "subscription", "customer_email": checkout_email,
    }
    record = PaymentRecord(business_id=business.id, subscription_id=sub.id, plan=plan, billing_interval=interval,
                            amount_kobo=amount_kobo, currency="NGN", paystack_reference=reference, status="initialized",
                            transaction_metadata=json.dumps(checkout_metadata))
    db.add(record); db.commit()

    import requests
    try:
        resp = requests.post(
            "https://api.paystack.co/transaction/initialize",
            headers={"Authorization": f"Bearer {PAYSTACK_SECRET_KEY}", "Content-Type": "application/json"},
            json={"email": checkout_email, "amount": amount_kobo, "currency": "NGN", "reference": reference,
                  "callback_url": paystack_callback_url(request),
                  "metadata": checkout_metadata},
            timeout=15,
        )
        payload = resp.json()
        if not resp.ok or not payload.get("status"):
            raise RuntimeError(payload.get("message", "Paystack rejected the request."))
    except Exception:
        record.status = "failed"; db.commit()
        raise HTTPException(status_code=502, detail="We couldn't start the payment right now. Please try again.")

    return {"authorization_url": payload["data"]["authorization_url"], "reference": reference, "amount_kobo": amount_kobo}


class SubscriptionCheckoutConfirmRequest(BaseModel):
    reference: str


@app.post("/subscription/checkout/confirm")
def confirm_subscription_checkout(
    data: SubscriptionCheckoutConfirmRequest,
    user: User = Depends(get_authenticated_user),
    db: Session = Depends(get_db),
):
    """Reconcile a Paystack browser callback through the webhook's authority.

    The caller cannot supply price, plan, tenant, or success. We lock and load
    only this tenant's server-created PaymentRecord and independently fetch the
    transaction from Paystack before applying the shared activation function.
    """
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can confirm a subscription payment.")
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured.")

    reference = str(data.reference or "").strip()
    if not reference:
        raise HTTPException(status_code=400, detail="Payment reference is required.")

    record = (
        db.query(PaymentRecord)
        .filter(
            PaymentRecord.paystack_reference == reference,
            PaymentRecord.business_id == user.business_id,
            PaymentRecord.purpose == "subscription",
        )
        .with_for_update()
        .first()
    )
    if not record:
        # Tenant-scoped 404 avoids revealing whether another business owns the
        # supplied reference.
        raise HTTPException(status_code=404, detail="We couldn't find that subscription payment.")
    business = db.query(BusinessProfile).filter(
        BusinessProfile.id == user.business_id,
    ).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found.")

    if record.status == "success":
        return {"status": "success", "already_processed": True, "reference": reference}

    try:
        verified_transaction = paystack_verify_transaction(reference)
    except Exception:
        db.rollback()
        raise HTTPException(status_code=502, detail="Payment verification is temporarily unavailable. Please retry.")

    result = reconcile_first_subscription_payment(db, record, business, verified_transaction)
    db.commit()
    payload = {**result, "reference": reference}
    if result["status"] == "pending":
        return JSONResponse(status_code=202, content=payload)
    if result["status"] == "failed":
        return JSONResponse(status_code=409, content={**payload, "detail": "Paystack reports that this payment failed."})
    if result["status"] == "flagged_verification_mismatch":
        return JSONResponse(status_code=409, content={**payload, "detail": "Payment verification did not match this checkout and was not applied."})
    return payload


# -----------------------------------------------------------------------------
# SUBSCRIPTION UPGRADES WITH UNUSED-TIME PRORATION
#
# Two-step flow, mirroring how a card-present checkout works elsewhere in this
# file: (1) a short-lived, server-computed quote the frontend can only ever
# display — never influence — and (2) a checkout step that can only charge
# the exact amount frozen into that quote. The webhook (the one and only
# place a subscription is actually upgraded) re-verifies both the payment
# AND that the underlying subscription hasn't changed since the quote was
# issued, before touching anything.
# -----------------------------------------------------------------------------
class UpgradeQuoteRequest(BaseModel):
    plan: str
    billing_interval: str

@app.post("/subscription/upgrade-quote")
def create_upgrade_quote(data: UpgradeQuoteRequest, request: Request, user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can request a subscription upgrade.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "upgrade-quote", f"business:{user.business_id}")
    check_rate_limit(db, "upgrade-quote-ip", client_ip)

    to_plan = str(data.plan or "").strip().lower()
    if to_plan not in PLAN_CONFIG:
        raise HTTPException(status_code=400, detail="Please choose a valid subscription plan.")
    to_interval = str(data.billing_interval or "").strip().lower()
    if to_interval not in ("monthly", "annual"):
        raise HTTPException(status_code=400, detail="Please choose a valid billing interval.")

    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)

    # Trials never go through paid proration — they keep using the existing
    # free plan-switch mechanism (/subscription/change-plan), and the trial
    # clock is never touched here.
    if sub.status == "trialing":
        raise HTTPException(status_code=400, detail="Your business is still on its free trial. Switch plans directly — no payment or proration is needed until the trial ends.")
    if sub.status not in ("active", "past_due"):
        raise HTTPException(status_code=409, detail="Your subscription must be active to upgrade. Please subscribe first.")
    if not sub.current_period_start or not sub.current_period_end:
        raise HTTPException(status_code=409, detail="Your current billing period isn't available yet. Please try again shortly or contact support.")

    from_plan = (sub.plan or "starter").strip().lower()
    from_interval = (sub.billing_interval or "monthly").strip().lower()

    # Only a genuine plan-tier increase counts as an "upgrade" here — a same-
    # plan interval change (e.g. Business Monthly -> Business Annual) is
    # deliberately out of scope for proration and stays on the existing
    # lateral-move path in /subscription/change-plan.
    if PLAN_RANK.get(to_plan, -1) <= PLAN_RANK.get(from_plan, -1):
        raise HTTPException(status_code=400, detail="That isn't a plan upgrade. Choose a higher plan to upgrade.")

    now = datetime.utcnow()
    total_seconds = (sub.current_period_end - sub.current_period_start).total_seconds()
    if total_seconds <= 0:
        raise HTTPException(status_code=409, detail="Your billing period looks invalid. Please contact support before upgrading.")
    remaining_seconds = max((sub.current_period_end - now).total_seconds(), 0.0)
    unused_fraction = min(remaining_seconds / total_seconds, 1.0)

    current_price_kobo = plan_amount_naira(from_plan, from_interval) * 100
    new_price_kobo = plan_amount_naira(to_plan, to_interval) * 100
    unused_credit_kobo = round(current_price_kobo * unused_fraction)
    # Never let the credit exceed the new plan's price (would mean a "free"
    # upgrade); floor at a small positive amount since Paystack requires one.
    amount_due_kobo = max(new_price_kobo - unused_credit_kobo, UPGRADE_MINIMUM_CHARGE_KOBO)

    # Only one live quote per business at a time — a later quote request
    # supersedes any earlier one rather than leaving multiple valid quotes
    # for different amounts floating around.
    db.query(SubscriptionUpgradeQuote).filter(
        SubscriptionUpgradeQuote.business_id == business.id, SubscriptionUpgradeQuote.status == "issued"
    ).update({"status": "invalidated"})

    quote = SubscriptionUpgradeQuote(
        quote_reference=f"upgquote_{secrets.token_hex(12)}", business_id=business.id,
        from_plan=from_plan, from_interval=from_interval, to_plan=to_plan, to_interval=to_interval,
        current_price_kobo=current_price_kobo, new_price_kobo=new_price_kobo,
        unused_credit_kobo=unused_credit_kobo, amount_due_kobo=amount_due_kobo,
        current_period_end_snapshot=sub.current_period_end, status="issued",
        expires_at=now + timedelta(minutes=UPGRADE_QUOTE_VALIDITY_MINUTES),
    )
    db.add(quote)
    add_audit(db, user, "SUBSCRIPTION_UPGRADE_QUOTED",
              f"Upgrade quote: {PLAN_CONFIG[from_plan]['label']} ({from_interval}) -> {PLAN_CONFIG[to_plan]['label']} ({to_interval}). "
              f"Unused credit {unused_credit_kobo/100:.2f} NGN, amount due {amount_due_kobo/100:.2f} NGN.", business_id=business.id)
    db.commit()

    return {
        "quote_reference": quote.quote_reference,
        "current_plan": from_plan, "current_plan_label": PLAN_CONFIG[from_plan]["label"], "current_interval": from_interval,
        "new_plan": to_plan, "new_plan_label": PLAN_CONFIG[to_plan]["label"], "new_interval": to_interval,
        "current_price": current_price_kobo / 100, "new_price": new_price_kobo / 100,
        "unused_credit": unused_credit_kobo / 100, "amount_due": amount_due_kobo / 100,
        "currency": "NGN",
        "current_period_end": to_utc_iso(sub.current_period_end),
        "next_renewal_amount": new_price_kobo / 100,
        "expires_at": to_utc_iso(quote.expires_at),
    }


class UpgradeCheckoutRequest(BaseModel):
    quote_reference: str

@app.post("/subscription/upgrade-checkout")
def start_upgrade_checkout(data: UpgradeCheckoutRequest, request: Request, user: User = Depends(get_authenticated_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an Admin can start a subscription payment.")
    client_ip = request.client.host if request and request.client else "unknown"
    check_rate_limit(db, "upgrade-checkout", f"business:{user.business_id}")
    check_rate_limit(db, "upgrade-checkout-ip", client_ip)
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payment processing is not configured yet. Please contact support to upgrade.")

    quote_reference = str(data.quote_reference or "").strip()
    quote = db.query(SubscriptionUpgradeQuote).filter(
        SubscriptionUpgradeQuote.quote_reference == quote_reference, SubscriptionUpgradeQuote.business_id == user.business_id
    ).first()
    if not quote:
        raise HTTPException(status_code=404, detail="We couldn't find that upgrade quote. Please request a new one.")
    if quote.status == "paid":
        raise HTTPException(status_code=409, detail="This upgrade has already been completed.")
    if quote.status != "issued":
        raise HTTPException(status_code=410, detail="This upgrade quote is no longer valid. Please request a new one.")
    now = datetime.utcnow()
    if now > quote.expires_at:
        quote.status = "expired"; db.commit()
        raise HTTPException(status_code=410, detail="This upgrade quote has expired. Please request a new one.")

    business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
    sub = get_or_create_subscription(db, business)
    # Revalidate against the LIVE subscription — if anything about it moved
    # since the quote was issued (another upgrade, a cancellation, a renewal),
    # the quote's numbers are potentially stale and must not be trusted.
    if (sub.plan or "").strip().lower() != quote.from_plan or (sub.billing_interval or "").strip().lower() != quote.from_interval \
            or sub.status not in ("active", "past_due") or sub.current_period_end != quote.current_period_end_snapshot:
        quote.status = "invalidated"; db.commit()
        raise HTTPException(status_code=409, detail="Your subscription changed since this quote was generated. Please request a new upgrade quote.")

    reference = f"cauldra_upgrade_{business.business_code}_{secrets.token_hex(8)}"
    record = PaymentRecord(
        business_id=business.id, subscription_id=sub.id, plan=quote.to_plan, billing_interval=quote.to_interval,
        amount_kobo=quote.amount_due_kobo, currency="NGN", paystack_reference=reference, status="initialized",
        purpose="subscription_upgrade",
        transaction_metadata=json.dumps({
            "purpose": "subscription_upgrade", "from_plan": quote.from_plan, "from_interval": quote.from_interval,
            "to_plan": quote.to_plan, "to_interval": quote.to_interval,
            "unused_credit_kobo": quote.unused_credit_kobo, "quoted_amount_kobo": quote.amount_due_kobo,
            "quote_reference": quote.quote_reference,
        }),
    )
    db.add(record)
    quote.paystack_reference = reference
    add_audit(db, user, "SUBSCRIPTION_UPGRADE_PAYMENT_STARTED",
              f"Started upgrade payment {PLAN_CONFIG[quote.from_plan]['label']} ({quote.from_interval}) -> {PLAN_CONFIG[quote.to_plan]['label']} ({quote.to_interval}), amount {quote.amount_due_kobo/100:.2f} NGN, reference {reference}.",
              business_id=business.id)
    db.commit()

    import requests
    try:
        resp = requests.post(
            "https://api.paystack.co/transaction/initialize",
            headers={"Authorization": f"Bearer {PAYSTACK_SECRET_KEY}", "Content-Type": "application/json"},
            json={"email": business.email or user.email, "amount": quote.amount_due_kobo, "currency": "NGN", "reference": reference,
                  "callback_url": paystack_callback_url(request),
                  "metadata": {"business_id": business.id, "purpose": "subscription_upgrade", "quote_reference": quote.quote_reference}},
            timeout=15,
        )
        payload = resp.json()
        if not resp.ok or not payload.get("status"):
            raise RuntimeError(payload.get("message", "Paystack rejected the request."))
    except Exception:
        record.status = "failed"
        add_audit(db, user, "SUBSCRIPTION_UPGRADE_PAYMENT_FAILED", f"Failed to initialize upgrade payment (reference {reference}). Existing plan and billing period unchanged.", business_id=business.id)
        db.commit()
        raise HTTPException(status_code=502, detail="We couldn't start the upgrade payment right now. Please try again.")

    return {"authorization_url": payload["data"]["authorization_url"], "reference": reference, "amount_kobo": quote.amount_due_kobo}


def find_subscription_by_customer_code(db: Session, customer_code: str) -> Optional[BusinessSubscription]:
    if not customer_code:
        return None
    return db.query(BusinessSubscription).filter(BusinessSubscription.paystack_customer_code == customer_code).first()

@app.post("/webhooks/paystack")
async def paystack_webhook(request: Request, db: Session = Depends(get_db)):
    raw_body = await request.body()
    signature = request.headers.get("x-paystack-signature", "")
    if not PAYSTACK_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Payments are not configured.")
    # Never treat a request as authoritative merely because it reached this URL —
    # only a signature computed with our own secret key proves it came from Paystack.
    expected = hmac.new(PAYSTACK_SECRET_KEY.encode("utf-8"), raw_body, "sha512").hexdigest()
    if not signature or not hmac.compare_digest(expected, signature):
        print("[webhook] Paystack signature verification failed — request rejected.")
        raise HTTPException(status_code=401, detail="Invalid webhook signature.")

    try:
        event = json.loads(raw_body.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Malformed webhook payload.")
    event_type = event.get("event", "")
    data = event.get("data", {}) or {}
    reference = data.get("reference", "")
    refund_transaction = data.get("transaction") or {}
    refund_transaction_reference = str(data.get("transaction_reference", "") or (refund_transaction.get("reference", "") if isinstance(refund_transaction, dict) else "")).strip()
    refund_reference = str(data.get("refund_reference", "") or "").strip()
    subscription_code = data.get("subscription_code") or (data.get("subscription") or {}).get("subscription_code", "")
    # Prefer the most specific stable identifier available; fall back to a hash
    # of the full payload so we never accidentally collide two distinct events
    # that both happen to lack a reference/id (still fully idempotent).
    natural_id = reference or refund_reference or refund_transaction_reference or subscription_code or str(data.get("id", ""))
    event_key = f"{event_type}:{natural_id}" if natural_id else f"{event_type}:{hash_text(raw_body.decode('utf-8', 'ignore'))}"

    # Paystack retries webhook delivery. Insert the unique marker inside the
    # SAME transaction as every database effect below. If processing raises,
    # dependency cleanup rolls the transaction back, including this marker, so
    # Paystack's retry can actually apply the event instead of being poisoned
    # forever as "already_processed". Concurrent duplicates block on the
    # unique key; only the committed winner can make the loser return early.
    if db.query(PaystackWebhookEvent).filter(PaystackWebhookEvent.event_key == event_key).first():
        return {"status": "already_processed"}
    db.add(PaystackWebhookEvent(event_key=event_key, event_type=event_type))
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        if db.query(PaystackWebhookEvent).filter(PaystackWebhookEvent.event_key == event_key).first():
            return {"status": "already_processed"}
        raise

    now = datetime.utcnow()

    if event_type in {"refund.pending", "refund.processing", "refund.needs-attention", "refund.processed", "refund.failed"}:
        # Refund webhooks are signed provider reconciliation evidence. Locate
        # only the server-owned transaction reference and update exactly one of
        # the two verification-charge record types; never infer success from
        # receiving the webhook itself.
        refund_row = None
        if refund_transaction_reference:
            refund_row = db.query(PaymentRecord).filter(
                PaymentRecord.paystack_reference == refund_transaction_reference,
            ).first()
            if refund_row is not None and refund_row.purpose != "card_verification":
                refund_row = None
            if refund_row is None:
                refund_row = db.query(OnboardingAuthorization).filter(
                    OnboardingAuthorization.paystack_reference == refund_transaction_reference,
                ).first()
        if refund_row is None and refund_reference:
            refund_row = db.query(PaymentRecord).filter(
                PaymentRecord.refund_provider_id == refund_reference,
                PaymentRecord.purpose == "card_verification",
            ).first()
            if refund_row is None:
                refund_row = db.query(OnboardingAuthorization).filter(
                    OnboardingAuthorization.refund_provider_id == refund_reference,
                ).first()
        if refund_row:
            provider_refund = dict(data)
            provider_refund["status"] = event_type.split(".", 1)[1]
            if not provider_refund.get("id") and refund_reference:
                provider_refund["id"] = refund_reference
            prior_status = refund_row.refund_status
            local_status = _apply_paystack_refund_state(refund_row, provider_refund, now)
            if isinstance(refund_row, PaymentRecord) and prior_status != local_status:
                action = {
                    "succeeded": "TRIAL_VERIFICATION_REFUND_SUCCEEDED",
                    "failed": "TRIAL_VERIFICATION_REFUND_FAILED",
                    "pending": "TRIAL_VERIFICATION_REFUND_PENDING",
                }[local_status]
                add_audit(
                    db, None, action,
                    f"Paystack verification-charge refund is {local_status} (provider status: {refund_row.refund_provider_status}).",
                    business_id=refund_row.business_id,
                )
            db.flush()

    elif event_type == "charge.success" and reference:
        record = db.query(PaymentRecord).filter(PaymentRecord.paystack_reference == reference).first()
        if record and record.status != "success" and record.purpose == "subscription_upgrade":
            # Upgrade payment, initiated by /subscription/upgrade-checkout. The
            # amount charged must match the frozen quote exactly — never trust
            # data.amount alone, always cross-check against our own quote record.
            meta = {}
            try:
                meta = json.loads(record.transaction_metadata or "{}")
            except Exception:
                meta = {}
            quote_reference = meta.get("quote_reference", "")
            quote = db.query(SubscriptionUpgradeQuote).filter(SubscriptionUpgradeQuote.quote_reference == quote_reference).first()
            business = db.query(BusinessProfile).filter(BusinessProfile.id == record.business_id).first()
            paid_amount_kobo = int(data.get("amount") or 0)
            paid_currency = (data.get("currency") or "").upper()

            if not quote or quote.status == "paid":
                # Idempotency: a duplicate webhook delivery for an already-applied
                # upgrade (or a quote we can no longer find) must do nothing further.
                record.status = "success"; record.paid_at = now; db.flush()
            elif paid_amount_kobo != quote.amount_due_kobo or paid_currency != "NGN" or data.get("status") != "success":
                record.status = "flagged_amount_mismatch"; db.flush()
                add_audit(db, None, "SUBSCRIPTION_UPGRADE_PAYMENT_FAILED",
                          f"Upgrade charge for reference {reference} did not match the quoted amount (expected {quote.amount_due_kobo} kobo, got {paid_amount_kobo} {paid_currency}). Not activated — flagged for review. Existing plan and billing period unchanged.",
                          business_id=business.id if business else None)
                db.flush()
            else:
                record.status = "success"; record.paid_at = now
                sub = get_or_create_subscription(db, business, commit=False)
                old_plan, old_interval = sub.plan, sub.billing_interval
                old_subscription_code = sub.paystack_subscription_code
                # THE key behavior this feature exists for: the paid-for portion
                # of the current billing period is honored exactly as-is. The new
                # plan activates immediately, but current_period_start/end are
                # left untouched — no reset, no accidental extra period.
                sub.plan = quote.to_plan; sub.billing_interval = quote.to_interval
                sub.status = "active"; sub.payment_status = "paid"; sub.paid_at = now
                sub.next_billing_at = sub.current_period_end
                sub.latest_transaction_reference = reference
                sub.cancel_at_period_end = False; sub.cancelled_at = None; sub.grace_period_ends_at = None
                resolve_notifications(db, f"sub_expired:{business.id}", business.id)
                resolve_notifications(db, f"payment_failed:{business.id}", business.id)
                # A genuine upgrade always supersedes any downgrade scheduled
                # earlier in the same period (see section 10 of the downgrade
                # design) — never let a stale lower-plan schedule re-apply later.
                clear_pending_downgrade(sub)
                business.subscription_plan = quote.to_plan; business.billing_interval = quote.to_interval
                quote.status = "paid"; quote.consumed_at = now
                db.flush()

                # Safely hand off the recurring Paystack subscription so the
                # customer is never charged by both the old and new plan: best-
                # effort disable the old one, then schedule the new recurring
                # subscription to start at the SAME preserved renewal boundary
                # (not "now") so the next charge still lands on the original date.
                try:
                    if old_subscription_code:
                        try:
                            fetched = paystack_fetch_subscription(old_subscription_code)
                            email_token = fetched.get("email_token")
                            if email_token:
                                paystack_disable_subscription(old_subscription_code, email_token)
                        except Exception:
                            add_audit(db, None, "SUBSCRIPTION_UPGRADE_OLD_PAYSTACK_DISABLE_FAILED", f"Upgrade applied, but disabling the prior {PLAN_CONFIG[old_plan]['label']} ({old_interval}) recurring Paystack subscription failed. Verify manually to ensure no duplicate charge occurs.", business_id=business.id)
                    new_plan_code = PLAN_CONFIG[quote.to_plan].get("paystack_annual_plan_code" if quote.to_interval == "annual" else "paystack_monthly_plan_code")
                    if new_plan_code and sub.paystack_authorization_code and sub.paystack_customer_code:
                        created = paystack_create_subscription(sub.paystack_customer_code, new_plan_code, sub.paystack_authorization_code, start_date=sub.current_period_end)
                        sub.paystack_subscription_code = created.get("subscription_code")
                        sub.paystack_plan_code = new_plan_code
                        db.flush()
                except Exception:
                    add_audit(db, None, "SUBSCRIPTION_UPGRADE_PAYSTACK_RESCHEDULE_FAILED", "Upgrade applied and billing period preserved, but scheduling the new recurring Paystack subscription failed. Automatic conversion at next renewal may not occur; verify manually.", business_id=business.id)
                    db.flush()

                add_audit(db, None, "SUBSCRIPTION_UPGRADED",
                          f"Upgraded {PLAN_CONFIG[old_plan]['label']} ({old_interval}) -> {PLAN_CONFIG[quote.to_plan]['label']} ({quote.to_interval}). "
                          f"Unused credit {quote.unused_credit_kobo/100:.2f} NGN applied; paid {quote.amount_due_kobo/100:.2f} NGN; reference {reference}. "
                          f"Billing period preserved: {to_utc_iso(sub.current_period_start) if sub.current_period_start else '?'} - {to_utc_iso(sub.current_period_end) if sub.current_period_end else '?'}.",
                          business_id=business.id)
                db.flush()
        elif record and record.status != "success" and record.purpose != "card_verification":
            # First-time subscribe, initiated by our own /subscription/checkout.
            # Lock the same PaymentRecord used by the authenticated callback so
            # webhook/callback races share one authority and one activation.
            record = (
                db.query(PaymentRecord)
                .filter(PaymentRecord.id == record.id)
                .with_for_update()
                .one()
            )
            if record.status != "success":
                business = db.query(BusinessProfile).filter(BusinessProfile.id == record.business_id).first()
                try:
                    verified_transaction = paystack_verify_transaction(reference)
                except Exception:
                    # This is retryable provider unavailability, not a failed charge.
                    # Explicit rollback removes the F-03 event marker as well as any
                    # pending effects so Paystack can safely retry the webhook.
                    db.rollback()
                    raise HTTPException(status_code=502, detail="Payment verification is temporarily unavailable.")
                result = reconcile_first_subscription_payment(
                    db, record, business, verified_transaction, now,
                )
                if result["status"] == "pending":
                    # Do not consume the webhook marker while Paystack's
                    # authoritative read is still pending; its retry remains useful.
                    db.rollback()
                    raise HTTPException(status_code=502, detail="Payment verification is still pending.")
        elif not record and reference:
            # No PaymentRecord we created means this charge was triggered by
            # Paystack's own recurring-subscription schedule (the automatic

            # post-trial conversion charge, or a later renewal) — not something
            # our checkout endpoint initiated. Resolve the business via the
            # Paystack customer_code we stored at trial start, then verify the
            # amount against OUR server-side authoritative price before
            # activating anything.
            customer_code = (data.get("customer") or {}).get("customer_code", "")
            sub = find_subscription_by_customer_code(db, customer_code)
            if sub:
                business = db.query(BusinessProfile).filter(BusinessProfile.id == sub.business_id).first()
                old_plan = (sub.plan or "starter").strip().lower()
                old_interval = (sub.billing_interval or "monthly").strip().lower()
                # Section 9 — automatic effective-date processing: a pending
                # downgrade was pre-arranged (see /subscription/downgrade) to
                # start billing at exactly this boundary, on a SEPARATE Paystack
                # subscription for the lower plan. If one is still on record here,
                # THIS is the verified payment signal that the boundary was
                # reached and nothing superseded it since (any legitimate change
                # in the meantime already clears pending_downgrade_plan) — so the
                # lower plan/interval is what Paystack actually billed for.
                applying_downgrade = bool(sub.pending_downgrade_plan)
                plan = sub.pending_downgrade_plan if applying_downgrade else old_plan
                interval = (sub.pending_downgrade_billing_interval if applying_downgrade else old_interval) or old_interval
                expected_amount_kobo = plan_amount_naira(plan, interval) * 100
                paid_amount_kobo = int(data.get("amount") or 0)
                paid_currency = (data.get("currency") or "").upper()
                new_record = PaymentRecord(business_id=business.id, subscription_id=sub.id, plan=plan, billing_interval=interval,
                                            amount_kobo=paid_amount_kobo, currency=paid_currency or "NGN", paystack_reference=reference,
                                            purpose="subscription_renewal",
                                            transaction_metadata=json.dumps({"source": "paystack_recurring_webhook", "event": event_type}))
                if paid_amount_kobo != expected_amount_kobo or paid_currency != "NGN" or data.get("status") != "success":
                    # Section 17: never silently activate on an unexpected amount —
                    # flag it for reconciliation instead. A pending downgrade is
                    # left untouched so it can still be applied once a matching
                    # charge is verified (or investigated manually).
                    new_record.status = "flagged_amount_mismatch"
                    db.add(new_record)
                    add_audit(db, None, "SUBSCRIPTION_PAYMENT_AMOUNT_MISMATCH", f"Recurring charge for reference {reference} did not match the expected amount for {PLAN_CONFIG[plan]['label']} ({interval}) (expected {expected_amount_kobo} kobo, got {paid_amount_kobo} {paid_currency}). Not activated — flagged for review.", business_id=business.id)
                    db.flush()
                else:
                    new_record.status = "success"; new_record.paid_at = now
                    db.add(new_record)
                    sub.status = "active"; sub.payment_status = "paid"; sub.paid_at = now
                    sub.current_period_start = now; sub.current_period_end = add_billing_interval(now, interval)
                    sub.next_billing_at = sub.current_period_end
                    sub.latest_transaction_reference = reference
                    sub.grace_period_ends_at = None
                    # Any charge landing successfully here (renewal or a
                    # downgrade taking effect) proves the subscription is no
                    # longer expired/past-due — free both dedup keys so a
                    # FUTURE recurrence of either condition notifies again
                    # instead of being silently swallowed by a stale entry.
                    resolve_notifications(db, f"sub_expired:{business.id}", business.id)
                    resolve_notifications(db, f"payment_failed:{business.id}", business.id)
                    if applying_downgrade:
                        sub.plan = plan; sub.billing_interval = interval
                        business.subscription_plan = plan; business.billing_interval = interval
                        if sub.pending_downgrade_paystack_subscription_code:
                            sub.paystack_subscription_code = sub.pending_downgrade_paystack_subscription_code
                            sub.paystack_plan_code = PLAN_CONFIG[plan].get("paystack_annual_plan_code" if interval == "annual" else "paystack_monthly_plan_code")
                        clear_pending_downgrade(sub)
                        add_audit(db, None, "SUBSCRIPTION_DOWNGRADE_APPLIED", f"Scheduled downgrade took effect: {PLAN_CONFIG[old_plan]['label']} ({old_interval}) -> {PLAN_CONFIG[plan]['label']} ({interval}). Existing business data was preserved; only new-resource creation is subject to the new plan's limits.", business_id=business.id)
                    else:
                        add_audit(db, None, "SUBSCRIPTION_RENEWED", f"Automatic recurring charge verified and applied for {PLAN_CONFIG[plan]['label']} ({interval}).", business_id=business.id)
                        create_notification(
                            db, business_id=business.id, category="subscription", severity="info", type="SUBSCRIPTION_RENEWED",
                            title="Subscription renewed", message="Your Cauldra subscription was renewed successfully.",
                            deep_link="subscription",
                        )
                    db.flush()

    elif event_type in ("subscription.not_renew", "subscription.disable"):
        sub = find_subscription_by_customer_code(db, (data.get("customer") or {}).get("customer_code", ""))
        if not sub and subscription_code:
            sub = db.query(BusinessSubscription).filter(BusinessSubscription.paystack_subscription_code == subscription_code).first()
        if sub and sub.status not in ("cancelled",):
            sub.status = "cancelled"
            sub.cancelled_at = sub.cancelled_at or now
            add_audit(db, None, "SUBSCRIPTION_CANCELLED", f"Paystack reported the recurring subscription ended ({event_type}).", business_id=sub.business_id)
            db.flush()

    elif event_type == "invoice.payment_failed":
        sub = find_subscription_by_customer_code(db, (data.get("customer") or {}).get("customer_code", ""))
        if sub:
            sub.status = "past_due"
            sub.payment_status = "failed"
            sub.grace_period_ends_at = sub.grace_period_ends_at or (now + timedelta(days=PAYSTACK_GRACE_PERIOD_DAYS))
            add_audit(db, None, "SUBSCRIPTION_PAYMENT_FAILED", f"Automatic recurring charge failed. Grace period until {to_utc_iso(sub.grace_period_ends_at)}.", business_id=sub.business_id)
            # dedup_key is business-scoped, not time-scoped, and stays
            # unresolved for the whole failure episode — a retried/duplicate
            # failure webhook for the SAME episode is deliberately a no-op
            # here (section 2's "do not repeatedly spam the same
            # notification"); the key only opens again once a renewal
            # resolves it (see the charge.success renewal branch above).
            create_notification(
                db, business_id=sub.business_id, category="subscription", severity="critical", type="SUBSCRIPTION_PAYMENT_FAILED",
                title="Subscription payment failed",
                message="We couldn't process your Cauldra subscription payment. Update your billing information to prevent service interruption.",
                deep_link="subscription", dedup_key=f"payment_failed:{sub.business_id}",
            )
            db.flush()

    # The event marker becomes durable only with all database effects.
    db.commit()
    return {"status": "received"}

def run_billable_ai(db: Session, user: User, operation: str, provider: str, model: str, callback):
    """Record all external-AI outcomes; only successful work consumes credits."""
    try:
        result = callback()
    except Exception:
        record_ai_usage(db, user, operation, False, provider, model)
        db.commit()
        raise
    credits = record_ai_usage(db, user, operation, True, provider, model)
    db.commit()
    try:
        business = db.query(BusinessProfile).filter(BusinessProfile.id == user.business_id).first()
        if business:
            check_ai_credit_notifications(db, business)
        db.commit()
    except Exception:
        db.rollback()  # never let a notification failure affect an already-successful AI call
    return result, credits

@app.post("/ai/suggest-margin")
def suggest_margin(req: MarginRequest, user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    text, credits = run_billable_ai(db, user, "margin_advisor", "gemini", GEMINI_MODEL, lambda: gemini_text_response("You are Cauldra's category-aware pricing advisor. Give useful, practical business pricing guidance without inventing facts.", f"Product: {req.name}\nCategory: {req.category}\nCost price: {req.cost_price}\nGive recommended wholesale and retail prices and explain the reasoning briefly."))
    nums=[float(x) for x in re.findall(r"(?<![A-Za-z])(?:\d+(?:\.\d+)?)", text)]
    return {"suggested_wholesale": round(nums[0],2) if nums else round(req.cost_price*1.15,2), "suggested_retail": round(nums[1],2) if len(nums)>1 else round(req.cost_price*1.30,2), "advice": text, "credits_consumed": credits}

@app.post("/ai/scan-invoice")
def scan_invoice(req: InvoiceScanRequest, user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    raw_bytes, content_type = decode_base64_upload(req.image_data, {"image/jpeg", "image/png", "image/webp"})
    upload = persist_upload(db, user, "invoice", req.file_name or "invoice-scan", content_type, raw_bytes)
    add_audit(db, user, "INVOICE_UPLOADED", f"Uploaded invoice image {upload.original_name} for review.")
    db.commit()
    data, credits = run_billable_ai(db, user, "invoice_ocr", "openai", OPENAI_MODEL, lambda: openai_json_response(
        "Extract this invoice or receipt into the exact schema. Do not invent values. If uncertain, use empty strings or zero. This endpoint is for review; do not make database mutations yourself.", req.image_data))
    return {"upload_id": upload.id, "supplier_name": data.get("supplier_name") or "", "invoice_number": data.get("invoice_number") or "", "invoice_date": data.get("invoice_date") or "", "items_count": len(data.get("items") or []), "items": data.get("items") or [], "subtotal": data.get("subtotal") or 0, "total": data.get("total") or 0, "requires_confirmation": True, "credits_consumed": credits}

@app.get("/uploads")
def list_uploads(kind: Optional[str] = Query(None), limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(StoredUpload).filter(StoredUpload.business_id == user.business_id)
    if kind:
        q = q.filter(StoredUpload.kind == kind.strip().lower())
    rows = q.order_by(StoredUpload.created_at.desc()).offset(offset).limit(limit).all()
    return [{"id": r.id, "kind": r.kind, "file_name": r.original_name, "content_type": r.content_type, "size_bytes": r.size_bytes, "uploaded_by_id": r.uploaded_by_id, "created_at": to_utc_iso(r.created_at), "download_url": f"/uploads/{r.id}/download"} for r in rows]

@app.get("/uploads/{upload_id}/download")
def download_upload(upload_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.query(StoredUpload).filter(StoredUpload.id == upload_id, StoredUpload.business_id == user.business_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Uploaded file is unavailable.")
    path = (UPLOAD_STORAGE_DIR / row.storage_key).resolve()
    if UPLOAD_STORAGE_DIR not in path.parents or not path.is_file():
        raise HTTPException(status_code=404, detail="Uploaded file is unavailable.")
    add_audit(db, user, "UPLOADED_FILE_DOWNLOADED", f"Downloaded retained {row.kind} file {row.original_name}.")
    db.commit()
    return FileResponse(str(path), media_type=row.content_type, filename=row.original_name, content_disposition_type="attachment")

@app.get("/ai/insights")
def ai_insights(user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    products=db.query(Product).filter(Product.business_id==user.business_id).all()
    snapshot=[{"name":p.name,"category":p.category,"qty":p.quantity,"min":p.min_stock_level,"retail":p.retail_price,"cost":p.cost_price} for p in products]
    insight, credits = run_billable_ai(db, user, "inventory_insight", "gemini", GEMINI_MODEL, lambda: gemini_text_response("You are Cauldra's inventory intelligence engine. Analyze only the provided inventory snapshot and give concise, actionable insights.", json.dumps(snapshot)))
    return {"insight": insight, "credits_consumed": credits}

@app.post("/ai/chat")
def ai_chat(req: ChatRequest, user: User = Depends(require_ai_access), db: Session = Depends(get_db)):
    products=db.query(Product).filter(Product.business_id==user.business_id).all()
    snapshot=[{"name":p.name,"sku":p.sku,"category":p.category,"qty":p.quantity,"min":p.min_stock_level,"retail":p.retail_price,"cost":p.cost_price} for p in products]
    reply, credits = run_billable_ai(db, user, "chat", "gemini", GEMINI_MODEL, lambda: gemini_text_response("You are Cauldra's business and inventory assistant. Use only the supplied business data. Be practical and concise; state when data is insufficient.", f"User question: {req.message}\nInventory snapshot: {json.dumps(snapshot)}"))
    return {"reply":reply, "credits_consumed": credits}

# -----------------------------------------------------------------------------
# PRODUCTION ROOT / HEALTH
# -----------------------------------------------------------------------------
@app.get("/health")
def health():
    """Runtime monitoring, not a replacement for verify_database_connectivity()
    above (which still gates startup) — this is what a load balancer/uptime
    monitor calls repeatedly, long after startup, to catch a database that
    was reachable when the process started but has since gone away. A
    single SELECT 1 against a short-lived, always-released connection: no
    transaction is held open, no schema is touched, no table is scanned.
    Never includes DATABASE_URL, a credential, or any other provider detail
    in the response — only the same fixed classification string startup
    failures use goes to the server log, for whoever is watching it."""
    try:
        _ping_database()
        return {"status": "ok", "database": "ok", "version": app.version, "refresh_enabled": True}
    except Exception as exc:
        print(f"[health] database check failed: {_classify_database_connection_error(exc)}")
        return JSONResponse(
            status_code=503,
            content={"status": "degraded", "database": "unavailable", "version": app.version, "refresh_enabled": True},
        )

@app.get("/health/database")
def health_database():
    """Manual diagnostic view — safe to hit from a browser or Railway's logs
    to confirm exactly which PostgreSQL database this running deployment is
    connected to (e.g. after a deploy, or when investigating "why does this
    look like the wrong environment"). Complements the lightweight /health
    above (which stays a fast, minimal check for load balancers/uptime
    monitors) rather than replacing it.

    Every value here comes from SQLAlchemy's already-parsed engine.url
    (host/port/database name) or a live read-only query (schema, server
    version) — never the raw DATABASE_URL string, and never a username or
    password. A Supabase hostname naturally contains the project reference
    (e.g. db.<ref>.supabase.co) — that's expected and not a secret on its
    own; nothing else about the connection is ever included."""
    url = engine.url
    try:
        with engine.connect() as conn:
            version = conn.execute(sql_text("SHOW server_version")).scalar()
            schema = conn.execute(sql_text("SELECT current_schema()")).scalar()
        return {
            "status": "ok", "database": "postgresql", "connected": True,
            "host": url.host, "port": url.port, "database_name": url.database,
            "schema": schema, "postgres_version": version,
        }
    except Exception as exc:
        print(f"[health/database] database check failed: {_classify_database_connection_error(exc)}")
        return JSONResponse(
            status_code=503,
            content={"status": "degraded", "database": "postgresql", "connected": False},
        )

@app.get("/")
def serve_index():
    return FileResponse(str(INDEX_PATH))

# Served from the root path (not /assets/) so its default scope covers the
# whole origin, not just the assets directory — required for it to be able to
# intercept and cache the app shell / navigation requests.
@app.get("/sw.js")
def serve_service_worker():
    return FileResponse(str(FRONTEND_DIR / "sw.js"), media_type="application/javascript", headers={"Service-Worker-Allowed": "/", "Cache-Control": "no-cache"})

# -----------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    # Production launcher should run without reload. Use an external process manager in production.
    uvicorn.run("main:app", host=os.getenv("HOST", "127.0.0.1"), port=int(os.getenv("PORT", "8000")), reload=False)
